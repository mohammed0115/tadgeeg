"""Pre-flight quota check for bulk uploads.

Before a BulkUploadJob is queued, we count the items the file contains
and compare against the org's remaining quota. Three outcomes:

  total ≤ remaining            → quota_status=allowed,        proceed
  total >  remaining > 0       → quota_status=quota_exceeded, reject
                                  (unless accept_partial=True, then
                                  quota_status=partially_allowed, cap
                                  to remaining and proceed)
  remaining = 0                → quota_status=quota_exceeded, reject

The per-item billing (reserve → consume / release for each row that
actually triggers an audit) is already enforced by the quota_gate
wrapped around run_audit_compat in Stage 5. This module's job is the
*upfront* check that stops a job before it consumes Celery throughput
on items that would all fail with QUOTA_NOT_ENOUGH at the gate.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import zipfile
from dataclasses import dataclass
from typing import Optional

from django.utils.translation import gettext as _


logger = logging.getLogger("billing.bulk_quota")


@dataclass
class BulkQuotaDecision:
    total_items: int
    quota_available: int
    quota_required: int
    allowed_items: int
    blocked_items: int
    quota_status: str    # "allowed" | "partially_allowed" | "quota_exceeded"
    accepted: bool       # True → caller may proceed with processing
    code: str            # machine-readable: "OK" | "QUOTA_NOT_ENOUGH" | "NO_SUBSCRIPTION"
    message: str         # localised, user-facing


# ─── format-specific counters ────────────────────────────────────────────────
def count_items(file_bytes: bytes, source_format: str, filename: str = "") -> int:
    """Return the number of logical items inside the file.

    For text/tabular formats we stream-count. For ZIP we count member files
    (not directories). Unknown formats fall back to 1 — the safe default
    (one upload = one document).
    """
    fmt = (source_format or "").lower()
    try:
        if fmt == "csv":
            return _count_csv(file_bytes)
        if fmt in ("xlsx", "xls"):
            return _count_xlsx(file_bytes)
        if fmt == "jsonl":
            return _count_jsonl(file_bytes)
        if fmt == "json":
            return _count_json(file_bytes)
        if fmt in ("zip", "mixed"):
            return _count_zip(file_bytes)
    except Exception:  # noqa: BLE001 — counting is best-effort
        logger.exception("count_items: failed to count %s (%s)", filename, fmt)
        return 1
    return 1


def _count_csv(data: bytes) -> int:
    """Header + N rows → N items. Empty rows are skipped."""
    text = data.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return 0
    # Drop the header. We don't try to guess if it's missing — a CSV
    # without a header is rare in this domain.
    data_rows = [r for r in rows[1:] if any((c or "").strip() for c in r)]
    return len(data_rows)


def _count_xlsx(data: bytes) -> int:
    """First worksheet rows minus header, blank rows skipped."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed — XLSX counter falls back to 1")
        return 1
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return 0
    count = 0
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if any(cell is not None and str(cell).strip() for cell in row):
            count += 1
    return count


def _count_jsonl(data: bytes) -> int:
    return sum(
        1 for line in data.decode("utf-8", errors="replace").splitlines()
        if line.strip()
    )


def _count_json(data: bytes) -> int:
    try:
        payload = json.loads(data.decode("utf-8", errors="replace") or "[]")
    except json.JSONDecodeError:
        return 0
    if isinstance(payload, list):
        return len(payload)
    if isinstance(payload, dict) and isinstance(payload.get("items"), list):
        return len(payload["items"])
    return 1  # single-object payload


def _count_zip(data: bytes) -> int:
    """Count member files (not directories). Skips macOS metadata."""
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        return sum(
            1 for info in zf.infolist()
            if not info.is_dir()
            and not info.filename.startswith("__MACOSX/")
            and not info.filename.endswith("/.DS_Store")
        )


# ─── decision builder ────────────────────────────────────────────────────────
def evaluate_bulk_quota(
    *, organization, total_items: int, accept_partial: bool = False,
) -> BulkQuotaDecision:
    """Build the decision the view + model need.

    organization     — Organization row, used to look up remaining quota.
    total_items      — count from ``count_items``.
    accept_partial   — when True and remaining > 0, allow the job to run
                       but cap it at remaining.
    """
    from apps.billing.services.quota_service import QuotaService
    svc = QuotaService()
    sub = svc.get_active_subscription(organization)
    if sub is None:
        return BulkQuotaDecision(
            total_items=total_items,
            quota_available=0,
            quota_required=total_items,
            allowed_items=0,
            blocked_items=total_items,
            quota_status="quota_exceeded",
            accepted=False,
            code="NO_SUBSCRIPTION",
            message=str(_("Please choose a subscription plan before uploading invoices for audit.")),
        )

    remaining = sub.remaining_invoices

    if remaining <= 0:
        return BulkQuotaDecision(
            total_items=total_items,
            quota_available=0,
            quota_required=total_items,
            allowed_items=0,
            blocked_items=total_items,
            quota_status="quota_exceeded",
            accepted=False,
            code="QUOTA_NOT_ENOUGH",
            message=str(_(
                "You have used all available invoices in your current plan. "
                "Please upgrade to continue."
            )),
        )

    if total_items <= remaining:
        return BulkQuotaDecision(
            total_items=total_items,
            quota_available=remaining,
            quota_required=total_items,
            allowed_items=total_items,
            blocked_items=0,
            quota_status="allowed",
            accepted=True,
            code="OK",
            message="",
        )

    # total > remaining > 0  → either reject with require_upgrade (default)
    # or partial-process if the caller opted in.
    if accept_partial:
        return BulkQuotaDecision(
            total_items=total_items,
            quota_available=remaining,
            quota_required=total_items,
            allowed_items=remaining,
            blocked_items=total_items - remaining,
            quota_status="partially_allowed",
            accepted=True,
            code="PARTIALLY_ALLOWED",
            message=str(_(
                "Processing the first %(allowed)d of %(total)d invoices — "
                "the remaining %(blocked)d need an upgraded plan."
            )) % {
                "allowed": remaining,
                "total":   total_items,
                "blocked": total_items - remaining,
            },
        )

    return BulkQuotaDecision(
        total_items=total_items,
        quota_available=remaining,
        quota_required=total_items,
        allowed_items=0,
        blocked_items=total_items,
        quota_status="quota_exceeded",
        accepted=False,
        code="QUOTA_NOT_ENOUGH",
        message=str(_(
            "You have %(remaining)d invoices remaining, and this file "
            "contains %(total)d. Upload only %(remaining)d (set accept_partial=true) "
            "or upgrade your plan."
        )) % {"remaining": remaining, "total": total_items},
    )
