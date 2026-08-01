"""Excel and PDF exports for the Trial Users Dashboard (§B.1).

Both exports take the **same** filtered queryset the dashboard is showing. That
is not politeness — an export that ignores the active filters and dumps every
row is a data-exposure bug, so filtering is applied by the caller once and both
renderers receive the result rather than re-deriving it.

Neither export includes ``registered_ip`` (ADR 0004 §2).

Dependencies are the ones already in the project: ``openpyxl`` (used by
apps/reports and apps/frontend) and WeasyPrint (used by apps/reports and
apps/audit). Nothing new is introduced.
"""

from __future__ import annotations

import io

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from .trial_selectors import row_values


#: (header, key) in display order. One definition, both renderers.
COLUMNS = (
    ("Name", "full_name"),
    ("Email", "email"),
    ("Phone", "phone"),
    ("Organization", "organization"),
    ("Country", "country"),
    ("City", "city"),
    ("Company", "company_name"),
    ("Client type", "primary_benefit"),
    ("Employees", "employee_count"),
    ("Sector", "sector"),
    ("Heard about", "heard_about"),
    ("Trial status", "trial_status"),
    ("Activity", "activity"),
    ("Invoices used", "invoices_used"),
    ("Registered at", "registered_at"),
    ("Last login", "last_login"),
)

#: Hard ceiling on exported rows. A dashboard export is an operational tool,
#: not a bulk data dump; without a cap one click could stream the entire lead
#: database into a spreadsheet. When it trips, the response says so (see the
#: `truncated` flag) instead of silently returning a partial file.
MAX_EXPORT_ROWS = 5000


def _rows(queryset):
    """Materialise at most MAX_EXPORT_ROWS + 1 rows, flattened.

    ``iterator()`` keeps the DB cursor streaming rather than loading the whole
    result set; the slice bounds memory regardless of table size.
    """
    rows = []
    truncated = False
    for index, profile in enumerate(queryset.iterator(chunk_size=500)):
        if index >= MAX_EXPORT_ROWS:
            truncated = True
            break
        rows.append(row_values(profile))
    return rows, truncated


def _filename(extension: str) -> str:
    stamp = timezone.now().strftime("%Y%m%d-%H%M")
    return f"trial-users-{stamp}.{extension}"


def export_xlsx(queryset) -> HttpResponse:
    """Excel workbook of the filtered trial registrants."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    rows, truncated = _rows(queryset)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Trial Users"

    header_font = Font(bold=True, color="FFFFFF")
    # Navy #003366 — the identity colour from static/src/css/tokens.css.
    header_fill = PatternFill("solid", start_color="003366", end_color="003366")

    for column_index, (label, _key) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(column_index)
        ].width = 20

    for row_index, row in enumerate(rows, start=2):
        for column_index, (_label, key) in enumerate(COLUMNS, start=1):
            sheet.cell(row=row_index, column=column_index, value=row.get(key, ""))

    sheet.freeze_panes = "A2"

    if truncated:
        sheet.cell(
            row=len(rows) + 3,
            column=1,
            value=(
                f"Truncated at {MAX_EXPORT_ROWS} rows. "
                "Narrow the filters to export the remainder."
            ),
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{_filename("xlsx")}"'
    response["X-Export-Truncated"] = "1" if truncated else "0"
    return response


def export_pdf(queryset, *, summary=None):
    """PDF of the filtered trial registrants.

    Returns ``None`` when the renderer is unavailable, so the caller can answer
    503 rather than 500. WeasyPrint imports *and renders* on this deployment
    (verified), but it depends on system libraries that a slimmer image may not
    carry, and a missing library must not read as a server fault.
    """
    try:
        from weasyprint import HTML
    except Exception:                                   # noqa: BLE001
        return None

    rows, truncated = _rows(queryset)
    # Flatten to ordered cells here rather than indexing a dict in the
    # template: Django templates cannot subscript by a variable key, and
    # adding a filter just for a print sheet is not worth the surface.
    html = render_to_string(
        "platform_admin/trial_users_export.html",
        {
            "headers": [label for label, _key in COLUMNS],
            "table_rows": [
                [row.get(key, "") for _label, key in COLUMNS] for row in rows
            ],
            "row_count": len(rows),
            "summary": summary or {},
            "truncated": truncated,
            "max_rows": MAX_EXPORT_ROWS,
            "generated_at": timezone.now(),
        },
    )

    try:
        pdf_bytes = HTML(string=html).write_pdf()
    except Exception:                                   # noqa: BLE001
        return None

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_filename("pdf")}"'
    response["X-Export-Truncated"] = "1" if truncated else "0"
    return response
