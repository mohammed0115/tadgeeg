"""Substantive Testing service (TADGEEG-FIN-AUDIT-9D).

Create test items, derive the independent "tested" value via deterministic
recompute helpers (straight-line depreciation for fixed assets; net pay for
payroll; counted quantity × unit cost for inventory), record the tested value,
and reconcile against the books (matched / variance by tolerance).

Deterministic (no AI); never writes to ``apps.ledger``; a variance is flagged,
never auto-corrected.
"""
from __future__ import annotations

import csv
import io
from decimal import Decimal, InvalidOperation

from django.db import IntegrityError, transaction
from django.db.models import Count, Q, Sum

from apps.audit.substantive_test_models import SubstantiveTestItem

_I = SubstantiveTestItem
_Area = _I.Area
_St = _I.Status
_ZERO = Decimal("0")

_MAX_IMPORT_ROWS = 5000


class SubstantiveTestError(Exception):
    """Invalid input or scoping violation."""


def _dec(value, default=_ZERO):
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return default


# ─────────────────────────────────────────────────────────────────────────────
# Deterministic recompute helpers (ISA 500/540 auditor re-performance)
# ─────────────────────────────────────────────────────────────────────────────
def straight_line_nbv(*, cost, salvage, useful_life_years, elapsed_years) -> dict:
    """Recompute accumulated depreciation + net book value (straight-line)."""
    cost = _dec(cost); salvage = _dec(salvage)
    life = _dec(useful_life_years); elapsed = _dec(elapsed_years)
    if life <= 0:
        raise SubstantiveTestError("useful_life_years must be > 0.")
    depreciable = max(cost - salvage, _ZERO)
    annual = (depreciable / life)
    accumulated = min(annual * elapsed, depreciable)
    nbv = cost - accumulated
    return {
        "annual_depreciation": annual.quantize(Decimal("0.0001")),
        "accumulated_depreciation": accumulated.quantize(Decimal("0.0001")),
        "net_book_value": nbv.quantize(Decimal("0.0001")),
    }


def net_pay(*, gross, deductions) -> Decimal:
    """Recompute net pay = gross − total deductions."""
    return (_dec(gross) - _dec(deductions)).quantize(Decimal("0.0001"))


def inventory_value(*, quantity, unit_cost) -> Decimal:
    """Recompute inventory value = counted quantity × unit cost."""
    return (_dec(quantity) * _dec(unit_cost)).quantize(Decimal("0.0001"))


def _tested_from_inputs(area, inputs) -> Decimal | None:
    """Derive the tested value from area-specific inputs (or None)."""
    if not inputs:
        return None
    try:
        if area == _Area.FIXED_ASSETS and "cost" in inputs:
            return straight_line_nbv(
                cost=inputs.get("cost", 0), salvage=inputs.get("salvage", 0),
                useful_life_years=inputs.get("useful_life_years", 0),
                elapsed_years=inputs.get("elapsed_years", 0))["net_book_value"]
        if area == _Area.PAYROLL and "gross" in inputs:
            return net_pay(gross=inputs.get("gross", 0),
                           deductions=inputs.get("deductions", 0))
        if area == _Area.INVENTORY and "unit_cost" in inputs and "quantity" in inputs:
            return inventory_value(quantity=inputs.get("quantity", 0),
                                   unit_cost=inputs.get("unit_cost", 0))
    except SubstantiveTestError:
        return None
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Workflow
# ─────────────────────────────────────────────────────────────────────────────
def _next_reference(organization) -> str:
    count = SubstantiveTestItem.objects.filter(organization=organization).count()
    return f"SUB-{count + 1:05d}"


def create_item(*, engagement, actor, area, book_value, item_reference="",
                description="", tested_value=None, tolerance=0, inputs=None,
                quantity_book=None, quantity_counted=None) -> SubstantiveTestItem:
    """Create a substantive-test item. Derives tested_value from inputs if given."""
    organization = engagement.organization
    inputs = inputs or {}
    derived = _tested_from_inputs(area, inputs)
    if derived is not None:
        tested_value = derived
    obj = SubstantiveTestItem(
        engagement=engagement, organization=organization, area=area,
        item_reference=item_reference, description=description,
        book_value=_dec(book_value),
        tested_value=_dec(tested_value) if tested_value not in (None, "") else None,
        tolerance=_dec(tolerance), inputs=inputs,
        quantity_book=_dec(quantity_book) if quantity_book not in (None, "") else None,
        quantity_counted=_dec(quantity_counted) if quantity_counted not in (None, "") else None,
        created_by=actor if getattr(actor, "pk", None) else None,
        status=_St.OPEN)
    obj.full_clean(exclude=["created_by", "reference"])
    with transaction.atomic():
        for attempt in range(5):
            obj.reference = _next_reference(organization)
            try:
                with transaction.atomic():
                    obj.save()
                break
            except IntegrityError:
                if attempt == 4:
                    raise
    # If a tested value is already present, classify immediately.
    if obj.tested_value is not None:
        _classify(obj)
    return obj


def record_tested(*, item, actor, tested_value, note="") -> SubstantiveTestItem:
    """Record the independently tested value and classify matched/variance."""
    if item.status == _St.CANCELLED:
        raise SubstantiveTestError("cannot test a cancelled item.")
    item.tested_value = _dec(tested_value)
    if note:
        item.notes = note
    item.save(update_fields=["tested_value", "notes", "updated_at"])
    return _classify(item)


def _classify(item) -> SubstantiveTestItem:
    within = item.is_within_tolerance
    item.status = _St.MATCHED if within else _St.VARIANCE
    item.save(update_fields=["status", "updated_at"])
    return item


def cancel(*, item, actor) -> SubstantiveTestItem:
    item.status = _St.CANCELLED
    item.save(update_fields=["status", "updated_at"])
    return item


def area_summary(*, organization, engagement=None) -> dict:
    """Per-area counts + total absolute variance for the dashboard."""
    qs = SubstantiveTestItem.objects.filter(organization=organization)
    if engagement is not None:
        qs = qs.filter(engagement=engagement)
    out = {}
    for area in _Area:
        aq = qs.filter(area=area.value)
        agg = aq.aggregate(
            total=Count("id"),
            matched=Count("id", filter=Q(status=_St.MATCHED)),
            variance=Count("id", filter=Q(status=_St.VARIANCE)),
            book=Sum("book_value"), tested=Sum("tested_value"))
        book = agg["book"] or _ZERO
        tested = agg["tested"] or _ZERO
        out[area.value] = {
            "total": agg["total"] or 0,
            "matched": agg["matched"] or 0,
            "variance": agg["variance"] or 0,
            "book_total": str(book),
            "tested_total": str(tested),
            "net_variance": str(book - tested),
        }
    totals = qs.aggregate(
        total=Count("id"),
        variance=Count("id", filter=Q(status=_St.VARIANCE)),
        matched=Count("id", filter=Q(status=_St.MATCHED)))
    out["_totals"] = {k: (v or 0) for k, v in totals.items()}
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Bulk import (CSV / XLSX) — count sheets, asset registers, payroll runs
# ─────────────────────────────────────────────────────────────────────────────
# Canonical field → accepted header aliases (normalised: lowercased, non-alnum
# collapsed to "_"). Lets auditors upload the register their client already has.
_COL_ALIASES = {
    "item_reference": ("item_reference", "reference", "ref", "sku", "asset_tag",
                       "asset_tag_no", "employee_id", "emp_id", "item", "code"),
    "description": ("description", "desc", "name", "item_name"),
    "book_value": ("book_value", "book", "recorded_value", "recorded",
                   "carrying_value", "carrying_amount", "gl_value", "net_book_value"),
    "tested_value": ("tested_value", "tested", "audited_value", "independent_value"),
    "tolerance": ("tolerance", "tol"),
    "quantity_book": ("quantity_book", "book_qty", "qty_book", "book_quantity"),
    "quantity_counted": ("quantity_counted", "counted_qty", "qty_counted",
                         "counted_quantity", "count", "physical_count"),
    "unit_cost": ("unit_cost", "cost_per_unit", "price", "unit_price"),
    "cost": ("cost", "acquisition_cost", "original_cost", "purchase_cost", "gross_cost"),
    "salvage": ("salvage", "residual", "salvage_value", "residual_value"),
    "useful_life_years": ("useful_life_years", "useful_life", "life", "life_years"),
    "elapsed_years": ("elapsed_years", "elapsed", "age_years", "years_elapsed", "age"),
    "gross": ("gross", "gross_pay", "gross_salary"),
    "deductions": ("deductions", "deduction", "total_deductions"),
}
_HEADER_TO_FIELD = {alias: field for field, aliases in _COL_ALIASES.items()
                    for alias in aliases}


def _norm_header(value) -> str:
    out = []
    for ch in str(value or "").strip().lower():
        out.append(ch if ch.isalnum() else "_")
    # collapse runs of "_"
    return "_".join(part for part in "".join(out).split("_") if part)


def _parse_amount(value):
    """Tolerant numeric cell → Decimal, or None if blank/unparseable."""
    if value in (None, ""):
        return None
    s = str(value).strip().replace(",", "").replace(" ", "")
    for sym in ("$", "£", "€", "ر.س", "SAR", "sar"):
        s = s.replace(sym, "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _read_csv(raw: bytes) -> list[dict]:
    text = raw.decode("utf-8-sig", errors="replace") if isinstance(raw, bytes) else raw
    return [dict(r) for r in csv.DictReader(io.StringIO(text))]


def _read_xlsx(file_obj) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl is a project dep
        raise SubstantiveTestError("openpyxl is required to read .xlsx files") from exc
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for values in rows[1:]:
        out.append({headers[i]: (values[i] if i < len(values) else None)
                    for i in range(len(headers))})
    return out


def _map_row(raw: dict) -> dict:
    """Map a raw upload row (arbitrary headers) to canonical fields."""
    mapped = {}
    for key, value in raw.items():
        field = _HEADER_TO_FIELD.get(_norm_header(key))
        if field and mapped.get(field) in (None, ""):
            mapped[field] = value
    return mapped


def _inputs_for_area(area, m) -> dict:
    """Build the recompute `inputs` dict for a mapped row, per area."""
    if area == _Area.FIXED_ASSETS and _parse_amount(m.get("cost")) is not None:
        return {"cost": str(_parse_amount(m.get("cost")) or 0),
                "salvage": str(_parse_amount(m.get("salvage")) or 0),
                "useful_life_years": str(_parse_amount(m.get("useful_life_years")) or 0),
                "elapsed_years": str(_parse_amount(m.get("elapsed_years")) or 0)}
    if area == _Area.PAYROLL and _parse_amount(m.get("gross")) is not None:
        return {"gross": str(_parse_amount(m.get("gross")) or 0),
                "deductions": str(_parse_amount(m.get("deductions")) or 0)}
    if (area == _Area.INVENTORY
            and _parse_amount(m.get("unit_cost")) is not None
            and _parse_amount(m.get("quantity_counted")) is not None):
        return {"quantity": str(_parse_amount(m.get("quantity_counted")) or 0),
                "unit_cost": str(_parse_amount(m.get("unit_cost")) or 0)}
    return {}


def import_items(*, engagement, actor, area, file_obj, filename) -> dict:
    """Bulk-create substantive-test items from a CSV/XLSX upload.

    Returns ``{"created": n, "skipped": n, "errors": [...]}``. Each row reuses
    ``create_item`` (so recompute + classification apply). Malformed rows are
    skipped with a reason; the import never partially corrupts — every created
    row is a valid item.
    """
    if area not in {a.value for a in _Area}:
        raise SubstantiveTestError("unknown area.")
    ext = (filename.rsplit(".", 1)[-1] if "." in (filename or "") else "").lower()
    raw = file_obj.read()
    if ext in ("xlsx", "xls"):
        rows = _read_xlsx(io.BytesIO(raw) if isinstance(raw, bytes) else raw)
    elif ext == "csv":
        rows = _read_csv(raw)
    else:
        raise SubstantiveTestError("unsupported file type (use .csv or .xlsx).")

    if len(rows) > _MAX_IMPORT_ROWS:
        raise SubstantiveTestError(
            f"too many rows ({len(rows)}); limit is {_MAX_IMPORT_ROWS}.")

    created, skipped, errors = 0, 0, []
    for idx, raw_row in enumerate(rows, start=2):  # row 1 = header
        m = _map_row(raw_row)
        book = _parse_amount(m.get("book_value"))
        # A wholly blank row is silently skipped; a row with data but no book
        # value is reported.
        if book is None:
            if any(v not in (None, "") for v in m.values()):
                skipped += 1
                errors.append(f"row {idx}: missing/blank book value")
            else:
                skipped += 1
            continue
        try:
            create_item(
                engagement=engagement, actor=actor, area=area, book_value=book,
                item_reference=str(m.get("item_reference") or "")[:128],
                description=str(m.get("description") or "")[:255],
                tested_value=_parse_amount(m.get("tested_value")),
                tolerance=_parse_amount(m.get("tolerance")) or _ZERO,
                inputs=_inputs_for_area(area, m),
                quantity_book=_parse_amount(m.get("quantity_book")),
                quantity_counted=_parse_amount(m.get("quantity_counted")))
            created += 1
        except Exception as exc:  # noqa: BLE001 - report, keep importing
            skipped += 1
            errors.append(f"row {idx}: {exc}")
    return {"created": created, "skipped": skipped, "errors": errors[:50]}
