"""Parser Service — thin wrapper delegating to core parser stack."""
from __future__ import annotations
import json
import logging
from datetime import date, datetime
from decimal import Decimal

logger = logging.getLogger(__name__)


def _format_cell(value) -> str:
    """Render a cell value for the AI: keep numbers/dates compact, strip noise."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float):
        # Drop trailing zeros for cleaner output
        return f"{value:.4f}".rstrip("0").rstrip(".") or "0"
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return text[:200]


class ParserService:
    """Delegates structured file parsing to the canonical core parser stack."""

    # Output cap so the model never receives more than ~50k chars from a single file.
    # GPT-4o context is 128k tokens (~500k chars) but most useful audit signal lives
    # in headers + a representative sample of rows, not in dumping millions of cells.
    _EXCEL_MAX_CHARS = 50_000
    _EXCEL_MAX_ROWS_PER_SHEET = 80
    _EXCEL_MAX_SHEETS = 8

    def parse_excel(self, file_path: str) -> str:
        """Smart-sampled Excel reader.

        Streams sheets in read-only mode (fast) and emits header + a bounded
        sample of rows per sheet, capped at ~50,000 chars total. This keeps the
        AI input within token limits while preserving structure.
        """
        try:
            return self._openpyxl_smart_sample(file_path)
        except Exception as exc:
            logger.warning("openpyxl smart-sample failed, falling back to pandas: %s", exc)
            return self._pandas_excel_fallback(file_path)

    def _openpyxl_smart_sample(self, file_path: str) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            parts: list[str] = []
            total_chars = 0
            for sheet_name in wb.sheetnames[: self._EXCEL_MAX_SHEETS]:
                ws = wb[sheet_name]
                lines: list[str] = [f"# Sheet: {sheet_name}"]

                # iter_rows yields per row; we only emit non-empty rows
                row_iter = ws.iter_rows(values_only=True)
                emitted = 0
                for row in row_iter:
                    if emitted >= self._EXCEL_MAX_ROWS_PER_SHEET:
                        lines.append(f"... [truncated rows after {emitted}] ...")
                        break
                    cells = [_format_cell(c) for c in row]
                    if not any(cells):
                        continue
                    lines.append(" | ".join(cells))
                    emitted += 1

                block = "\n".join(lines)
                if total_chars + len(block) > self._EXCEL_MAX_CHARS:
                    parts.append(block[: self._EXCEL_MAX_CHARS - total_chars])
                    parts.append("\n... [Excel content truncated at output cap] ...")
                    break
                parts.append(block)
                total_chars += len(block) + 2  # +2 for the joining newlines

            return "\n\n".join(parts)
        finally:
            wb.close()

    def _pandas_excel_fallback(self, file_path: str) -> str:
        try:
            import pandas as pd
            xls = pd.ExcelFile(file_path)
            parts: list[str] = []
            total = 0
            for sheet_name in xls.sheet_names[: self._EXCEL_MAX_SHEETS]:
                df = pd.read_excel(xls, sheet_name=sheet_name, nrows=self._EXCEL_MAX_ROWS_PER_SHEET)
                block = f"# Sheet: {sheet_name}\n{df.to_string(index=False, max_rows=self._EXCEL_MAX_ROWS_PER_SHEET)}"
                if total + len(block) > self._EXCEL_MAX_CHARS:
                    parts.append(block[: self._EXCEL_MAX_CHARS - total])
                    break
                parts.append(block)
                total += len(block) + 2
            return "\n\n".join(parts)
        except Exception as exc:
            logger.warning("Pandas Excel fallback failed: %s", exc)
            return ""

    def parse_csv(self, file_path: str) -> str:
        try:
            from core.services.parsers.csv_parser import CsvParser
            parser = CsvParser()
            result = parser.parse(file_path)
            return result.text if hasattr(result, "text") else str(result)
        except Exception as exc:
            logger.warning("CSV parse via core failed, falling back to pandas: %s", exc)
            return self._pandas_csv_fallback(file_path)

    def _pandas_csv_fallback(self, file_path: str) -> str:
        try:
            import pandas as pd
            df = pd.read_csv(file_path, nrows=500)
            return df.to_string(index=False, max_rows=300)
        except Exception as exc:
            logger.warning("Pandas CSV fallback failed: %s", exc)
            return ""

    def parse_json(self, file_path: str) -> str:
        try:
            from core.services.parsers.json_parser import JsonParser
            parser = JsonParser()
            result = parser.parse(file_path)
            return result.text if hasattr(result, "text") else str(result)
        except Exception as exc:
            logger.warning("JSON parse via core failed, falling back to stdlib: %s", exc)
            return self._json_fallback(file_path)

    def _json_fallback(self, file_path: str) -> str:
        try:
            with open(file_path, encoding="utf-8", errors="replace") as f:
                data = json.load(f)
            return json.dumps(data, ensure_ascii=False, indent=2)[:8000]
        except Exception as exc:
            logger.warning("JSON stdlib fallback failed: %s", exc)
            return ""

    def parse_xml(self, file_path: str) -> str:
        """Parse XML (incl. ZATCA UBL invoices). Returns flat key:value text."""
        try:
            from lxml import etree
            tree = etree.parse(file_path)
            root = tree.getroot()

            lines: list[str] = []
            for el in root.iter():
                tag = etree.QName(el).localname
                text = (el.text or "").strip()
                if text:
                    lines.append(f"{tag}: {text}")
                for k, v in (el.attrib or {}).items():
                    lines.append(f"{tag}@{etree.QName(k).localname}: {v}")
                if len(lines) >= 800:
                    break
            return "\n".join(lines)[:12000]
        except Exception as exc:
            logger.warning("XML parse via lxml failed, falling back to text: %s", exc)
            try:
                with open(file_path, encoding="utf-8", errors="replace") as f:
                    return f.read()[:12000]
            except Exception as e2:
                logger.warning("XML text fallback failed: %s", e2)
                return ""

    def parse_text(self, file_path: str) -> str:
        """Read a plain-text file (.txt/.md/.log)."""
        try:
            with open(file_path, encoding="utf-8", errors="replace") as f:
                return f.read()[:12000]
        except Exception as exc:
            logger.warning("Text read failed: %s", exc)
            return ""

    def parse_pdf_text(self, file_path: str) -> str:
        """Extract embedded text from a PDF using PyMuPDF (no OCR)."""
        try:
            import fitz  # PyMuPDF
            with fitz.open(file_path) as pdf:
                pages = []
                for page in pdf:
                    txt = page.get_text("text")
                    if txt:
                        pages.append(txt)
            return "\n".join(pages)[:20000]
        except Exception as exc:
            logger.warning("PDF text extraction failed: %s", exc)
            return ""

    def render_pdf_pages(self, file_path: str, max_pages: int = 5, dpi: int = 200):
        """Render the first N PDF pages as PNG bytes for Vision input."""
        try:
            import fitz
            images = []
            with fitz.open(file_path) as pdf:
                for i, page in enumerate(pdf):
                    if i >= max_pages:
                        break
                    pix = page.get_pixmap(dpi=dpi)
                    images.append(pix.tobytes("png"))
            return images
        except Exception as exc:
            logger.warning("PDF render failed: %s", exc)
            return []
