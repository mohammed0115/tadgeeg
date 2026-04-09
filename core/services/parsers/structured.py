"""
core/services/parsers/structured.py
=====================================
Structured-file record iterators (CSV, TSV, Excel, JSON).

These were previously defined as private module-level functions inside
``apps/invoices/views.py``, making them untestable and unreachable from
Celery tasks without importing the entire view module.

Public API
----------
iter_structured_records(uploaded_file, filename)
    Yields (row_number, payload_dict) for each data row in the file.
    Returns None if the file extension is not a supported structured type.

detect_csv_delimiter(sample)
    Returns the most likely CSV delimiter character for the given text sample.

normalize_row(raw_mapping)
    Cleans a raw dict row (strips keys, removes NaN/None/empty values).
"""
from __future__ import annotations

import csv
import io
import json
import os
from typing import Generator, Iterator, Optional

_SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".json", ".xlsx", ".xls"}


# ── Public helpers ─────────────────────────────────────────────────────────────

def normalize_row(raw_mapping: dict) -> dict:
    """
    Clean a raw row dict coming from a CSV/Excel/JSON reader.

    - Strips whitespace from string keys and values
    - Drops keys starting with "Unnamed" (pandas artefact)
    - Drops entries whose cleaned value is empty, "none", "nan", or "nat"
    """
    normalized: dict = {}
    for key, value in (raw_mapping or {}).items():
        key_name = str(key or "").strip()
        if not key_name or key_name.lower().startswith("unnamed"):
            continue
        cleaned = _serialize_value(value)
        if cleaned != "":
            normalized[key_name] = cleaned
    return normalized


def detect_csv_delimiter(sample: str) -> str:
    """Sniff the delimiter from a text sample; fall back to comma."""
    try:
        dialect = csv.Sniffer().sniff(sample or "", delimiters=",;\t|")
        return dialect.delimiter
    except Exception:
        counts = {
            ",": sample.count(","),
            ";": sample.count(";"),
            "\t": sample.count("\t"),
            "|": sample.count("|"),
        }
        return max(counts, key=counts.get) if any(counts.values()) else ","


def iter_structured_records(
    uploaded_file,
    filename: str,
) -> Optional[Iterator[tuple[int, dict]]]:
    """
    Yield ``(row_number, payload)`` pairs for every data row in the file.

    Returns ``None`` if the file extension is not a supported structured type
    (e.g. PDFs and images are not handled here — route those through the
    document engine instead).
    """
    ext = os.path.splitext(filename)[1].lower()
    if ext in {".csv", ".tsv"}:
        return _iter_csv_records(uploaded_file)
    if ext in {".xlsx", ".xls"}:
        return _iter_excel_records(uploaded_file, ext)
    if ext == ".json":
        return _iter_json_records(uploaded_file)
    return None


# ── Internal iterators ─────────────────────────────────────────────────────────

def _serialize_value(value) -> str:
    """Convert a cell value to a clean string; return '' for empty/null values."""
    if value is None:
        return ""
    from decimal import Decimal
    if isinstance(value, Decimal):
        return format(value, "f")
    if hasattr(value, "isoformat"):
        return value.isoformat()
    cleaned = str(value).strip()
    return "" if cleaned.lower() in {"", "none", "nan", "nat"} else cleaned


def _iter_csv_records(csv_file) -> Generator[tuple[int, dict], None, None]:
    csv_file.seek(0)
    sample = csv_file.read(4096).decode("utf-8", errors="replace")
    csv_file.seek(0)
    delimiter = detect_csv_delimiter(sample)
    wrapper = io.TextIOWrapper(csv_file, encoding="utf-8-sig", newline="")
    try:
        reader = csv.DictReader(wrapper, delimiter=delimiter)
        for row_number, row in enumerate(reader, start=2):
            payload = normalize_row(row)
            if payload:
                yield row_number, payload
    finally:
        try:
            wrapper.detach()
        except Exception:
            pass
        csv_file.seek(0)


def _iter_excel_records(spreadsheet_file, ext: str) -> Generator[tuple[int, dict], None, None]:
    spreadsheet_file.seek(0)
    if ext == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(spreadsheet_file, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                row_iter = worksheet.iter_rows(values_only=True)
                headers = next(row_iter, None)
                if not headers:
                    continue
                header_names = [
                    str(h).strip() if h not in (None, "") else f"column_{i + 1}"
                    for i, h in enumerate(headers)
                ]
                for row_number, row_values in enumerate(row_iter, start=2):
                    payload = normalize_row(dict(zip(header_names, row_values)))
                    if payload:
                        yield row_number, payload
        finally:
            workbook.close()
            spreadsheet_file.seek(0)
        return

    import pandas as pd

    data_frame = pd.read_excel(spreadsheet_file, dtype=str)
    spreadsheet_file.seek(0)
    for row_number, row in enumerate(data_frame.to_dict(orient="records"), start=2):
        payload = normalize_row(row)
        if payload:
            yield row_number, payload


def _iter_json_records(json_file) -> Generator[tuple[int, dict], None, None]:
    try:
        raw = json_file.read()
        json_file.seek(0)
        data = json.loads(raw.decode("utf-8", errors="replace").lstrip("\xef\xbb\xbf"))
    except Exception:
        return

    if not isinstance(data, list):
        return

    for row_number, item in enumerate(data, start=1):
        if isinstance(item, dict):
            payload = normalize_row(item)
            if payload:
                yield row_number, payload
