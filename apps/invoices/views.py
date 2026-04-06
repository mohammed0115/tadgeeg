"""
Invoice Auditing Views
Supports: single file, multiple files, ZIP archive upload.
Runs all 30 validation rules + AI analysis on each invoice.
"""

import csv
import io
import json
import logging
import os
import time
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.files.base import ContentFile
from django.db.models import Avg, Count, F, Max, Min, Q, Sum
from django.http import FileResponse
from django.shortcuts import render
from rest_framework import mixins
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.translation import gettext as _
from drf_spectacular.utils import OpenApiParameter, extend_schema
from rest_framework import generics, status
from rest_framework.authentication import SessionAuthentication
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication

from apps.audit.services import AuditSessionService
from apps.audit.serializers import AuditFindingSerializer
from apps.authentication.permissions import IsSeniorAuditorOrAbove, IsOwnOrganization
from core.services.invoice_ai_service import analyze_invoice_risk, extract_invoice_with_ai
from core.services.invoice_validator import RULES, TOTAL_RULES, compute_file_hash
from core.services.invoice_validator import run_all_rules as core_run_all_rules
from core.services.normalization import NormalizationService
from core.services.ocr_service import pdf_to_images
from core.services.validation_pipeline import ValidationPipelineService
from core.services.zip_validator import validate_zip_bomb, ZipValidationError
from core.utils.audit import log_action
from apps.authentication.models import AuditLog

from .models import (
    Invoice, InvoiceAuditEvent, InvoiceBatch, InvoiceValidationResult, VendorProfile
)
from .serializers import (
    InvoiceAuditEventSerializer,
    InvoiceBatchSerializer, InvoiceDetailSerializer,
    InvoiceListSerializer, InvoiceValidationResultSerializer,
    VendorProfileSerializer,
)

logger = logging.getLogger("finai")
normalizer = NormalizationService()


def run_all_rules(invoice, organization=None, file_hash: str = "") -> dict:
    """Backward-compatible validation entry point used by older callers and tests."""
    return core_run_all_rules(invoice, organization=organization, file_hash=file_hash)

ALLOWED_MIME = {"application/pdf", "image/jpeg", "image/png", "image/tiff",
                "application/zip", "application/x-zip-compressed",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel", "application/json", "text/json",
                "text/csv", "text/plain", "text/tab-separated-values",
                "application/csv", "application/octet-stream"}
ALLOWED_EXT  = {".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".zip",
                ".xlsx", ".xls", ".json", ".csv", ".tsv"}

DEFAULT_BULK_CHUNK_SIZE = 250
MIN_BULK_CHUNK_SIZE = 25
MAX_BULK_CHUNK_SIZE = 1000
AUTO_ASYNC_FILE_SIZE = 512 * 1024
STRUCTURED_BULK_EXTENSIONS = {".csv", ".tsv", ".json", ".xlsx", ".xls"}

# Loaded lazily to avoid circular imports while still allowing tests to patch it.
process_invoice_rows_chunk_task = None


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _get_client_ip(request):
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    return xff.split(",")[0].strip() if xff else request.META.get("REMOTE_ADDR", "")


def _save_audit_event(invoice, user, event_type, description="", before=None, after=None, request=None):
    InvoiceAuditEvent.objects.create(
        invoice=invoice,
        user=user,
        event_type=event_type,
        description=description,
        before_data=before or {},
        after_data=after or {},
        ip_address=_get_client_ip(request) if request else None,
    )


def _first_non_empty(*values):
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned:
                return cleaned
            continue
        return value
    return ""


def _to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError, ArithmeticError):
        return default


def _risk_level_from_score(score):
    numeric_score = max(0.0, min(100.0, _to_float(score)))
    if numeric_score >= 70:
        return "high"
    if numeric_score >= 40:
        return "medium"
    return "low"


def _fallback_risk_score(validation_result):
    validation_score = max(0.0, min(100.0, _to_float(validation_result.get("validation_score"), 0.0)))
    failed_codes = set(validation_result.get("failed_rule_codes") or [])
    fallback_score = round(100.0 - validation_score, 2)

    if any(code.startswith("DUP-") for code in failed_codes):
        fallback_score = max(fallback_score, 78.0)
    elif any(code.startswith("ANO-") for code in failed_codes):
        fallback_score = max(fallback_score, 72.0)
    elif any(code.startswith("VAT-") for code in failed_codes):
        fallback_score = max(fallback_score, 58.0)

    return round(max(0.0, min(100.0, fallback_score)), 2)


def _merge_risk_assessment(invoice, validation_result, risk_result):
    risk_result = risk_result or {}
    fallback_score = _fallback_risk_score(validation_result)
    ai_score = max(0.0, min(100.0, _to_float(risk_result.get("overall_risk_score"), 0.0)))
    final_score = max(ai_score, fallback_score)
    final_level = _risk_level_from_score(final_score)

    invoice.risk_score = round(final_score, 2)
    invoice.risk_level = final_level
    invoice.ai_recommendations = risk_result.get("recommendations", [])
    if not invoice.ai_summary:
        invoice.ai_summary = str(risk_result.get("ai_summary", "") or "")

    invoice.is_duplicate = any(
        code in validation_result.get("failed_rule_codes", []) for code in ["DUP-001", "DUP-002", "DUP-003", "DUP-004"]
    )
    if invoice.status not in [Invoice.Status.APPROVED, Invoice.Status.REJECTED]:
        invoice.status = Invoice.Status.FLAGGED if invoice.risk_score >= 70 else Invoice.Status.VALIDATED

    return {
        "overall_risk_score": invoice.risk_score,
        "risk_level": invoice.risk_level,
        "recommendations": invoice.ai_recommendations,
        "ai_summary": invoice.ai_summary,
    }


REVIEW_FIELD_META = [
    ("vendor_name", _("Vendor Name"), "text"),
    ("vendor_vat_number", _("Vendor VAT Number"), "text"),
    ("invoice_number", _("Invoice Number"), "text"),
    ("invoice_date", _("Invoice Date"), "date"),
    ("due_date", _("Due Date"), "date"),
    ("subtotal", _("Subtotal"), "amount"),
    ("vat_amount", _("VAT Amount"), "amount"),
    ("total_amount", _("Total Amount"), "amount"),
    ("currency", _("Currency"), "text"),
    ("customer_name", _("Customer Name"), "text"),
    ("customer_vat_number", _("Customer VAT Number"), "text"),
    ("cost_center", _("Cost Center"), "text"),
    ("account_code", _("Account Code"), "text"),
    ("budget_code", _("Budget Code"), "text"),
]


def _serialize_review_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def _coerce_review_value(field_name: str, raw_value):
    if raw_value in (None, ""):
        return None

    if field_name in {"invoice_date", "due_date"}:
        if hasattr(raw_value, "isoformat"):
            return raw_value
        parsed = parse_date(str(raw_value).strip())
        if parsed:
            return parsed
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
            try:
                return datetime.strptime(str(raw_value).strip(), fmt).date()
            except ValueError:
                continue
        raise ValueError(_("Invalid date for %(field_name)s.") % {"field_name": field_name})

    if field_name in {"subtotal", "vat_amount", "total_amount"}:
        try:
            return Decimal(str(raw_value).replace(",", "").strip())
        except (InvalidOperation, ValueError, TypeError, AttributeError) as exc:
            raise ValueError(_("Invalid amount for %(field_name)s.") % {"field_name": field_name}) from exc

    value = str(raw_value).strip()
    return value or None


def _build_review_payload(invoice):
    extracted = dict(invoice.extracted_data or {})
    normalized = dict(extracted.get("normalized") or {})
    manual_review = dict(extracted.get("review") or {})
    manual_corrections = dict(manual_review.get("corrections") or {})
    ai_payload = {
        key: value
        for key, value in extracted.items()
        if key not in {"normalized", "review"} and not str(key).startswith("_")
    }

    field_rows = []
    for field_name, label, field_type in REVIEW_FIELD_META:
        field_rows.append(
            {
                "field": field_name,
                "label": label,
                "type": field_type,
                "current": _serialize_review_value(getattr(invoice, field_name, "")),
                "normalized": _serialize_review_value(normalized.get(field_name)),
                "ai": _serialize_review_value(ai_payload.get(field_name)),
                "manual": _serialize_review_value(manual_corrections.get(field_name)),
            }
        )

    review_events = invoice.audit_events.filter(
        event_type=InvoiceAuditEvent.EventType.EDITED,
        description__icontains="Manual review",
    ).order_by("-timestamp")

    return {
        "fields": field_rows,
        "raw_text": invoice.raw_text or "",
        "normalized": normalized,
        "ai_extracted": ai_payload,
        "manual_review": manual_review,
        "manual_corrections": manual_corrections,
        "last_review_event": InvoiceAuditEventSerializer(review_events.first()).data if review_events.exists() else None,
    }


def _run_invoice_revalidation(invoice, acting_user, request=None):
    file_hash = (invoice.extracted_data or {}).get("file_hash", "")
    validation_result = run_all_rules(
        invoice,
        organization=invoice.organization,
        file_hash=file_hash,
    )
    validation_result = ValidationPipelineService.persist_validation_result(
        invoice=invoice,
        validation_result=validation_result,
        created_by=acting_user,
    )
    risk_result = {}
    try:
        risk_result = analyze_invoice_risk(
            invoice.extracted_data or {},
            _get_vendor_history(invoice.organization, invoice.vendor_name),
        )
    except Exception as exc:
        logger.warning("AI risk re-analysis failed during manual review: %s", exc)

    _merge_risk_assessment(invoice, validation_result, risk_result)
    invoice.save(update_fields=["risk_score", "risk_level", "ai_recommendations", "ai_summary", "is_duplicate", "status", "updated_at"])
    _save_audit_event(
        invoice,
        acting_user,
        InvoiceAuditEvent.EventType.REPROCESSED,
        f"Re-validated after manual review: score={validation_result['validation_score']}%",
        request=request,
    )
    return validation_result


def _process_single_file(file_obj, filename: str, org, user, batch=None, request=None, audit_session=None) -> dict:
    """
    Full invoice processing pipeline:
      1. Upload File        — save raw file, create Invoice record
      2. Document Engine    — MIME detection, route to correct parser
      3. File Parser        — PDF/Image/Excel/JSON/CSV parser
      4. OCR/Text Extract   — Tesseract (+ OpenAI OCR upgrade if confidence < 60%)
      5. OpenAI Extraction  — ZATCA-specific invoice field extraction (GPT-4o)
      6. Financial AI       — classify → dup → fraud → compliance → risk
      7. Audit Engine       — run 6 audit rules, persist AuditCase
      8. Risk Engine        — merge scores from val + AI → final risk level
      9. Save to DB         — update Invoice, InvoiceValidationResult, VendorProfile

    Returns:
        dict with invoice_id, validation, risk, errors.
    """
    from core.services.document_engine import DocumentEngine
    from core.services.financial_ai_engine import FinancialAIEngine
    from apps.audit.audit_engine import run_audit

    start = time.time()
    ext = os.path.splitext(filename)[1].lower()
    file_data = file_obj.read()
    file_obj.seek(0)

    # ── Step 1: Upload File — save raw bytes, create initial Invoice record ───
    file_hash = compute_file_hash(io.BytesIO(file_data))
    if audit_session:
        AuditSessionService.advance_to_extracting(audit_session)
        if AuditSessionService.has_file_hash(audit_session, file_hash):
            raise ValueError(_("Duplicate file already processed in this audit session."))

    invoice = Invoice.objects.create(
        organization=org,
        uploaded_by=user,
        audit_session=audit_session,
        batch=batch,
        file=ContentFile(file_data, name=filename),
        original_filename=filename,
        file_size=len(file_data),
        mime_type=file_obj.content_type if hasattr(file_obj, "content_type") else "",
        extracted_data={"file_hash": file_hash},
    )
    _save_audit_event(invoice, user, InvoiceAuditEvent.EventType.UPLOADED,
                      f"Uploaded: {filename}", request=request)

    file_path = invoice.file.path

    # ── Steps 2 & 3: Document Engine → File Parser ───────────────────────────
    # MIME detection → routes to PDF / Image / Excel / JSON / CSV parser
    doc_engine = DocumentEngine(use_ai=True)
    ingestion = doc_engine.ingest(file_path)
    if audit_session:
        AuditSessionService.advance_to_normalizing(audit_session)

    if ingestion.fatal_error:
        invoice.status = Invoice.Status.FLAGGED
        invoice.processing_error = ingestion.fatal_error
        invoice.extracted_data = {"file_hash": file_hash, "error": ingestion.fatal_error}
        invoice.save(update_fields=["status", "processing_error", "extracted_data", "updated_at"])
        return {
            "invoice_id": str(invoice.id), "filename": filename,
            "success": False, "error": ingestion.fatal_error, "risk_level": "high",
        }

    # ── Step 4: OCR / Text Extraction (handled inside the parser) ────────────
    # image_parser:  Tesseract → OpenAI OCR upgrade when confidence < 60%
    # pdf_parser:    PyMuPDF page render → Tesseract → OpenAI OCR upgrade
    # excel_parser:  pandas direct extraction (no OCR needed)
    # json_parser:   direct text extraction (no OCR needed)
    raw_text = ingestion.raw_text
    ocr_confidence = float(ingestion.metadata.get("ocr_confidence", 0.0))

    _save_audit_event(invoice, user, InvoiceAuditEvent.EventType.PROCESSED,
                      f"Parser: {ingestion.extraction_method} | OCR confidence: {ocr_confidence:.0f}%",
                      request=request)

    # ── Step 5: OpenAI Extraction — ZATCA-specific invoice field extraction ───
    # For vision: images use file_path directly; PDFs render first page
    if ext == ".pdf":
        try:
            img_pages = pdf_to_images(file_path)
            img_for_ai = img_pages[0] if img_pages else file_path
        except Exception:
            img_for_ai = file_path
    else:
        img_for_ai = file_path

    try:
        ai_data = extract_invoice_with_ai(img_for_ai, raw_text)
    except Exception as e:
        logger.warning(f"OpenAI extraction failed for {filename}: {e}")
        from core.services.invoice_ai_service import _fallback_extraction
        ai_data = _fallback_extraction(raw_text)

    normalized_ingestion = ingestion.normalized or {}
    normalization = normalizer.normalize(
        {
            **normalized_ingestion,
            **(ingestion.structured or {}),
            **(ai_data or {}),
            "raw_text": raw_text,
            "extraction_method": ingestion.extraction_method,
            "language": ai_data.get("language") or normalized_ingestion.get("language") or ingestion.metadata.get("language", "unknown"),
        },
        default_currency=getattr(org, "currency", "SAR") or "SAR",
    )
    normalized = normalization.normalized_data
    serialized_normalized = normalization.to_serializable_dict()

    # Populate Invoice fields (normalized canonical data takes priority)
    def _safe_decimal(val):
        try:
            return Decimal(str(val)) if val else Decimal("0")
        except Exception:
            return Decimal("0")

    def _safe_date(val):
        if not val:
            return None
        try:
            from datetime import datetime
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
                        "%Y/%m/%d", "%d %b %Y", "%d %B %Y"):
                try:
                    return datetime.strptime(str(val), fmt).date()
                except ValueError:
                    continue
        except Exception:
            pass
        return None

    invoice.invoice_number    = str(_first_non_empty(normalized.get("invoice_number"), ai_data.get("invoice_number"), filename) or "")
    invoice.invoice_date      = _safe_date(normalized.get("invoice_date") or ai_data.get("invoice_date"))
    invoice.due_date          = _safe_date(normalized.get("due_date") or ai_data.get("due_date"))
    invoice.vendor_name       = str(_first_non_empty(
        normalized.get("vendor_name"), ai_data.get("vendor_name"), ai_data.get("vendor_name_ar"),
        ai_data.get("supplier_name"), ai_data.get("merchant_name"),
    ) or "")
    invoice.vendor_name_ar    = str(_first_non_empty(ai_data.get("vendor_name_ar"), ai_data.get("vendor_name")) or "")
    invoice.vendor_vat_number = str(normalized.get("vendor_vat_number") or ai_data.get("vendor_vat_number", "") or "")
    invoice.vendor_cr_number  = str(ai_data.get("vendor_cr_number", "") or "")
    invoice.vendor_address    = str(ai_data.get("vendor_address", "") or "")
    invoice.vendor_phone      = str(ai_data.get("vendor_phone", "") or "")
    invoice.customer_name     = str(normalized.get("customer_name") or ai_data.get("customer_name", "") or "")
    invoice.customer_vat_number = str(normalized.get("customer_vat_number") or ai_data.get("customer_vat_number", "") or "")
    invoice.currency          = str(normalized.get("currency") or ai_data.get("currency", "SAR") or "SAR")
    invoice.subtotal          = _safe_decimal(normalized.get("subtotal"))
    invoice.vat_rate          = _safe_decimal(normalized.get("vat_rate", 15))
    invoice.vat_amount        = _safe_decimal(normalized.get("vat_amount"))
    invoice.discount          = _safe_decimal(normalized.get("discount"))
    invoice.total_amount      = _safe_decimal(normalized.get("total_amount"))
    invoice.line_items        = serialized_normalized.get("line_items") or ai_data.get("line_items", [])
    invoice.has_qr_code       = bool(normalized.get("has_qr_code", False))
    invoice.qr_code_valid     = bool(ai_data.get("qr_code_valid", False))
    invoice.is_handwritten    = bool(ai_data.get("is_handwritten") or ingestion.metadata.get("is_handwritten", False))
    invoice.is_clear          = bool(ai_data.get("is_clear", True))
    invoice.has_alterations   = bool(ai_data.get("has_alterations", False))
    invoice.language          = str(normalized.get("language") or ai_data.get("language", "unknown"))
    invoice.raw_text          = raw_text
    invoice.ocr_confidence    = ocr_confidence
    invoice.ai_summary        = str(ai_data.get("ai_summary", ""))
    invoice.extracted_data    = {
        **ai_data,
        "file_hash": file_hash,
        "_extraction_method": ingestion.extraction_method,
        "normalized": serialized_normalized,
    }
    invoice.status            = Invoice.Status.PROCESSING
    invoice.save()

    # ── Step 6: Financial AI Engine ──────────────────────────────────────────
    # classify → field extract → duplicate detection → fraud detection
    # → ZATCA compliance → Risk Engine (internal final step)
    ai_engine = FinancialAIEngine(organization_id=org.id, country_code="SA", use_ai=True)
    analysis = ai_engine.analyse(ingestion)

    # ── Step 7: Audit Engine — 6 structural audit rules + AuditCase persist ───
    try:
        run_audit(
            analysis.to_dict(),
            organization_id=org.id,
            invoice_id=invoice.id,
            persist=True,
        )
    except Exception as e:
        logger.warning(f"Audit engine failed for {filename}: {e}")

    # ── Step 7b: New Rule Engine — full 97-rule AuditPipeline (async) ────────
    try:
        from apps.rule_engine.tasks.audit_tasks import run_audit_task
        run_audit_task.delay(
            document_id=str(invoice.id),
            document_type="sales_invoice",
            organization_id=str(org.id),
            triggered_by="upload",
        )
    except Exception as e:
        logger.warning(f"Rule engine pipeline trigger failed for {filename}: {e}")

    # ── Also run 30 ZATCA validation rules against the Invoice model fields ───
    if audit_session:
        AuditSessionService.advance_to_validating(audit_session)
    val_result = ValidationPipelineService.validate_invoice(
        invoice=invoice,
        organization=org,
        file_hash=file_hash,
        created_by=user,
    )

    _save_audit_event(invoice, user, InvoiceAuditEvent.EventType.VALIDATED,
                      f"Validation: {val_result['validation_score']:.0f}% | Failed: {val_result['failed_rule_codes']}",
                      request=request)

    # ── Step 8: Risk Engine — merge validation + AI scores → final risk level ─
    # analyze_invoice_risk provides vendor-history-aware risk from invoice_ai_service
    # FinancialAIEngine already ran its own Risk Engine; we take the max score
    vendor_hist = _get_vendor_history(org, invoice.vendor_name)
    try:
        ai_risk = analyze_invoice_risk({k: v for k, v in invoice.extracted_data.items() if not k.startswith("_")}, vendor_hist)
    except Exception as e:
        logger.warning(f"AI risk analysis failed for {filename}: {e}")
        ai_risk = {}

    # Final score = max(Financial AI risk, invoice AI risk, fallback from val rules)
    ai_risk["overall_risk_score"] = max(
        _to_float(ai_risk.get("overall_risk_score"), 0.0),
        float(analysis.risk_score),
    )
    _merge_risk_assessment(invoice, val_result, ai_risk)

    # ── Step 9: Save to DB ────────────────────────────────────────────────────
    invoice.is_duplicate = invoice.is_duplicate or analysis.is_duplicate
    invoice.save()

    _save_canonical_invoice(invoice)
    _update_vendor_profile(org, invoice)

    elapsed_ms = int((time.time() - start) * 1000)
    logger.info(
        "[Upload] %s | risk=%s | score=%.0f%% | dup=%s | fraud=%.2f | %dms",
        filename, invoice.risk_level, val_result["validation_score"],
        invoice.is_duplicate, analysis.fraud_score, elapsed_ms,
    )
    if audit_session:
        AuditSessionService.record_success(
            audit_session,
            invoice,
            review_required=bool(
                getattr(analysis, "requires_review", False)
                or invoice.status == Invoice.Status.FLAGGED
                or bool(val_result.get("failed_rule_codes"))
            ),
        )

    return {
        "invoice_id":       str(invoice.id),
        "filename":         filename,
        "success":          True,
        "validation_score": val_result["validation_score"],
        "rules_failed":     val_result["failed_rule_codes"],
        "risk_level":       invoice.risk_level,
        "is_duplicate":     invoice.is_duplicate,
        "fraud_score":      round(analysis.fraud_score, 2),
        "findings_summary": val_result.get("findings_summary", {}),
        "status":           invoice.status,
        "processing_ms":    elapsed_ms,
    }


def _get_vendor_history(org, vendor_name: str) -> dict:
    if not vendor_name:
        return {"is_new": True, "invoice_count": 0}
    try:
        vp = VendorProfile.objects.get(organization=org, vendor_name=vendor_name)
        return {
            "invoice_count": vp.invoice_count,
            "avg_amount": float(vp.avg_invoice_amount),
            "max_amount": float(vp.max_invoice_amount),
            "flagged_count": vp.flagged_count,
            "is_new": vp.is_new,
        }
    except VendorProfile.DoesNotExist:
        return {"is_new": True, "invoice_count": 0}


def _save_canonical_invoice(invoice: Invoice) -> None:
    """
    Persist a DocumentCanonicalData record for an invoice after it is saved.
    Builds raw_data from invoice model fields so the mapper gets clean values
    (normalized data already took priority over raw AI output at this point).
    Never raises — failure is logged and silently swallowed.
    """
    try:
        from core.services.canonical_mapper import CanonicalMapper
        ed = invoice.extracted_data or {}
        raw_data = {
            # Identity
            "invoice_number":      invoice.invoice_number,
            "invoice_date":        str(invoice.invoice_date) if invoice.invoice_date else None,
            "due_date":            str(invoice.due_date)     if invoice.due_date     else None,
            # Counterparty
            "vendor_name":         invoice.vendor_name,
            "vendor_vat_number":   invoice.vendor_vat_number,
            "vendor_cr_number":    invoice.vendor_cr_number,
            "customer_name":       invoice.customer_name,
            "customer_vat_number": invoice.customer_vat_number,
            # Financial
            "currency":            invoice.currency,
            "subtotal":            float(invoice.subtotal    or 0),
            "vat_rate":            float(invoice.vat_rate    or 15),
            "vat_amount":          float(invoice.vat_amount  or 0),
            "discount":            float(invoice.discount    or 0),
            "total_amount":        float(invoice.total_amount or 0),
            # Compliance
            "qr_code_valid":       invoice.qr_code_valid,
            "cost_center":         ed.get("cost_center", ""),
            "department":          ed.get("department",  ""),
            "account_code":        ed.get("account_code",""),
        }
        CanonicalMapper().save_canonical(
            raw_data        = raw_data,
            document_type   = "invoice",
            typed_model_name= "Invoice",
            typed_object_id = invoice.id,
        )
    except Exception as exc:
        logger.warning("[canonical] invoice save failed for %s: %s", invoice.id, exc)


def _update_vendor_profile(org, invoice: Invoice):
    """Update vendor intelligence after processing an invoice."""
    if not invoice.vendor_name:
        return
    try:
        from apps.invoices.services.vendor_intelligence import VendorIntelligenceService

        VendorIntelligenceService().update_from_invoice(org, invoice)
    except Exception as e:
        logger.warning(f"Vendor profile update failed: {e}")


def _handle_processing_result(result: dict, filename: str, batch, results: list, errors: list, audit_session=None):
    if result.get("success"):
        results.append(result)
        batch.processed_files += 1
        return

    message = result.get("error") or "Processing failed"
    errors.append({"filename": filename, "error": message})
    batch.failed_files += 1
    if audit_session:
        AuditSessionService.record_failure(audit_session, message)


def _is_truthy(value) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _parse_chunk_size(raw_value) -> int:
    try:
        numeric = int(raw_value or DEFAULT_BULK_CHUNK_SIZE)
    except (TypeError, ValueError):
        numeric = DEFAULT_BULK_CHUNK_SIZE
    return max(MIN_BULK_CHUNK_SIZE, min(numeric, MAX_BULK_CHUNK_SIZE))


def _should_queue_async_processing(request=None, uploaded_file=None) -> bool:
    if request is not None:
        if _is_truthy(getattr(request, "data", {}).get("async") if hasattr(request, "data") else None):
            return True
        if _is_truthy(getattr(request, "data", {}).get("use_async") if hasattr(request, "data") else None):
            return True
        if hasattr(request, "query_params"):
            if _is_truthy(request.query_params.get("async")) or _is_truthy(request.query_params.get("use_async")):
                return True
    return bool(uploaded_file and getattr(uploaded_file, "size", 0) >= AUTO_ASYNC_FILE_SIZE)


def _serialize_structured_value(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format(value, "f")
    if hasattr(value, "isoformat"):
        return value.isoformat()
    cleaned = str(value).strip()
    return "" if cleaned.lower() in {"", "none", "nan", "nat"} else cleaned


def _normalize_structured_mapping(raw_mapping: dict) -> dict:
    normalized = {}
    for key, value in (raw_mapping or {}).items():
        key_name = str(key or "").strip()
        if not key_name or key_name.lower().startswith("unnamed"):
            continue
        cleaned_value = _serialize_structured_value(value)
        if cleaned_value != "":
            normalized[key_name] = cleaned_value
    return normalized


def _detect_csv_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample or "", delimiters=",;\t|")
        return dialect.delimiter
    except Exception:
        counts = {",": sample.count(","), ";": sample.count(";"), "\t": sample.count("\t"), "|": sample.count("|")}
        return max(counts, key=counts.get) if any(counts.values()) else ","


def _iter_csv_records(csv_file):
    csv_file.seek(0)
    sample = csv_file.read(4096).decode("utf-8", errors="replace")
    csv_file.seek(0)
    delimiter = _detect_csv_delimiter(sample)
    wrapper = io.TextIOWrapper(csv_file, encoding="utf-8-sig", newline="")
    try:
        reader = csv.DictReader(wrapper, delimiter=delimiter)
        for row_number, row in enumerate(reader, start=2):
            payload = _normalize_structured_mapping(row)
            if payload:
                yield row_number, payload
    finally:
        try:
            wrapper.detach()
        except Exception:
            pass
        csv_file.seek(0)


def _iter_excel_records(spreadsheet_file, ext: str):
    spreadsheet_file.seek(0)
    if ext == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(spreadsheet_file, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                row_iter = worksheet.iter_rows(values_only=True)
                headers = next(row_iter, None)
                if not headers:
                    continue
                header_names = [str(h).strip() if h not in (None, "") else f"column_{i + 1}" for i, h in enumerate(headers)]
                for row_number, row_values in enumerate(row_iter, start=2):
                    payload = _normalize_structured_mapping(dict(zip(header_names, row_values)))
                    if payload:
                        yield row_number, payload
        finally:
            workbook.close()
            spreadsheet_file.seek(0)
        return

    import pandas as pd

    data_frame = pd.read_excel(spreadsheet_file, dtype=str)
    spreadsheet_file.seek(0)
    for row_number, row in enumerate(data_frame.to_dict(orient="records"), start=2):
        payload = _normalize_structured_mapping(row)
        if payload:
            yield row_number, payload


def _iter_json_records(json_file):
    try:
        raw = json_file.read()
        json_file.seek(0)
        data = json.loads(raw.decode("utf-8", errors="replace").lstrip("\xef\xbb\xbf"))
    except Exception:
        return

    if not isinstance(data, list):
        return

    for row_number, item in enumerate(data, start=1):
        if isinstance(item, dict):
            payload = _normalize_structured_mapping(item)
            if payload:
                yield row_number, payload


def _iter_structured_records(uploaded_file, filename: str):
    ext = os.path.splitext(filename)[1].lower()
    if ext in {".csv", ".tsv"}:
        return _iter_csv_records(uploaded_file)
    if ext in {".xlsx", ".xls"}:
        return _iter_excel_records(uploaded_file, ext)
    if ext == ".json":
        return _iter_json_records(uploaded_file)
    return None


def _get_invoice_chunk_task():
    global process_invoice_rows_chunk_task
    if process_invoice_rows_chunk_task is None:
        from .tasks import process_invoice_rows_chunk_task as imported_task
        process_invoice_rows_chunk_task = imported_task
    return process_invoice_rows_chunk_task


def _process_structured_rows_chunk(
    rows,
    *,
    base_name: str,
    org,
    user,
    batch,
    request=None,
    audit_session=None,
    persist_batch_progress: bool = True,
):
    results, errors = [], []
    success_count = failure_count = duplicate_count = high_risk_count = review_required_count = 0
    last_error = ""

    for item in rows:
        row_number = item.get("row_number") or (success_count + failure_count + 1)
        payload = item.get("payload") or {}
        if not payload:
            continue

        row_name = f"{base_name}_row{row_number}.json"
        item_bytes = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        file_like = io.BytesIO(item_bytes)
        file_like.name = row_name
        file_like.size = len(item_bytes)
        file_like.content_type = "application/json"

        try:
            result = _process_single_file(file_like, row_name, org, user, batch, request, audit_session)
        except Exception as exc:
            logger.error("Structured row %s failed: %s", row_name, exc)
            result = {"success": False, "error": str(exc)}

        if result.get("success"):
            results.append(result)
            success_count += 1
            if result.get("is_duplicate"):
                duplicate_count += 1
            if result.get("risk_level") in {"high", "critical"}:
                high_risk_count += 1
            if result.get("status") == Invoice.Status.FLAGGED or bool(result.get("rules_failed")):
                review_required_count += 1
        else:
            message = result.get("error") or "Processing failed"
            errors.append({"filename": row_name, "error": message})
            failure_count += 1
            last_error = message
            if audit_session:
                AuditSessionService.record_failure(audit_session, message)

    if persist_batch_progress and batch is not None and (success_count or failure_count):
        batch.processed_files += success_count
        batch.failed_files += failure_count

    return {
        "results": results,
        "errors": errors,
        "success_count": success_count,
        "failure_count": failure_count,
        "duplicate_count": duplicate_count,
        "high_risk_count": high_risk_count,
        "review_required_count": review_required_count,
        "last_error": last_error,
    }


def _process_structured_upload(uploaded_file, filename: str, org, user, batch, request=None, audit_session=None):
    row_iter = _iter_structured_records(uploaded_file, filename)
    if row_iter is None:
        return None

    base_name = os.path.splitext(filename)[0]
    chunk_size = _parse_chunk_size(getattr(request, "data", {}).get("chunk_size") if request is not None and hasattr(request, "data") else None)
    prefer_async = _should_queue_async_processing(request, uploaded_file)

    results, errors = [], []
    pending_chunks = []
    current_chunk = []
    total_rows = 0

    for row_number, payload in row_iter:
        total_rows += 1
        current_chunk.append({"row_number": row_number, "payload": payload})
        if len(current_chunk) >= chunk_size:
            if prefer_async:
                pending_chunks.append(current_chunk)
            else:
                outcome = _process_structured_rows_chunk(
                    current_chunk,
                    base_name=base_name,
                    org=org,
                    user=user,
                    batch=batch,
                    request=request,
                    audit_session=audit_session,
                )
                results.extend(outcome["results"])
                errors.extend(outcome["errors"])
            current_chunk = []

    if current_chunk:
        if prefer_async:
            pending_chunks.append(current_chunk)
        else:
            outcome = _process_structured_rows_chunk(
                current_chunk,
                base_name=base_name,
                org=org,
                user=user,
                batch=batch,
                request=request,
                audit_session=audit_session,
            )
            results.extend(outcome["results"])
            errors.extend(outcome["errors"])

    if total_rows == 0:
        return None

    batch.total_files += max(total_rows - 1, 0)
    if audit_session:
        AuditSessionService.sync_expected_total(audit_session, batch.total_files)

    if prefer_async and pending_chunks:
        batch.status = InvoiceBatch.BatchStatus.PROCESSING
        if audit_session:
            AuditSessionService.advance_to_extracting(audit_session)

        queued_task_ids = []
        for chunk in pending_chunks:
            try:
                task = _get_invoice_chunk_task()
                async_result = task.delay(
                    rows=chunk,
                    base_name=base_name,
                    org_id=str(org.id),
                    user_id=str(user.id),
                    batch_id=str(batch.id),
                    audit_session_id=str(audit_session.id) if audit_session else None,
                )
                queued_task_ids.append(getattr(async_result, "id", ""))
            except Exception as exc:
                logger.warning("Async chunk dispatch failed for %s; falling back to inline processing: %s", filename, exc)
                inline = _process_structured_rows_chunk(
                    chunk,
                    base_name=base_name,
                    org=org,
                    user=user,
                    batch=batch,
                    request=request,
                    audit_session=audit_session,
                )
                results.extend(inline["results"])
                errors.extend(inline["errors"])

        if queued_task_ids:
            log_entry = {
                "mode": "async_chunked",
                "source_file": filename,
                "queued_rows": total_rows,
                "queued_chunks": len(queued_task_ids),
                "chunk_size": chunk_size,
                "task_ids": queued_task_ids,
                "streaming": True,
            }
            batch.processing_log = list(batch.processing_log or []) + [log_entry]
            batch.save(update_fields=["status", "total_files", "processed_files", "failed_files", "processing_log"])
            return {
                "handled": True,
                "mode": "async_chunked",
                "results": results,
                "errors": errors,
                "queued_rows": total_rows,
                "queued_chunks": len(queued_task_ids),
                "task_ids": queued_task_ids,
                "chunk_size": chunk_size,
                "streaming": True,
            }

    return {
        "handled": True,
        "mode": "sync_chunked",
        "results": results,
        "errors": errors,
        "queued_rows": 0,
        "queued_chunks": 0,
        "task_ids": [],
        "chunk_size": chunk_size,
        "streaming": True,
        "processed_rows": total_rows,
    }


# ─── Views ────────────────────────────────────────────────────────────────────

class InvoiceUploadView(APIView):
    """
    Upload invoices for auditing.

    Supports:
    - Single file (PDF / JPG / PNG / TIFF)
    - Multiple files (multipart with multiple 'files' fields)
    - ZIP archive (contains multiple invoice files)
    """
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Invoices"],
        summary="Upload one or more invoices (PDF/image/ZIP) — runs all 30 audit rules",
        request={"type": "object", "properties": {
            "files": {"type": "array", "items": {"type": "string", "format": "binary"},
                      "description": "One or more invoice files (PDF, JPG, PNG, TIFF, or ZIP)"},
            "batch_name": {"type": "string"},
        }},
    )
    def post(self, request):
        org = request.user.organization
        if not org:
            return Response({"error": _("User has no organization.")}, status=400)

        uploaded_files = request.FILES.getlist("files")
        if not uploaded_files:
            # Try single file key
            single = request.FILES.get("file")
            if single:
                uploaded_files = [single]
            else:
                return Response({"error": _("No files uploaded. Use 'files' or 'file' field.")}, status=400)

        batch_name = request.data.get("batch_name", f"Batch {timezone.now().strftime('%Y-%m-%d %H:%M')}")

        # Create batch
        batch = InvoiceBatch.objects.create(
            organization=org,
            uploaded_by=request.user,
            batch_name=batch_name,
            total_files=len(uploaded_files),
        )
        audit_session = AuditSessionService.create_session(
            organization=org,
            created_by=request.user,
            name=batch_name,
            total_count=len(uploaded_files),
            context={"source": "invoice_upload"},
        )
        batch.audit_session = audit_session
        batch.save(update_fields=["audit_session"])

        results = []
        errors  = []
        async_jobs = []

        for uploaded_file in uploaded_files:
            filename = uploaded_file.name
            ext = os.path.splitext(filename)[1].lower()

            if ext not in ALLOWED_EXT:
                errors.append({"filename": filename, "error": _("Unsupported file type: %(ext)s") % {"ext": ext}})
                batch.failed_files += 1
                AuditSessionService.record_failure(audit_session, f"Unsupported file type: {ext}")
                continue

            # ── ZIP: extract and process each file inside ─────────────────────
            if ext == ".zip":
                zip_results, zip_errors, zip_async_jobs = _process_zip(uploaded_file, org, request.user, batch, request, audit_session)
                results.extend(zip_results)
                errors.extend(zip_errors)
                async_jobs.extend(zip_async_jobs)
                continue

            # ── Structured bulk uploads: CSV / TSV / JSON list / Excel ───────
            if ext in STRUCTURED_BULK_EXTENSIONS:
                structured = _process_structured_upload(uploaded_file, filename, org, request.user, batch, request, audit_session)
                if structured and structured.get("handled"):
                    results.extend(structured.get("results", []))
                    errors.extend(structured.get("errors", []))
                    if structured.get("mode") == "async_chunked":
                        async_jobs.append({
                            "filename": filename,
                            "queued_rows": structured.get("queued_rows", 0),
                            "queued_chunks": structured.get("queued_chunks", 0),
                            "task_ids": structured.get("task_ids", []),
                            "chunk_size": structured.get("chunk_size"),
                            "streaming": structured.get("streaming", True),
                        })
                    continue

            try:
                r = _process_single_file(uploaded_file, filename, org, request.user, batch, request, audit_session)
                _handle_processing_result(r, filename, batch, results, errors, audit_session)
            except Exception as e:
                logger.error(f"Failed processing {filename}: {e}")
                errors.append({"filename": filename, "error": str(e)})
                batch.failed_files += 1
                AuditSessionService.record_failure(audit_session, str(e))

        # Finalize or keep running when async chunks were queued.
        batch.processing_log = list(batch.processing_log or []) + results + errors
        if async_jobs:
            batch.status = InvoiceBatch.BatchStatus.PROCESSING if not errors else InvoiceBatch.BatchStatus.PARTIAL
            batch.save(update_fields=["status", "total_files", "processed_files", "failed_files", "processing_log"])
            AuditSessionService.sync_expected_total(audit_session, batch.total_files)

            log_action(request, AuditLog.Action.DOCUMENT_UPLOAD, "invoice_batch", str(batch.id), {
                "files": len(results),
                "errors": len(errors),
                "async_jobs": len(async_jobs),
                "chunked": True,
            })

            return Response({
                "batch_id": str(batch.id),
                "audit_session_id": str(audit_session.id),
                "batch_name": batch_name,
                "total_files": batch.total_files,
                "processed": batch.processed_files,
                "failed": batch.failed_files,
                "status": batch.status,
                "processing_mode": "async_chunked",
                "streaming": True,
                "queued_chunks": sum(job.get("queued_chunks", 0) for job in async_jobs),
                "queued_rows": sum(job.get("queued_rows", 0) for job in async_jobs),
                "async_jobs": async_jobs,
                "results": results,
                "errors": errors,
            }, status=status.HTTP_202_ACCEPTED)

        batch.status = (
            InvoiceBatch.BatchStatus.COMPLETED if not errors else
            InvoiceBatch.BatchStatus.PARTIAL if results else
            InvoiceBatch.BatchStatus.FAILED
        )
        batch.completed_at = timezone.now()
        batch.save(update_fields=["status", "completed_at", "total_files", "processed_files", "failed_files", "processing_log"])
        AuditSessionService.sync_expected_total(audit_session, batch.total_files)
        AuditSessionService.finalize_if_ready(audit_session)

        log_action(request, AuditLog.Action.DOCUMENT_UPLOAD, "invoice_batch", str(batch.id),
                   {"files": len(results), "errors": len(errors), "chunked": True})

        return Response({
            "batch_id":      str(batch.id),
            "audit_session_id": str(audit_session.id),
            "batch_name":    batch_name,
            "total_files":   batch.total_files,
            "processed":     len(results),
            "failed":        len(errors),
            "status":        batch.status,
            "processing_mode": "sync_chunked",
            "streaming": True,
            "results":       results,
            "errors":        errors,
        }, status=status.HTTP_201_CREATED)


def _process_zip(zip_file, org, user, batch, request, audit_session=None) -> tuple[list, list, list]:
    """Extract and process all invoice files inside a ZIP archive."""
    results, errors, async_jobs = [], [], []
    member_count = 0
    try:
        # Validate ZIP before extraction (bomb detection)
        zip_file.seek(0)
        try:
            validate_zip_bomb(zip_file)
        except ZipValidationError as e:
            errors.append({"filename": zip_file.name, "error": str(e)})
            if audit_session:
                AuditSessionService.record_failure(audit_session, str(e))
            return results, errors, async_jobs

        zip_file.seek(0)
        with zipfile.ZipFile(io.BytesIO(zip_file.read()), "r") as zf:
            for member in zf.infolist():
                if member.is_dir():
                    continue
                name = os.path.basename(member.filename)
                ext = os.path.splitext(name)[1].lower()
                if ext not in {".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".xlsx", ".xls", ".json", ".csv", ".tsv"}:
                    continue
                member_count += 1
                try:
                    data = zf.read(member)
                    file_like = io.BytesIO(data)
                    file_like.name = name
                    file_like.size = len(data)
                    file_like.content_type = _guess_mime(ext)

                    if ext in STRUCTURED_BULK_EXTENSIONS:
                        structured = _process_structured_upload(file_like, name, org, user, batch, request, audit_session)
                        if structured and structured.get("handled"):
                            results.extend(structured.get("results", []))
                            errors.extend(structured.get("errors", []))
                            if structured.get("mode") == "async_chunked":
                                async_jobs.append({
                                    "filename": name,
                                    "queued_rows": structured.get("queued_rows", 0),
                                    "queued_chunks": structured.get("queued_chunks", 0),
                                    "task_ids": structured.get("task_ids", []),
                                    "chunk_size": structured.get("chunk_size"),
                                    "streaming": structured.get("streaming", True),
                                })
                            continue

                    r = _process_single_file(file_like, name, org, user, batch, request, audit_session)
                    _handle_processing_result(r, name, batch, results, errors, audit_session)
                except Exception as e:
                    logger.error(f"ZIP member {name} failed: {e}")
                    errors.append({"filename": name, "error": str(e)})
                    batch.failed_files += 1
                    if audit_session:
                        AuditSessionService.record_failure(audit_session, str(e))
    except zipfile.BadZipFile:
        errors.append({"filename": zip_file.name, "error": _("Invalid ZIP file")})
        if audit_session:
            AuditSessionService.record_failure(audit_session, "Invalid ZIP file")
    except ZipValidationError as e:
        errors.append({"filename": zip_file.name, "error": str(e)})
        if audit_session:
            AuditSessionService.record_failure(audit_session, str(e))

    if member_count:
        batch.total_files += max(member_count - 1, 0)
        if audit_session:
            AuditSessionService.sync_expected_total(audit_session, batch.total_files)

    return results, errors, async_jobs


def _process_csv_rows(csv_file, filename: str, org, user, batch, request, audit_session=None) -> tuple[list, list]:
    """
    Parse a CSV/TSV file and process each row as a separate invoice.
    Returns (results, errors). Returns ([], []) if the file has ≤1 data row (caller handles as single).
    """
    results, errors = [], []
    try:
        import pandas as pd
        raw = csv_file.read()
        csv_file.seek(0)

        # Detect delimiter
        sample = raw[:4096].decode("utf-8", errors="replace")
        delimiter = max({",": sample.count(","), ";": sample.count(";"),
                         "\t": sample.count("\t"), "|": sample.count("|")},
                        key=lambda k: {",": sample.count(","), ";": sample.count(";"),
                                       "\t": sample.count("\t"), "|": sample.count("|")}[k])

        df = pd.read_csv(
            io.BytesIO(raw), sep=delimiter, dtype=str,
            keep_default_na=False, on_bad_lines="skip",
            encoding="utf-8-sig",
        )
        df = df.dropna(how="all")

        if len(df) <= 1:
            return [], []  # Single row — caller processes as one file

    except Exception as e:
        logger.warning(f"CSV pre-parse failed for {filename}: {e}")
        return [], []

    base_name = os.path.splitext(filename)[0]
    for idx, row in df.iterrows():
        row_dict = {k: v for k, v in row.items() if str(v).strip()}
        if not row_dict:
            continue
        row_name = f"{base_name}_row{idx + 1}.json"
        item_bytes = __import__("json").dumps(row_dict, ensure_ascii=False).encode("utf-8")
        file_like = io.BytesIO(item_bytes)
        file_like.name = row_name
        file_like.content_type = "application/json"
        try:
            r = _process_single_file(file_like, row_name, org, user, batch, request, audit_session)
            _handle_processing_result(r, row_name, batch, results, errors, audit_session)
        except Exception as e:
            logger.error(f"CSV row {row_name} failed: {e}")
            errors.append({"filename": row_name, "error": str(e)})
            batch.failed_files += 1
            if audit_session:
                AuditSessionService.record_failure(audit_session, str(e))

    return results, errors


def _process_json_list(json_file, filename: str, org, user, batch, request, audit_session=None) -> tuple[list, list]:
    """
    If a JSON file contains a list[dict], process each dict as a separate invoice.
    Returns (results, errors). Returns ([], []) if the JSON is NOT a list (caller handles it).
    """
    import json as _json
    results, errors = [], []
    try:
        raw = json_file.read()
        json_file.seek(0)
        data = _json.loads(raw.decode("utf-8", errors="replace").lstrip("\xef\xbb\xbf"))
    except Exception:
        return [], []

    if not isinstance(data, list):
        return [], []  # Not a list — caller will process as single file

    base_name = os.path.splitext(filename)[0]
    for idx, item in enumerate(data):
        if not isinstance(item, dict):
            continue
        item_name = f"{base_name}_item{idx + 1}.json"
        item_bytes = _json.dumps(item, ensure_ascii=False).encode("utf-8")
        file_like = io.BytesIO(item_bytes)
        file_like.name = item_name
        file_like.content_type = "application/json"
        try:
            r = _process_single_file(file_like, item_name, org, user, batch, request, audit_session)
            _handle_processing_result(r, item_name, batch, results, errors, audit_session)
        except Exception as e:
            logger.error(f"JSON item {item_name} failed: {e}")
            errors.append({"filename": item_name, "error": str(e)})
            batch.failed_files += 1
            if audit_session:
                AuditSessionService.record_failure(audit_session, str(e))

    return results, errors


def _guess_mime(ext: str) -> str:
    return {
        "pdf": "application/pdf", "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "png": "image/png", "tiff": "image/tiff", "tif": "image/tiff",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xls": "application/vnd.ms-excel", "json": "application/json",
    }.get(ext.lstrip("."), "application/octet-stream")


class InvoiceListView(generics.ListAPIView):
    serializer_class = InvoiceListSerializer
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Invoices"],
        summary="List all invoices for the organization",
        parameters=[
            OpenApiParameter("status",       description="Filter by status"),
            OpenApiParameter("risk_level",   description="Filter by risk level"),
            OpenApiParameter("vendor_name",  description="Filter by vendor name"),
            OpenApiParameter("is_duplicate", type=bool),
            OpenApiParameter("date_from"),
            OpenApiParameter("date_to"),
            OpenApiParameter("min_amount",   type=float),
            OpenApiParameter("max_amount",   type=float),
            OpenApiParameter("search",       description="Search vendor/invoice number/notes"),
            OpenApiParameter("batch_id",     description="Filter by batch ID"),
        ],
    )
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Invoice.objects.filter(
            organization=self.request.user.organization,
            is_deleted=False  # Exclude soft-deleted invoices
        ).select_related(
            "uploaded_by", "approved_by", "batch"
        )
        p = self.request.query_params
        if v := p.get("status"):        qs = qs.filter(status=v)
        if v := p.get("risk_level"):    qs = qs.filter(risk_level=v)
        if v := p.get("vendor_name"):   qs = qs.filter(vendor_name__icontains=v)
        if v := p.get("is_duplicate"):  qs = qs.filter(is_duplicate=v.lower() == "true")
        if v := p.get("date_from"):     qs = qs.filter(invoice_date__gte=v)
        if v := p.get("date_to"):       qs = qs.filter(invoice_date__lte=v)
        if v := p.get("min_amount"):    qs = qs.filter(total_amount__gte=v)
        if v := p.get("max_amount"):    qs = qs.filter(total_amount__lte=v)
        if v := p.get("batch_id") or p.get("batch"):
            qs = qs.filter(batch_id=v)
        if v := p.get("search"):
            qs = qs.filter(
                Q(vendor_name__icontains=v) | Q(invoice_number__icontains=v) |
                Q(notes__icontains=v) | Q(vendor_vat_number__icontains=v)
            )
        return qs.order_by("-created_at")


class InvoiceDetailView(mixins.DestroyModelMixin, generics.RetrieveUpdateAPIView):
    serializer_class = InvoiceDetailSerializer
    permission_classes = [IsAuthenticated, IsOwnOrganization]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    @extend_schema(tags=["Invoices"], summary="Get full invoice details with validation results")
    def get(self, request, *args, **kwargs):
        invoice = self.get_object()
        
        # Generate ZATCA QR code if not already present and invoice has required fields
        if not invoice.qr_code_image and invoice.vendor_vat_number and invoice.total_amount:
            try:
                invoice.generate_zatca_qr()
                invoice.save(update_fields=['qr_code_image', 'qr_code_data', 'has_qr_code', 'qr_code_valid'])
            except Exception as e:
                import logging
                logging.warning(f"Failed to generate QR code for invoice {invoice.id}: {str(e)}")
        
        if request.path.startswith("/invoices/") and "sessionid" in request.COOKIES:
            from apps.frontend.page_views import _build_invoice_display, _ctx

            audit_trail = invoice.audit_events.select_related("user").order_by("-timestamp")[:40]
            return render(
                request._request,
                "invoices/detail_premium.html",
                _ctx(
                    request._request,
                    "invoices",
                    invoice=invoice,
                    invoice_display=_build_invoice_display(invoice),
                    audit_trail=audit_trail,
                ),
            )

        data = InvoiceDetailSerializer(invoice).data
        try:
            data["validation"] = InvoiceValidationResultSerializer(invoice.validation).data
        except InvoiceValidationResult.DoesNotExist:
            data["validation"] = None
        data["findings"] = AuditFindingSerializer(
            invoice.audit_findings.order_by("-last_detected_at")[:20],
            many=True,
        ).data
        data["audit_trail"] = list(
            invoice.audit_events.order_by("-timestamp").values(
                "event_type", "description", "timestamp", "user__full_name", "ip_address"
            )[:20]
        )
        data["review"] = _build_review_payload(invoice)
        return Response(data)

    def perform_update(self, serializer):
        invoice = self.get_object()
        if invoice.status == Invoice.Status.APPROVED:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Cannot edit an approved invoice. (Rule CTL-004)")
        before = InvoiceDetailSerializer(invoice).data
        updated = serializer.save()
        _save_audit_event(updated, self.request.user, InvoiceAuditEvent.EventType.EDITED,
                          "Invoice fields updated", before=before,
                          after=InvoiceDetailSerializer(updated).data,
                          request=self.request)

    def perform_destroy(self, instance):
        """Soft delete: mark as deleted with audit trail (GDPR Article 17)."""
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.deleted_by = self.request.user
        instance.save(update_fields=['is_deleted', 'deleted_at', 'deleted_by', 'updated_at'])
        
        # Log deletion in audit trail
        _save_audit_event(
            instance, 
            self.request.user, 
            InvoiceAuditEvent.EventType.DELETED,
            f"Invoice soft-deleted by {self.request.user.full_name} (GDPR Article 17)",
            request=self.request
        )

    def get_queryset(self):
        return Invoice.objects.filter(
            organization=self.request.user.organization,
            is_deleted=False
        ).select_related(
            "validation",
            "approved_by",
            "duplicate_of",
        )


class InvoiceApproveView(APIView):
    """Approve or reject an invoice (Rule CTL-005: must have approver)."""
    permission_classes = [IsAuthenticated, IsSeniorAuditorOrAbove]

    @extend_schema(
        tags=["Invoices"],
        summary="Approve or reject an invoice",
        request={"type": "object", "properties": {
            "action": {"type": "string", "enum": ["approve", "reject"]},
            "reason": {"type": "string", "description": "Required for rejection"},
        }},
    )
    def post(self, request, pk):
        try:
            invoice = Invoice.objects.get(pk=pk, organization=request.user.organization)
        except Invoice.DoesNotExist:
            return Response({"error": _("Invoice not found.")}, status=404)

        action = request.data.get("action")
        if action not in ("approve", "reject"):
            return Response({"error": _("action must be 'approve' or 'reject'")}, status=400)

        before = {"status": invoice.status}

        if action == "approve":
            invoice.status      = Invoice.Status.APPROVED
            invoice.approved_by = request.user
            invoice.approved_at = timezone.now()
            event_type          = InvoiceAuditEvent.EventType.APPROVED
            msg                 = f"Approved by {request.user.full_name}"
        else:
            reason = request.data.get("reason", "")
            if not reason:
                return Response({"error": _("reason is required for rejection.")}, status=400)
            invoice.status          = Invoice.Status.REJECTED
            invoice.rejected_reason = reason
            event_type              = InvoiceAuditEvent.EventType.REJECTED
            msg                     = f"Rejected: {reason}"

        invoice.save()
        _save_audit_event(invoice, request.user, event_type, msg,
                          before=before, after={"status": invoice.status}, request=request)

        return Response({"invoice_id": str(invoice.id), "status": invoice.status, "message": msg})


class InvoiceRevalidateView(APIView):
    """Re-run all 30 validation rules on an existing invoice."""
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="Re-run all 30 validation rules on an existing invoice")
    def post(self, request, pk):
        try:
            invoice = Invoice.objects.get(pk=pk, organization=request.user.organization)
        except Invoice.DoesNotExist:
            return Response({"error": _("Invoice not found.")}, status=404)
        val_result = _run_invoice_revalidation(invoice, request.user, request=request)
        return Response(val_result)


class InvoiceManualReviewView(APIView):
    """Persist reviewer corrections and optional revalidation for an invoice."""

    permission_classes = [IsAuthenticated, IsSeniorAuditorOrAbove]
    authentication_classes = [JWTAuthentication, SessionAuthentication]

    @extend_schema(
        tags=["Invoices"],
        summary="Save manual review corrections for an invoice",
        request={
            "type": "object",
            "properties": {
                "corrections": {"type": "object"},
                "note": {"type": "string"},
                "revalidate": {"type": "boolean"},
            },
        },
    )
    def post(self, request, pk):
        try:
            invoice = Invoice.objects.get(pk=pk, organization=request.user.organization)
        except Invoice.DoesNotExist:
            return Response({"error": _("Invoice not found.")}, status=404)

        corrections = request.data.get("corrections") or {}
        if not isinstance(corrections, dict):
            return Response({"error": _("corrections must be an object.")}, status=400)

        note = str(request.data.get("note", "") or "").strip()
        should_revalidate = bool(request.data.get("revalidate"))

        before = {field: _serialize_review_value(getattr(invoice, field, "")) for field, _, _ in REVIEW_FIELD_META}
        applied = {}
        validation_errors = {}

        for field_name, _, _ in REVIEW_FIELD_META:
            if field_name not in corrections:
                continue
            try:
                coerced_value = _coerce_review_value(field_name, corrections.get(field_name))
            except ValueError as exc:
                validation_errors[field_name] = str(exc)
                continue
            if coerced_value is None and field_name in {"invoice_date", "due_date"}:
                setattr(invoice, field_name, None)
            elif coerced_value is None and field_name in {"subtotal", "vat_amount", "total_amount"}:
                setattr(invoice, field_name, Decimal("0"))
            else:
                setattr(invoice, field_name, coerced_value if coerced_value is not None else "")
            applied[field_name] = _serialize_review_value(coerced_value)

        if validation_errors:
            return Response({"errors": validation_errors}, status=400)

        extracted_data = dict(invoice.extracted_data or {})
        extracted_data["review"] = {
            "corrections": {**(extracted_data.get("review", {}).get("corrections", {})), **applied},
            "note": note,
            "reviewed_by": str(request.user.id),
            "reviewed_by_name": request.user.full_name,
            "reviewed_at": timezone.now().isoformat(),
        }
        invoice.extracted_data = extracted_data
        update_fields = [field for field in applied] + ["extracted_data", "updated_at"]
        invoice.save(update_fields=list(dict.fromkeys(update_fields)))

        after = {field: _serialize_review_value(getattr(invoice, field, "")) for field, _, _ in REVIEW_FIELD_META}
        _save_audit_event(
            invoice,
            request.user,
            InvoiceAuditEvent.EventType.EDITED,
            "Manual review corrections applied",
            before=before,
            after={"fields": after, "note": note, "applied": applied},
            request=request,
        )

        validation_result = None
        if should_revalidate:
            validation_result = _run_invoice_revalidation(invoice, request.user, request=request)

        return Response(
            {
                "invoice_id": str(invoice.id),
                "status": invoice.status,
                "applied_fields": applied,
                "review": _build_review_payload(invoice),
                "validation": validation_result,
            }
        )


class InvoiceBatchListView(generics.ListAPIView):
    serializer_class = InvoiceBatchSerializer
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="List upload batches")
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        return InvoiceBatch.objects.filter(organization=self.request.user.organization)


class InvoiceBatchDetailView(APIView):
    """Get batch details with all invoice summaries."""
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="Get batch details with invoice summaries")
    def get(self, request, pk):
        try:
            batch = InvoiceBatch.objects.get(pk=pk, organization=request.user.organization)
        except InvoiceBatch.DoesNotExist:
            return Response({"error": _("Batch not found.")}, status=404)

        invoices = Invoice.objects.filter(batch=batch, is_deleted=False).values(
            "id", "original_filename", "vendor_name", "total_amount", "currency",
            "invoice_date", "status", "risk_level", "is_duplicate", "ocr_confidence",
        )
        stats = Invoice.objects.filter(batch=batch, is_deleted=False).aggregate(
            total_amount=Sum("total_amount"),
            avg_score=Avg("ocr_confidence"),
            flagged=Count("id", filter=Q(status="flagged")),
            approved=Count("id", filter=Q(status="approved")),
            duplicates=Count("id", filter=Q(is_duplicate=True)),
            critical=Count("id", filter=Q(risk_level="critical")),
        )
        return Response({
            "batch": InvoiceBatchSerializer(batch).data,
            "audit_session_id": str(batch.audit_session_id) if batch.audit_session_id else None,
            "stats": stats,
            "invoices": list(invoices),
        })


class InvoiceDownloadView(APIView):
    """Securely download an invoice file for the current organization."""
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="Download an invoice file")
    def get(self, request, pk):
        try:
            invoice = Invoice.objects.get(
                pk=pk,
                organization=request.user.organization,
                is_deleted=False  # Exclude soft-deleted invoices
            )
        except Invoice.DoesNotExist:
            return Response({"error": _("Invoice not found.")}, status=404)

        if not invoice.file:
            return Response({"error": _("Invoice file is unavailable.")}, status=404)

        filename = invoice.original_filename or os.path.basename(invoice.file.name)
        return FileResponse(invoice.file.open("rb"), as_attachment=True, filename=filename)


# ─── Reports ─────────────────────────────────────────────────────────────────

class InvoiceRiskReportView(APIView):
    """Report: High-risk invoices."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Invoices"],
        summary="Risk report — flagged and high-risk invoices",
        parameters=[
            OpenApiParameter("date_from"), OpenApiParameter("date_to"),
            OpenApiParameter("risk_level", description="low|medium|high|critical"),
        ],
    )
    def get(self, request):
        org = request.user.organization
        qs = Invoice.objects.filter(organization=org)
        if v := request.query_params.get("date_from"):
            qs = qs.filter(invoice_date__gte=v)
        if v := request.query_params.get("date_to"):
            qs = qs.filter(invoice_date__lte=v)

        high_risk_filter = (
            Q(risk_level__in=["high", "critical"])
            | Q(risk_score__gte=60)
            | Q(status=Invoice.Status.FLAGGED)
        )
        if v := request.query_params.get("risk_level"):
            normalized = str(v).strip().lower()
            if normalized in {"high_risk", "high", "critical"}:
                qs = qs.filter(high_risk_filter)
            else:
                qs = qs.filter(risk_level=normalized)
        else:
            qs = qs.filter(high_risk_filter)

        try:
            page_size = min(max(int(request.query_params.get("page_size", 20) or 20), 1), 100)
        except (TypeError, ValueError):
            page_size = 20

        stats = qs.aggregate(
            count=Count("id"), total=Sum("total_amount"), avg_risk=Avg("risk_score")
        )
        items = list(
            qs.order_by("-risk_score", "-invoice_date").values(
                "id", "invoice_number", "vendor_name", "total_amount", "currency",
                "invoice_date", "risk_level", "risk_score", "ai_summary",
                "is_duplicate", "status", "ai_recommendations",
            )[:page_size]
        )
        return Response({
            "report_type": "risk_report",
            "generated_at": timezone.now().isoformat(),
            "count": stats.get("count") or 0,
            "stats": stats,
            "results": items,
            "invoices": items,
        })


class DuplicateInvoiceReportView(APIView):
    """Report: Duplicate invoices."""
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="Duplicate invoices report")
    def get(self, request):
        org = request.user.organization
        qs = Invoice.objects.filter(organization=org, is_duplicate=True).order_by("-created_at")
        invoices = qs.values(
            "id", "invoice_number", "vendor_name", "total_amount", "currency",
            "invoice_date", "status", "duplicate_of_id", "created_at",
        )
        return Response({
            "report_type": "duplicate_report",
            "generated_at": timezone.now().isoformat(),
            "total_duplicates": qs.count(),
            "invoices": list(invoices),
        })


class VendorRiskReportView(APIView):
    """Report: Vendor intelligence and supplier risk analysis."""
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="Vendor intelligence risk analysis report")
    def get(self, request):
        org = request.user.organization
        if not org:
            return Response({"report_type": "vendor_risk_report", "summary": {}, "vendors": []})

        vendors = VendorProfile.objects.filter(organization=org).order_by("-risk_score", "-total_amount")
        if search := request.query_params.get("search"):
            vendors = vendors.filter(vendor_name__icontains=search)
        if risk_tier := request.query_params.get("risk_tier"):
            vendors = vendors.filter(risk_tier=risk_tier)

        limit = min(int(request.query_params.get("limit", 100) or 100), 250)
        summary = vendors.aggregate(
            total_vendors=Count("id"),
            high_risk_vendors=Count("id", filter=Q(risk_tier__in=["high", "blocked"])),
            blocked_vendors=Count("id", filter=Q(risk_tier="blocked")),
            unapproved_vendors=Count("id", filter=Q(is_approved=False)),
            avg_risk_score=Avg("risk_score"),
            avg_frequency_30d=Avg("transaction_frequency_30d"),
            compliance_issue_total=Sum("compliance_issue_count"),
        )
        top_risky = vendors[:5]

        return Response({
            "report_type": "vendor_risk_report",
            "generated_at": timezone.now().isoformat(),
            "summary": {
                "total_vendors": summary.get("total_vendors") or 0,
                "high_risk_vendors": summary.get("high_risk_vendors") or 0,
                "blocked_vendors": summary.get("blocked_vendors") or 0,
                "unapproved_vendors": summary.get("unapproved_vendors") or 0,
                "avg_risk_score": round(float(summary.get("avg_risk_score") or 0), 2),
                "avg_frequency_30d": round(float(summary.get("avg_frequency_30d") or 0), 2),
                "compliance_issue_total": summary.get("compliance_issue_total") or 0,
            },
            "top_risky": VendorProfileSerializer(top_risky, many=True).data,
            "vendors": VendorProfileSerializer(vendors[:limit], many=True).data,
        })


class VendorListView(APIView):
    """Compatibility vendor list endpoint for dashboard and API consumers."""

    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="List vendor intelligence profiles for the current organization")
    def get(self, request):
        org = request.user.organization
        if not org:
            return Response({"count": 0, "results": []})

        vendors = VendorProfile.objects.filter(organization=org).order_by("-risk_score", "-total_amount", "vendor_name")
        if search := request.query_params.get("search"):
            vendors = vendors.filter(vendor_name__icontains=search)
        if risk_tier := request.query_params.get("risk_tier"):
            vendors = vendors.filter(risk_tier=risk_tier)
        if approved := request.query_params.get("is_approved"):
            if approved.lower() in {"true", "1", "yes"}:
                vendors = vendors.filter(is_approved=True)
            elif approved.lower() in {"false", "0", "no"}:
                vendors = vendors.filter(is_approved=False)

        return Response({"count": vendors.count(), "results": VendorProfileSerializer(vendors[:100], many=True).data})


class HighRiskVendorListView(APIView):
    """List only high-risk vendor intelligence profiles for the current organization."""

    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="List high-risk vendors for the current organization")
    def get(self, request):
        org = request.user.organization
        if not org:
            return Response({"count": 0, "results": []})

        vendors = VendorProfile.objects.filter(
            organization=org,
            risk_tier__in=["high", "blocked"],
        ).order_by("-risk_score", "-updated_at")
        return Response({"count": vendors.count(), "results": VendorProfileSerializer(vendors[:100], many=True).data})


class VendorDetailView(APIView):
    """Detailed Vendor Intelligence profile for a single tenant-scoped vendor."""

    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="Get vendor intelligence details for a single supplier")
    def get(self, request, pk):
        org = request.user.organization
        vendor = VendorProfile.objects.filter(organization=org, pk=pk).first()
        if not vendor:
            return Response({"detail": "Not found."}, status=404)

        payload = VendorProfileSerializer(vendor).data
        payload["recent_documents"] = (vendor.transaction_history or {}).get("recent_documents", [])
        payload["compliance_summary"] = vendor.compliance_history or {}
        return Response(payload)


class SpendAnalysisReportView(APIView):
    """Report: Spend analysis by vendor and category."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Invoices"],
        summary="Spend analysis report",
        parameters=[
            OpenApiParameter("date_from"), OpenApiParameter("date_to"),
        ],
    )
    def get(self, request):
        org = request.user.organization
        qs = Invoice.objects.filter(organization=org)
        if v := request.query_params.get("date_from"):
            qs = qs.filter(invoice_date__gte=v)
        if v := request.query_params.get("date_to"):
            qs = qs.filter(invoice_date__lte=v)

        # By vendor
        by_vendor = qs.values("vendor_name").annotate(
            total=Sum("total_amount"), count=Count("id"),
            avg=Avg("total_amount"), flagged=Count("id", filter=Q(is_duplicate=True)),
        ).order_by("-total")[:20]

        # By currency
        by_currency = qs.values("currency").annotate(
            total=Sum("total_amount"), count=Count("id"),
        ).order_by("-total")

        # Monthly trend
        from django.db.models.functions import TruncMonth
        monthly = qs.annotate(month=TruncMonth("invoice_date")).values("month").annotate(
            total=Sum("total_amount"), count=Count("id"),
            flagged=Count("id", filter=Q(risk_level__in=["high", "critical"])),
        ).order_by("month")

        # Overall stats
        stats = qs.aggregate(
            grand_total=Sum("total_amount"),
            total_vat=Sum("vat_amount"),
            total_invoices=Count("id"),
            avg_invoice=Avg("total_amount"),
            flagged_total=Sum("total_amount", filter=Q(risk_level__in=["high","critical"])),
        )

        return Response({
            "report_type": "spend_analysis",
            "generated_at": timezone.now().isoformat(),
            "overall": stats,
            "by_vendor": list(by_vendor),
            "by_currency": list(by_currency),
            "monthly_trend": [
                {**m, "month": str(m["month"])[:7] if m["month"] else None}
                for m in monthly
            ],
        })


class ValidationRulesListView(APIView):
    """List all 30 validation rules with descriptions."""
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Invoices"], summary="List all 30 invoice audit rules")
    def get(self, request):
        groups = {
            "Group 1 — Invoice Validation": {k: v for k, v in RULES.items() if k.startswith("INV")},
            "Group 2 — Duplicate Detection": {k: v for k, v in RULES.items() if k.startswith("DUP")},
            "Group 3 — VAT Validation": {k: v for k, v in RULES.items() if k.startswith("VAT")},
            "Group 4 — Anomaly Detection": {k: v for k, v in RULES.items() if k.startswith("ANO")},
            "Group 5 — Financial Controls": {k: v for k, v in RULES.items() if k.startswith("CTL")},
            "Group 6 — Document Quality": {k: v for k, v in RULES.items() if k.startswith("DOC")},
        }
        return Response({"total_rules": TOTAL_RULES, "rule_groups": groups})
