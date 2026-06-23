"""Trial Balance import parser + validation service (TADGEEG-FIN-AUDIT-1A).

Reads an uploaded CSV/XLSX trial balance, maps client headers (Arabic/English)
to canonical columns, normalises amounts, validates row- and import-level, and
persists ``TrialBalanceRow`` staging rows + a summary on the ``TrialBalanceImport``.

It NEVER writes to ``apps.ledger``. It only touches the two staging tables.
"""
from __future__ import annotations

import csv
import hashlib
import io
from decimal import Decimal, InvalidOperation
from typing import IO

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from apps.audit.trial_balance_models import TrialBalanceImport, TrialBalanceRow

# Balance tolerance: TBs commonly round to 2 dp; allow a small slack.
BALANCE_TOLERANCE = Decimal("0.01")
_ZERO = Decimal("0")

# Canonical column → accepted header aliases (lower-cased, whitespace-collapsed).
HEADER_ALIASES: dict[str, list[str]] = {
    "account_code": [
        "account code", "account_code", "account no", "account number",
        "acct code", "code", "gl code", "كود الحساب", "رقم الحساب", "الحساب",
    ],
    "account_name": [
        "account name", "account_name", "account title", "name", "description",
        "اسم الحساب", "الوصف", "البيان",
    ],
    "account_type": [
        "account type", "account_type", "account category", "category", "type",
        "نوع الحساب", "تصنيف الحساب", "التصنيف",
    ],
    "opening_debit":  ["opening debit", "opening_debit", "رصيد افتتاحي مدين", "افتتاحي مدين"],
    "opening_credit": ["opening credit", "opening_credit", "رصيد افتتاحي دائن", "افتتاحي دائن"],
    "period_debit":   ["period debit", "period_debit", "movement debit", "حركة مدين", "مدين الفترة"],
    "period_credit":  ["period credit", "period_credit", "movement credit", "حركة دائن", "دائن الفترة"],
    "closing_debit":  ["closing debit", "closing_debit", "الرصيد المدين", "ختامي مدين", "رصيد مدين"],
    "closing_credit": ["closing credit", "closing_credit", "الرصيد الدائن", "ختامي دائن", "رصيد دائن"],
    "debit":  ["debit", "dr", "مدين"],
    "credit": ["credit", "cr", "دائن"],
    "balance": ["balance", "closing balance", "net balance", "الرصيد", "الرصيد الختامي"],
    "currency": ["currency", "ccy", "العملة"],
}

SUPPORTED_FORMATS = {"csv", "xlsx", "xls"}


class TrialBalanceImportError(Exception):
    """Raised for unrecoverable problems (bad format, unreadable file)."""


def _norm_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _build_column_mapping(headers: list[str]) -> dict[str, str]:
    """Map canonical key → original header string, using the alias table."""
    mapping: dict[str, str] = {}
    for original in headers:
        key = _norm_header(original)
        for canonical, aliases in HEADER_ALIASES.items():
            if canonical in mapping:
                continue
            if key == canonical or key in aliases:
                mapping[canonical] = original
                break
    return mapping


def _normalize_amount(value) -> Decimal | None:
    """Parse a cell into a Decimal. Returns None if non-numeric (not blank)."""
    if value is None:
        return _ZERO
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip()
    if text == "":
        return _ZERO
    negative = False
    if text.startswith("(") and text.endswith(")"):  # accounting negatives
        negative = True
        text = text[1:-1]
    # Strip thousands separators, spaces, and common currency symbols.
    for ch in (",", " ", " ", "ر.س", "SAR", "$", "ريال"):
        text = text.replace(ch, "")
    if text in ("", "-"):
        return _ZERO
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def _read_csv(file_obj: IO[bytes]) -> list[dict]:
    raw = file_obj.read()
    if isinstance(raw, bytes):
        # utf-8-sig strips a BOM that Excel-exported CSVs often carry.
        raw = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(raw))
    return [dict(r) for r in reader]


def _read_xlsx(file_obj: IO[bytes]) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl is a project dep
        raise TrialBalanceImportError("openpyxl is required to read .xlsx files") from exc
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for values in rows[1:]:
        out.append({headers[i]: values[i] if i < len(values) else None
                    for i in range(len(headers))})
    return out


def _resolve_amounts(row: dict, mapping: dict[str, str]) -> tuple[dict, list[str]]:
    """Resolve closing debit/credit for a row using whatever columns exist."""
    errors: list[str] = []

    def cell(canonical):
        col = mapping.get(canonical)
        return row.get(col) if col else None

    parsed: dict[str, Decimal] = {}
    for key in ("opening_debit", "opening_credit", "period_debit",
                "period_credit", "closing_debit", "closing_credit",
                "debit", "credit", "balance"):
        if key in mapping:
            amt = _normalize_amount(cell(key))
            if amt is None:
                errors.append(f"non-numeric value in '{mapping[key]}'")
                amt = _ZERO
            parsed[key] = amt

    # Priority: explicit closing d/c → debit/credit → signed balance.
    if "closing_debit" in parsed or "closing_credit" in parsed:
        cd = parsed.get("closing_debit", _ZERO)
        cc = parsed.get("closing_credit", _ZERO)
    elif "debit" in parsed or "credit" in parsed:
        cd = parsed.get("debit", _ZERO)
        cc = parsed.get("credit", _ZERO)
    elif "balance" in parsed:
        bal = parsed["balance"]
        cd, cc = (bal, _ZERO) if bal >= 0 else (_ZERO, -bal)
    else:
        cd = cc = _ZERO
        errors.append("no debit/credit/balance column found")

    return {
        "opening_debit":  parsed.get("opening_debit", _ZERO),
        "opening_credit": parsed.get("opening_credit", _ZERO),
        "period_debit":   parsed.get("period_debit", _ZERO),
        "period_credit":  parsed.get("period_credit", _ZERO),
        "closing_debit":  cd,
        "closing_credit": cc,
    }, errors


def parse_and_validate(import_batch: TrialBalanceImport, file_obj: IO[bytes] | None = None):
    """Parse + validate the file, persist rows, update the import summary.

    Returns the refreshed ``import_batch``. Raises ``TrialBalanceImportError``
    for unrecoverable issues (unsupported format / unreadable file) and
    ``ValidationError`` if the import crosses tenants.
    """
    # ── Tenant safety: import org must equal engagement org. ────────────
    if import_batch.engagement.organization_id != import_batch.organization_id:
        raise ValidationError(
            "cross-tenant import denied: organization != engagement.organization"
        )

    fmt = (import_batch.source_format or "").lower()
    if fmt not in SUPPORTED_FORMATS:
        import_batch.status = TrialBalanceImport.Status.VALIDATION_FAILED
        import_batch.errors = [f"unsupported format: {import_batch.source_format!r}"]
        import_batch.save(update_fields=["status", "errors", "updated_at"])
        raise TrialBalanceImportError(f"unsupported format: {import_batch.source_format!r}")

    if file_obj is None:
        if not import_batch.uploaded_file:
            raise TrialBalanceImportError("no file attached to import")
        file_obj = import_batch.uploaded_file.open("rb")

    # Capture integrity hash (mirrors the documents-app pattern).
    data = file_obj.read()
    if isinstance(data, str):
        data = data.encode("utf-8")
    import_batch.file_sha256 = hashlib.sha256(data).hexdigest()

    import_batch.status = TrialBalanceImport.Status.VALIDATING
    import_batch.save(update_fields=["status", "file_sha256", "updated_at"])

    buf = io.BytesIO(data)
    try:
        records = _read_csv(buf) if fmt == "csv" else _read_xlsx(buf)
    except TrialBalanceImportError:
        raise
    except Exception as exc:  # malformed file
        import_batch.status = TrialBalanceImport.Status.VALIDATION_FAILED
        import_batch.errors = [f"could not parse file: {exc}"]
        import_batch.save(update_fields=["status", "errors", "updated_at"])
        raise TrialBalanceImportError(str(exc)) from exc

    headers = list(records[0].keys()) if records else []
    mapping = _build_column_mapping(headers)

    import_errors: list[str] = []
    import_warnings: list[str] = []
    if "account_code" not in mapping:
        import_errors.append("required column 'account_code' not found")
    if "account_name" not in mapping:
        import_warnings.append("no 'account_name' column detected")
    has_amount_cols = any(k in mapping for k in
                          ("closing_debit", "closing_credit", "debit", "credit", "balance"))
    if not has_amount_cols:
        import_errors.append("no debit/credit/balance columns found")

    # Period sanity.
    if (import_batch.period_start and import_batch.period_end
            and import_batch.period_start > import_batch.period_end):
        import_errors.append("period_start is after period_end")

    total_debit = _ZERO
    total_credit = _ZERO
    valid_count = 0
    invalid_count = 0
    row_number = 0
    new_rows: list[TrialBalanceRow] = []

    code_col = mapping.get("account_code")
    name_col = mapping.get("account_name")
    type_col = mapping.get("account_type")
    ccy_col = mapping.get("currency")

    for rec in records:
        # Skip fully-blank rows (don't count them).
        if not any(str(v).strip() for v in rec.values() if v is not None):
            continue
        row_number += 1

        row_errors: list[str] = []
        account_code = str(rec.get(code_col, "") or "").strip() if code_col else ""
        if not account_code:
            row_errors.append("missing account_code")

        amounts, amt_errors = _resolve_amounts(rec, mapping)
        row_errors.extend(amt_errors)

        cd, cc = amounts["closing_debit"], amounts["closing_credit"]
        net_movement = amounts["period_debit"] - amounts["period_credit"]
        closing_balance = cd - cc
        is_valid = not row_errors

        if is_valid:
            valid_count += 1
            total_debit += cd
            total_credit += cc
        else:
            invalid_count += 1

        new_rows.append(TrialBalanceRow(
            import_batch=import_batch,
            engagement=import_batch.engagement,
            organization=import_batch.organization,
            row_number=row_number,
            account_code=account_code[:64],
            account_name=(str(rec.get(name_col, "") or "").strip()[:255] if name_col else ""),
            account_type=(str(rec.get(type_col, "") or "").strip()[:64] if type_col else ""),
            opening_debit=amounts["opening_debit"],
            opening_credit=amounts["opening_credit"],
            period_debit=amounts["period_debit"],
            period_credit=amounts["period_credit"],
            closing_debit=cd,
            closing_credit=cc,
            net_movement=net_movement,
            closing_balance=closing_balance,
            currency=(str(rec.get(ccy_col, "") or "").strip()[:8] if ccy_col else import_batch.currency),
            is_valid=is_valid,
            validation_errors=row_errors,
            raw_data={str(k): (str(v) if v is not None else None) for k, v in rec.items()},
        ))

    difference = total_debit - total_credit
    is_balanced = abs(difference) <= BALANCE_TOLERANCE

    import_batch.row_count = row_number
    import_batch.valid_row_count = valid_count
    import_batch.invalid_row_count = invalid_count
    import_batch.total_debit = total_debit
    import_batch.total_credit = total_credit
    import_batch.difference = difference
    import_batch.is_balanced = is_balanced
    import_batch.column_mapping = mapping
    import_batch.errors = import_errors
    import_batch.warnings = import_warnings
    import_batch.validation_summary = {
        "row_count": row_number,
        "valid_row_count": valid_count,
        "invalid_row_count": invalid_count,
        "total_debit": str(total_debit),
        "total_credit": str(total_credit),
        "difference": str(difference),
        "is_balanced": is_balanced,
        "balance_tolerance": str(BALANCE_TOLERANCE),
        "columns_detected": list(mapping.keys()),
    }

    # Status: hard errors or any invalid row → validation_failed; else validated.
    if import_errors or invalid_count:
        import_batch.status = TrialBalanceImport.Status.VALIDATION_FAILED
    else:
        import_batch.status = TrialBalanceImport.Status.VALIDATED
        import_batch.validated_at = timezone.now()

    # Persist rows + summary together so a batch is never half-written.
    # (This atomic block is scoped to persistence only; the early
    # validation-failure saves above are committed outside it and survive.)
    with transaction.atomic():
        import_batch.rows.all().delete()  # idempotent re-validation
        TrialBalanceRow.objects.bulk_create(new_rows)
        import_batch.save()
    return import_batch
