"""Tadgeeg AI frontend page views."""

import logging
import secrets
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from urllib.parse import urlencode

from django.conf import settings as django_settings
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.dateparse import parse_date
from django.utils.translation import gettext as _, get_language
from django.views.decorators.csrf import ensure_csrf_cookie

from apps.authentication.forms import EmailOTPResendForm, EmailOTPVerifyForm
from apps.authentication.serializers import LoginSerializer, RegisterSerializer
from django_countries import countries

from apps.authentication.models import Organization

logger = logging.getLogger("frontend.pages")
from apps.leads.attribution import remember_campaign
from apps.leads.models import TrialLeadProfile
from apps.authentication.services.email_otp import (
    EmailOTPError,
    clear_pending_verification,
    complete_verified_login,
    get_challenge_state,
    get_latest_pending_challenge,
    get_pending_verification_user,
    has_pending_verification,
    issue_email_otp,
    mask_email_address,
    resend_email_otp,
    verify_email_otp,
)
from apps.authentication.services.google_oauth import (
    GoogleOAuthError,
    build_google_oauth_authorization_url,
    exchange_google_code_for_tokens,
    fetch_google_user_profile,
    get_or_create_local_user_from_google_profile,
    is_google_oauth_configured,
)


GOOGLE_OAUTH_STATE_SESSION_KEY = "google_oauth_state"
POST_LOGIN_TOKENS_SESSION_KEY = "post_login_tokens"


def _consume_post_login_tokens(request):
    if not getattr(request.user, "is_authenticated", False):
        return None
    tokens = request.session.pop(POST_LOGIN_TOKENS_SESSION_KEY, None)
    if tokens is not None:
        request.session.modified = True
    return tokens


def _google_auth_error_message(code: str) -> str:
    return {
        "missing_code": _("Unable to complete Google sign-in. Please try again."),
        "token_exchange_failed": _("Unable to validate the Google session right now. Please try again shortly."),
        "invalid_client": _("Google OAuth settings are currently invalid. Check the Client ID and Client Secret."),
        "invalid_grant": _("The Google sign-in code has expired or was already used. Start the flow again."),
        "redirect_uri_mismatch": _("The Google redirect URI does not match the current configuration. Verify GOOGLE_REDIRECT_URI."),
        "userinfo_failed": _("Unable to fetch the Google account profile. Please try again."),
        "no_email": _("The Google account does not provide an email address for registration."),
        "invalid_state": _("The Google sign-in session has expired. Please start again."),
        "inactive_user": _("This account is currently inactive. Contact your administrator."),
        "oauth_not_configured": _("Google sign-in is not configured right now."),
    }.get(code, _("Unable to sign in with Google right now. Please try again."))


def _redirect_to_login_with_auth_error(code: str):
    return redirect(f"{reverse('frontend:login')}?{urlencode({'auth_error': code})}")


def _ctx(request, active="dashboard", **extra):
    pending_count = 0
    try:
        from apps.invoices.models import Invoice

        organization = getattr(request.user, "organization", None)
        if organization:
            pending_count = Invoice.objects.filter(organization=organization, status="flagged").count()
    except Exception:
        pass
    return {
        "pending_count": pending_count,
        "active": active,
        "bootstrap_tokens": _consume_post_login_tokens(request),
        "can_manage_users": getattr(request.user, "can_manage_users", False),
        "can_generate_reports": getattr(request.user, "can_generate_reports", False),
        **extra,
    }


def _paginate(qs, request, per_page=25):
    """Return (page_obj, page_kwargs) for any list view.

    `page_kwargs` carries the keys list templates expect:
      - rows / invoices: page-bound iterable (so `{% for r in rows %}` works)
      - page_obj:        Django Page object for pagination controls
      - paginator:       Paginator (also exposed for partials that want it)
      - per_page:        the size we used (so the template can show "showing N of M")
      - querystring:     the request querystring with `page` removed
                          (handy for keeping filters when changing page)
      - total_count:     paginator.count
    """
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    paginator = Paginator(qs, per_page)
    page_num = request.GET.get("page") or 1
    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages or 1)

    # Build a querystring without the `page` parameter so pagination links
    # preserve other filters (q/status/risk).
    qd = request.GET.copy()
    qd.pop("page", None)
    qs_str = qd.urlencode()

    return {
        "page_obj":     page_obj,
        "paginator":    paginator,
        "per_page":     per_page,
        "querystring":  qs_str,
        "total_count":  paginator.count,
    }


def _report_types():
    return [
        {
            "type": "invoice_audit",
            "lang": "ar",
            "label": _("Invoice Audit"),
            "desc": _("30 rules + risks + vendors"),
            "icon": "file-check-2",
            "bg": "bg-blue-100 dark:bg-blue-900/30",
            "color": "text-blue-600 dark:text-blue-400",
        },
        {
            "type": "executive_summary",
            "lang": "ar",
            "label": _("Executive Summary"),
            "desc": _("A concise view for leadership"),
            "icon": "bar-chart-3",
            "bg": "bg-violet-100 dark:bg-violet-900/30",
            "color": "text-violet-600 dark:text-violet-400",
        },
        {
            "type": "risk_assessment",
            "lang": "ar",
            "label": _("Risk Assessment"),
            "desc": _("High-risk invoices and vendors"),
            "icon": "shield-alert",
            "bg": "bg-red-100 dark:bg-red-900/30",
            "color": "text-red-600 dark:text-red-400",
        },
        {
            "type": "vendor_analysis",
            "lang": "ar",
            "label": _("Vendor Analysis"),
            "desc": _("Spending patterns and risk indicators"),
            "icon": "building-2",
            "bg": "bg-emerald-100 dark:bg-emerald-900/30",
            "color": "text-emerald-600 dark:text-emerald-400",
        },
        # ── Document-level audit report cards (one per document type) ──────────
        {
            "type": "document_audit_sales_invoice",
            "lang": "ar",
            "label": _("Sales Invoice Audit"),
            "desc": _("VAT, totals, and ZATCA compliance"),
            "icon": "receipt",
            "bg": "bg-indigo-100 dark:bg-indigo-900/30",
            "color": "text-indigo-600 dark:text-indigo-400",
        },
        {
            "type": "document_audit_purchase_order",
            "lang": "ar",
            "label": _("Purchase Order Audit"),
            "desc": _("Vendor, budget, and approval flow"),
            "icon": "package",
            "bg": "bg-orange-100 dark:bg-orange-900/30",
            "color": "text-orange-600 dark:text-orange-400",
        },
        {
            "type": "document_audit_bank_statement",
            "lang": "ar",
            "label": _("Bank Statement Audit"),
            "desc": _("Transactions, reconciliation, and Benford analysis"),
            "icon": "landmark",
            "bg": "bg-teal-100 dark:bg-teal-900/30",
            "color": "text-teal-600 dark:text-teal-400",
        },
        {
            "type": "document_audit_payroll",
            "lang": "ar",
            "label": _("Payroll Audit"),
            "desc": _("Salaries, deductions, and insurance compliance"),
            "icon": "users",
            "bg": "bg-pink-100 dark:bg-pink-900/30",
            "color": "text-pink-600 dark:text-pink-400",
        },
        {
            "type": "document_audit_expense_report",
            "lang": "ar",
            "label": _("Expense Report Audit"),
            "desc": _("Policy compliance, receipts, and categories"),
            "icon": "credit-card",
            "bg": "bg-amber-100 dark:bg-amber-900/30",
            "color": "text-amber-600 dark:text-amber-400",
        },
        {
            "type": "document_audit_vat_return",
            "lang": "ar",
            "label": _("VAT Return Audit"),
            "desc": _("Input/output tax and reconciliation"),
            "icon": "percent",
            "bg": "bg-cyan-100 dark:bg-cyan-900/30",
            "color": "text-cyan-600 dark:text-cyan-400",
        },
        {
            "type": "document_audit_fixed_asset",
            "lang": "ar",
            "label": _("Fixed Asset Audit"),
            "desc": _("Depreciation, disposal, and register checks"),
            "icon": "hard-drive",
            "bg": "bg-stone-100 dark:bg-stone-900/30",
            "color": "text-stone-600 dark:text-stone-400",
        },
        {
            "type": "document_audit_sales_receipt",
            "lang": "ar",
            "label": _("Sales Receipt Audit"),
            "desc": _("Payment, cash limits, and ZATCA QR"),
            "icon": "scroll-text",
            "bg": "bg-fuchsia-100 dark:bg-fuchsia-900/30",
            "color": "text-fuchsia-600 dark:text-fuchsia-400",
        },
    ]


#: Countries shown at the top of the registration dropdown. Derived from the
#: billing enum so the two can never drift apart. These are the markets the
#: product bills in; the rest of the ISO list is still selectable below them.
_GCC_CODES = set(Organization.Country.values)


def _public_ctx(request, **extra):
    # Campaign attribution is captured on the FIRST public page a visitor sees
    # (landing, pricing, auth portal...), not at /register/. By the time the
    # form is submitted the utm parameters are several navigations in the past,
    # so they are stashed in the session here. First value wins — see
    # apps.leads.attribution.remember_campaign.
    remember_campaign(request)

    auth_error = (request.GET.get("auth_error") or "").strip()
    return {
        # Registration dropdowns (§A.1/§A.2/§A.3). Sourced from the same
        # choices the serializer validates against, so the form can never drift
        # from what the server accepts.
        #
        # Full ISO list, NOT Organization.Country: the latter is the billing
        # jurisdiction (paired with a currency, GCC-only). Using it here locked
        # every non-GCC prospect out of registration entirely.
        "country_choices": list(countries),
        # GCC members shown first — they are the overwhelming majority of
        # registrants, and 249 options with no ordering is a usability problem.
        "priority_country_choices": [
            (code, name) for code, name in countries if code in _GCC_CODES
        ],
        "primary_benefit_choices": TrialLeadProfile.PrimaryBenefit.choices,
        "employee_count_choices": TrialLeadProfile.EmployeeCount.choices,
        "heard_about_choices": TrialLeadProfile.HeardAbout.choices,
        "google_client_id": getattr(django_settings, "GOOGLE_CLIENT_ID", ""),
        "google_oauth_enabled": is_google_oauth_configured(),
        "auth_error": auth_error,
        "auth_error_message": _google_auth_error_message(auth_error) if auth_error else "",
        **extra,
    }


def _post_auth_redirect(user):
    if not getattr(user, "is_email_verified", False):
        return "/verify-email/"
    # After email verification, route through billing onboarding when the
    # org has no usable subscription. Superusers + staff bypass — they
    # use the dashboard for support work.
    if not (user.is_superuser or user.is_staff):
        org = getattr(user, "organization", None)
        if org is not None:
            from apps.billing.services.quota_service import QuotaService
            if QuotaService().get_active_subscription(org) is None:
                return "/billing/plans/"
    return "/dashboard/"


def _otp_error_status(error: EmailOTPError) -> int:
    if error.code in {"resend_cooldown", "resend_limit", "attempts_exceeded"}:
        return 429
    if error.code == "send_failed":
        return 503
    if error.code == "already_verified":
        return 409
    return 400


def _otp_pending_payload(user, challenge, *, sent: bool):
    state = get_challenge_state(challenge)
    return {
        "success": True,
        "requires_verification": True,
        "redirect": "/verify-email/",
        "masked_email": mask_email_address(user.email),
        "message": (
            _("A verification code has been sent to your email address.")
            if sent
            else _("An active verification code already exists. Please check your email.")
        ),
        "otp_expires_in_seconds": state["expires_in"],
        "resend_cooldown_seconds": state["resend_available_in"],
        "attempts_remaining": state["attempts_remaining"],
    }


def _verified_login_payload(request, user):
    return {
        "success": True,
        "redirect": _post_auth_redirect(user),
        "tokens": complete_verified_login(request, user),
    }


def _stringify_error(value):
    if isinstance(value, list):
        return str(value[0]) if value else ""
    return str(value)


def _first_error(errors, *, flow="generic"):
    if not errors:
        return _("Unable to complete the request right now.")

    if isinstance(errors, dict):
        if flow == "login":
            if "email" in errors:
                first_email_error = _stringify_error(errors["email"])
                lowered = first_email_error.lower()
                if "valid email" in lowered:
                    return _("The email address is invalid.")
                if "no account found" in lowered or "not found" in lowered:
                    return _("No account was found for this email address.")
                return _("The email address or password is incorrect.")
            if "non_field_errors" in errors:
                first_non_field_error = _stringify_error(errors["non_field_errors"])
                lowered = first_non_field_error.lower()
                if "locked" in lowered:
                    return _("The account has been temporarily locked. Please try again later.")
                if "inactive" in lowered:
                    return _("This account is currently inactive.")
                return _("The email address or password is incorrect.")
        register_email_error = errors.get("email")
        if register_email_error:
            first_email_error = _stringify_error(register_email_error)
            lowered = first_email_error.lower()
            if "valid email" in lowered:
                return _("The email address is invalid.")
            return _("This email address is already in use.")
        if "email" in errors:
            return _("This email address is already in use.") if errors["email"] else _("The email address is invalid.")
        if "password" in errors:
            first_password_error = _stringify_error(errors["password"])
            if "match" in str(first_password_error).lower():
                return _("Passwords do not match.")
            return _("Please review the password and try again.")
        if "full_name" in errors:
            return _("Full name is required.")

        first_key = next(iter(errors))
        first_value = errors[first_key]
        return _stringify_error(first_value)

    if isinstance(errors, list) and errors:
        return str(errors[0])

    return str(errors)


def _first_present(*values):
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            stripped = value.strip()
            if stripped:
                return stripped
            continue
        return value
    return None


def _to_decimal(value):
    if value in (None, "", False):
        return None
    try:
        decimal_value = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError, TypeError, AttributeError):
        return None
    return decimal_value


def _first_non_zero_amount(*values):
    for value in values:
        decimal_value = _to_decimal(value)
        if decimal_value is not None and decimal_value > 0:
            return decimal_value
    return None


def _to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _coerce_date(value):
    if not value:
        return None
    if hasattr(value, "isoformat"):
        return value

    text = str(value).strip()
    parsed = parse_date(text)
    if parsed:
        return parsed

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _clean_ai_summary(summary):
    cleaned = str(summary or "").strip()
    if not cleaned:
        return ""

    lowered = cleaned.lower()
    if "openai unavailable" in lowered or "openai not configured" in lowered:
        return _("Data was extracted automatically using local OCR. Some fields may still need human review.")
    return cleaned


def _risk_band(score):
    numeric_score = max(0.0, min(100.0, _to_float(score)))
    if numeric_score >= 70:
        return {
            "key": "high",
            "label_ar": _("High Risk"),
            "label_en": "High Risk",
            "badge_class": "badge-high",
            "text_class": "text-red-600 dark:text-red-400",
            "ring_color": "#ef4444",
        }
    if numeric_score >= 40:
        return {
            "key": "medium",
            "label_ar": _("Medium Risk"),
            "label_en": "Medium Risk",
            "badge_class": "badge-medium",
            "text_class": "text-amber-600 dark:text-amber-400",
            "ring_color": "#f59e0b",
        }
    return {
        "key": "low",
        "label_ar": _("Low Risk"),
        "label_en": "Low Risk",
        "badge_class": "badge-low",
        "text_class": "text-emerald-600 dark:text-emerald-400",
        "ring_color": "#22c55e",
    }


def _status_meta(status):
    # (label_en, label_ar, badge_class)
    mapping = {
        "approved":   (_("Approved"),   "تمت الموافقة",  "status-approved"),
        "rejected":   (_("Rejected"),   "مرفوض",         "status-rejected"),
        "flagged":    (_("Needs Review"), "تحتاج مراجعة", "status-flagged"),
        "validated":  (_("Validated"),  "تم التحقق",     "status-validated"),
        "processing": (_("Processing"), "قيد المعالجة",  "status-processing"),
        "pending":    (_("Pending"),    "قيد الانتظار",  "status-pending"),
    }
    label_en, label_ar, badge = mapping.get(
        status, (status or _("Unknown"), status or "", "status-pending")
    )
    return label_en, label_ar, badge


def _build_invoice_display(invoice):
    extracted = invoice.extracted_data or {}
    validation = getattr(invoice, "validation", None)
    text_fallback = {}

    if invoice.raw_text:
        try:
            from core.services.invoice_ai_service import _fallback_extraction

            text_fallback = _fallback_extraction(invoice.raw_text)
        except Exception:
            text_fallback = {}

    audit_score = _to_float(getattr(validation, "validation_score", 0) or extracted.get("validation_score") or 0)
    risk_score = _to_float(invoice.risk_score, 0.0)
    if risk_score <= 0 and audit_score > 0:
        risk_score = round(max(0.0, min(100.0, 100.0 - audit_score)), 2)

    risk = _risk_band(risk_score)
    effective_status = invoice.status
    if invoice.status not in {"approved", "rejected"} and risk_score >= 70:
        effective_status = "flagged"
    status_label, status_label_ar, status_badge_class = _status_meta(effective_status)

    vendor_name = _first_present(
        invoice.vendor_name,
        invoice.vendor_name_ar,
        extracted.get("vendor_name"),
        extracted.get("vendor_name_ar"),
        extracted.get("supplier_name"),
        extracted.get("merchant_name"),
        text_fallback.get("vendor_name"),
        text_fallback.get("vendor_name_ar"),
        _("Unspecified"),
    )
    vendor_vat_number = _first_present(
        invoice.vendor_vat_number,
        extracted.get("vendor_vat_number"),
        text_fallback.get("vendor_vat_number"),
        "",
    )
    invoice_number = _first_present(
        invoice.invoice_number,
        extracted.get("invoice_number"),
        text_fallback.get("invoice_number"),
        invoice.original_filename,
        str(invoice.id),
    )
    issue_date = _coerce_date(
        _first_present(invoice.invoice_date, extracted.get("invoice_date"), text_fallback.get("invoice_date"))
    )
    due_date = _coerce_date(
        _first_present(invoice.due_date, extracted.get("due_date"), text_fallback.get("due_date"))
    )
    currency = _first_present(invoice.currency, extracted.get("currency"), text_fallback.get("currency"), "SAR")
    subtotal = _first_non_zero_amount(invoice.subtotal, extracted.get("subtotal"), text_fallback.get("subtotal"))
    vat_amount = _first_non_zero_amount(invoice.vat_amount, extracted.get("vat_amount"), text_fallback.get("vat_amount"))
    total_amount = _first_non_zero_amount(invoice.total_amount, extracted.get("total_amount"), text_fallback.get("total_amount"))
    discount = _first_non_zero_amount(invoice.discount, extracted.get("discount"), text_fallback.get("discount"))
    vat_rate = _first_present(invoice.vat_rate, extracted.get("vat_rate"), text_fallback.get("vat_rate"), 15)

    if total_amount is None and subtotal is not None and vat_amount is not None:
        total_amount = subtotal + vat_amount - (discount or Decimal("0"))
    if subtotal is None and total_amount is not None and vat_amount is not None:
        subtotal = total_amount - vat_amount + (discount or Decimal("0"))

    ai_summary = _first_present(
        _clean_ai_summary(invoice.ai_summary),
        _clean_ai_summary(extracted.get("ai_summary")),
        _clean_ai_summary(text_fallback.get("ai_summary")),
        _("Data was extracted automatically. Review sensitive fields before final approval."),
    )
    extraction_method = _first_present(
        extracted.get("_extraction_method"),
        extracted.get("method"),
        text_fallback.get("_extraction_method"),
        "tesseract_fallback",
    )
    extraction_method_label = {
        "pdf_text_layer": _("PDF Text Layer"),
        "openai_vision": _("AI Vision (GPT-4o)"),
        "ocr_fallback": _("OCR Fallback (Tesseract)"),
        "openai_text": _("AI Text Extraction (GPT-4o)"),
        "openai_ocr": _("AI OCR (GPT-4o)"),
        "tesseract_fallback": _("OCR (Tesseract)"),
        "image_ai_vision": _("AI Vision (GPT-4o)"),
        "image_ocr": _("Image OCR"),
        "pdf_ocr": _("PDF OCR"),
        "excel_parser": _("Excel Parsing"),
        "json_parser": _("JSON Parsing"),
        "csv_parser": _("CSV Parsing"),
    }.get(extraction_method, _("Automatic Extraction"))

    def _format_amount(amount):
        if amount is None:
            return "—"
        try:
            return f"{Decimal(str(amount)):.2f} {currency}"
        except (InvalidOperation, TypeError, ValueError):
            return "—"

    def _format_date(date_value):
        return date_value.strftime("%Y-%m-%d") if date_value else "—"

    # ── Risk-gate fields for approval scenarios ───────────────────────────────
    from apps.rule_engine.models import RiskScoreSummary  # lazy to avoid circular
    _risk_summary   = RiskScoreSummary.objects.filter(document_id=invoice.id).first()
    _has_blocking   = _risk_summary.blocks_approval   if _risk_summary else False
    _blocking_count = _risk_summary.blocking_failures if _risk_summary else 0
    _fraud_score    = float(_risk_summary.score_breakdown.get("fraud_score", 0)) if _risk_summary else 0.0
    _risk_level_key = _risk_summary.risk_level        if _risk_summary else "low"
    _blocking_rules = (
        [r for r in _risk_summary.top_failed_rules if r.get("rule_code")]
        if _risk_summary else []
    )
    _rules_passed_n = int(getattr(validation, "rules_passed", 0) or 0)
    _rules_failed_n = int(getattr(validation, "rules_failed", 0) or 0)
    _has_real_audit = (_rules_passed_n + _rules_failed_n) > 0
    _has_warnings   = _rules_failed_n > 0 and not _has_blocking
    _needs_senior   = not _has_blocking and (
        _fraud_score > 70 or _risk_level_key in ("high", "critical")
    )

    return {
        "vendor_name": vendor_name,
        "vendor_vat_number": vendor_vat_number,
        "invoice_number": invoice_number,
        "invoice_date": issue_date,
        "invoice_date_display": _format_date(issue_date),
        "due_date": due_date,
        "due_date_display": _format_date(due_date),
        "currency": currency,
        "subtotal": subtotal,
        "subtotal_display": _format_amount(subtotal),
        "vat_amount": vat_amount,
        "vat_amount_display": _format_amount(vat_amount),
        "total_amount": total_amount,
        "total_amount_display": _format_amount(total_amount),
        "discount": discount,
        "discount_display": _format_amount(discount),
        "vat_rate": _to_float(vat_rate, 15),
        "ocr_confidence": _to_float(invoice.ocr_confidence, 0.0),
        "risk_score": round(risk_score, 2),
        "risk_label_ar": risk["label_ar"],
        "risk_label_en": risk["label_en"],
        "risk_badge_class": risk["badge_class"],
        "risk_text_class": risk["text_class"],
        "risk_ring_color": risk["ring_color"],
        "validation_score": round(audit_score, 2),
        "rules_passed": _rules_passed_n,
        "rules_failed": _rules_failed_n,
        # ── Approval gate fields (7-scenario logic) ───────────────────────────
        "has_blocking":    _has_blocking,
        "blocking_count":  _blocking_count,
        "fraud_score":     _fraud_score,
        "risk_level":      _risk_level_key,
        "has_warnings":    _has_warnings,
        "needs_senior":    _needs_senior,
        "has_real_audit":  _has_real_audit,
        "blocking_rules":  _blocking_rules,
        # ─────────────────────────────────────────────────────────────────────
        "status_label": status_label,
        "status_label_ar": status_label_ar,
        "status_badge_class": status_badge_class,
        "effective_status": effective_status,
        "status_mismatch": effective_status != invoice.status,
        "source_method": extraction_method,
        "source_method_label": extraction_method_label,
        "ai_summary": ai_summary,
    }


def landing(request):
    return render(request, "landing/index.html", _public_ctx(request))


def _render_marketing_page(request, *, page_key, title, eyebrow, description, bullets):
    return render(
        request,
        "landing/page.html",
        _public_ctx(
            request,
            page_key=page_key,
            page_title=title,
            page_eyebrow=eyebrow,
            page_description=description,
            page_bullets=bullets,
        ),
    )


def _render_legal_page(request, *, page_key, title, eyebrow, description, sections, updated=None):
    """Render a long-form legal / informational page (terms, privacy, refund,
    services) through the shared ``landing/page.html`` template.

    ``sections`` is a list of dicts shaped like
    ``{"heading": str, "body": str, "bullets": [str, ...]}`` (``body`` and
    ``bullets`` are both optional). ``bullets`` is deliberately **not** named
    ``items`` — Django template lookup of ``section.items`` falls through to the
    dict's ``.items()`` method when the key is absent, rendering raw
    ``('heading', ...)`` tuples. The template renders each section as a titled
    prose block with an optional bullet list, while still honouring the existing
    ``page_bullets`` / ``contact_info`` paths used by the other marketing pages.
    """
    return render(
        request,
        "landing/page.html",
        _public_ctx(
            request,
            page_key=page_key,
            page_title=title,
            page_eyebrow=eyebrow,
            page_description=description,
            page_sections=sections,
            page_updated=updated,
        ),
    )


def _is_arabic(request) -> bool:
    """Whether the active request language is Arabic.

    Legal copy below is authored bilingually and selected at render time so the
    full professional Arabic text is guaranteed to appear on the Arabic site
    (the default) without depending on compiled ``.po``/``.mo`` catalogues for
    these long, page-specific paragraphs."""
    return (get_language() or "").lower().startswith("ar")


def pricing(request):
    """Public pricing page — renders the four canonical plans LIVE from
    the Plan table so a price change in the DB shows up here without a
    redeploy. Falls back to a static layout if apps.billing isn't
    installed (so unrelated deployments don't break)."""
    plans, business_plans, accounting_plans = [], [], []
    try:
        from apps.billing.choices import ACCOUNTING_PLAN_CODES
        from apps.billing.services.plan_service import (
            addon_savings,
            list_purchasable_plans,
        )

        plans = list(list_purchasable_plans())
        # §I.4 — computed from the add-on prices at render time, never written
        # as text. A hardcoded "27%" stops being true at the first price edit,
        # silently, because nothing ties the sentence to the number.
        savings = addon_savings()
        # §L.4 — accounting-firm plans get their own section rather than being
        # mixed into the main row: different buyer, different pricing basis.
        # Split here rather than in the template so the grouping rule lives in
        # one place and the template just renders.
        business_plans = [p for p in plans if p.code not in ACCOUNTING_PLAN_CODES]
        accounting_plans = [p for p in plans if p.code in ACCOUNTING_PLAN_CODES]
    except Exception:                       # noqa: BLE001 — degrade gracefully
        plans, business_plans, accounting_plans = [], [], []
        savings = {"users_percent": None, "invoices_percent": None}

    ctx = _public_ctx(
        request,
        page_key="pricing",
        page_title=str(_("Pricing & plans")),
        page_eyebrow=str(_("Pricing")),
        page_description=str(_(
            "Choose the plan that fits your audit volume. Every plan includes "
            "the full Tadgeeg AI pipeline, fraud detection, and ZATCA-ready exports."
        )),
        plans=plans,
        business_plans=business_plans,
        accounting_plans=accounting_plans,
        savings=savings,
        is_authenticated=request.user.is_authenticated,
    )
    return render(request, "landing/pricing.html", ctx)


def partners(request):
    """Public «شركاؤنا» page (§C / §L.2).

    Sections come from apps.partners.selectors, which builds them off
    ``Partner.published`` — the publish gate lives in the data layer, so an
    unpublished partner is unreachable here regardless of what the template
    does (§D4).

    _public_ctx() is called deliberately: besides the shared context it is the
    campaign-attribution chokepoint Phase 1 depends on, so skipping it would
    silently break utm capture for anyone landing on this page first.
    """
    from apps.partners.selectors import get_public_sections, get_strategic_partners

    return render(
        request,
        "landing/partners.html",
        _public_ctx(
            request,
            page_title=str(_("Our Partners")),
            page_description=str(_(
                "The Tadgeeg partner ecosystem — strategic partners, tiered "
                "partners, and authorized distributors."
            )),
            strategic_partners=list(get_strategic_partners()),
            partner_sections=get_public_sections(),
        ),
    )


def partner_apply(request):
    """Public partner application form (§E / §L.3).

    Renders only; the POST goes to /api/v1/partners/applications/, which is
    throttled per IP and validates everything server-side. Keeping the write on
    the API means the form has no privileged path of its own.
    """
    from django.conf import settings

    from apps.partners.models import BusinessArea, PartnerType
    from apps.partners.uploads import ALLOWED_EXTENSIONS

    return render(
        request,
        "landing/partner_apply.html",
        _public_ctx(
            request,
            page_title=str(_("Join as a Partner")),
            page_description=str(_(
                "Apply to join the Tadgeeg partner ecosystem."
            )),
            partner_type_choices=PartnerType.choices,
            business_area_choices=BusinessArea.choices,
            # The wizard pre-checks files before upload starts. These come from
            # the server's own rules rather than being restated in JavaScript,
            # so the convenience check cannot drift away from the authority.
            # The server still validates everything; this only saves the user a
            # failed upload.
            upload_max_mb=settings.PARTNER_DOC_MAX_FILE_MB,
            upload_max_files=settings.PARTNER_DOC_MAX_FILES,
            upload_allowed_extensions=sorted(ALLOWED_EXTENSIONS),
        ),
    )


def partner_detail(request, slug):
    """Public partner profile. Published partners only.

    404 (not 403) for an unpublished slug: a 403 would confirm the record
    exists, telling an anonymous visitor about partners the company has not
    announced.
    """
    from django.http import Http404

    from apps.partners.selectors import get_public_partner_by_slug

    partner = get_public_partner_by_slug(slug)
    if partner is None:
        raise Http404("Partner not found.")

    return render(
        request,
        "landing/partner_detail.html",
        _public_ctx(
            request,
            # The <title> and <meta description> are what a search engine
            # indexes and what a shared link previews, so they follow the
            # active language like the page body does. Reading the raw columns
            # here put a Latin name on the Arabic page's title bar and an
            # Arabic sentence in the English page's meta description.
            page_title=partner.display_name,
            page_description=partner.display_short_description,
            partner=partner,
            partner_public=partner.public_payload(),
        ),
    )


def about(request):
    return _render_marketing_page(
        request,
        page_key="about",
        title=_("A financial auditing platform built for the GCC market"),
        eyebrow=_("About %(product)s") % {"product": django_settings.PRODUCT_NAME},
        description=_("%(product)s focuses on auditing invoices and financial documents with tax and financial compliance support across GCC business environments.") % {"product": django_settings.PRODUCT_NAME},
        bullets=[
            _("Arabic and English support throughout the product journey."),
            _("A focus on ZATCA, VAT, and document review."),
            _("A practical experience for finance, audit, and compliance teams."),
        ],
    )


def contact(request):
    ar = _is_arabic(request)
    contact_info = {
        # Company identity (bilingual, selected at render time) — the public
        # contact page presents the operating company, not an individual, for
        # payment-gateway review and public trust.
        "rep_name": "شركة احصل الحل" if ar else "Get Solution Company",
        "rep_role": (
            "المالك والمشغّل الرسمي لمنصة تدقيق" if ar
            else "Official owner and operator of Tadgeeg AI"
        ),
        "rep_avatar": "ت" if ar else "GS",
        "phone_display": "+966 54 054 1719",
        "phone_link": "tel:+966540541719",
        "email": "contact@tadgeeg.com",
        "email_link": "mailto:contact@tadgeeg.com",
        "website_display": "www.tadgeeg.com",
        "website_link": "https://www.tadgeeg.com",
        "whatsapp_link": "https://wa.me/966540541719",
        "cta_call_label": _("Call now"),
        "cta_whatsapp_label": _("Message on WhatsApp"),
        "cta_email_label": _("Email us"),
    }
    return render(
        request,
        "landing/page.html",
        _public_ctx(
            request,
            page_key="contact",
            page_title=_("Talk to the %(product)s team") % {"product": django_settings.PRODUCT_NAME},
            page_eyebrow=_("Contact"),
            page_description=_("If you need a tailored demo, enterprise onboarding, or technical guidance, our team can arrange an introductory session and rollout plan."),
            page_bullets=[
                _("Tailored responses for finance and compliance teams."),
                _("Support for pilot and enterprise launches."),
                _("Technical assistance and onboarding guidance."),
            ],
            contact_info=contact_info,
        ),
    )


def services(request):
    """Public services page — a detailed catalogue of what the platform does."""
    product = django_settings.PRODUCT_NAME
    ar = _is_arabic(request)

    if ar:
        title = "خدمات منصة تدقيق"
        eyebrow = "الخدمات"
        description = (
            "تُقدّم منصة «تدقيق» مجموعة متكاملة من خدمات التدقيق المالي الذكي "
            "تساعد فرق المالية والمراجعة والامتثال على فحص المستندات المالية بدقة "
            "وسرعة، وكشف المخاطر والأخطاء والاحتيال قبل اعتماد الصرف."
        )
        sections = [
            {
                "heading": "التدقيق الذكي للفواتير",
                "body": (
                    "نقوم بقراءة الفواتير تلقائياً واستخراج بياناتها، ثم تطبيق عشرات "
                    "قواعد التدقيق عليها للكشف عن المشكلات قبل اعتماد الدفع."
                ),
                "bullets": [
                    "كشف الفواتير المكرّرة والمدفوعة مسبقاً.",
                    "التحقق من صحة احتساب ضريبة القيمة المضافة (15%) والإجماليات.",
                    "رصد مؤشرات الاحتيال والتلاعب في المبالغ والتواريخ.",
                    "التحقق من اكتمال البيانات الإلزامية ومطابقة متطلبات هيئة الزكاة والضريبة والجمارك (ZATCA).",
                ],
            },
            {
                "heading": "تدقيق أوامر الشراء",
                "body": (
                    "مطابقة أوامر الشراء مع الفواتير وإشعارات الاستلام لضمان أن ما تم "
                    "التعاقد عليه هو ما تم استلامه وفوترته فعلياً."
                ),
                "bullets": [
                    "مطابقة ثلاثية بين أمر الشراء والفاتورة وإشعار الاستلام.",
                    "كشف الفروقات في الكميات والأسعار والشروط.",
                    "متابعة تجاوزات الموازنة وحدود الصلاحيات.",
                ],
            },
            {
                "heading": "تدقيق المصروفات وكشف الشذوذ",
                "body": (
                    "تحليل تقارير المصروفات للتأكد من توافقها مع سياسات المنشأة "
                    "والكشف عن الأنماط غير الاعتيادية."
                ),
                "bullets": [
                    "التحقق من الالتزام بسياسات المصروفات وحدود الفئات.",
                    "كشف المصروفات المكرّرة أو المبالغ فيها.",
                    "تحليل الشذوذ الإحصائي لرصد المعاملات المشبوهة.",
                ],
            },
            {
                "heading": "تدقيق كشوف الحسابات البنكية",
                "body": (
                    "فحص كشوف الحسابات البنكية ومطابقتها مع السجلات المحاسبية "
                    "للكشف عن الفروقات والمعاملات غير المبرّرة."
                ),
                "bullets": [
                    "مطابقة الحركات البنكية مع القيود المحاسبية.",
                    "كشف المعاملات غير المبرّرة أو المتكررة.",
                    "تطبيق تحليل بنفورد (Benford) للكشف عن التلاعب في الأرقام.",
                ],
            },
            {
                "heading": "الذكاء الاصطناعي المالي",
                "body": (
                    "محرّك ذكاء اصطناعي يقرأ المستندات بلغات متعددة، ويستخرج البيانات، "
                    "ويولّد ملخصات ذكية وتوصيات عملية لفرق المالية."
                ),
                "bullets": [
                    "استخراج آلي للبيانات من ملفات PDF والصور والملفات الإلكترونية.",
                    "ملخصات وتنبيهات ذكية حول المخاطر والملاحظات.",
                    "دعم اللغتين العربية والإنجليزية في القراءة والتحليل.",
                ],
            },
            {
                "heading": "التقارير التنفيذية",
                "body": (
                    "تقارير جاهزة للإدارة العليا والمراجعين تغطّي المخاطر والامتثال "
                    "والحوكمة، قابلة للتصدير ومتوافقة مع المعايير المهنية."
                ),
                "bullets": [
                    "تقارير المخاطر وتقييم المنشآت والموردين.",
                    "تقارير الامتثال الضريبي ومتطلبات ZATCA.",
                    "تقارير الحوكمة وأوراق العمل وفق معايير التدقيق الدولية.",
                ],
            },
        ]
    else:
        title = "%s platform services" % product
        eyebrow = "Services"
        description = (
            "%s delivers an integrated suite of intelligent financial-audit "
            "services that help finance, audit, and compliance teams review "
            "financial documents accurately, and surface risk, errors, and fraud "
            "before payments are approved." % product
        )
        sections = [
            {
                "heading": "Intelligent invoice auditing",
                "body": (
                    "We read invoices automatically, extract their data, and apply "
                    "dozens of audit rules to catch issues before payment approval."
                ),
                "bullets": [
                    "Detect duplicate and previously paid invoices.",
                    "Validate VAT (15%) calculations and totals.",
                    "Flag fraud and tampering indicators in amounts and dates.",
                    "Check mandatory fields and ZATCA compliance.",
                ],
            },
            {
                "heading": "Purchase order auditing",
                "body": (
                    "Match purchase orders against invoices and goods-receipt notes "
                    "to ensure what was contracted is what was received and billed."
                ),
                "bullets": [
                    "Three-way match between PO, invoice, and receipt note.",
                    "Detect variances in quantities, prices, and terms.",
                    "Track budget overruns and approval limits.",
                ],
            },
            {
                "heading": "Expense auditing & anomaly detection",
                "body": (
                    "Analyse expense reports for policy compliance and surface "
                    "unusual patterns."
                ),
                "bullets": [
                    "Verify compliance with expense policies and category limits.",
                    "Detect duplicate or inflated expenses.",
                    "Statistical anomaly analysis for suspicious transactions.",
                ],
            },
            {
                "heading": "Bank statement auditing",
                "body": (
                    "Review bank statements and reconcile them against accounting "
                    "records to surface variances and unexplained transactions."
                ),
                "bullets": [
                    "Reconcile bank movements with accounting entries.",
                    "Detect unexplained or repeated transactions.",
                    "Apply Benford analysis to detect number manipulation.",
                ],
            },
            {
                "heading": "Financial artificial intelligence",
                "body": (
                    "An AI engine that reads documents in multiple languages, "
                    "extracts data, and produces smart summaries and recommendations."
                ),
                "bullets": [
                    "Automatic data extraction from PDFs, images, and digital files.",
                    "Smart summaries and alerts about risks and findings.",
                    "Arabic and English support in reading and analysis.",
                ],
            },
            {
                "heading": "Executive reporting",
                "body": (
                    "Ready-made reports for leadership and auditors covering risk, "
                    "compliance, and governance, exportable and standards-aligned."
                ),
                "bullets": [
                    "Risk reports and entity/vendor assessments.",
                    "Tax compliance and ZATCA reporting.",
                    "Governance reports and working papers aligned to ISA standards.",
                ],
            },
        ]

    return _render_legal_page(
        request,
        page_key="services",
        title=title,
        eyebrow=eyebrow,
        description=description,
        sections=sections,
    )


def privacy(request):
    """Detailed privacy & data-protection policy (payment-gateway compliant)."""
    ar = _is_arabic(request)
    updated = "2026-05-31"

    if ar:
        title = "سياسة الخصوصية"
        eyebrow = "الخصوصية"
        description = (
            "نلتزم في منصة «تدقيق» بحماية بياناتك ومعالجتها بأعلى معايير العناية "
            "والسرية، بما يتوافق مع نظام حماية البيانات الشخصية في المملكة العربية "
            "السعودية. توضّح هذه السياسة البيانات التي نجمعها وكيفية استخدامها وحمايتها وحقوقك تجاهها."
        )
        sections = [
            {
                "heading": "البيانات التي نجمعها",
                "body": "نجمع الحد اللازم من البيانات لتقديم الخدمة، وتشمل:",
                "bullets": [
                    "بيانات الحساب: الاسم، البريد الإلكتروني، رقم الجوال، وكلمة المرور (مشفّرة).",
                    "بيانات المنشأة: اسم المنشأة، السجل التجاري، الرقم الضريبي، وبيانات الاتصال.",
                    "بيانات المستندات: الفواتير، أوامر الشراء، كشوف الحسابات البنكية، تقارير المصروفات وما يرتبط بها من بيانات مالية ترفعها بنفسك.",
                    "بيانات الاستخدام التقنية: سجلّات الدخول وعنوان IP ونوع المتصفح، لأغراض الأمان وتحسين الخدمة.",
                ],
            },
            {
                "heading": "كيفية استخدام البيانات",
                "body": "نستخدم بياناتك للأغراض التالية فقط:",
                "bullets": [
                    "تشغيل خدمات التدقيق المالي الذكي وتحليل المستندات وإصدار التقارير.",
                    "إدارة حسابك واشتراكك وتقديم الدعم الفني.",
                    "تحسين دقة المنصة وأمنها وأدائها.",
                    "إرسال الإشعارات التشغيلية والتنبيهات المتعلقة بحسابك.",
                ],
            },
            {
                "heading": "حماية البيانات",
                "body": "نطبّق ضوابط تقنية وتنظيمية صارمة لحماية بياناتك:",
                "bullets": [
                    "تشفير البيانات أثناء النقل عبر بروتوكول SSL/TLS.",
                    "نسخ احتياطية دورية لضمان استمرارية الخدمة واستعادة البيانات.",
                    "صلاحيات وصول قائمة على الأدوار (RBAC) وعزل بيانات كل منشأة عن غيرها.",
                    "سجلّات تدقيق (Audit Logs) لتتبّع العمليات الحساسة على البيانات.",
                ],
            },
            {
                "heading": "مشاركة البيانات",
                "body": (
                    "لا نبيع بياناتك ولا نشاركها مع أي طرف ثالث لأغراض تسويقية. "
                    "تقتصر المشاركة على الحالات التالية:"
                ),
                "bullets": [
                    "بموافقتك الصريحة كعميل.",
                    "عند وجود أمر قضائي أو التزام نظامي يفرض ذلك.",
                    "مع مزوّدي الخدمات التقنية اللازمين لتشغيل المنصة (مثل الاستضافة وبوابات الدفع)، وفق اتفاقيات سرية ومعالجة بيانات.",
                ],
            },
            {
                "heading": "حقوق العميل",
                "body": "تملك في أي وقت الحق في:",
                "bullets": [
                    "الوصول إلى بياناتك وتصحيحها أو تحديثها.",
                    "تصدير بياناتك ومستنداتك بصيغ قابلة للقراءة.",
                    "حذف ملفاتك المرفوعة أو حذف حسابك بالكامل.",
                    "سحب الموافقة على المعالجة ضمن الحدود النظامية.",
                ],
            },
            {
                "heading": "الاحتفاظ بالبيانات والامتثال",
                "body": (
                    "نحتفظ ببياناتك طوال مدة اشتراكك وللمدة اللازمة للوفاء بالالتزامات "
                    "النظامية، ثم نحذفها أو نجعلها مجهولة المصدر. تخضع هذه السياسة "
                    "لنظام حماية البيانات الشخصية في المملكة العربية السعودية ولوائحه التنفيذية."
                ),
            },
            {
                "heading": "التواصل بشأن الخصوصية",
                "body": (
                    "لأي استفسار أو طلب يتعلق بخصوصية بياناتك، يمكنك التواصل معنا عبر "
                    "البريد الإلكتروني: contact@tadgeeg.com."
                ),
            },
        ]
    else:
        title = "Privacy Policy"
        eyebrow = "Privacy"
        description = (
            "Tadgeeg is committed to protecting your data and processing it with "
            "the highest standards of care and confidentiality, in line with the "
            "Saudi Personal Data Protection Law (PDPL). This policy explains what "
            "data we collect, how we use and protect it, and your rights."
        )
        sections = [
            {
                "heading": "Data we collect",
                "body": "We collect only the data needed to provide the service:",
                "bullets": [
                    "Account data: name, email, mobile number, and password (encrypted).",
                    "Organization data: company name, commercial registration, VAT number, and contact details.",
                    "Document data: invoices, purchase orders, bank statements, expense reports, and related financial data you upload.",
                    "Technical usage data: login logs, IP address, and browser type, for security and service improvement.",
                ],
            },
            {
                "heading": "How we use data",
                "body": "We use your data only for the following purposes:",
                "bullets": [
                    "Operating intelligent audit services, analysing documents, and generating reports.",
                    "Managing your account and subscription and providing support.",
                    "Improving platform accuracy, security, and performance.",
                    "Sending operational notifications and alerts related to your account.",
                ],
            },
            {
                "heading": "Data protection",
                "body": "We apply strict technical and organizational controls:",
                "bullets": [
                    "Encryption in transit via SSL/TLS.",
                    "Regular backups for service continuity and recovery.",
                    "Role-based access control (RBAC) and isolation of each organization's data.",
                    "Audit logs to trace sensitive operations on data.",
                ],
            },
            {
                "heading": "Data sharing",
                "body": (
                    "We do not sell your data or share it with third parties for "
                    "marketing. Sharing is limited to:"
                ),
                "bullets": [
                    "With your explicit consent as the customer.",
                    "When required by a court order or legal obligation.",
                    "With technical service providers necessary to operate the platform (e.g. hosting, payment gateways) under confidentiality and data-processing agreements.",
                ],
            },
            {
                "heading": "Customer rights",
                "body": "At any time you have the right to:",
                "bullets": [
                    "Access, correct, or update your data.",
                    "Export your data and documents in readable formats.",
                    "Delete your uploaded files or your entire account.",
                    "Withdraw consent to processing within legal limits.",
                ],
            },
            {
                "heading": "Data retention & compliance",
                "body": (
                    "We retain your data for the duration of your subscription and "
                    "as long as needed to meet legal obligations, then delete or "
                    "anonymise it. This policy is governed by the Saudi Personal "
                    "Data Protection Law and its regulations."
                ),
            },
            {
                "heading": "Privacy contact",
                "body": (
                    "For any privacy-related question or request, contact us at "
                    "contact@tadgeeg.com."
                ),
            },
        ]

    return _render_legal_page(
        request,
        page_key="privacy",
        title=title,
        eyebrow=eyebrow,
        description=description,
        sections=sections,
        updated=updated,
    )


def terms(request):
    """Terms & Conditions page (payment-gateway compliant)."""
    ar = _is_arabic(request)
    updated = "2026-05-31"
    product = django_settings.PRODUCT_NAME

    if ar:
        title = "الأحكام والشروط"
        eyebrow = "الأحكام والشروط"
        description = (
            "توضّح هذه الأحكام والشروط القواعد المنظِّمة لاستخدام منصة «تدقيق»، "
            "ويُعدّ استخدامك للمنصة أو الاشتراك فيها موافقةً صريحةً على هذه الأحكام."
        )
        sections = [
            {
                "heading": "تعريف المنصة وطبيعة الخدمة",
                "body": (
                    "منصة «تدقيق» هي أداة برمجية للتدقيق المالي الذكي تساعد المستخدم "
                    "على فحص المستندات المالية وكشف الأخطاء والمخاطر. وهي أداة مساعِدة "
                    "تدعم اتخاذ القرار، ولا تُعدّ بديلاً عن المحاسب القانوني أو المراجع "
                    "المعتمد، ولا تتحمّل المنصة مسؤولية القرارات المالية أو القانونية "
                    "التي يتخذها المستخدم بناءً على نتائجها."
                ),
            },
            {
                "heading": "أهلية الاستخدام",
                "body": (
                    "يجب أن يكون المستخدم شخصاً ذا أهلية نظامية أو ممثلاً مفوّضاً عن "
                    "منشأة، وأن تكون البيانات المقدَّمة عند التسجيل صحيحة وكاملة. "
                    "تُستخدم المنصة لأغراض تجارية ومهنية مشروعة فقط."
                ),
            },
            {
                "heading": "مسؤوليات المستخدم",
                "bullets": [
                    "ضمان صحة واكتمال البيانات والمستندات التي يرفعها على المنصة.",
                    "المحافظة على سرية بيانات الدخول وكلمة المرور وعدم مشاركتها.",
                    "إقرار المستخدم بملكيته أو تفويضه باستخدام المستندات التي يرفعها.",
                    "استخدام المنصة بما يتوافق مع الأنظمة المعمول بها وعدم إساءة استخدامها.",
                ],
            },
            {
                "heading": "مسؤوليات المنصة",
                "bullets": [
                    "بذل العناية اللازمة لتوفير الخدمة وأمن البيانات واستمراريتها.",
                    "تطبيق ضوابط حماية وتشفير وعزل لبيانات كل منشأة.",
                    "توفير الدعم الفني ومعالجة الأعطال خلال مدة معقولة.",
                ],
            },
            {
                "heading": "الاشتراكات والفوترة",
                "bullets": [
                    "الدفع يكون مقدّماً عن مدة الاشتراك المختارة (شهرية أو سنوية).",
                    "يتجدّد الاشتراك تلقائياً ما لم يقم المستخدم بإيقاف التجديد قبل نهاية المدة.",
                    "يمكن إيقاف التجديد التلقائي في أي وقت مع بقاء الاشتراك فعّالاً حتى نهاية المدة المدفوعة.",
                    "تخضع عمليات الاسترداد لسياسة الاسترجاع المنشورة على المنصة.",
                ],
            },
            {
                "heading": "تعليق الحساب أو إلغاؤه",
                "body": (
                    "يحقّ للمنصة تعليق أو إنهاء الحساب في حال مخالفة هذه الأحكام أو "
                    "استخدام الخدمة بشكل غير مشروع أو يهدد أمن المنصة أو المستخدمين "
                    "الآخرين، مع إشعار المستخدم متى أمكن ذلك."
                ),
            },
            {
                "heading": "حدود المسؤولية",
                "body": (
                    "تُقدَّم الخدمة «كما هي» وفق أفضل الجهود. ولا تتحمّل المنصة "
                    "المسؤولية عن أي أضرار غير مباشرة أو تبعية، أو عن قرارات اتُّخذت "
                    "بناءً على نتائج التحليل، أو عن أخطاء ناتجة عن بيانات غير صحيحة "
                    "قدّمها المستخدم. وتقتصر المسؤولية القصوى للمنصة على قيمة الاشتراك "
                    "المدفوع عن المدة محل النزاع."
                ),
            },
            {
                "heading": "القانون الواجب التطبيق",
                "body": (
                    "تخضع هذه الأحكام والشروط وتُفسَّر وفقاً لأنظمة المملكة العربية "
                    "السعودية، وتختص الجهات القضائية المختصة في المملكة بالفصل في أي نزاع ينشأ عنها."
                ),
            },
            {
                "heading": "تعديل الأحكام",
                "body": (
                    "يجوز لـ %s تحديث هذه الأحكام من وقت لآخر، ويُعدّ استمرار "
                    "استخدامك للمنصة بعد نشر التعديلات موافقةً عليها." % product
                ),
            },
        ]
    else:
        title = "Terms & Conditions"
        eyebrow = "Terms & Conditions"
        description = (
            "These Terms & Conditions set out the rules governing use of the "
            "Tadgeeg platform. By using or subscribing to the platform, you "
            "expressly agree to these terms."
        )
        sections = [
            {
                "heading": "Definition and nature of the service",
                "body": (
                    "Tadgeeg is an intelligent financial-audit software tool that "
                    "helps users review financial documents and surface errors and "
                    "risks. It is an assistive decision-support tool and is not a "
                    "substitute for a certified accountant or licensed auditor. The "
                    "platform is not responsible for financial or legal decisions "
                    "users make based on its results."
                ),
            },
            {
                "heading": "Eligibility",
                "body": (
                    "Users must have legal capacity or be an authorised "
                    "representative of an organization, and must provide accurate, "
                    "complete information at registration. The platform is for "
                    "lawful business and professional use only."
                ),
            },
            {
                "heading": "User responsibilities",
                "bullets": [
                    "Ensure the accuracy and completeness of data and documents uploaded.",
                    "Keep login credentials and passwords confidential and not share them.",
                    "Confirm ownership of, or authorisation to use, uploaded documents.",
                    "Use the platform in compliance with applicable laws and not misuse it.",
                ],
            },
            {
                "heading": "Platform responsibilities",
                "bullets": [
                    "Exercise due care to provide the service and secure data and continuity.",
                    "Apply protection, encryption, and isolation for each organization's data.",
                    "Provide technical support and address faults within a reasonable time.",
                ],
            },
            {
                "heading": "Subscriptions and billing",
                "bullets": [
                    "Payment is made in advance for the chosen subscription period (monthly or annual).",
                    "Subscriptions renew automatically unless the user cancels renewal before the period ends.",
                    "Auto-renewal can be cancelled anytime; the subscription stays active until the end of the paid period.",
                    "Refunds are subject to the refund policy published on the platform.",
                ],
            },
            {
                "heading": "Account suspension or termination",
                "body": (
                    "The platform may suspend or terminate an account in case of "
                    "breach of these terms, unlawful use, or use that threatens the "
                    "security of the platform or other users, with notice where possible."
                ),
            },
            {
                "heading": "Limitation of liability",
                "body": (
                    "The service is provided 'as is' on a best-effort basis. The "
                    "platform is not liable for indirect or consequential damages, "
                    "for decisions made based on analysis results, or for errors "
                    "arising from inaccurate data provided by the user. The "
                    "platform's maximum liability is limited to the subscription "
                    "fee paid for the disputed period."
                ),
            },
            {
                "heading": "Governing law",
                "body": (
                    "These terms are governed by and construed in accordance with "
                    "the laws of the Kingdom of Saudi Arabia, and the competent "
                    "Saudi judicial authorities have jurisdiction over any dispute."
                ),
            },
            {
                "heading": "Changes to the terms",
                "body": (
                    "%s may update these terms from time to time. Continued use of "
                    "the platform after changes are published constitutes acceptance." % product
                ),
            },
        ]

    return _render_legal_page(
        request,
        page_key="terms",
        title=title,
        eyebrow=eyebrow,
        description=description,
        sections=sections,
        updated=updated,
    )


def refund_policy(request):
    """Refund / cancellation policy (Saudi e-commerce compliant)."""
    ar = _is_arabic(request)
    updated = "2026-05-31"

    if ar:
        title = "سياسة الاسترجاع والاستبدال"
        eyebrow = "سياسة الاسترجاع"
        description = (
            "توضّح هذه السياسة الحالات التي يحقّ فيها استرداد قيمة الاشتراك في منصة "
            "«تدقيق» والحالات التي لا يُستحق فيها الاسترداد، بما يتوافق مع نظام "
            "التجارة الإلكترونية في المملكة العربية السعودية."
        )
        sections = [
            {
                "heading": "طبيعة الخدمة",
                "body": (
                    "منصة «تدقيق» خدمة رقمية تُقدَّم عبر الإنترنت بنظام الاشتراك "
                    "(SaaS)، ويبدأ تقديم الخدمة فور تفعيل الاشتراك وإتاحة الوصول إلى "
                    "ميزات المنصة."
                ),
            },
            {
                "heading": "الفترة التجريبية",
                "body": (
                    "في حال توفّر فترة تجربة مجانية، يمكنك خلالها تقييم الخدمة دون "
                    "أي رسوم، ولا يتم احتساب أي مبلغ إلا بعد انتهاء الفترة التجريبية "
                    "وبدء الاشتراك المدفوع."
                ),
            },
            {
                "heading": "الاسترداد الكامل قبل التفعيل",
                "body": (
                    "يحقّ لك طلب استرداد كامل لقيمة الاشتراك إذا قدّمت الطلب قبل "
                    "البدء الفعلي في استخدام الخدمة وقبل رفع أو معالجة أي مستندات على "
                    "المنصة."
                ),
            },
            {
                "heading": "حالات عدم الاسترداد",
                "body": (
                    "نظراً لطبيعة الخدمة الرقمية، لا يُستحق الاسترداد في الحالات التالية:"
                ),
                "bullets": [
                    "بعد البدء الفعلي في استخدام الخدمة أو معالجة المستندات على المنصة.",
                    "عن المدة المنقضية من الاشتراك التي تم استخدام الخدمة خلالها.",
                    "في حال إيقاف التجديد التلقائي، إذ يبقى الاشتراك فعّالاً حتى نهاية المدة المدفوعة دون استرداد للمدة المتبقية.",
                    "عند مخالفة المستخدم لأحكام وشروط الاستخدام بما أدى إلى إيقاف الحساب.",
                ],
            },
            {
                "heading": "الحالات الاستثنائية للاسترداد",
                "body": "يجوز النظر في طلب الاسترداد في الحالات الاستثنائية التالية:",
                "bullets": [
                    "تعذُّر تشغيل الخدمة بشكل كامل لأسباب فنية من جانب المنصة.",
                    "وجود خطأ تقني جسيم يمنع الاستفادة من الخدمة ولم تتم معالجته خلال مدة معقولة.",
                    "حدوث خصم مالي بالخطأ أو تكرار غير مقصود لعملية الدفع.",
                ],
            },
            {
                "heading": "إلغاء التجديد التلقائي",
                "body": (
                    "يمكنك إلغاء التجديد التلقائي في أي وقت من إعدادات حسابك. يؤدي "
                    "الإلغاء إلى إيقاف الفوترة المستقبلية فقط، ويبقى اشتراكك الحالي "
                    "فعّالاً حتى نهاية المدة المدفوعة."
                ),
            },
            {
                "heading": "آلية تقديم طلب الاسترداد",
                "body": (
                    "لتقديم طلب استرداد ضمن الحالات المؤهّلة، تواصل معنا عبر البريد "
                    "الإلكتروني contact@tadgeeg.com موضّحاً بيانات الاشتراك وسبب "
                    "الطلب. تتم معالجة الطلبات المؤهّلة خلال مدة معقولة عبر وسيلة الدفع الأصلية."
                ),
            },
        ]
    else:
        title = "Refund & Cancellation Policy"
        eyebrow = "Refund Policy"
        description = (
            "This policy explains when a Tadgeeg subscription is eligible for a "
            "refund and when refunds are not due, in line with the Saudi "
            "E-Commerce Law."
        )
        sections = [
            {
                "heading": "Nature of the service",
                "body": (
                    "Tadgeeg is a digital subscription service (SaaS) delivered "
                    "online. The service begins as soon as the subscription is "
                    "activated and access to platform features is granted."
                ),
            },
            {
                "heading": "Free trial",
                "body": (
                    "Where a free trial is available, you may evaluate the service "
                    "at no charge during it; no amount is charged until the trial "
                    "ends and a paid subscription begins."
                ),
            },
            {
                "heading": "Full refund before activation",
                "body": (
                    "You may request a full refund if you submit the request before "
                    "actually using the service and before uploading or processing "
                    "any documents on the platform."
                ),
            },
            {
                "heading": "Cases where refunds are not due",
                "body": (
                    "Given the digital nature of the service, refunds are not due in "
                    "the following cases:"
                ),
                "bullets": [
                    "After actual use of the service or processing of documents on the platform.",
                    "For the elapsed portion of a subscription during which the service was used.",
                    "When auto-renewal is cancelled — the subscription stays active until the end of the paid period with no refund for the remaining time.",
                    "When the user breaches the terms of use in a way that leads to account suspension.",
                ],
            },
            {
                "heading": "Exceptional refund cases",
                "body": "A refund may be considered in the following exceptional cases:",
                "bullets": [
                    "The service could not operate fully due to technical issues on the platform's side.",
                    "A material technical defect prevents use of the service and was not resolved within a reasonable time.",
                    "An incorrect charge or unintended duplicate payment occurred.",
                ],
            },
            {
                "heading": "Cancelling auto-renewal",
                "body": (
                    "You can cancel auto-renewal anytime from your account settings. "
                    "Cancellation stops future billing only; your current "
                    "subscription stays active until the end of the paid period."
                ),
            },
            {
                "heading": "How to request a refund",
                "body": (
                    "To request a refund within eligible cases, contact us at "
                    "contact@tadgeeg.com with your subscription details and the "
                    "reason. Eligible requests are processed within a reasonable "
                    "time via the original payment method."
                ),
            },
        ]

    return _render_legal_page(
        request,
        page_key="refund_policy",
        title=title,
        eyebrow=eyebrow,
        description=description,
        sections=sections,
        updated=updated,
    )


def blog(request):
    return _render_marketing_page(
        request,
        page_key="blog",
        title=_("Audit and compliance insights"),
        eyebrow=_("Insights"),
        description=_("A place to share operating insights and practical guidance on financial auditing, tax compliance, and digital transformation for finance teams."),
        bullets=[
            _("Best practices for reviewing invoices and vendors."),
            _("Updates on GCC compliance requirements."),
            _("Operational guides for enabling AI inside finance teams."),
        ],
    )


def integrations(request):
    return _render_marketing_page(
        request,
        page_key="integrations",
        title=_("Integrations that are ready and scalable"),
        eyebrow=_("Integrations"),
        description=_("Integration plans cover email, approval flows, and connectivity with your current finance operating environment."),
        bullets=[
            _("Gradual connectivity with your existing company systems."),
            _("Internal APIs that can be extended for integrations."),
            _("Tailored setup for organizations with specialized workflows."),
        ],
    )


def api_page(request):
    return _render_marketing_page(
        request,
        page_key="api",
        title=_("APIs ready for enterprise integration"),
        eyebrow=_("API"),
        description=_("%(product)s provides secure, organization-scoped APIs for document upload, session tracking, and access to audit results and reports.")
        % {"product": django_settings.PRODUCT_NAME},
        bullets=[
            _("API authentication with clear tenant isolation."),
            _("Endpoints for uploads, tracking, and reports."),
            _("Suitable for ERP systems and internal business portals."),
        ],
    )


def careers(request):
    return _render_marketing_page(
        request,
        page_key="careers",
        title=_("Working at %(product)s") % {"product": django_settings.PRODUCT_NAME},
        eyebrow=_("Careers"),
        description=_("We are building a financial product that demands precise execution, and we look for teams that combine technical depth with financial and regulatory understanding."),
        bullets=[
            _("Opportunities in engineering, product, and user experience."),
            _("A focus on trust-sensitive B2B products."),
            _("A practical environment for shipping real operational products."),
        ],
    )


@ensure_csrf_cookie
def login_view(request):
    if request.method == "GET" and request.GET.get("cancel") == "1":
        clear_pending_verification(request)
        return redirect("frontend:login")

    if request.user.is_authenticated:
        if not getattr(request.user, "is_email_verified", False):
            return redirect("frontend:otp_verify")
        return redirect("frontend:dashboard")

    if request.method == "GET" and has_pending_verification(request):
        return redirect("frontend:otp_verify")

    if request.method == "POST":
        serializer = LoginSerializer(
            data={
                "email": request.POST.get("email", "").strip(),
                "password": request.POST.get("password", ""),
            }
        )
        if not serializer.is_valid():
            return JsonResponse({"success": False, "error": _first_error(serializer.errors, flow="login")}, status=400)

        user = serializer.validated_data["user"]
        if not user.is_email_verified:
            try:
                challenge, sent = issue_email_otp(user, request)
            except EmailOTPError as exc:
                return JsonResponse(
                    {"success": False, "error": exc.message, "redirect": "/verify-email/"},
                    status=_otp_error_status(exc),
                )
            return JsonResponse(_otp_pending_payload(user, challenge, sent=sent))

        return JsonResponse(_verified_login_payload(request, user))

    return render(request, "auth/portal.html", _public_ctx(request, auth_mode="login"))


@ensure_csrf_cookie
def register_view(request):
    if request.method == "GET" and request.GET.get("cancel") == "1":
        clear_pending_verification(request)
        return redirect("frontend:register")

    if request.user.is_authenticated:
        if not getattr(request.user, "is_email_verified", False):
            return redirect("frontend:otp_verify")
        return redirect("frontend:dashboard")

    if request.method == "GET" and has_pending_verification(request):
        return redirect("frontend:otp_verify")

    if request.method == "POST":
        payload = {
            "full_name": request.POST.get("full_name", "").strip(),
            "email": request.POST.get("email", "").strip(),
            "password": request.POST.get("password", ""),
            "password_confirm": request.POST.get("password_confirm", ""),
            # §A.1/§A.2 — required. Passed through even when blank so the
            # serializer produces the field-specific error rather than a
            # generic one; validation belongs to the serializer, not here.
            "phone": request.POST.get("phone", "").strip(),
            "country": request.POST.get("country", "").strip(),
            "primary_benefit": request.POST.get("primary_benefit", "").strip(),
        }

        # §A.3 — optional. Only forwarded when supplied, so an omitted field
        # stays "not provided" rather than becoming an empty-string choice.
        for optional_field in ("city", "company_name", "employee_count", "sector", "heard_about"):
            value = request.POST.get(optional_field, "").strip()
            if value:
                payload[optional_field] = value

        organization_name = request.POST.get("organization_name", "").strip()
        if organization_name:
            payload["organization_name"] = organization_name

        # request in context: the serializer derives the auto-captured block
        # (IP, device, language, referrer, campaign) from it — those values are
        # never read from POST, so a client cannot forge them.
        serializer = RegisterSerializer(data=payload, context={"request": request})
        if not serializer.is_valid():
            return JsonResponse({"success": False, "error": _first_error(serializer.errors, flow="register")}, status=400)

        user = serializer.save()
        try:
            challenge, sent = issue_email_otp(user, request)
        except EmailOTPError as exc:
            return JsonResponse(
                {"success": False, "error": exc.message, "redirect": "/verify-email/"},
                status=_otp_error_status(exc),
            )

        return JsonResponse(_otp_pending_payload(user, challenge, sent=sent))

    return render(request, "auth/portal.html", _public_ctx(request, auth_mode="register"))


def logout_view(request):
    clear_pending_verification(request)
    logout(request)
    return redirect("frontend:home")


def google_oauth_login(request):
    if request.user.is_authenticated:
        return redirect(_post_auth_redirect(request.user))

    try:
        state = secrets.token_urlsafe(32)
        request.session[GOOGLE_OAUTH_STATE_SESSION_KEY] = state
        request.session.modified = True
        return redirect(build_google_oauth_authorization_url(state))
    except GoogleOAuthError as exc:
        return _redirect_to_login_with_auth_error(exc.code)


def google_oauth_callback(request):
    if request.user.is_authenticated:
        return redirect(_post_auth_redirect(request.user))

    code = (request.GET.get("code") or "").strip()
    state = (request.GET.get("state") or "").strip()
    expected_state = request.session.pop(GOOGLE_OAUTH_STATE_SESSION_KEY, "")

    if not code:
        return _redirect_to_login_with_auth_error("missing_code")
    if not state or not expected_state or state != expected_state:
        return _redirect_to_login_with_auth_error("invalid_state")

    try:
        token_payload = exchange_google_code_for_tokens(code)
        profile = fetch_google_user_profile(token_payload["access_token"])
        user, _ = get_or_create_local_user_from_google_profile(profile)
    except GoogleOAuthError as exc:
        return _redirect_to_login_with_auth_error(exc.code)

    if not user.is_active:
        return _redirect_to_login_with_auth_error("inactive_user")

    request.session[POST_LOGIN_TOKENS_SESSION_KEY] = complete_verified_login(request, user)
    request.session.modified = True
    return redirect("frontend:dashboard")


def google_pending(request):
    if request.user.is_authenticated and request.user.is_email_verified:
        return redirect("frontend:dashboard")
    if has_pending_verification(request) or (request.user.is_authenticated and not request.user.is_email_verified):
        return redirect("frontend:otp_verify")
    return redirect("frontend:login")


@ensure_csrf_cookie
def otp_verify(request):
    pending_user = get_pending_verification_user(request)
    if pending_user is None:
        return redirect("frontend:login")

    if pending_user.is_email_verified:
        clear_pending_verification(request)
        return redirect("frontend:dashboard")

    challenge = get_latest_pending_challenge(pending_user)
    initial_status = 200
    initial_error_message = ""
    if challenge is None:
        try:
            challenge, _sent = issue_email_otp(pending_user, request)
        except EmailOTPError as exc:
            initial_status = _otp_error_status(exc)
            initial_error_message = exc.message
            challenge = None

    state = get_challenge_state(challenge)
    form = EmailOTPVerifyForm(request.POST or None)
    context = _public_ctx(
        request,
        auth_mode="otp",
        otp_form=form,
        otp_user=pending_user,
        masked_email=mask_email_address(pending_user.email),
        resend_cooldown_seconds=state["resend_available_in"],
        otp_expires_in_seconds=state["expires_in"],
        otp_attempts_remaining=state["attempts_remaining"],
        otp_resend_remaining=state["resend_remaining"],
        error_message=initial_error_message,
    )

    if request.method == "POST":
        if not form.is_valid():
            error_message = _first_error(form.errors)
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": error_message}, status=400)
            context["error_message"] = error_message
            return render(request, "auth/otp_verify.html", context, status=400)

        try:
            verify_email_otp(pending_user, form.cleaned_data["otp_code"])
        except EmailOTPError as exc:
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": exc.message}, status=_otp_error_status(exc))
            context["error_message"] = exc.message
            return render(request, "auth/otp_verify.html", context, status=_otp_error_status(exc))

        payload = _verified_login_payload(request, pending_user)
        payload["message"] = _("Your email address has been verified successfully.")
        return JsonResponse(payload)

    return render(request, "auth/otp_verify.html", context, status=initial_status)


@ensure_csrf_cookie
def otp_resend(request):
    if request.method != "POST":
        return redirect("frontend:otp_verify")

    pending_user = get_pending_verification_user(request)
    if pending_user is None:
        return JsonResponse({"success": False, "error": _("The verification session has expired. Please sign in again.")}, status=400)

    form = EmailOTPResendForm(request.POST)
    if not form.is_valid():
        return JsonResponse({"success": False, "error": _("Unable to resend the code right now.")}, status=400)

    try:
        challenge = resend_email_otp(pending_user, request)
    except EmailOTPError as exc:
        return JsonResponse(
            {
                "success": False,
                "error": exc.message,
                "retry_after": exc.wait_seconds,
            },
            status=_otp_error_status(exc),
        )

    state = get_challenge_state(challenge)
    return JsonResponse(
        {
            "success": True,
            "message": _("A new verification code has been sent to your email address."),
            "masked_email": mask_email_address(pending_user.email),
            "resend_cooldown_seconds": state["resend_available_in"],
            "otp_expires_in_seconds": state["expires_in"],
            "attempts_remaining": state["attempts_remaining"],
        }
    )


# ✅ PHASE 1 FIX: MFA Login Verification View 
@ensure_csrf_cookie
def mfa_login_verify(request):
    """
    Handle MFA verification during login flow.
    POST with temp_token + code → Issue full JWT tokens
    """
    if request.method != "POST":
        return JsonResponse({"error": _("Method not allowed")}, status=405)
    
    try:
        import json
        import pyotp
        from datetime import timedelta
        from django.utils import timezone
        from rest_framework_simplejwt.tokens import RefreshToken
        from apps.authentication.models import User, AuditLog
        from core.utils.audit import log_action
        
        data = json.loads(request.body)
        temp_token_str = data.get("temp_token", "").strip()
        totp_code = str(data.get("code", "")).strip()
        
        if not temp_token_str or not totp_code or len(totp_code) != 6:
            return JsonResponse(
                {"error": _("Temporary token and 6-digit TOTP code are required")},
                status=400
            )
        
        # Verify temporary token and get user
        try:
            temp_token = RefreshToken(temp_token_str)
            user_id = temp_token.get("user_id")
            user = User.objects.get(id=user_id)
        except Exception:
            return JsonResponse(
                {"error": _("Invalid or expired temporary token")},
                status=401
            )
        
        # Verify TOTP code
        if not user.mfa_enabled or not user.mfa_secret:
            return JsonResponse(
                {"error": _("MFA not enabled for this user")},
                status=400
            )
        
        totp = pyotp.TOTP(user.mfa_secret)
        if not totp.verify(totp_code, valid_window=1):
            # Track failed MFA attempts
            user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
            if user.failed_login_attempts >= 5:
                user.locked_until = timezone.now() + timedelta(minutes=30)
            user.save()
            
            return JsonResponse(
                {"error": _("Invalid or expired TOTP code")},
                status=401
            )
        
        # MFA successful - issue full tokens
        user.failed_login_attempts = 0
        user.last_login = timezone.now()
        user.save()
        
        # Log the action
        log_action(request, AuditLog.Action.LOGIN, "user", str(user.id), details={"mfa": "verified"})
        
        # Generate JWT tokens
        refresh = RefreshToken.for_user(user)
        payload = {
            "success": True,
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "redirect": reverse("frontend:dashboard"),
            "message": _("Login successful!")
        }
        return JsonResponse(payload, status=200)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse(
            {"error": _("An error occurred during MFA verification")},
            status=500
        )


@login_required(login_url="/login/")
def dashboard(request):
    """Tenant-scoped dashboard.

    Shape: the KPI tiles + risk chart + monthly trend + recent activity + top
    risky vendors. Performance contract: ≤ 6 DB round trips per render
    (was 17+), 60-second cache per (org, day-bucket) so re-renders within
    the same minute are free.

    Correctness fixes vs the previous version:
      - total_amount is computed PER CURRENCY rather than blindly summing
        across SAR / USD / EUR (which it never converted).
      - extraction_accuracy_pct uses Avg() with NULL exclusion (was
        Sum/Count which biased low when ocr_confidence was NULL).
      - recent_invoices stays as model instances (not .values()) so the
        template's get_status_display() / get_risk_level_display() actually
        resolve. Bumped to 10 to match the "Last 10 analyzed documents"
        copy in the template.
      - doc_counts now covers all 21 typed doc types (was 7).
      - risk_breakdown aggregates across all typed docs (was invoice-only).
      - top_risky_vendors prefers the VendorProfile registry when present
        (deduplicated by (org, vendor_name) at write time) and falls back
        to invoice-name grouping only when the registry is empty.
    """
    # TADGEEG-G4.2 — a client (auditee) user's home is their evidence portal,
    # never the auditor dashboard.
    if getattr(request.user, "is_client", False):
        return redirect("frontend:client_evidence_list")

    from datetime import timedelta
    from django.core.cache import cache
    from django.db.models import Count, Sum, Q, Avg
    from django.utils import timezone
    from apps.invoices.models import Invoice, VendorProfile

    org = getattr(request.user, "organization", None)
    now = timezone.now()

    # Empty-state when the user has no org. Render the same template; the
    # template surfaces an "you're not in an organization yet" banner.
    if org is None:
        return render(request, "dashboard/index.html", _ctx(
            request, "dashboard",
            kpis=_empty_dashboard_kpis(),
            risk_breakdown={"low": 0, "medium": 0, "high": 0, "critical": 0},
            chart_series={"labels": [], "counts": [], "amounts": []},
            top_risky_vendors=[],
            recent_invoices=[],
            monthly_growth=0,
            no_organization=True,
        ))

    # TADGEEG-FIN-AUDIT-6B — evidence widget. Computed OUTSIDE
    # _build_dashboard_payload so that function's documented query budget is
    # unchanged; status_counts() is a single aggregate query.
    evidence_counts = _dashboard_evidence_counts(org)
    # 6C: review-throughput cards (waiting / pending / avg review time).
    evidence_summary = _dashboard_evidence_summary(org)
    # 6D: assurance widgets (integrity % / coverage % / verification status).
    evidence_assurance = _dashboard_evidence_assurance(org)

    cache_key = f"dashboard:v2:{org.id}:{now.strftime('%Y-%m-%d-%H')}"
    cached = cache.get(cache_key)
    if cached is not None:
        return render(request, "dashboard/index.html", _ctx(
            request, "dashboard", **cached, no_organization=False,
            evidence_counts=evidence_counts, evidence_summary=evidence_summary,
            evidence_assurance=evidence_assurance,
        ))

    payload = _build_dashboard_payload(org, now)
    # Fraud analysis may still run internally for audit evidence, but its
    # customer-facing dashboard signal is a package capability.  Do not show a
    # partial counter as though the organisation bought fraud detection.
    from apps.billing.services.features import feature_decision
    fraud_enabled = feature_decision(org, "fraud_detection", minimum_tier="advanced").enabled
    payload["fraud_detection_enabled"] = fraud_enabled
    if not fraud_enabled:
        payload["kpis"] = {**payload["kpis"], "fraud_alerts": None}
    # 60 s feels right: KPIs aren't real-time-critical and a fresh upload
    # already invalidates downstream report caches separately.
    cache.set(cache_key, payload, 60)
    return render(request, "dashboard/index.html", _ctx(
        request, "dashboard", **payload, no_organization=False,
        evidence_counts=evidence_counts, evidence_summary=evidence_summary,
        evidence_assurance=evidence_assurance,
    ))



# Dashboard aggregation moved to apps/frontend/selectors/dashboard.py — a view
# module was the wrong home for functions that take (org, now) and return a
# dict. Re-exported so existing callers and tests keep working.
from apps.frontend.selectors.dashboard import (  # noqa: E402,F401
    _build_dashboard_payload,
    _dashboard_evidence_counts,
)



def _dashboard_evidence_summary(org) -> dict:
    """Evidence review-throughput cards for the dashboard (6C).

    Never raises: a widget must not be able to break the dashboard.
    """
    try:
        from apps.audit.services import evidence_lifecycle as lc_service
        return lc_service.dashboard_summary(organization=org)
    except Exception:  # pragma: no cover - defensive
        return {}


def _dashboard_evidence_assurance(org) -> dict:
    """Evidence assurance widgets for the dashboard (6D).

    Never raises: a widget must not be able to break the dashboard.
    """
    try:
        from apps.audit.services import evidence_assurance as assurance
        return assurance.assurance_dashboard(organization=org)
    except Exception:  # pragma: no cover - defensive
        return {}


def _empty_dashboard_kpis() -> dict:
    return {
        "total_invoices": 0, "total_pos": 0, "total_amount_by_currency": {},
        "primary_currency": "SAR",
        "high_risk_count": 0, "fraud_alerts": 0, "compliance_alerts": 0,
        "pending_review": 0, "automation_pct": 0, "extraction_accuracy_pct": 0,
        "monthly_growth": 0, "vat_total": 0,
        "doc_counts": {},
    }




@login_required(login_url="/login/")
def upload(request):
    return render(request, "invoices/upload.html", _ctx(request, "upload"))


@login_required(login_url="/login/")
def invoices(request):
    """List invoices for the current user's organization, with optional filters."""
    from apps.invoices.models import Invoice

    org = getattr(request.user, "organization", None)
    # Tenant-isolation: a user with no organization MUST NOT see invoices
    # belonging to any other tenant. Empty queryset is the fail-safe default.
    qs = (Invoice.objects.filter(organization=org).select_related("organization")
          .order_by("-created_at")) if org else Invoice.objects.none()

    # Optional filters
    status_filter = (request.GET.get("status") or "").strip().lower()
    if status_filter and status_filter != "all":
        qs = qs.filter(status=status_filter)

    risk_filter = (request.GET.get("risk") or "").strip().lower()
    if risk_filter and risk_filter != "all":
        qs = qs.filter(risk_level=risk_filter)

    search = (request.GET.get("q") or "").strip()
    if search:
        from django.db.models import Q
        qs = qs.filter(
            Q(invoice_number__icontains=search)
            | Q(vendor_name__icontains=search)
            | Q(vendor_name_ar__icontains=search)
        )

    # Paginate — 25 per page is a good balance for a dense table.
    page_kwargs = _paginate(qs, request, per_page=25)
    invoices_list = list(page_kwargs["page_obj"].object_list)

    # Status counters for the tab pills (full org scope, not filtered)
    counters = {"all": 0, "pending": 0, "approved": 0, "flagged": 0, "rejected": 0}
    counter_qs = Invoice.objects.filter(organization=org) if org else Invoice.objects.none()
    counters["all"] = counter_qs.count()
    for s in ("pending", "approved", "flagged", "rejected"):
        counters[s] = counter_qs.filter(status=s).count()

    return render(
        request,
        "invoices/list.html",
        _ctx(
            request,
            "invoices",
            invoices=invoices_list,
            counters=counters,
            search=search,
            status_filter=status_filter or "all",
            risk_filter=risk_filter or "all",
            **page_kwargs,
        ),
    )


@login_required(login_url="/login/")
def invoice_detail(request, pk):
    try:
        from apps.invoices.models import Invoice, InvoiceAuditEvent

        organization = getattr(request.user, "organization", None)
        invoice = Invoice.objects.select_related("approved_by", "duplicate_of", "validation").get(pk=pk)
        if organization and invoice.organization != organization:
            return redirect("frontend:invoices")
        audit_trail = InvoiceAuditEvent.objects.filter(invoice=invoice).select_related("user").order_by("-timestamp")[:40]
    except Exception:
        return redirect("frontend:invoices")

    # Cross-doc linkage: PO ↔ GRN ↔ Payment + 3-way match (parity with Phase-2 detail pages).
    try:
        from core.services.cross_doc_linker import find_links
        cross_links = find_links("invoice", invoice, organization)
    except Exception:
        cross_links = {}

    invoice_display   = _build_invoice_display(invoice)
    user_can_override = request.user.has_perm("invoices.can_override_approval")

    # NOTE: /invoices/<uuid>/ is NOT served by this function — apps.invoices.urls
    # is included ahead of the frontend catch-all, so the DRF InvoiceDetailView
    # content-negotiates and renders detail_premium.html itself. The manual
    # review panel is wired up there. This function stays as the fallback.
    return render(
        request,
        "invoices/detail_premium.html",
        _ctx(
            request, "invoices",
            invoice=invoice,
            invoice_display=invoice_display,
            audit_trail=audit_trail,
            user_can_override=user_can_override,
            cross_links=cross_links,
        ),
    )


@login_required(login_url="/login/")
def invoice_pdf(request, pk):
    """Render the invoice's audit detail as a downloadable PDF."""
    from django.http import HttpResponse, Http404
    from django.template.loader import render_to_string
    from django.utils.translation import get_language
    from apps.invoices.models import Invoice, InvoiceAuditEvent

    organization = getattr(request.user, "organization", None)
    try:
        invoice = Invoice.objects.select_related("approved_by", "validation").get(pk=pk)
    except Invoice.DoesNotExist:
        raise Http404("Invoice not found")
    if organization and invoice.organization_id != organization.id:
        raise Http404("Invoice not found")

    audit_trail = list(
        InvoiceAuditEvent.objects.filter(invoice=invoice).select_related("user").order_by("-timestamp")[:40]
    )
    try:
        from core.services.cross_doc_linker import find_links
        cross_links = find_links("invoice", invoice, organization)
    except Exception:
        cross_links = {}

    # Language-aware AI fields + a rich failed-rules table built from
    # validation_details (each rule carries description / message / severity
    # in Arabic). Compute here so the template stays purely presentational.
    lang = (get_language() or "ar")[:2]
    other = "en" if lang == "ar" else "ar"

    def _pick(base: str):
        for suffix in (f"_{lang}", f"_{other}", ""):
            val = getattr(invoice, f"{base}{suffix}", None)
            if isinstance(val, str) and val.strip():
                return val
            if isinstance(val, (list, tuple)) and any(v for v in val):
                return val
        return getattr(invoice, base, None)

    ai_summary_localized = _pick("ai_summary") or ""
    ai_recommendations_localized = _pick("ai_recommendations") or []
    if isinstance(ai_recommendations_localized, str):
        ai_recommendations_localized = [ai_recommendations_localized]

    # Build failed_rules_table from validation_details.
    _SEV_WEIGHT = {"critical": 25, "high": 15, "medium": 8, "low": 3, "info": 1}
    _SEV_AR = {"critical": "حرج", "high": "مرتفع", "medium": "متوسط",
               "low": "منخفض", "info": "للعلم"}
    _SEV_EN = {"critical": "Critical", "high": "High", "medium": "Medium",
               "low": "Low", "info": "Info"}
    _SEV_LABELS = _SEV_AR if lang == "ar" else _SEV_EN
    try:
        from core.services.invoice_validator import rule_description as _rule_desc
    except Exception:
        _rule_desc = lambda code: code  # noqa: E731
    failed_rules_table: list[dict] = []
    v = getattr(invoice, "validation", None)
    details = (getattr(v, "validation_details", None) or {}) if v else {}
    failed_codes = (getattr(v, "failed_rule_codes", None) or []) if v else []
    for code in failed_codes:
        d = details.get(code) if isinstance(details, dict) else None
        severity = ((d or {}).get("severity") or "medium").lower()
        # Translate the rule description to the active language; fall back to
        # the stored Arabic message, then to the rule code.
        text = _rule_desc(code) or (d or {}).get("message") or code
        weight = _SEV_WEIGHT.get(severity, _SEV_WEIGHT["medium"])
        failed_rules_table.append({
            "code":           code,
            "text":           text,
            "severity":       severity,
            "severity_label": _SEV_LABELS.get(severity, severity),
            "score":          weight,
        })
    # Sort: critical first, then high, medium, low. Stable on rule_code.
    _SEV_RANK = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
    failed_rules_table.sort(key=lambda r: (_SEV_RANK.get(r["severity"], 9), r["code"]))

    html = render_to_string(
        "invoices/invoice_pdf.html",
        _ctx(
            request, "invoices",
            invoice=invoice,
            invoice_display=_build_invoice_display(invoice),
            audit_trail=audit_trail,
            cross_links=cross_links,
            ai_summary_localized=ai_summary_localized,
            ai_recommendations_localized=ai_recommendations_localized,
            failed_rules_table=failed_rules_table,
        ),
        request=request,
    )

    try:
        from apps.reports.views import _render_report_pdf_bytes
        pdf_bytes = _render_report_pdf_bytes(html, request.build_absolute_uri("/"))
    except Exception as exc:
        # WeasyPrint unavailable / failed → return inline HTML so the user gets *something*
        return HttpResponse(html, content_type="text/html; charset=utf-8")

    safe_id = (invoice.invoice_number or str(invoice.id))[:60].replace(" ", "_")
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    from apps.reports.views import _attachment_disposition
    response["Content-Disposition"] = _attachment_disposition(f"invoice_{safe_id}.pdf")
    return response


@login_required(login_url="/login/")
def run_audit(request, doc_type, pk):
    """
    Manually re-run validators on a single document.
    POST only. Returns JSON for AJAX callers; redirects back to detail page on form POST.
    """
    from django.http import JsonResponse, Http404
    from django.shortcuts import redirect as _redirect
    from apps.documents.models import Document
    from apps.documents.typed_models import (
        PurchaseOrder, BankStatement, PayrollSheet, ExpenseReport,
        VATReturn, FixedAsset, SalesReceipt, GoodsReceiptNote, PaymentVoucher,
    )
    from apps.documents.typed_models_v2 import (
        SalesOrder, Quotation, ProformaInvoice, ReceiptVoucher, CashVoucher,
        GeneralLedger, Ledger, Contract, SupplierStatement, CustomerStatement, JournalEntry,
    )
    from apps.invoices.models import Invoice
    from core.services.doc_validators.doc_validators import run_document_validation

    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    org = getattr(request.user, "organization", None)
    if not org:
        return JsonResponse({"error": "no_organization"}, status=400)

    MODEL_MAP = {
        "invoice": Invoice, "sales_invoice": Invoice,
        "purchase_order": PurchaseOrder, "bank_statement": BankStatement,
        "payroll": PayrollSheet, "expense_report": ExpenseReport,
        "vat_return": VATReturn, "fixed_asset": FixedAsset, "sales_receipt": SalesReceipt,
        "goods_receipt_note": GoodsReceiptNote, "payment_voucher": PaymentVoucher,
        "sales_order": SalesOrder, "quotation": Quotation,
        "proforma_invoice": ProformaInvoice, "receipt_voucher": ReceiptVoucher,
        "cash_voucher": CashVoucher, "general_ledger": GeneralLedger, "ledger": Ledger,
        "contract": Contract, "supplier_statement": SupplierStatement,
        "customer_statement": CustomerStatement, "journal_entry": JournalEntry,
    }
    Model = MODEL_MAP.get(doc_type)
    if Model is None:
        return JsonResponse({"error": f"unknown doc_type: {doc_type}"}, status=400)

    obj = Model.objects.filter(organization=org, pk=pk).first()
    if obj is None:
        raise Http404(f"{doc_type} not found")

    try:
        result = run_document_validation(doc_type, obj)
        # Persist results to the row when fields exist (AuditMixin schema)
        for field in ("validation_score", "risk_level", "rules_passed", "rules_failed", "failed_rule_codes"):
            if hasattr(obj, field) and field in result:
                setattr(obj, field, result[field])
        obj.save(update_fields=[f for f in ("validation_score","risk_level","rules_passed","rules_failed","failed_rule_codes") if hasattr(obj, f)])
    except Exception as exc:
        return JsonResponse({"error": str(exc)[:200]}, status=500)

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if is_ajax:
        return JsonResponse({
            "ok": True, "doc_type": doc_type, "id": str(pk),
            "rules_passed": result.get("rules_passed", 0),
            "rules_failed": result.get("rules_failed", 0),
            "risk_level": result.get("risk_level", "low"),
        })
    # form POST → bounce back to detail page
    detail_paths = {
        "invoice": f"/invoices/{pk}/", "sales_invoice": f"/invoices/{pk}/",
        "purchase_order": f"/documents/purchase-orders/{pk}/",
        "bank_statement": f"/documents/bank-statements/{pk}/",
        "payroll": f"/documents/payroll/{pk}/",
        "expense_report": f"/documents/expense-reports/{pk}/",
        "vat_return": f"/documents/vat-returns/{pk}/",
        "fixed_asset": f"/documents/fixed-assets/{pk}/",
        "sales_receipt": f"/documents/sales-receipts/{pk}/",
        "goods_receipt_note": f"/documents/grns/{pk}/",
        "payment_voucher": f"/documents/payment-vouchers/{pk}/",
        "sales_order": f"/documents/sales-orders/{pk}/",
        "quotation": f"/documents/quotations/{pk}/",
        "proforma_invoice": f"/documents/proforma-invoices/{pk}/",
        "receipt_voucher": f"/documents/receipt-vouchers/{pk}/",
        "cash_voucher": f"/documents/cash-vouchers/{pk}/",
        "general_ledger": f"/documents/general-ledgers/{pk}/",
        "ledger": f"/documents/ledgers/{pk}/",
        "contract": f"/documents/contracts/{pk}/",
        "supplier_statement": f"/documents/supplier-statements/{pk}/",
        "customer_statement": f"/documents/customer-statements/{pk}/",
        "journal_entry": f"/documents/journal-entries/{pk}/",
    }
    return _redirect(detail_paths.get(doc_type, "/documents/"))


@login_required(login_url="/login/")
def vendor_detail(request, vendor_name):
    """Vendor 360 — every invoice / PO / payment / contract for one vendor."""
    from urllib.parse import unquote
    from django.db.models import Sum, Count, Q
    from datetime import timedelta
    from django.utils import timezone
    from apps.invoices.models import Invoice
    from apps.documents.typed_models import PurchaseOrder, GoodsReceiptNote, PaymentVoucher
    from apps.documents.typed_models_v2 import Contract

    org = getattr(request.user, "organization", None)
    name = unquote(vendor_name).strip()

    if not org:
        return render(request, "vendors/detail.html", _ctx(
            request, "vendors", vendor=None, message="No organization"))

    # Aggregate everything tied to this vendor name (case-insensitive match)
    invs = Invoice.objects.filter(organization=org, vendor_name__iexact=name)
    pos  = PurchaseOrder.objects.filter(organization=org, vendor_name__iexact=name)
    grns = GoodsReceiptNote.objects.filter(organization=org, vendor_name__iexact=name)
    pays = PaymentVoucher.objects.filter(organization=org, payee_name__iexact=name)
    contracts = Contract.objects.filter(organization=org, party_b__iexact=name)

    inv_total = invs.aggregate(s=Sum("total_amount"))["s"] or 0
    pay_total = pays.aggregate(s=Sum("total_amount"))["s"] or 0
    high_risk = invs.filter(risk_level__in=["high", "critical"]).count()
    duplicates = invs.filter(is_duplicate=True).count()

    # Risk score: weighted average from invoices
    if invs.exists():
        avg_risk = invs.aggregate(a=Sum("risk_score"))["a"] or 0
        risk_score = int(avg_risk / invs.count())
    else:
        risk_score = 0

    summary = {
        "name": name,
        "invoice_count": invs.count(),
        "po_count": pos.count(),
        "grn_count": grns.count(),
        "payment_count": pays.count(),
        "contract_count": contracts.count(),
        "total_invoiced": float(inv_total or 0),
        "total_paid": float(pay_total or 0),
        "outstanding": float((inv_total or 0) - (pay_total or 0)),
        "high_risk_count": high_risk,
        "duplicate_count": duplicates,
        "risk_score": risk_score,
        "risk_level": "critical" if risk_score >= 81 else "high" if risk_score >= 51
                      else "medium" if risk_score >= 21 else "low",
        "first_seen": invs.order_by("created_at").values_list("created_at", flat=True).first(),
        "last_seen":  invs.order_by("-created_at").values_list("created_at", flat=True).first(),
        "vat_numbers": list(invs.exclude(vendor_vat_number="").values_list("vendor_vat_number", flat=True).distinct()[:5]),
    }
    recent_invoices = list(invs.order_by("-invoice_date")[:25].values(
        "id", "invoice_number", "invoice_date", "total_amount", "currency",
        "status", "risk_level", "is_duplicate"
    ))
    return render(request, "vendors/detail.html", _ctx(
        request, "vendors",
        vendor=summary,
        invoices=recent_invoices,
        pos=list(pos.order_by("-po_date")[:10].values("id","po_number","po_date","total_amount","currency")),
        contracts=list(contracts.values("id","contract_number","start_date","end_date","contract_value","status")[:10]),
    ))


@login_required(login_url="/login/")
def audit_inbox(request):
    """Approval inbox — every flagged document waiting for a decision."""
    from apps.invoices.models import Invoice
    from apps.documents.typed_models import (PurchaseOrder, BankStatement, PayrollSheet,
        ExpenseReport, VATReturn, FixedAsset, SalesReceipt,
        GoodsReceiptNote, PaymentVoucher)
    from apps.documents.typed_models_v2 import (SalesOrder, Quotation, ProformaInvoice,
        ReceiptVoucher, CashVoucher, Contract, SupplierStatement, CustomerStatement)

    org = getattr(request.user, "organization", None)
    inbox = []
    if org:
        # Invoices flagged or critical
        for inv in Invoice.objects.filter(organization=org).filter(
            status__in=["flagged", "pending"]
        ).order_by("-created_at")[:50]:
            inbox.append({
                "doc_type": "invoice",
                "id": str(inv.id),
                "number": inv.invoice_number or "—",
                "party": inv.vendor_name or "—",
                "amount": float(inv.total_amount or 0),
                "currency": inv.currency or "SAR",
                "risk_level": inv.risk_level or "low",
                "rules_failed": getattr(getattr(inv, "validation", None), "rules_failed", 0) or 0,
                "url": f"/invoices/{inv.id}/",
                "created_at": inv.created_at,
            })
        # All Phase-2 typed docs flagged
        type_map = [
            ("purchase_order", PurchaseOrder, "po_number", "vendor_name"),
            ("bank_statement", BankStatement, "account_number", "bank_name"),
            ("goods_receipt_note", GoodsReceiptNote, "grn_number", "vendor_name"),
            ("payment_voucher", PaymentVoucher, "payment_number", "payee_name"),
            ("contract", Contract, "contract_number", "party_b"),
            ("sales_order", SalesOrder, "so_number", "customer_name"),
            ("expense_report", ExpenseReport, "report_number", "employee_name"),
        ]
        for dtype, Model, num_field, party_field in type_map:
            for obj in Model.objects.filter(organization=org, audit_status__in=["flagged","pending"]).order_by("-created_at")[:30]:
                inbox.append({
                    "doc_type": dtype,
                    "id": str(obj.id),
                    "number": getattr(obj, num_field, "") or "—",
                    "party":  getattr(obj, party_field, "") or "—",
                    "amount": float(getattr(obj, "total_amount", 0) or 0),
                    "currency": getattr(obj, "currency", "SAR") or "SAR",
                    "risk_level": getattr(obj, "risk_level", "low") or "low",
                    "rules_failed": getattr(obj, "rules_failed", 0) or 0,
                    "url": f"/documents/{dtype.replace('_','-')}s/{obj.id}/" if dtype != "goods_receipt_note" else f"/documents/grns/{obj.id}/",
                    "created_at": obj.created_at,
                })
    # Sort: critical first, then high, then by date
    severity_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    inbox.sort(key=lambda x: (severity_order.get(x["risk_level"], 9),
                              -(x.get("created_at").timestamp() if x.get("created_at") else 0)))
    return render(request, "audit/inbox.html", _ctx(
        request, "audit",
        inbox=inbox,
        total_pending=len(inbox),
        critical_count=sum(1 for x in inbox if x["risk_level"] == "critical"),
        high_count=sum(1 for x in inbox if x["risk_level"] == "high"),
    ))


@login_required(login_url="/login/")
def load_demo_data(request):
    """
    POST-only — populate the caller's organisation with a small set of
    demo invoices, POs, GRNs, payments, contracts, and statements so first-
    time users see filled dashboards / charts / reports instead of zeros.

    Idempotent: no-op (with a warning message) if the org already has data.
    """
    from django.http import JsonResponse, Http404
    from django.shortcuts import redirect as _redirect
    from datetime import date, timedelta
    from decimal import Decimal
    from django.core.files.uploadedfile import SimpleUploadedFile

    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    org = getattr(request.user, "organization", None)
    if not org:
        return JsonResponse({"error": "no_organization"}, status=400)

    from apps.invoices.models import Invoice
    if Invoice.objects.filter(organization=org).count() > 5:
        msg = "Org already has data — demo loader skipped to avoid duplicates."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"ok": False, "skipped": True, "message": msg})
        return _redirect("/dashboard/")

    from apps.documents.models import Document
    from apps.documents.typed_models import (
        PurchaseOrder, GoodsReceiptNote, PaymentVoucher, BankStatement,
    )
    from apps.documents.typed_models_v2 import (
        Contract, SupplierStatement, SalesOrder,
    )

    def _doc(name):
        return Document.objects.create(
            organization=org,
            file=SimpleUploadedFile(name, b"demo"),
            original_filename=name, file_size=10,
            mime_type="application/octet-stream",
            document_type=Document.DocumentType.OTHER,
            uploaded_by=request.user,
        )

    today = date.today()
    created = {"invoices": 0, "pos": 0, "grns": 0, "payments": 0,
               "contracts": 0, "supplier_statements": 0, "sales_orders": 0}

    # 6 invoices spanning low/medium/high risk + one duplicate pair
    vendors = [
        ("Acme Corp",   "300000111100003", Decimal("12500"), "low"),
        ("Acme Corp",   "300000111100003", Decimal("12500"), "high"),  # duplicate of #1
        ("Beta LLC",    "300000222200003", Decimal("87000"), "medium"),
        ("Gamma Inc",   "300000333300003", Decimal("4500"),  "low"),
        ("Delta Trade", "300000444400003", Decimal("250000"),"high"),
        ("Epsilon",     "300000555500003", Decimal("3200"),  "critical"),
    ]
    invoices = []
    for i, (vendor, vat, amount, risk) in enumerate(vendors, 1):
        inv = Invoice.objects.create(
            organization=org,
            invoice_number=f"DEMO-INV-{i:03d}",
            invoice_date=today - timedelta(days=i * 5),
            vendor_name=vendor, vendor_vat_number=vat,
            currency="SAR",
            subtotal=amount, vat_amount=amount * Decimal("0.15"),
            total_amount=amount * Decimal("1.15"),
            risk_score={"low": 12, "medium": 35, "high": 65, "critical": 88}[risk],
            risk_level=risk, status="validated",
            is_duplicate=(i == 2),
        )
        invoices.append(inv)
        created["invoices"] += 1

    # 3 POs matching the first 3 invoices (clean 3-way match for #1, mismatched for #3)
    for i, inv in enumerate(invoices[:3], 1):
        po = PurchaseOrder.objects.create(
            organization=org, document=_doc(f"po-{i}.pdf"), uploaded_by=request.user,
            po_number=f"DEMO-PO-{i:03d}",
            po_date=inv.invoice_date - timedelta(days=10),
            vendor_name=inv.vendor_name,
            total_amount=inv.total_amount if i != 3 else inv.total_amount * Decimal("0.9"),
            currency="SAR",
        )
        created["pos"] += 1
        # GRN for the first PO only (clean match)
        if i == 1:
            GoodsReceiptNote.objects.create(
                organization=org, document=_doc(f"grn-{i}.pdf"), uploaded_by=request.user,
                grn_number=f"DEMO-GRN-{i:03d}",
                grn_date=inv.invoice_date - timedelta(days=3),
                po_number=po.po_number, po_id=po.id,
                invoice_number=inv.invoice_number,
                vendor_name=inv.vendor_name,
                total_amount=inv.total_amount,
            )
            created["grns"] += 1
        # Payment for invoice #1 only
        if i == 1:
            PaymentVoucher.objects.create(
                organization=org, document=_doc(f"pv-{i}.pdf"), uploaded_by=request.user,
                payment_number=f"DEMO-PV-{i:03d}",
                payment_date=inv.invoice_date + timedelta(days=14),
                payee_name=inv.vendor_name,
                amount=inv.total_amount, total_amount=inv.total_amount,
                linked_invoice_id=inv.id, linked_invoice_number=inv.invoice_number,
                linked_po_number=po.po_number, payment_method="bank_transfer",
                approval_status="approved",
            )
            created["payments"] += 1

    # 1 contract covering Acme
    Contract.objects.create(
        organization=org, document=_doc("ct.pdf"), uploaded_by=request.user,
        contract_number="DEMO-CT-001", title="Annual Services Agreement",
        party_a=org.name, party_b="Acme Corp", party_b_type="vendor",
        start_date=date(today.year, 1, 1), end_date=date(today.year, 12, 31),
        is_signed=True, status="active",
        contract_value=Decimal("500000"), currency="SAR",
        payment_terms="Net 30",
    )
    created["contracts"] += 1

    # 1 supplier statement
    SupplierStatement.objects.create(
        organization=org, document=_doc("ss.pdf"), uploaded_by=request.user,
        supplier_name="Acme Corp",
        period_from=today - timedelta(days=90), period_to=today,
        opening_balance=Decimal("0"), closing_balance=Decimal("28750"),
        total_invoiced=Decimal("28750"), total_paid=Decimal("0"),
    )
    created["supplier_statements"] += 1

    # 1 sales order
    SalesOrder.objects.create(
        organization=org, document=_doc("so.pdf"), uploaded_by=request.user,
        so_number="DEMO-SO-001", so_date=today - timedelta(days=2),
        customer_name="Big Customer Co",
        currency="SAR", subtotal=Decimal("10000"),
        vat_amount=Decimal("1500"), total_amount=Decimal("11500"),
        status="confirmed",
    )
    created["sales_orders"] += 1

    msg = f"Loaded demo data: {created}"
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"ok": True, "created": created, "message": msg})
    return _redirect("/dashboard/")


@login_required(login_url="/login/")
def doc_list_export_excel(request, doc_type):
    """Export the current org's rows for a doc-type as XLSX."""
    from django.http import HttpResponse, Http404
    from apps.documents.typed_models import (
        PurchaseOrder, BankStatement, PayrollSheet, ExpenseReport,
        VATReturn, FixedAsset, SalesReceipt, GoodsReceiptNote, PaymentVoucher,
    )
    from apps.documents.typed_models_v2 import (
        SalesOrder, Quotation, ProformaInvoice, ReceiptVoucher, CashVoucher,
        GeneralLedger, Ledger, Contract, SupplierStatement, CustomerStatement, JournalEntry,
    )
    MODEL_MAP = {
        "purchase_order": PurchaseOrder, "bank_statement": BankStatement,
        "payroll": PayrollSheet, "expense_report": ExpenseReport,
        "vat_return": VATReturn, "fixed_asset": FixedAsset, "sales_receipt": SalesReceipt,
        "goods_receipt_note": GoodsReceiptNote, "payment_voucher": PaymentVoucher,
        "sales_order": SalesOrder, "quotation": Quotation,
        "proforma_invoice": ProformaInvoice, "receipt_voucher": ReceiptVoucher,
        "cash_voucher": CashVoucher, "general_ledger": GeneralLedger, "ledger": Ledger,
        "contract": Contract, "supplier_statement": SupplierStatement,
        "customer_statement": CustomerStatement, "journal_entry": JournalEntry,
    }
    Model = MODEL_MAP.get(doc_type)
    if Model is None:
        raise Http404("unknown doc_type")

    org = getattr(request.user, "organization", None)
    qs = Model.objects.filter(organization=org).order_by("-created_at")[:5000] if org else Model.objects.none()

    # Pick the most useful columns from the model's concrete fields (skip JSON / FK / AuditMixin internals)
    skip = {"organization", "uploaded_by", "document", "validation_details", "ai_recommendations"}
    cols = [
        f.name for f in Model._meta.get_fields()
        if getattr(f, "concrete", False) and not f.is_relation
        and f.name not in skip
        and f.get_internal_type() not in {"JSONField"}
    ][:25]

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return HttpResponse("openpyxl not installed on server", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = doc_type[:31]
    # Header
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col.replace("_", " ").title())
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="003366")
    # Rows
    for r, obj in enumerate(qs, 2):
        for c, col in enumerate(cols, 1):
            v = getattr(obj, col, None)
            if hasattr(v, "isoformat"):
                v = v.isoformat()
            ws.cell(row=r, column=c, value=v if v is None or isinstance(v, (int, float, str, bool)) else str(v))

    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{doc_type}_export.xlsx"'
    return response


@login_required(login_url="/login/")
def invoices_export(request, fmt):
    """
    Export the org's invoices as CSV or Excel.

    GET /invoices/export.csv  → text/csv
    GET /invoices/export.xlsx → application/vnd...spreadsheet
    Same status/risk/q filters as the list view apply.
    """
    from django.http import HttpResponse, Http404
    from apps.invoices.models import Invoice

    org = getattr(request.user, "organization", None)
    qs = (Invoice.objects.filter(organization=org).select_related("organization")
          .order_by("-created_at")) if org else Invoice.objects.none()

    status_filter = (request.GET.get("status") or "").strip().lower()
    if status_filter and status_filter != "all":
        qs = qs.filter(status=status_filter)
    risk_filter = (request.GET.get("risk") or "").strip().lower()
    if risk_filter and risk_filter != "all":
        qs = qs.filter(risk_level=risk_filter)
    search = (request.GET.get("q") or "").strip()
    if search:
        from django.db.models import Q
        qs = qs.filter(
            Q(invoice_number__icontains=search)
            | Q(vendor_name__icontains=search)
            | Q(vendor_name_ar__icontains=search)
        )

    qs = qs[:5000]
    cols = [
        ("invoice_number", "Invoice #"),
        ("invoice_date",   "Date"),
        ("vendor_name",    "Vendor"),
        ("vendor_vat_number", "Vendor VAT"),
        ("customer_name",  "Customer"),
        ("subtotal",       "Subtotal"),
        ("vat_amount",     "VAT"),
        ("total_amount",   "Total"),
        ("currency",       "Currency"),
        ("status",         "Status"),
        ("risk_level",     "Risk"),
        ("risk_score",     "Risk Score"),
        ("is_duplicate",   "Duplicate"),
        ("has_qr_code",    "Has QR"),
        ("qr_code_valid",  "QR Valid"),
        ("created_at",     "Uploaded"),
    ]

    def _val(obj, name):
        v = getattr(obj, name, None)
        if hasattr(v, "isoformat"):
            return v.isoformat()
        if v is None:
            return ""
        return v

    if fmt == "csv":
        import csv, io
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([h for _, h in cols])
        for inv in qs:
            w.writerow([_val(inv, k) for k, _ in cols])
        response = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="invoices.csv"'
        return response

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return HttpResponse("openpyxl not installed", status=500)
        import io as _io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"
        for c, (_, h) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="003366")
        for r, inv in enumerate(qs, 2):
            for c, (k, _) in enumerate(cols, 1):
                v = _val(inv, k)
                ws.cell(row=r, column=c, value=v if isinstance(v, (int, float, str, bool)) else str(v))
        buf = _io.BytesIO()
        wb.save(buf)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="invoices.xlsx"'
        return response

    raise Http404("unsupported format")


@login_required(login_url="/login/")
def invoice_subreport(request, kind):
    """
    HTML wrapper for the JSON sub-reports under /api/v1/invoices/reports/.
    Renders a server-side page with the same data the JSON endpoint returns,
    so auditors can read the report instead of parsing JSON.

    kind ∈ {risk, duplicates, vendors, spend}
    """
    from django.http import Http404
    from apps.invoices.views import (
        InvoiceRiskReportView, DuplicateInvoiceReportView,
        VendorRiskReportView, SpendAnalysisReportView,
    )

    view_cls = {
        "risk":       InvoiceRiskReportView,
        "duplicates": DuplicateInvoiceReportView,
        "vendors":    VendorRiskReportView,
        "spend":      SpendAnalysisReportView,
    }.get(kind)
    if view_cls is None:
        raise Http404("unknown report")

    # Re-use the API view's data path so HTML and JSON stay in sync.
    api_response = view_cls.as_view()(request)
    data = api_response.data if hasattr(api_response, "data") else {}

    return render(
        request,
        "invoices/subreport.html",
        _ctx(
            request, "reports",
            report_kind=kind,
            report_data=data,
            report_title={
                "risk":       "تقرير المخاطر",
                "duplicates": "تقرير الفواتير المكررة",
                "vendors":    "تقرير مخاطر الموردين",
                "spend":      "تحليل الإنفاق",
            }[kind],
        ),
    )


@login_required(login_url="/login/")
def batches(request):
    return render(request, "invoices/batches.html", _ctx(request, "batches"))


@login_required(login_url="/login/")
def batch_detail(request, pk):
    return render(request, "invoices/batch_detail.html", _ctx(request, "batches", batch_id=str(pk)))


@login_required(login_url="/login/")
def audit_session_detail(request, pk):
    return render(request, "invoices/session_detail.html", _ctx(request, "batches", audit_session_id=str(pk)))


@login_required(login_url="/login/")
def ledger_dashboard(request):
    """Phase 7.1 — General Ledger landing page.

    Shows the trial balance + the most recent journal entries. Posting +
    voiding go via the JSON API at /api/v1/ledger/*.
    """
    from apps.ledger import reports as gl_reports
    from apps.ledger import services as gl
    from apps.ledger.models import Account, JournalEntry

    org = getattr(request.user, "organization", None)
    tb = {"rows": [], "totals": {}}
    entries = []
    accounts_count = 0
    if org:
        gl.ensure_default_accounts(org)
        tb = gl_reports.trial_balance(org)[0]
        entries = list(
            JournalEntry.objects.filter(organization=org)
            .order_by("-entry_date", "-created_at")[:30]
        )
        accounts_count = Account.objects.filter(organization=org, is_active=True).count()

    return render(request, "ledger/dashboard.html", _ctx(
        request, "compliance",
        trial_balance=tb,
        entries=entries,
        accounts_count=accounts_count,
    ))


@login_required(login_url="/login/")
def banking_dashboard(request):
    """Phase 5 — Bank Connectors landing page.

    Lists existing connections + the reconciliation queue. Sync + onboarding
    actions hit the JSON API at /api/v1/banking/*.
    """
    from apps.banking.models import BankConnection, Reconciliation
    from apps.banking.connectors.registry import REGISTRY

    org = getattr(request.user, "organization", None)
    connections, recons, counts = [], [], {}
    if org:
        connections = list(
            BankConnection.objects.filter(organization=org)
            .prefetch_related("accounts").order_by("bank_code", "-updated_at")
        )
        recons = list(
            Reconciliation.objects.filter(organization=org,
                                          status=Reconciliation.Status.SUGGESTED)
            .select_related("transaction", "invoice")
            .order_by("-score", "-created_at")[:25]
        )
        for s in ("suggested", "confirmed", "rejected", "manual"):
            counts[s] = Reconciliation.objects.filter(organization=org, status=s).count()

    return render(request, "banking/dashboard.html", _ctx(
        request, "compliance",
        connections=connections,
        recons=recons,
        counts=counts,
        available_banks=[(c, cls.display_name) for c, cls in REGISTRY.items()],
    ))


@login_required(login_url="/login/")
def zatca_dashboard(request):
    """Phase 4 — ZATCA Phase 2 compliance dashboard.

    Server-renders a shell that hits the JSON dashboard endpoint for the
    real numbers. Devices + submissions tables come straight from the DB.
    """
    from apps.zatca.models import EGSDevice, InvoiceSubmission, RejectionCode
    from apps.zatca.rejection_codes import seed_rejection_codes

    org = getattr(request.user, "organization", None)
    seed_rejection_codes()

    devices = []
    submissions = []
    if org:
        devices = list(
            EGSDevice.objects.filter(organization=org).order_by("-updated_at")
        )
        submissions = list(
            InvoiceSubmission.objects.filter(organization=org)
            .order_by("-created_at")[:25]
        )

    return render(request, "zatca/dashboard.html", _ctx(
        request, "compliance",
        devices=devices,
        submissions=submissions,
    ))


@login_required(login_url="/login/")
def alerts_dashboard(request):
    """Phase 3.2 — manage alert rules + see the events log + acknowledge.

    Server-rendered shell; the form-driven interactions hit the JSON API
    under ``/api/v1/alerts/*``.
    """
    from apps.alerts.models import AlertEvent, AlertRule

    org = getattr(request.user, "organization", None)
    rules: list = []
    events: list = []
    counts = {"sent": 0, "failed": 0, "suppressed": 0, "acknowledged": 0}
    if org:
        rules = list(AlertRule.objects.filter(organization=org).order_by("-updated_at"))
        events = list(
            AlertEvent.objects.filter(organization=org)
            .select_related("rule")
            .order_by("-sent_at")[:100]
        )
        for s in counts:
            counts[s] = AlertEvent.objects.filter(organization=org, status=s).count()

    return render(request, "alerts/index.html", _ctx(
        request, "audit",
        rules=rules, events=events, counts=counts,
    ))


@login_required(login_url="/login/")
def streaming_ops(request):
    """Phase 3.1 — live ops dashboard for the continuous-auditing stream.

    Shows throughput, p95 latency, error rate, recent anomaly hits, and
    bus stats (queue depth + DLQ length) so on-call can see the pipeline's
    health at a glance.
    """
    from apps.streaming.worker import metrics
    from apps.streaming.models import AnomalyHit

    org = getattr(request.user, "organization", None)
    window = int(request.GET.get("window") or 30)
    m = metrics(window_minutes=window)

    recent_hits = []
    if org:
        recent_hits = list(
            AnomalyHit.objects.filter(organization=org)
            .order_by("-occurred_at")[:25]
        )

    return render(request, "streaming/ops.html", _ctx(
        request, "audit",
        metrics=m,
        recent_hits=recent_hits,
        window_minutes=window,
    ))


@login_required(login_url="/login/")
def rule_builder(request):
    """Visual custom-rule builder page.

    Server-side just renders the shell — the form is fully client-side
    (Alpine.js + fetch against /api/v1/audit/rule-builder/*) so the auditor
    can build, test, and publish rules without a page reload.
    """
    return render(request, "audit/rule_builder.html", _ctx(request, "audit"))


@login_required(login_url="/login/")
def working_papers(request):
    """List all working papers in the user's org with filters."""
    from apps.audit.models import WorkingPaper

    org = getattr(request.user, "organization", None)
    qs = (WorkingPaper.objects.filter(organization=org)
          .select_related("prepared_by", "reviewed_by", "partner_signed_by")
          .prefetch_related("signatures")
          .order_by("-updated_at")) if org else WorkingPaper.objects.none()

    status_filter = (request.GET.get("status") or "").strip().lower()
    if status_filter and status_filter != "all":
        qs = qs.filter(status=status_filter)

    type_filter = (request.GET.get("type") or "").strip().lower()
    if type_filter and type_filter != "all":
        qs = qs.filter(paper_type=type_filter)

    search = (request.GET.get("q") or "").strip()
    if search:
        from django.db.models import Q
        qs = qs.filter(Q(reference__icontains=search) | Q(title__icontains=search))

    counts = {s: 0 for s in ("draft", "ready_for_review", "reviewed", "locked", "archived")}
    if org:
        for s in counts:
            counts[s] = WorkingPaper.objects.filter(organization=org, status=s).count()

    return render(request, "working_papers/list.html", _ctx(
        request, "audit",
        papers=list(qs[:200]),
        counts=counts,
        status_filter=status_filter or "all",
        type_filter=type_filter or "all",
        search=search,
        paper_types=WorkingPaper.PaperType.choices,
    ))


@login_required(login_url="/login/")
def working_paper_detail(request, pk):
    """Detail page — preparer edits, reviewer/partner sign, anyone reads."""
    from django.http import Http404
    from apps.audit.models import WorkingPaper

    org = getattr(request.user, "organization", None)
    try:
        paper = (WorkingPaper.objects
                 .select_related("prepared_by", "reviewed_by", "partner_signed_by")
                 .prefetch_related("signatures__user", "related_invoices",
                                   "related_papers", "attachments")
                 .get(pk=pk))
    except WorkingPaper.DoesNotExist:
        raise Http404("Working paper not found")

    if org and paper.organization_id != org.id:
        raise Http404("Working paper not found")

    return render(request, "working_papers/detail.html", _ctx(
        request, "audit",
        paper=paper,
        signatures=list(paper.signatures.all()),
        can_submit=(paper.status == WorkingPaper.Status.DRAFT
                    and paper.prepared_by_id == request.user.id),
        can_review=(paper.status == WorkingPaper.Status.READY_FOR_REVIEW
                    and (request.user.is_superuser or
                         request.user.role in {request.user.Role.SENIOR_AUDITOR,
                                               request.user.Role.CHIEF_AUDIT_OFFICER,
                                               request.user.Role.ADMIN})),
        can_partner_sign=(paper.status == WorkingPaper.Status.REVIEWED
                          and (request.user.is_superuser or
                               request.user.role in {request.user.Role.CHIEF_AUDIT_OFFICER,
                                                     request.user.Role.ADMIN,
                                                     request.user.Role.EXTERNAL_AUDITOR})),
    ))


@login_required(login_url="/login/")
def working_paper_create(request):
    """Create a new draft working paper."""
    from django.shortcuts import redirect as _redirect
    from apps.audit.models import WorkingPaper
    from apps.audit.services.working_papers import next_reference

    org = getattr(request.user, "organization", None)
    if not org:
        return _redirect("frontend:working_papers")

    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        paper_type = (request.POST.get("paper_type") or "").strip()
        content_text = (request.POST.get("content") or "").strip()
        if not title or paper_type not in dict(WorkingPaper.PaperType.choices):
            return render(request, "working_papers/form.html", _ctx(
                request, "audit",
                paper_types=WorkingPaper.PaperType.choices,
                title=title, paper_type=paper_type, content=content_text,
                error="Title and a valid paper type are required.",
            ))
        paper = WorkingPaper.objects.create(
            organization=org,
            reference=next_reference(org, paper_type),
            title=title,
            paper_type=paper_type,
            content={"notes": content_text} if content_text else {},
            status=WorkingPaper.Status.DRAFT,
            prepared_by=request.user,
        )
        return _redirect("frontend:working_paper_detail", pk=paper.id)

    return render(request, "working_papers/form.html", _ctx(
        request, "audit",
        paper_types=WorkingPaper.PaperType.choices,
    ))


@login_required(login_url="/login/")
def working_paper_action(request, pk):
    """POST endpoint: submit / review-approve / review-reject / partner-sign."""
    from django.http import HttpResponse, JsonResponse, Http404
    from django.shortcuts import redirect as _redirect
    from apps.audit.models import WorkingPaper
    from django.core.exceptions import PermissionDenied
    from apps.audit.services.working_papers import (
        submit_for_review, review_paper, partner_sign, WorkingPaperWorkflowError,
    )

    if request.method != "POST":
        return HttpResponse(status=405)

    org = getattr(request.user, "organization", None)
    try:
        paper = WorkingPaper.objects.get(pk=pk)
    except WorkingPaper.DoesNotExist:
        raise Http404("Working paper not found")
    if org and paper.organization_id != org.id:
        raise Http404("Working paper not found")

    action = (request.POST.get("action") or "").strip().lower()
    notes = (request.POST.get("notes") or "").strip()
    typed_name = (request.POST.get("typed_name") or "").strip() or request.user.full_name
    ip = request.META.get("REMOTE_ADDR")
    sig_data = {"name": typed_name}

    try:
        if action == "submit":
            submit_for_review(paper, request.user)
        elif action == "review_approve":
            review_paper(paper, request.user, decision="approve",
                         notes=notes, signature_data=sig_data, ip_address=ip)
        elif action == "review_reject":
            review_paper(paper, request.user, decision="reject",
                         notes=notes, signature_data=sig_data, ip_address=ip)
        elif action == "partner_sign":
            partner_sign(paper, request.user, notes=notes,
                         signature_data=sig_data, ip_address=ip)
        else:
            return JsonResponse({"error": "unknown action"}, status=400)
    except WorkingPaperWorkflowError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except PermissionDenied as exc:
        return JsonResponse({"error": str(exc)}, status=403)

    return _redirect("frontend:working_paper_detail", pk=paper.id)


@login_required(login_url="/login/")
def audit_integrity(request):
    """Walk every audit hash chain in the user's org and surface any break.

    Phase 1.1 of the Enterprise Roadmap — this is the tripwire that fires
    when a row in the audit trail is mutated outside the application (e.g.
    via a SQL shell or backup-restore that lost a row).
    """
    from apps.audit.integrity import verify_chain, HashChainMixin
    from apps.invoices.models import InvoiceAuditEvent

    org = getattr(request.user, "organization", None)
    if not org:
        return render(request, "audit/integrity.html", _ctx(
            request, "audit", reports=[], chain_models=[], total_rows=0,
            total_breaks=0, all_intact=True,
        ))

    # Collect every concrete subclass of HashChainMixin so the page grows
    # automatically as Working Papers, AuditCase, etc. join the chain in
    # later phases.
    chain_models: list[type] = []
    stack = list(HashChainMixin.__subclasses__())
    seen = set()
    while stack:
        cls = stack.pop()
        if cls in seen:
            continue
        seen.add(cls)
        if not getattr(cls._meta, "abstract", False):
            chain_models.append(cls)
        stack.extend(cls.__subclasses__())

    reports = []
    total_rows   = 0
    total_breaks = 0
    for Model in chain_models:
        try:
            rep = verify_chain(Model, str(org.id))
        except Exception as exc:
            reports.append({
                "model": Model.__name__,
                "rows_checked": 0,
                "head_hash": "",
                "head_hash_full": "",
                "is_intact": False,
                "break_count": 1,
                "breaks": [{"chain_position": 0, "row_id": "",
                            "reason": f"verification_failed: {exc}",
                            "expected": "", "actual": ""}],
            })
            continue
        d = rep.to_dict()
        reports.append(d)
        total_rows   += d["rows_checked"]
        total_breaks += d["break_count"]

    return render(request, "audit/integrity.html", _ctx(
        request, "audit",
        reports=reports,
        chain_models=[c.__name__ for c in chain_models],
        total_rows=total_rows,
        total_breaks=total_breaks,
        all_intact=(total_breaks == 0),
    ))


@login_required(login_url="/login/")
def audit_tools(request):
    """ISA 320 Materiality calculator + ISA 530 Sampling engine — single page.

    Reads benchmark + sample-size from query string when supplied; otherwise
    renders an empty form. The page exposes the same maths the Big-4 firms use
    so an external auditor can run a rough plan without leaving Tadgeeg.

    Phase 1.2 extensions:
      • judgment factors that pull the percentage inside the band
        (industry_risk, control_environment, prior_misstatements,
         going_concern_doubt, first_year_audit)
      • multi-component allocation across segments (`?components=Riyadh:60,Jeddah:40`)
      • sample error projection (most-likely error + upper-error-limit)
        when the auditor enters per-item misstatements as `?errors=100,250,...`
    """
    from apps.audit.services import materiality as M
    from apps.audit.services import sampling as S
    from apps.invoices.models import Invoice
    from decimal import Decimal

    org = getattr(request.user, "organization", None)
    benchmark_key = request.GET.get("benchmark") or "profit_before_tax"
    benchmark_amount = request.GET.get("amount") or ""
    sample_method = request.GET.get("method") or "random"
    sample_size = int(request.GET.get("size") or 0) or None
    seed = int(request.GET.get("seed") or 42)
    confidence_pct = int(request.GET.get("confidence") or 95)

    # Phase 1.2: judgment factors. The form posts each as its own param
    # (?industry_risk=high&control_environment=weak&...) so they round-trip
    # naturally through the GET query string.
    judgment_factors = {
        k: v for k, v in (
            ("industry_risk",         request.GET.get("industry_risk")),
            ("control_environment",   request.GET.get("control_environment")),
            ("prior_misstatements",   request.GET.get("prior_misstatements")),
            ("going_concern_doubt",   request.GET.get("going_concern_doubt")),
            ("first_year_audit",      request.GET.get("first_year_audit")),
        ) if v
    }

    # Multi-component allocation: "Riyadh:60,Jeddah:30,Dammam:10".
    components_param = (request.GET.get("components") or "").strip()
    components = []
    if components_param:
        for piece in components_param.split(","):
            if ":" not in piece:
                continue
            name, weight = piece.split(":", 1)
            try:
                components.append({
                    "name": name.strip(),
                    "weight_pct": Decimal(weight.strip()),
                })
            except Exception:
                continue

    # Per-item sample errors: "100,250.50,1200".
    errors_param = (request.GET.get("errors") or "").strip()
    sample_errors = []
    if errors_param:
        for piece in errors_param.split(","):
            try:
                sample_errors.append(Decimal(piece.strip()))
            except Exception:
                continue

    materiality_result = None
    flagged_above = []
    sampling_result = None
    error_projection = None
    pop_count = 0

    if org:
        invoice_qs = Invoice.objects.filter(organization=org).order_by("-created_at")
        pop_count = invoice_qs.count()

        if benchmark_amount:
            try:
                materiality_result = M.calculate(
                    benchmark_amount=Decimal(benchmark_amount),
                    benchmark_key=benchmark_key,
                    judgment_factors=judgment_factors or None,
                    components=components or None,
                ).to_dict()
                flagged_above = M.flag_invoices_above_threshold(
                    invoice_qs[:1000],
                    Decimal(str(materiality_result["performance_materiality"])),
                )[:50]
            except Exception:
                # ISA 320 materiality. A blank panel where a number should be
                # is indistinguishable from "not calculated yet", so the
                # auditor cannot tell a bad benchmark from a broken feature.
                # At minimum it must be findable in the log.
                logger.exception(
                    "materiality calculation failed for organization=%s "
                    "benchmark=%s amount=%s — the panel will render empty",
                    getattr(org, "pk", "?"), benchmark_key, benchmark_amount,
                )
                materiality_result = None

        if pop_count and sample_size:
            pop = list(invoice_qs[:5000])
            try:
                if sample_method == "systematic":
                    sampling_result = S.systematic_sample(pop, sample_size, seed).to_dict()
                elif sample_method == "monetary_unit":
                    interval = (
                        sum(float(getattr(i, "total_amount", 0) or 0) for i in pop) / sample_size
                        if sample_size > 0 else 1
                    )
                    sampling_result = S.monetary_unit_sample(pop, interval, seed=seed).to_dict()
                else:
                    sampling_result = S.random_sample(pop, sample_size, seed).to_dict()
            except Exception:
                # ISA 530 sampling. Same reasoning as materiality above: an
                # empty result is not evidence of anything, and silently
                # producing one in an audit tool is worse than an error page.
                logger.exception(
                    "sampling failed for organization=%s method=%s "
                    "population=%s sample_size=%s",
                    getattr(org, "pk", "?"), sample_method, pop_count, sample_size,
                )
                sampling_result = None

            # Phase 1.2: project errors to the population if the auditor
            # supplied per-item misstatements. The interval comes from the
            # sampling result (MUS) or pop-total ÷ sample-size for attribute
            # methods.
            if sample_errors and sampling_result:
                pm = (Decimal(str(materiality_result["performance_materiality"]))
                      if materiality_result else None)
                if sample_method == "monetary_unit":
                    interval = sampling_result.get("notes", [])
                    # Pull the numeric interval out of the notes line.
                    interval_val = sum(float(getattr(i, "total_amount", 0) or 0)
                                       for i in pop) / sample_size
                else:
                    interval_val = sum(float(getattr(i, "total_amount", 0) or 0)
                                       for i in pop) / sample_size
                try:
                    error_projection = S.project_error(
                        sample_errors=sample_errors,
                        sampling_interval=interval_val,
                        population_size=pop_count,
                        sample_size=sample_size,
                        confidence_pct=confidence_pct,
                        performance_materiality=pm,
                        method=sample_method,
                    ).to_dict()
                except Exception:
                    logger.exception(
                        "error projection failed for organization=%s method=%s "
                        "— the auditor supplied misstatements and gets no projection",
                        getattr(org, "pk", "?"), sample_method,
                    )
                    error_projection = None

    return render(request, "audit/tools.html", _ctx(
        request, "reports",
        benchmark_key=benchmark_key,
        benchmark_amount=benchmark_amount,
        sample_method=sample_method,
        sample_size=sample_size or "",
        seed=seed,
        confidence_pct=confidence_pct,
        components_param=components_param,
        errors_param=errors_param,
        judgment_factors=judgment_factors,
        materiality_result=materiality_result,
        flagged_above=flagged_above,
        sampling_result=sampling_result,
        error_projection=error_projection,
        pop_count=pop_count,
        benchmarks=[
            {"key": k, "name": v.name,
             "pct_low": float(v.pct_low * 100), "pct_high": float(v.pct_high * 100),
             "rationale": v.rationale}
            for k, v in M.BENCHMARKS.items()
        ],
        suggested_size=S.suggest_sample_size(pop_count) if pop_count else 0,
    ))


@login_required(login_url="/login/")
def reports(request):
    """Reports landing — real KPIs + per-doc-type report links."""
    org = getattr(request.user, "organization", None)
    kpis = {
        "total_invoices": 0, "total_amount": 0, "avg_risk_score": 0,
        "high_risk_count": 0, "duplicate_count": 0, "rules_failed_30d": 0,
        "doc_counts_total": 0,
    }
    recent_reports = []
    if org:
        from django.db.models import Sum, Avg, Count, Q
        from datetime import timedelta
        from django.utils import timezone
        from apps.invoices.models import Invoice
        from apps.documents.typed_models import (
            PurchaseOrder, BankStatement, PayrollSheet, ExpenseReport,
            VATReturn, FixedAsset, SalesReceipt, GoodsReceiptNote, PaymentVoucher,
        )
        from apps.documents.typed_models_v2 import (
            SalesOrder, Quotation, ProformaInvoice, ReceiptVoucher, CashVoucher,
            GeneralLedger, Ledger, Contract, SupplierStatement, CustomerStatement,
            JournalEntry,
        )

        invs = Invoice.objects.filter(organization=org)
        kpis["total_invoices"]   = invs.count()
        kpis["total_amount"]     = float(invs.aggregate(s=Sum("total_amount"))["s"] or 0)
        kpis["avg_risk_score"]   = round(float(invs.aggregate(a=Avg("risk_score"))["a"] or 0), 1)
        kpis["high_risk_count"]  = invs.filter(risk_level__in=["high", "critical"]).count()
        kpis["duplicate_count"]  = invs.filter(is_duplicate=True).count()

        # Sum across all 21 doc-type tables
        all_typed_models = [
            PurchaseOrder, BankStatement, PayrollSheet, ExpenseReport, VATReturn,
            FixedAsset, SalesReceipt, GoodsReceiptNote, PaymentVoucher,
            SalesOrder, Quotation, ProformaInvoice, ReceiptVoucher, CashVoucher,
            GeneralLedger, Ledger, Contract, SupplierStatement, CustomerStatement,
            JournalEntry,
        ]
        kpis["doc_counts_total"] = (
            kpis["total_invoices"] +
            sum(M.objects.filter(organization=org).count() for M in all_typed_models)
        )

        # ZATCA Phase 2 compliance, MEASURED. Phase 2 requires a cryptographic
        # QR on every invoice, so "share of this org's invoices carrying a QR
        # that validated" is a defensible number. The template used to print a
        # hardcoded 95% here, under a label that told the auditor it was their
        # organisation's figure — a fabricated measurement inside an audit
        # product. None means "no invoices yet", which the template must render
        # as "—", never as 0% or 95%.
        kpis["zatca_qr_valid_count"] = invs.filter(qr_code_valid=True).count()
        kpis["zatca_compliance_pct"] = (
            round(100.0 * kpis["zatca_qr_valid_count"] / kpis["total_invoices"], 1)
            if kpis["total_invoices"] else None
        )

        # Top 5 most-recent flagged invoices (anchor for "recent reports")
        recent_reports = list(
            invs.filter(Q(risk_level__in=["high", "critical"]) | Q(is_duplicate=True))
                .order_by("-updated_at")
                .values("id", "invoice_number", "vendor_name", "total_amount",
                        "risk_level", "is_duplicate", "updated_at")[:5]
        )

    return render(request, "reports/index.html", _ctx(
        request, "reports",
        report_types=_report_types(),
        selected_type=request.GET.get("type", "invoice"),
        report_kpis=kpis,
        recent_reports=recent_reports,
        # Extraction accuracy is NOT computable from tenant data — it needs a
        # labelled ground-truth set. The one honest source is a validation run
        # (apps.auditing.models.AIValidationRun, populated by the
        # validate_ai_claim command). Until one is approved this is None and
        # the card says so, rather than printing the 98.5% that used to be
        # hardcoded in the template.
        extraction_accuracy=_latest_measured_extraction_accuracy(),
        # Rule precision, in contrast, IS computable from this tenant's own
        # data — the auditors' verdicts are the labels. It answers a different
        # question from extraction accuracy (did the rule fire correctly, not
        # did OCR read the field correctly), so it is a separate figure and
        # never a substitute.
        rule_accuracy=_measured_rule_accuracy(org),
    ))


def _measured_rule_accuracy(organization):
    """Precision across all judged findings, with the coverage behind it.

    Returns None when nothing has been judged. Coverage travels with the
    number and is not optional: 100% precision over four judged findings is
    not a fact about the engine, and a template that shows the ratio without
    the sample size rebuilds the unsourced claim this work removed.
    """
    if organization is None:
        return None
    try:
        from apps.audit.services.finding_feedback import FindingFeedbackService
    except ImportError:  # pragma: no cover - app not installed
        # ImportError only. A broad catch here would also swallow a real fault
        # inside the service module and render it as "no data" — the exact
        # failure shape that cost hours on the billing menu.
        return None

    service = FindingFeedbackService()
    rows = service.rule_precision(organization)
    judged = sum(r["judged"] for r in rows)
    if not judged:
        return None

    true_positives = sum(r["true_positives"] for r in rows)
    coverage = service.coverage(organization)
    # Worst measured rule first — the actionable end of the list. Rules with
    # no judgements are excluded rather than shown as perfect.
    worst = next((r for r in rows if r["precision"] is not None), None)
    return {
        "pct": round(100.0 * true_positives / judged, 1),
        "judged": judged,
        "coverage_pct": coverage["percent"],
        "worst_rule": worst["rule_code"] if worst else "",
        "worst_pct": round(100.0 * worst["precision"], 1) if worst else None,
    }


def _latest_measured_extraction_accuracy():
    """Approved OCR/extraction accuracy from the validation harness, or None.

    None is the correct answer when nothing has been measured, and it must
    stay distinguishable from 0.0 — the same unlimited-vs-zero trap the billing
    quotas hit. Callers render None as "not measured", not as a number.
    """
    try:
        from apps.auditing.models import AIValidationRun
    except ImportError:  # pragma: no cover - app not installed
        return None

    run = (
        AIValidationRun.objects
        .filter(component=AIValidationRun.Component.OCR,
                decision=AIValidationRun.Decision.APPROVED)
        .exclude(field_accuracy=None)
        .order_by("-created_at")
        .first()
    )
    if run is None:
        return None
    return {
        "pct": round(100.0 * run.field_accuracy, 1),
        "model_version": run.model_version,
        "dataset": run.dataset.name,
    }


def _weight_for_rule(rule_code):
    prefix = (rule_code or "").split("-", 1)[0]
    if prefix in {"DUP", "VAT", "ANO"}:
        return "مرتفع"
    if prefix in {"INV", "CTL"}:
        return "متوسط"
    return "منخفض"


def _failed_rules_with_invoice_refs(top_failed_rules, validations):
    """Enrich each failed rule with affected-invoice list AND severity/score.

    score = severity_weight × failure_count, using the same weight scheme as
    `core/services/doc_validators/doc_validators.SEVERITY_WEIGHTS`:
        critical = 25, high = 15, medium = 8, low = 3.

    severity is derived from the rule_code prefix when the catalog doesn't
    resolve it. Both the bilingual `severity_label` and the numeric
    `score` are exposed so the PDF / web report can render either form.
    """
    _SEV_WEIGHT = {"critical": 25, "high": 15, "medium": 8, "low": 3, "info": 1}
    _SEV_AR = {"critical": "حرج", "high": "مرتفع", "medium": "متوسط",
               "low": "منخفض", "info": "للعلم"}
    # Heuristic prefix → severity. Tuned to match the doc_validators rule
    # catalog used by the invoice audit pipeline. Unknown prefixes default
    # to "medium" so the score is non-zero but doesn't dominate.
    _PREFIX_SEV = {
        "DUP": "critical", "VAT": "critical",
        "ANO": "high",     "CTL": "high",     "DOC": "high",
        "INV": "medium",   "HDR": "medium",
    }

    top_codes = [r.get("rule_code") for r in (top_failed_rules or []) if r.get("rule_code")]
    top_codes_set = set(top_codes)
    buckets = {code: [] for code in top_codes}

    for vr in validations:
        inv_no = vr.invoice.invoice_number or str(vr.invoice_id)[:8]
        for code in (vr.failed_rule_codes or []):
            if code not in top_codes_set:
                continue
            if inv_no not in buckets[code]:
                buckets[code].append(inv_no)

    # Try to resolve severity from the canonical rule catalog when possible.
    try:
        from apps.rule_engine.catalog import resolve_rule_catalog_metadata
        def _catalog_severity(c):
            try:
                return getattr(resolve_rule_catalog_metadata(c), "severity", "") or ""
            except Exception:
                return ""
    except Exception:
        def _catalog_severity(c):  # noqa: E306
            return ""

    enriched = []
    for rule in (top_failed_rules or []):
        code = rule.get("rule_code") or ""
        invoice_numbers = buckets.get(code, [])
        failure_count = int(rule.get("failures") or 0)
        severity = (rule.get("severity")
                    or _catalog_severity(code)
                    or _PREFIX_SEV.get(code.split("-", 1)[0], "medium"))
        severity = str(severity).lower()
        weight = _SEV_WEIGHT.get(severity, _SEV_WEIGHT["medium"])
        enriched.append({
            "rule_code":      code,
            "description":    rule.get("description") or code,
            "failure_count":  failure_count,
            "invoice_numbers": invoice_numbers,
            "invoice_count":  len(invoice_numbers),
            # New: severity + numeric score for the failed-rules table.
            "severity":        severity,
            "severity_label":  _SEV_AR.get(severity, severity),
            "score":           weight * failure_count,
        })
    return enriched


def _build_high_risk_violations(invoice_rows, validation_map, rule_catalog):
    """Enrich each high-risk invoice with its full violation list (severity-aware).

    Each invoice row carries every failed rule (not just the first 4) so an
    auditor can see the complete picture for the document. Violations include
    severity so the template can color-code per-row severity badges.
    """
    from apps.reports.services.findings_service import severity_for, recommendation_for

    rows = []
    for row in (invoice_rows or []):
        inv_id = row.get("id")
        vr = validation_map.get(inv_id)
        violations = []

        if vr:
            details = vr.validation_details or {}
            for code in (vr.failed_rule_codes or []):  # ALL violations, not just first 4
                detail = details.get(code, {}) if isinstance(details, dict) else {}
                reason = (
                    detail.get("reason")
                    or detail.get("message")
                    or detail.get("note")
                    or str(rule_catalog.get(code, code))
                )
                violations.append({
                    "rule_code": code,
                    "description": rule_catalog.get(code, code),
                    "reason": reason,
                    "severity": severity_for(code),
                    "recommendation": str(recommendation_for(code)),
                })

        score = float(row.get("risk_score") or 0)
        rows.append({
            "id": inv_id,
            "invoice_number": row.get("invoice_number") or "-",
            "vendor_name": row.get("vendor_name") or "",
            "amount": row.get("total_amount") or 0,
            "currency": row.get("currency") or "SAR",
            "date": row.get("invoice_date") or "-",
            "risk_level": row.get("risk_level") or "low",
            "risk_score": score,
            "violations": violations,
            "violation_count": len(violations),
        })

    return rows


# Sections that each report kind should expose. Used by the template to
# conditionally render content. "audit_summary" is the canonical full report.
_REPORT_KIND_SECTIONS = {
    "audit_summary": {"executive", "compliance", "high_risk", "failed_rules", "risk", "duplicates", "vendor", "recommendations", "rule_engine"},
    "risk_analysis": {"executive", "risk", "high_risk", "duplicates"},
    "compliance":    {"executive", "compliance", "rule_engine"},
    "vendor":        {"executive", "vendor", "high_risk"},
    "trend":         {"executive", "risk", "vendor"},
    "isa700":        {"executive", "isa700", "compliance", "recommendations"},
}

# Pretty labels for each kind, shown in the report header.
_REPORT_KIND_LABELS = {
    "audit_summary": ("تقرير التدقيق الشامل", "Full Audit Summary"),
    "risk_analysis": ("تحليل المخاطر", "Risk Analysis"),
    "compliance":    ("تقرير الامتثال", "Compliance Report"),
    "vendor":        ("تحليل الموردين", "Vendor Analysis"),
    "trend":         ("تحليل الاتجاهات", "Trend Analysis"),
    "isa700":        ("ISA 700 Opinion", "ISA 700 Opinion"),
}

# Document type → display label. Used when type ≠ invoice to show a friendly
# banner. Wrapped in gettext_lazy so the localized form is picked up on
# dereference — this keeps strings like "Total Purchase Orders" properly
# translated in narrative text. Use the canonical name `gettext_lazy` so
# Django's makemessages xgettext pass recognises the call.
from django.utils.translation import gettext_lazy
_DOC_TYPE_LABELS = {
    # Original 8 types
    "invoice":            gettext_lazy("Invoices"),
    "purchase_order":     gettext_lazy("Purchase Orders"),
    "bank_statement":     gettext_lazy("Bank Statements"),
    "payroll":            gettext_lazy("Payroll"),
    "expense_report":     gettext_lazy("Expense Reports"),
    "vat_return":         gettext_lazy("VAT Returns"),
    "tax_declaration":    gettext_lazy("VAT Returns"),
    "fixed_asset":        gettext_lazy("Fixed Assets"),
    "sales_receipt":      gettext_lazy("Sales Receipts"),
    # Phase-1 additions (10 new types)
    "sales_invoice":      gettext_lazy("Sales Invoices"),
    "purchase_invoice":   gettext_lazy("Purchase Invoices"),
    "sales_order":        gettext_lazy("Sales Orders"),
    "quotation":          gettext_lazy("Quotations"),
    "proforma_invoice":   gettext_lazy("Proforma Invoices"),
    "grn":                gettext_lazy("Goods Receipt Notes"),
    "payment_voucher":    gettext_lazy("Payment Vouchers"),
    "receipt_voucher":    gettext_lazy("Receipt Vouchers"),
    "cash_voucher":       gettext_lazy("Cash Vouchers"),
    "journal_entry":      gettext_lazy("Journal Entries"),
    "general_ledger":     gettext_lazy("General Ledgers"),
    "ledger":             gettext_lazy("Ledgers"),
    "contract":           gettext_lazy("Contracts"),
    "tax_vat_document":   gettext_lazy("Tax / VAT Documents"),
    "supplier_statement": gettext_lazy("Supplier Statements"),
    "customer_statement": gettext_lazy("Customer Statements"),
}

# Singular noun for "Total {type}" KPI / heading copy. Wrapped in
# gettext_lazy (already imported above) so makemessages picks them up and
# the template gets the right localized form via `{{ type_singular }}`.
_DOC_TYPE_SINGULAR = {
    # Original 8 types
    "invoice":            gettext_lazy("Invoice"),
    "purchase_order":     gettext_lazy("Purchase Order"),
    "bank_statement":     gettext_lazy("Bank Statement"),
    "payroll":            gettext_lazy("Payroll Sheet"),
    "expense_report":     gettext_lazy("Expense Report"),
    "vat_return":         gettext_lazy("VAT Return"),
    "tax_declaration":    gettext_lazy("VAT Return"),
    "fixed_asset":        gettext_lazy("Fixed Asset"),
    "sales_receipt":      gettext_lazy("Sales Receipt"),
    # Phase-1 additions
    "sales_invoice":      gettext_lazy("Sales Invoice"),
    "purchase_invoice":   gettext_lazy("Purchase Invoice"),
    "sales_order":        gettext_lazy("Sales Order"),
    "quotation":          gettext_lazy("Quotation"),
    "proforma_invoice":   gettext_lazy("Proforma Invoice"),
    "grn":                gettext_lazy("Goods Receipt Note"),
    "payment_voucher":    gettext_lazy("Payment Voucher"),
    "receipt_voucher":    gettext_lazy("Receipt Voucher"),
    "cash_voucher":       gettext_lazy("Cash Voucher"),
    "journal_entry":      gettext_lazy("Journal Entry"),
    "general_ledger":     gettext_lazy("General Ledger"),
    "ledger":             gettext_lazy("Ledger"),
    "contract":           gettext_lazy("Contract"),
    "tax_vat_document":   gettext_lazy("Tax / VAT Document"),
    "supplier_statement": gettext_lazy("Supplier Statement"),
    "customer_statement": gettext_lazy("Customer Statement"),
}

# Per-type report title for <title>, breadcrumb, and report header.
_DOC_TYPE_REPORT_TITLE = {
    # Original 8 types
    "invoice":            gettext_lazy("Invoice Audit Report"),
    "purchase_order":     gettext_lazy("Purchase Order Audit Report"),
    "bank_statement":     gettext_lazy("Bank Statement Audit Report"),
    "payroll":            gettext_lazy("Payroll Audit Report"),
    "expense_report":     gettext_lazy("Expense Report Audit"),
    "vat_return":         gettext_lazy("VAT Return Audit Report"),
    "tax_declaration":    gettext_lazy("VAT Return Audit Report"),
    "fixed_asset":        gettext_lazy("Fixed Asset Audit Report"),
    "sales_receipt":      gettext_lazy("Sales Receipt Audit Report"),
    # Phase-1 additions
    "sales_invoice":      gettext_lazy("Sales Invoice Audit Report"),
    "purchase_invoice":   gettext_lazy("Purchase Invoice Audit Report"),
    "sales_order":        gettext_lazy("Sales Order Audit Report"),
    "quotation":          gettext_lazy("Quotation Audit Report"),
    "proforma_invoice":   gettext_lazy("Proforma Invoice Audit Report"),
    "grn":                gettext_lazy("Goods Receipt Note Audit Report"),
    "payment_voucher":    gettext_lazy("Payment Voucher Audit Report"),
    "receipt_voucher":    gettext_lazy("Receipt Voucher Audit Report"),
    "cash_voucher":       gettext_lazy("Cash Voucher Audit Report"),
    "journal_entry":      gettext_lazy("Journal Entry Audit Report"),
    "general_ledger":     gettext_lazy("General Ledger Audit Report"),
    "ledger":             gettext_lazy("Ledger Audit Report"),
    "contract":           gettext_lazy("Contract Audit Report"),
    "tax_vat_document":   gettext_lazy("Tax / VAT Document Audit Report"),
    "supplier_statement": gettext_lazy("Supplier Statement Audit Report"),
    "customer_statement": gettext_lazy("Customer Statement Audit Report"),
}


@login_required(login_url="/login/")
def invoice_audit_report(request):
    from django.db.models import Avg, Count, Sum
    from apps.reports.models import Report
    from apps.invoices.models import InvoiceValidationResult
    from core.services.invoice_validator import RULES

    org = getattr(request.user, "organization", None)
    if not org:
        return redirect("frontend:reports")

    # ── Read query params ──────────────────────────────────────────────────
    selected_kind = request.GET.get("kind", "audit_summary").lower()
    if selected_kind not in _REPORT_KIND_SECTIONS:
        selected_kind = "audit_summary"

    selected_type = (request.GET.get("type") or "invoice").lower()
    type_label = _DOC_TYPE_LABELS.get(selected_type, "Documents")
    is_invoice_report = selected_type in ("invoice", "")

    # Per-doc-type detail URL prefix. The Findings Register and high-risk
    # invoices table use this to deep-link to the right detail page —
    # /invoices/<uuid>/ for invoices, /documents/<slug>/<uuid>/ for everything
    # else. Always end with a trailing slash since the views expect it.
    _DOC_DETAIL_PREFIX = {
        # Original 8
        "invoice":            "/invoices/",
        "purchase_order":     "/documents/purchase-orders/",
        "bank_statement":     "/documents/bank-statements/",
        "payroll":            "/documents/payroll/",
        "expense_report":     "/documents/expense-reports/",
        "vat_return":         "/documents/vat-returns/",
        "fixed_asset":        "/documents/fixed-assets/",
        "sales_receipt":      "/documents/sales-receipts/",
        # Phase-2 — list pages don't exist yet (Phase 4); detail URLs will be
        # added when each type gets its UI surface. For now the deep-link
        # falls back to the audit-cases page so users always land somewhere
        # sensible.
        "sales_invoice":      "/invoices/",
        "purchase_invoice":   "/invoices/",
        "sales_order":        "/audit/",
        "quotation":          "/audit/",
        "proforma_invoice":   "/audit/",
        "grn":                "/audit/",
        "payment_voucher":    "/audit/",
        "receipt_voucher":    "/audit/",
        "cash_voucher":       "/audit/",
        "journal_entry":      "/audit/",
        "general_ledger":     "/audit/",
        "ledger":             "/audit/",
        "contract":           "/audit/",
        "tax_vat_document":   "/audit/",
        "supplier_statement": "/audit/",
        "customer_statement": "/audit/",
    }
    detail_url_prefix = _DOC_DETAIL_PREFIX.get(selected_type, "/audit/")
    # `is_first_class_report` covers any doc type whose data source we can
    # fully populate — invoices natively, plus every type the multi-doc
    # adapter knows how to handle. The "coming soon" banner only renders
    # when neither path is available.
    from apps.reports.services.multi_doc_audit_adapter import is_supported as _multi_doc_supported
    is_first_class_report = is_invoice_report or _multi_doc_supported(selected_type)

    sections = _REPORT_KIND_SECTIONS[selected_kind]
    kind_label_ar, kind_label_en = _REPORT_KIND_LABELS[selected_kind]

    # ── Find latest report record for this org ─────────────────────────────
    report_id = request.GET.get("report_id")
    report_obj = None
    base_qs = Report.objects.filter(organization=org, report_type="invoice_audit").order_by("-created_at")
    if report_id:
        report_obj = base_qs.filter(id=report_id).first()
    if not report_obj:
        report_obj = base_qs.first()

    # v2 reports (generated by InvoiceAuditReportService) have a "report_header" key
    # at the root of data — redirect to the dedicated v2 HTML view.
    if report_id and report_obj and isinstance(report_obj.data, dict) and "report_header" in report_obj.data:
        from django.shortcuts import redirect as _redirect
        return _redirect(f"/api/v1/reports/invoice-audit/{report_obj.id}/html/")

    raw_data = (report_obj.data or {}) if report_obj else {}
    source_data = raw_data.get("invoice_audit") if isinstance(raw_data.get("invoice_audit"), dict) else raw_data
    narrative = (report_obj.narrative or {}) if report_obj else {}

    # ── Pick data source based on doc type ────────────────────────────────
    # Invoices use the rich `InvoiceValidationResult` (failed_rule_codes array
    # + validation_details map). Other doc types use per-type validation
    # models with boolean fields, surfaced through the multi-doc adapter.
    from apps.reports.services.multi_doc_audit_adapter import build_for_doc_type, is_supported

    multi_doc_payload = None
    if selected_type and selected_type != "invoice" and is_supported(selected_type):
        multi_doc_payload = build_for_doc_type(org, selected_type)

    if multi_doc_payload is not None:
        # Multi-doc path: synthesize the same shape `_failed_rules_with_invoice_refs`
        # and `_build_high_risk_violations` already consume.
        validations = multi_doc_payload["validations"]
        validation_map = {str(v.invoice.id): v for v in validations}
        top_risk = multi_doc_payload["top_risk"]
        top_failed = multi_doc_payload["top_failed"]
        rules_applied = multi_doc_payload["rules_applied"]
        rules_passed = multi_doc_payload["rules_passed"]
        rules_failed = multi_doc_payload["rules_failed"]
        compliance_pct = multi_doc_payload["compliance_pct"]
        total_invoices = multi_doc_payload["documents"]
        # Use the doc-type-specific rule catalog so high-risk violation
        # tooltips render the right title (e.g. PO-008 → "يوجد موافق على الأمر").
        active_rule_catalog = multi_doc_payload["rule_catalog"]
        risk_avg = round(
            sum(r.get("risk_score", 0) for r in top_risk) / len(top_risk), 1
        ) if top_risk else 0
        high_risk_count = sum(1 for r in top_risk if r["risk_level"] in ("high", "critical"))
        overall = {
            "total_amount": multi_doc_payload.get("total_amount", 0.0),
            "currency":     multi_doc_payload.get("currency", "SAR"),
        }
        validation_summary = {}
        # Live counterparty/vendor analysis from the doc model — replaces the
        # previous hardcoded empty list so the "Vendor Analysis" section shows
        # real spend distribution per vendor for PO/Receipt/etc.
        vendor_rows = multi_doc_payload.get("vendor_analysis", [])
    else:
        # Invoice path (default).
        validations = list(
            InvoiceValidationResult.objects.filter(invoice__organization=org)
            .select_related("invoice")
            .only(
                "invoice_id", "invoice__invoice_number", "invoice__total_amount",
                "invoice__vendor_name", "invoice__invoice_date",
                "failed_rule_codes", "validation_details",
            )
        )
        validation_map = {str(vr.invoice_id): vr for vr in validations}

        overall = source_data.get("overall_stats", {})
        validation_summary = source_data.get("validation_summary", {})
        top_risk = source_data.get("top_risk_invoices", [])
        vendor_rows = source_data.get("vendor_analysis", [])
        top_failed = source_data.get("top_failed_rules", [])

        compliance_agg = InvoiceValidationResult.objects.filter(invoice__organization=org).aggregate(
            invoices=Count("id"),
            rules_applied=Sum("total_rules"),
            rules_passed=Sum("rules_passed"),
            rules_failed=Sum("rules_failed"),
            avg_score=Avg("validation_score"),
        )
        rules_applied = int(compliance_agg.get("rules_applied") or 0)
        rules_passed = int(compliance_agg.get("rules_passed") or 0)
        rules_failed = int(compliance_agg.get("rules_failed") or 0)
        compliance_pct = round((rules_passed / rules_applied) * 100, 1) if rules_applied else 0

        total_invoices = int(overall.get("total_invoices") or 0)
        risk_avg = float(overall.get("avg_risk_score") or 0)
        high_risk_count = int((overall.get("critical_count") or 0) + (overall.get("high_count") or 0))
        active_rule_catalog = RULES

    # ── Live risk distribution + duplicates (works on both paths) ──────────
    # The saved report's `overall_stats` is often empty / stale. Compute the
    # risk-level distribution and the duplicate count directly from the
    # active doc model so the "Risk Analysis" and "Duplicates" sections
    # always reflect what's actually in the database.
    risk_dist_live = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    duplicate_count_live = 0
    if multi_doc_payload is not None:
        # Sum across the multi-doc validation rows we already loaded.
        for v in validations:
            lvl = (v.invoice.risk_level or "low").lower()
            if lvl in risk_dist_live:
                risk_dist_live[lvl] += 1
        # Document models other than Invoice don't have an `is_duplicate`
        # boolean — count duplicate rule codes (DUP-* / *DUPLICATE*) instead.
        for v in validations:
            if any("DUP" in (c or "").upper() for c in (v.failed_rule_codes or [])):
                duplicate_count_live += 1
    else:
        # Invoice path — query the live tables.
        from apps.invoices.models import Invoice
        risk_qs = Invoice.objects.filter(organization=org).values("risk_level").annotate(n=Count("id"))
        for r in risk_qs:
            lvl = (r["risk_level"] or "low").lower()
            if lvl in risk_dist_live:
                risk_dist_live[lvl] += int(r["n"])
        duplicate_count_live = Invoice.objects.filter(organization=org, is_duplicate=True).count()

    # Live average risk score (across all docs in the active path).
    if multi_doc_payload is not None:
        scores = [v.invoice.risk_score for v in validations
                  if getattr(v.invoice, "risk_score", None) is not None]
        risk_avg_live = round(sum(scores) / len(scores), 1) if scores else 0
    else:
        from apps.invoices.models import Invoice
        risk_avg_live = float(
            Invoice.objects.filter(organization=org).aggregate(a=Avg("risk_score")).get("a") or 0
        )

    # Promote live values when the saved report didn't carry them.
    if not high_risk_count:
        high_risk_count = risk_dist_live["high"] + risk_dist_live["critical"]
    if not risk_avg:
        risk_avg = risk_avg_live

    # ── Real-time fallback (INVOICE PATH ONLY) ────────────────────────────
    # If we're on the invoice path and the saved report's top_failed_rules /
    # top_risk_invoices are empty but live validation data shows failures,
    # rebuild from the database so KPIs and tables stay in sync.
    # We must NOT touch top_failed/top_risk on the multi-doc path — those
    # already reflect the correct doc type and overwriting with Invoice data
    # would re-introduce the very bug this guard prevents.
    if multi_doc_payload is None:
        if not top_failed and rules_failed > 0:
            rule_failure_counts = {}
            for vr in validations:
                for code in (vr.failed_rule_codes or []):
                    rule_failure_counts[code] = rule_failure_counts.get(code, 0) + 1
            top_failed = [
                {"rule_code": code, "failures": cnt, "description": RULES.get(code, code)}
                for code, cnt in sorted(rule_failure_counts.items(), key=lambda kv: -kv[1])[:15]
            ]

        if not top_risk:
            from apps.invoices.models import Invoice
            risky_qs = (
                Invoice.objects.filter(organization=org, risk_level__in=["high", "critical"])
                .order_by("-risk_score")
                .values(
                    "id", "invoice_number", "vendor_name", "total_amount", "currency",
                    "invoice_date", "risk_level", "risk_score",
                )[:15]
            )
            top_risk = []
            for r in risky_qs:
                r["id"] = str(r["id"])
                r["total_amount"] = float(r["total_amount"] or 0)
                r["invoice_date"] = str(r["invoice_date"]) if r["invoice_date"] else "-"
                top_risk.append(r)

        # Refresh totals from real-time when the saved report is empty
        if total_invoices == 0:
            from apps.invoices.models import Invoice
            live_inv_count = Invoice.objects.filter(organization=org).count()
            if live_inv_count:
                total_invoices = live_inv_count
                high_risk_count = Invoice.objects.filter(
                    organization=org, risk_level__in=["high", "critical"]
                ).count()
                live_total_amount = float(
                    Invoice.objects.filter(organization=org).aggregate(
                        s=Sum("total_amount")
                    )["s"] or 0
                )
                if live_total_amount and not overall.get("total_amount"):
                    overall = {**overall, "total_amount": live_total_amount}

    # ── Empty-state detection ──────────────────────────────────────────────
    # The page is "empty" only when there is genuinely no live data anywhere.
    is_empty = (total_invoices == 0 and rules_applied == 0 and not vendor_rows and not top_failed)

    if total_invoices == 0:
        report_status = "بدون بيانات"
    elif risk_avg >= 70 or high_risk_count > 0:
        report_status = "عالي المخاطر"
    elif risk_avg >= 40:
        report_status = "يحتاج مراجعة"
    else:
        report_status = "متوافق"

    failed_rules = _failed_rules_with_invoice_refs(top_failed, validations)
    high_risk_invoices = _build_high_risk_violations(top_risk, validation_map, active_rule_catalog)

    # ── Senior-auditor-grade Findings Register ────────────────────────────
    # Groups every failed rule by severity, attaches financial impact and
    # clickable invoice references. Drives the new "Findings Register" section
    # which is the canonical place reviewers act from.
    from apps.reports.services.findings_service import build_findings_register, build_narrative
    findings_register = build_findings_register(
        top_failed_rules=top_failed,
        validations=validations,
        rule_catalog=active_rule_catalog,
    )
    # Derive executive narrative + recommendations from the findings register
    # so they always reflect the actual data and current doc type — replaces
    # the previous hardcoded invoice-flavoured copy.
    narrative_payload = build_narrative(
        findings_register=findings_register,
        type_label=str(_DOC_TYPE_LABELS.get(selected_type, "Documents")),
        type_singular=str(_DOC_TYPE_SINGULAR.get(selected_type, "Document")),
    )
    # When OPENAI_API_KEY is configured, upgrade the narrative to LLM-synthesized
    # prose grounded in the same findings register. Returns None when the key
    # is missing, the call fails, or the response shape is invalid — in which
    # case we keep the deterministic template above.
    from apps.reports.services.ai_narrative_service import build_ai_narrative
    ai_narrative = build_ai_narrative(
        findings_register=findings_register,
        type_label=str(_DOC_TYPE_LABELS.get(selected_type, "Documents")),
        type_singular=str(_DOC_TYPE_SINGULAR.get(selected_type, "Document")),
        language=(get_language() or "ar")[:2],
        organization=org,
    )
    if ai_narrative is not None:
        narrative_payload = ai_narrative

    # Aggregate the per-document AI fields (`ai_summary`, `ai_recommendations`,
    # `anomalies_found`) into a single section so reports surface the AI work
    # the upload pipeline already does. Empty when no docs have AI content.
    from apps.reports.services.ai_insights_service import build_ai_insights
    ai_insights = build_ai_insights(org, selected_type)

    dominant_vendor = vendor_rows[0] if vendor_rows else {}
    # Prefer the saved-report aggregate when present, otherwise fall back to
    # the live count we just computed. Same approach for missing-QR /
    # handwritten / new-vendor — these come from the saved report only since
    # the live invoice schema doesn't carry them.
    duplicate_count = int(overall.get("duplicate_count") or duplicate_count_live or 0)
    missing_qr_count = int(overall.get("missing_qr_count") or 0)
    handwritten_count = int(overall.get("handwritten_count") or 0)
    new_vendor_count = int(overall.get("new_vendor_count") or 0)

    # ── Compliance rule matrix — only real data, NO mock fallback ──────────
    compliance_matrix = []
    for row in top_failed[:6]:
        compliance_matrix.append({
            "rule_code": row.get("rule_code"),
            "status": "مخالف",
            "weight": _weight_for_rule(row.get("rule_code")),
            "note": row.get("description") or "فشل متكرر يتطلب إجراء فوري.",
        })

    # ── Smart Audit Rule Engine summary (from rule_engine.AuditRun) ────────
    rule_engine_summary = None
    try:
        from apps.rule_engine.models import AuditRun
        runs = AuditRun.objects.filter(organization=org)
        if selected_type and selected_type != "invoice":
            runs = runs.filter(document_type=selected_type)
        runs_total = runs.count()
        if runs_total > 0:
            agg = runs.aggregate(
                total_rules=Sum("total_rules"),
                passed=Sum("passed_rules"),
                failed=Sum("failed_rules"),
                warnings=Sum("warning_rules"),
                avg_risk=Avg("risk_score"),
            )
            risk_dist = {
                "critical": runs.filter(risk_level="critical").count(),
                "high":     runs.filter(risk_level="high").count(),
                "medium":   runs.filter(risk_level="medium").count(),
                "low":      runs.filter(risk_level="low").count(),
            }
            re_total_rules = int(agg.get("total_rules") or 0)
            re_passed = int(agg.get("passed") or 0)
            re_compliance = round((re_passed / re_total_rules) * 100, 1) if re_total_rules else 0
            rule_engine_summary = {
                "total_runs": runs_total,
                "total_rules_applied": re_total_rules,
                "passed_rules": re_passed,
                "failed_rules": int(agg.get("failed") or 0),
                "warning_rules": int(agg.get("warnings") or 0),
                "compliance_pct": re_compliance,
                "avg_risk_score": float(agg.get("avg_risk") or 0),
                "risk_distribution": risk_dist,
            }
    except Exception:  # pragma: no cover — defensive
        rule_engine_summary = None

    # ── ISA 700 opinion (real if rule_engine has data, else None) ──────────
    isa700_opinion = None
    if rule_engine_summary and "isa700" in sections:
        crit = rule_engine_summary["risk_distribution"]["critical"]
        high = rule_engine_summary["risk_distribution"]["high"]
        comp = rule_engine_summary["compliance_pct"]
        if crit > 0:
            opinion, opinion_ar = "Adverse", "رأي معاكس"
        elif high > 5 or comp < 70:
            opinion, opinion_ar = "Disclaimer", "إخلاء مسؤولية"
        elif high > 0 or comp < 95:
            opinion, opinion_ar = "Qualified", "رأي متحفظ"
        else:
            opinion, opinion_ar = "Unqualified", "رأي غير متحفظ"
        isa700_opinion = {
            "opinion": opinion,
            "opinion_ar": opinion_ar,
            "compliance_pct": comp,
            "critical_findings": crit,
            "high_findings": high,
            "basis": (
                "Based on automated audit runs across all uploaded documents using the "
                "Tadgeeg AI rule engine — covering ZATCA Phase 2, VAT validation, fraud "
                "detection, and document quality controls."
            ),
            "basis_ar": (
                "بناءً على التشغيل الآلي لقواعد التدقيق على جميع المستندات المرفوعة عبر "
                "محرك Tadgeeg الذكي — بما يشمل ZATCA Phase 2، صحة الضريبة، كشف الاحتيال، "
                "وضوابط جودة المستندات."
            ),
        }

    # ── Build the report payload ───────────────────────────────────────────
    report_payload = {
        "title": kind_label_ar,
        "title_en": kind_label_en,
        "organization_name": getattr(org, "name", "-") or "-",
        "generated_at": report_obj.created_at if report_obj else datetime.now(),
        "report_type": kind_label_en,
        "status": report_status,
        "summary": {
            "total_invoices": total_invoices,
            "total_amount": float(overall.get("total_amount") or 0),
            "compliance_rate": float(validation_summary.get("vat_compliance_pct") or compliance_pct),
            "avg_risk_score": risk_avg,
            "high_risk_invoices": high_risk_count,
            "duplicate_count": duplicate_count,
        },
        "executive_summary": {
            # When narrative explicitly came from the saved report, prefer it;
            # otherwise build dynamically from the findings register so
            # conclusions/recommendations always reflect real data + doc type.
            "conclusion": narrative.get("executive_summary") or narrative_payload["conclusion"],
            "key_findings": narrative_payload["key_findings"],
            "recommendations": narrative_payload["exec_recs"],
        },
        "compliance_engine": {
            "rules_applied": rules_applied,
            "rules_passed": rules_passed,
            "rules_failed": rules_failed,
            "compliance_rate": compliance_pct,
            "rules_matrix": compliance_matrix,
        },
        "high_risk_invoices": high_risk_invoices,
        "failed_rules": failed_rules,
        "risk_analysis": {
            # Use the saved-report aggregate when it carries data, otherwise
            # the live distribution we computed above. This guarantees the
            # "Risk Analysis" section always reflects what's in the DB.
            "avg_risk_score": risk_avg,
            "high":   int((overall.get("critical_count") or 0) + (overall.get("high_count") or 0)) or (risk_dist_live["critical"] + risk_dist_live["high"]),
            "review": int(overall.get("medium_count") or 0) or risk_dist_live["medium"],
            "safe":   int(overall.get("low_count") or 0) or risk_dist_live["low"],
        },
        "duplicates_anomalies": {
            "duplicates": duplicate_count,
            "dominant_vendor": dominant_vendor.get("vendor_name") or "-",
            "vendor_dependency_pct": float(dominant_vendor.get("spend_share_pct") or 0),
            "patterns": [
                {"label": "غياب رمز QR", "count": missing_qr_count, "severity": "high" if missing_qr_count else "low"},
                {"label": "فواتير بخط يدوي", "count": handwritten_count, "severity": "review" if handwritten_count else "low"},
                {"label": "موردون جدد غير معروفين", "count": new_vendor_count, "severity": "review" if new_vendor_count else "low"},
            ],
        },
        "vendor_analysis": vendor_rows,
        # Action plan derived from real findings — `immediate` collects the
        # recommendations from critical/high findings; `future` collects the
        # rest (or generic improvements when there are no non-blocking ones).
        "recommendations": {
            "immediate": narrative_payload["immediate"],
            "future":    narrative_payload["future"],
        },
        "rule_engine": rule_engine_summary,
        "isa700": isa700_opinion,
        "findings_register": findings_register,
        "ai_insights": ai_insights,
    }

    context = _ctx(
        request,
        "reports",
        report=report_payload,
        report_record=report_obj,
        today=date.today(),
        # New context flags for the template:
        is_empty=is_empty,
        selected_kind=selected_kind,
        selected_type=selected_type,
        type_label=type_label,
        is_invoice_report=is_invoice_report,
        is_first_class_report=is_first_class_report,
        detail_url_prefix=detail_url_prefix,
        report_title=_DOC_TYPE_REPORT_TITLE.get(selected_type, "Audit Report"),
        type_singular=_DOC_TYPE_SINGULAR.get(selected_type, "Document"),
        type_plural=type_label,
        sections=sections,
        kind_label=kind_label_ar,
    )
    return render(request, "reports/invoice_audit_report.html", context)


@login_required(login_url="/login/")
def vendors(request):
    """
    Aggregate every distinct vendor we've seen across invoices and purchase
    orders, even if no `VendorProfile` row exists yet (auto-extraction lags
    behind multi-record uploads). Each row shows volume + VAT number + last
    activity, computed on the fly.
    """
    from collections import defaultdict
    from apps.invoices.models import Invoice, VendorProfile
    from apps.documents.typed_models import PurchaseOrder

    org = getattr(request.user, "organization", None)
    rows = []
    stats = {"total": 0, "verified": 0, "needs_review": 0}

    if org:
        # Aggregate by vendor_name from both invoice + PO tables.
        agg = defaultdict(lambda: {
            "vendor_name": "",
            "vendor_vat_number": "",
            "invoice_count": 0,
            "po_count": 0,
            "total_invoice_amount": 0,
            "total_po_amount": 0,
            "last_activity": None,
            "country": "",
        })

        for inv in Invoice.objects.filter(organization=org).exclude(vendor_name="").only(
            "vendor_name", "vendor_vat_number", "total_amount", "invoice_date", "created_at",
        ):
            key = (inv.vendor_name or "").strip()
            if not key:
                continue
            row = agg[key]
            row["vendor_name"] = key
            if inv.vendor_vat_number and not row["vendor_vat_number"]:
                row["vendor_vat_number"] = inv.vendor_vat_number
            row["invoice_count"] += 1
            row["total_invoice_amount"] += float(inv.total_amount or 0)
            ts = inv.invoice_date or inv.created_at
            if ts and (not row["last_activity"] or ts > row["last_activity"]):
                row["last_activity"] = ts

        for po in PurchaseOrder.objects.filter(organization=org).exclude(vendor_name="").only(
            "vendor_name", "vendor_vat_number", "total_amount", "po_date", "created_at",
        ):
            key = (po.vendor_name or "").strip()
            if not key:
                continue
            row = agg[key]
            row["vendor_name"] = key
            if po.vendor_vat_number and not row["vendor_vat_number"]:
                row["vendor_vat_number"] = po.vendor_vat_number
            row["po_count"] += 1
            row["total_po_amount"] += float(po.total_amount or 0)
            ts = po.po_date or po.created_at
            if ts and (not row["last_activity"] or ts > row["last_activity"]):
                row["last_activity"] = ts

        # Filter by search query
        q = (request.GET.get("q") or "").strip().lower()
        if q:
            agg = {k: v for k, v in agg.items() if q in k.lower() or q in v["vendor_vat_number"].lower()}

        # Saudi ZATCA TRN format: 15 digits starting with 3 and ending with 3.
        def _is_zatca_valid(vat: str) -> bool:
            return bool(vat) and len(vat) == 15 and vat.isdigit() and vat.startswith("3") and vat.endswith("3")

        # Vendor risk lives on VendorProfile and was never joined here, so the
        # list a user browses showed none of it — the risk score existed, the
        # API served it, and vendors/detail.html displayed it, but you had to
        # already know which vendor to open. Risk that is only visible after
        # you have chosen cannot tell you what to choose.
        #
        # One query into a dict rather than a lookup per row: this page caps at
        # 200 vendors and a per-row query would be 200 round trips.
        profiles = {
            (p.vendor_name or "").strip(): p
            for p in VendorProfile.objects.filter(organization=org).only(
                "vendor_name", "risk_score", "risk_tier", "transaction_frequency_30d",
                "compliance_issue_count", "duplicate_count", "is_suspicious",
            )
        }

        rows = sorted(agg.values(), key=lambda r: r["invoice_count"] + r["po_count"], reverse=True)
        for r in rows:
            r["total_doc_count"] = r["invoice_count"] + r["po_count"]
            r["total_amount"] = r["total_invoice_amount"] + r["total_po_amount"]
            r["is_verified"] = _is_zatca_valid(r["vendor_vat_number"])
            r["needs_review"] = not r["vendor_vat_number"] or not r["is_verified"]
            r["initials"] = "".join(w[0] for w in r["vendor_name"].split()[:2]).upper() or "?"

            # None, not 0: a vendor with no profile has not been scored, and
            # "unscored" must not render as "lowest possible risk" — that is
            # the same unmeasured-is-not-zero trap as the quota and precision
            # code, and here it would hide exactly the vendors nobody has
            # looked at yet.
            profile = profiles.get(r["vendor_name"])
            r["risk_score"] = profile.risk_score if profile else None
            r["risk_tier"] = profile.risk_tier if profile else ""
            r["frequency_30d"] = profile.transaction_frequency_30d if profile else None
            r["compliance_issues"] = profile.compliance_issue_count if profile else None
            r["is_suspicious"] = bool(profile and profile.is_suspicious)

        stats["total"] = len(rows)
        stats["verified"] = sum(1 for r in rows if r["is_verified"])
        stats["needs_review"] = sum(1 for r in rows if r["needs_review"])

        # Cap at 200 rows on the page (matches the typed-list pages).
        rows = rows[:200]

    return render(request, "vendors/index.html", _ctx(
        request, "vendors",
        rows=rows, stats=stats, search=request.GET.get("q", ""),
    ))


@login_required(login_url="/login/")
def analytics(request):
    """Time-series + breakdown charts of invoice volume / VAT / vendor concentration."""
    from datetime import timedelta
    from django.utils import timezone
    from django.db.models import Count, Sum
    from django.db.models.functions import TruncMonth
    from apps.invoices.models import Invoice
    from apps.documents.typed_models import PurchaseOrder

    org = getattr(request.user, "organization", None)
    monthly = []
    risk_dist = []
    top_vendors = []
    status_dist = []
    vat_monthly = []
    rule_failures = []
    doc_type_mix = []
    summary = {"total_invoices": 0, "total_pos": 0, "total_amount": 0, "avg_invoice": 0}

    if org:
        cutoff = timezone.now() - timedelta(days=365)
        inv_qs = Invoice.objects.filter(organization=org, created_at__gte=cutoff)

        # Monthly invoice + amount series for the last 12 months.
        by_month = (
            inv_qs.annotate(m=TruncMonth("created_at"))
            .values("m").annotate(c=Count("id"), s=Sum("total_amount"), v=Sum("vat_amount"))
            .order_by("m")
        )
        monthly = [
            {"month": r["m"].strftime("%Y-%m") if r["m"] else "?",
             "count": r["c"], "amount": float(r["s"] or 0)}
            for r in by_month
        ]
        vat_monthly = [
            {"month": r["m"].strftime("%Y-%m") if r["m"] else "?",
             "vat": float(r["v"] or 0)}
            for r in by_month
        ]

        # Risk distribution (current snapshot, not just last 12mo)
        all_inv = Invoice.objects.filter(organization=org)
        risk_dist = [
            {"level": level,
             "count": all_inv.filter(risk_level=level).count()}
            for level in ("low", "medium", "high", "critical")
        ]

        # Status distribution (workflow snapshot)
        status_dist = [
            {"status": s, "count": all_inv.filter(status=s).count()}
            for s in ("pending", "processing", "flagged", "approved", "rejected")
        ]

        # Top 10 vendors by invoice + PO total
        from collections import defaultdict
        agg = defaultdict(float)
        for v, t in all_inv.exclude(vendor_name="").values_list("vendor_name", "total_amount"):
            agg[v] += float(t or 0)
        for v, t in PurchaseOrder.objects.filter(organization=org).exclude(vendor_name="").values_list("vendor_name", "total_amount"):
            agg[v] += float(t or 0)
        top_vendors = sorted(
            ({"name": k, "total": v} for k, v in agg.items()),
            key=lambda r: r["total"], reverse=True,
        )[:10]

        # Rule-failure histogram from AuditCases
        try:
            from apps.audit.models import AuditCase
            cases = AuditCase.objects.filter(organization=org).values_list("title", flat=True)
            ctr = defaultdict(int)
            for t in cases:
                # Title looks like "[R001] Duplicate Invoice Detection"
                code = (t or "").split("]", 1)[0].lstrip("[")
                if code.startswith("R") and code[1:].isdigit():
                    ctr[code] += 1
            rule_failures = [{"rule": k, "count": v} for k, v in sorted(ctr.items())]
        except Exception:
            rule_failures = []

        # Document-type mix across all 21 typed models + invoice
        try:
            from apps.documents.typed_models import (
                BankStatement, PayrollSheet, ExpenseReport, VATReturn,
                FixedAsset, SalesReceipt, GoodsReceiptNote, PaymentVoucher,
            )
            from apps.documents.typed_models_v2 import (
                SalesOrder, Quotation, ProformaInvoice, ReceiptVoucher, CashVoucher,
                GeneralLedger, Ledger, Contract, SupplierStatement, CustomerStatement,
                JournalEntry,
            )
            doc_models = [
                ("Invoice",            Invoice),
                ("Purchase Order",     PurchaseOrder),
                ("Bank Statement",     BankStatement),
                ("Payroll",            PayrollSheet),
                ("Expense Report",     ExpenseReport),
                ("VAT Return",         VATReturn),
                ("Fixed Asset",        FixedAsset),
                ("Sales Receipt",      SalesReceipt),
                ("GRN",                GoodsReceiptNote),
                ("Payment Voucher",    PaymentVoucher),
                ("Sales Order",        SalesOrder),
                ("Quotation",          Quotation),
                ("Proforma Invoice",   ProformaInvoice),
                ("Receipt Voucher",    ReceiptVoucher),
                ("Cash Voucher",       CashVoucher),
                ("General Ledger",     GeneralLedger),
                ("Sub-Ledger",         Ledger),
                ("Contract",           Contract),
                ("Supplier Statement", SupplierStatement),
                ("Customer Statement", CustomerStatement),
                ("Journal Entry",      JournalEntry),
            ]
            doc_type_mix = [
                {"name": label, "count": M.objects.filter(organization=org).count()}
                for label, M in doc_models
            ]
            doc_type_mix = [d for d in doc_type_mix if d["count"] > 0]
            doc_type_mix.sort(key=lambda r: r["count"], reverse=True)
        except Exception:
            doc_type_mix = []

        summary["total_invoices"] = all_inv.count()
        summary["total_pos"] = PurchaseOrder.objects.filter(organization=org).count()
        summary["total_amount"] = float(all_inv.aggregate(s=Sum("total_amount"))["s"] or 0)
        summary["avg_invoice"] = round(summary["total_amount"] / summary["total_invoices"], 2) if summary["total_invoices"] else 0

    import json as _json
    return render(request, "analytics/index.html", _ctx(
        request, "analytics",
        monthly=_json.dumps(monthly),
        risk_dist=_json.dumps(risk_dist),
        status_dist=_json.dumps(status_dist),
        vat_monthly=_json.dumps(vat_monthly),
        rule_failures=_json.dumps(rule_failures),
        doc_type_mix=_json.dumps(doc_type_mix),
        top_vendors=top_vendors,
        summary=summary,
    ))


@login_required(login_url="/login/")
def audit(request):
    """Audit cases / runs list page — aggregates audit activity from every
    document model that carries validation results (Invoice + the 7 typed
    document models). The legacy `rule_engine.AuditRun` table is mostly
    empty, so we read from where the data actually lives.
    """
    org = getattr(request.user, "organization", None)
    rows: list[dict] = []
    counts = {"total": 0, "completed": 0, "running": 0, "failed": 0}
    if not org:
        return render(request, "audit/index.html", _ctx(
            request, "audit", rows=rows, counts=counts,
        ))

    # ── Pull AuditRun first (when present) ────────────────────────────────
    try:
        from apps.rule_engine.models import AuditRun
        run_qs = AuditRun.objects.filter(organization=org)
        if run_qs.exists():
            counts["total"]     += run_qs.count()
            counts["completed"] += run_qs.filter(status="completed").count()
            counts["running"]   += run_qs.filter(status__in=["pending", "running"]).count()
            counts["failed"]    += run_qs.filter(status="failed").count()
            rows.extend(run_qs.order_by("-started_at")[:50].values(
                "id", "document_type", "document_id", "status", "started_at",
                "completed_at", "rules_passed", "rules_failed", "risk_level",
            ))
    except Exception:  # AuditRun may not be migrated
        pass

    # ── Aggregate per-document validation activity ────────────────────────
    # Each typed model writes `failed_rule_codes`, `rules_passed`,
    # `rules_failed`, `validation_score`, `risk_level` directly. A row is
    # treated as "completed" when validation has happened (score>0 OR any
    # rule was applied), and "failed" when at least one rule failed AND it
    # wasn't completed yet (we won't double-count those that completed).
    from apps.invoices.models import Invoice
    from apps.documents.models import (
        PurchaseOrder, BankStatement, PayrollSheet, ExpenseReport,
        VATReturn, FixedAsset, SalesReceipt,
    )
    DOC_SOURCES = [
        ("invoice",        Invoice,        "invoice_number",        "created_at"),
        ("purchase_order", PurchaseOrder,  "po_number",             "created_at"),
        ("bank_statement", BankStatement,  "account_number",        "created_at"),
        ("payroll",        PayrollSheet,   "payroll_period_from",   "created_at"),
        ("expense_report", ExpenseReport,  "report_number",         "created_at"),
        ("vat_return",     VATReturn,      "vat_number",            "created_at"),
        ("fixed_asset",    FixedAsset,     "fiscal_year",           "created_at"),
        ("sales_receipt",  SalesReceipt,   "receipt_number",        "created_at"),
    ]

    from django.db.models import Q

    for doc_type, Model, num_field, ts_field in DOC_SOURCES:
        fields = {f.name for f in Model._meta.get_fields()}
        if "validation_score" not in fields or "rules_failed" not in fields:
            continue
        base = Model.objects.filter(organization=org)
        # "Completed" = something was actually validated (score recorded OR
        # at least one rule applied). Catches both clean (rules_failed=0)
        # and flagged docs.
        completed_q = Q(validation_score__gt=0) | Q(rules_passed__gt=0) | Q(rules_failed__gt=0)
        completed_n = base.filter(completed_q).count()
        # "Failed" — completed AND at least one rule failed
        failed_n = base.filter(completed_q & Q(rules_failed__gt=0)).count()
        total_n = base.count()
        running_n = max(0, total_n - completed_n)  # uploaded but not yet validated
        counts["total"]     += total_n
        counts["completed"] += completed_n
        counts["running"]   += running_n
        counts["failed"]    += failed_n

        # Append recent rows for the activity table — newest 30 per type.
        only_fields = [
            "id", "validation_score", "rules_passed", "rules_failed",
            "risk_level", num_field,
        ]
        if ts_field in fields:
            only_fields.append(ts_field)
        for d in base.only(*only_fields).order_by(f"-{ts_field}" if ts_field in fields else "-id")[:30]:
            score = float(getattr(d, "validation_score", 0) or 0)
            r_failed = int(getattr(d, "rules_failed", 0) or 0)
            r_passed = int(getattr(d, "rules_passed", 0) or 0)
            if r_failed or r_passed or score:
                status = "failed" if r_failed > 0 else "completed"
            else:
                status = "running"
            rows.append({
                "id":          str(d.id),
                "document_type": doc_type,
                "document_id":   str(getattr(d, num_field, "") or "")[:24] or str(d.id)[:8],
                "status":        status,
                "started_at":    getattr(d, ts_field, None),
                "completed_at":  getattr(d, ts_field, None) if status == "completed" else None,
                "rules_passed":  r_passed,
                "rules_failed":  r_failed,
                "risk_level":    getattr(d, "risk_level", None) or "low",
            })

    # Sort all rows by started_at desc (None last) and keep the most recent 200
    def _sort_key(r):
        ts = r.get("started_at")
        return ts or 0
    rows.sort(key=_sort_key, reverse=True)
    rows = rows[:200]

    return render(request, "audit/index.html", _ctx(
        request, "audit", rows=rows, counts=counts,
    ))


@login_required(login_url="/login/")
def audit_detail(request, pk):
    return render(request, "audit/detail.html", _ctx(request, "audit", case_id=str(pk)))


@login_required(login_url="/login/")
def compliance(request):
    """ZATCA / VAT / regulatory health dashboard for the org."""
    org = getattr(request.user, "organization", None)
    health = {
        "total_invoices": 0,
        "qr_valid_pct": 0,
        "vat_validated_pct": 0,
        "missing_vat_count": 0,
        "high_risk_count": 0,
        "duplicate_count": 0,
        "compliance_score": 0,
    }
    issues = []
    if org:
        from apps.invoices.models import Invoice
        qs = Invoice.objects.filter(organization=org)
        n = qs.count()
        health["total_invoices"] = n
        if n:
            qr_ok = qs.filter(qr_code_valid=True).count()
            vat_ok = qs.exclude(vendor_vat_number="").count()
            health["qr_valid_pct"] = int((qr_ok / n) * 100)
            health["vat_validated_pct"] = int((vat_ok / n) * 100)
            health["missing_vat_count"] = qs.filter(vendor_vat_number="").count()
            health["high_risk_count"] = qs.filter(risk_level__in=["high", "critical"]).count()
            health["duplicate_count"] = qs.filter(is_duplicate=True).count()
            # Composite score: average of the three health signals.
            health["compliance_score"] = int(
                (health["qr_valid_pct"] + health["vat_validated_pct"] +
                 (100 - int(health["high_risk_count"] / n * 100))) / 3
            )

        # Surface the actual problem invoices so the user can drill down.
        issues = list(
            qs.filter(qr_code_valid=False).order_by("-created_at")[:10].values(
                "id", "invoice_number", "vendor_name", "total_amount", "created_at",
            )
        )

    return render(request, "compliance/index.html", _ctx(
        request, "compliance", health=health, issues=issues,
    ))


@login_required(login_url="/login/")
def documents(request):
    org = getattr(request.user, "organization", None)
    counts = {
        # Phase 1
        "invoices": 0, "bank_statements": 0, "vat_returns": 0, "payroll": 0,
        "purchase_orders": 0, "expense_reports": 0, "fixed_assets": 0, "sales_receipts": 0,
        # Phase 2
        "sales_orders": 0, "quotations": 0, "proforma_invoices": 0,
        "receipt_vouchers": 0, "cash_vouchers": 0,
        "general_ledgers": 0, "ledgers": 0,
        "contracts": 0, "supplier_statements": 0, "customer_statements": 0,
        # Late additions to complete the 20-doc-type catalog
        "goods_receipt_notes": 0, "payment_vouchers": 0, "journal_entries": 0,
    }
    if org:
        from apps.invoices.models import Invoice
        from apps.documents.typed_models import (
            PurchaseOrder, BankStatement, PayrollSheet, ExpenseReport,
            VATReturn, FixedAsset, SalesReceipt,
        )
        from apps.documents.typed_models_v2 import (
            SalesOrder, Quotation, ProformaInvoice, ReceiptVoucher, CashVoucher,
            GeneralLedger, Ledger, Contract, SupplierStatement, CustomerStatement,
            JournalEntry,
        )
        from apps.documents.typed_models import GoodsReceiptNote, PaymentVoucher
        counts["invoices"]            = Invoice.objects.filter(organization=org).count()
        counts["purchase_orders"]     = PurchaseOrder.objects.filter(organization=org).count()
        counts["bank_statements"]     = BankStatement.objects.filter(organization=org).count()
        counts["payroll"]             = PayrollSheet.objects.filter(organization=org).count()
        counts["expense_reports"]     = ExpenseReport.objects.filter(organization=org).count()
        counts["vat_returns"]         = VATReturn.objects.filter(organization=org).count()
        counts["fixed_assets"]        = FixedAsset.objects.filter(organization=org).count()
        counts["sales_receipts"]      = SalesReceipt.objects.filter(organization=org).count()
        # Phase 2 — typed-models v2
        counts["sales_orders"]        = SalesOrder.objects.filter(organization=org).count()
        counts["quotations"]          = Quotation.objects.filter(organization=org).count()
        counts["proforma_invoices"]   = ProformaInvoice.objects.filter(organization=org).count()
        counts["receipt_vouchers"]    = ReceiptVoucher.objects.filter(organization=org).count()
        counts["cash_vouchers"]       = CashVoucher.objects.filter(organization=org).count()
        counts["general_ledgers"]     = GeneralLedger.objects.filter(organization=org).count()
        counts["ledgers"]             = Ledger.objects.filter(organization=org).count()
        counts["contracts"]           = Contract.objects.filter(organization=org).count()
        counts["supplier_statements"] = SupplierStatement.objects.filter(organization=org).count()
        counts["customer_statements"] = CustomerStatement.objects.filter(organization=org).count()
        # GRN, Payment Voucher, Journal Entry — late additions completing the 20
        counts["goods_receipt_notes"] = GoodsReceiptNote.objects.filter(organization=org).count()
        counts["payment_vouchers"]    = PaymentVoucher.objects.filter(organization=org).count()
        counts["journal_entries"]     = JournalEntry.objects.filter(organization=org).count()
    response = render(request, "documents/index.html", _ctx(request, "documents", doc_counts=counts))
    # Force the browser to fetch fresh markup whenever counts change rather than
    # serving a stale cached page where the grid layout looks broken.
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


@login_required(login_url="/login/")
def transactions(request):
    """Unified transactions feed across invoices + POs + bank lines."""
    org = getattr(request.user, "organization", None)
    rows = []
    if org:
        from apps.invoices.models import Invoice
        from apps.documents.typed_models import PurchaseOrder

        for inv in Invoice.objects.filter(organization=org).order_by("-created_at")[:100]:
            rows.append({
                "id": str(inv.id),
                "type": "invoice",
                "ref": inv.invoice_number or "—",
                "party": inv.vendor_name or "—",
                "amount": float(inv.total_amount or 0),
                "currency": inv.currency or "SAR",
                "date": inv.invoice_date or inv.created_at.date(),
                "status": inv.status,
                "url": f"/invoices/{inv.id}/",
            })
        for po in PurchaseOrder.objects.filter(organization=org).order_by("-created_at")[:100]:
            rows.append({
                "id": str(po.id),
                "type": "purchase_order",
                "ref": po.po_number or "—",
                "party": po.vendor_name or "—",
                "amount": float(po.total_amount or 0),
                "currency": po.currency or "SAR",
                "date": po.po_date or po.created_at.date(),
                "status": po.audit_status,
                "url": f"/documents/purchase-orders/{po.id}/",
            })
        rows.sort(key=lambda r: r["date"] or "", reverse=True)
        rows = rows[:200]
    return render(request, "transactions.html", _ctx(request, "transactions", rows=rows))


@login_required(login_url="/login/")
def users(request):
    if not getattr(request.user, "can_manage_users", False):
        return redirect("frontend:dashboard")
    return render(request, "users/index.html", _ctx(request, "users"))


@login_required(login_url="/login/")
def settings(request):
    return render(request, "settings/index.html", _ctx(request, "settings"))


@login_required(login_url="/login/")
def doc_upload(request):
    return render(request, "documents/upload.html", _ctx(request, "documents", selected_type=request.GET.get("type", "")))


# ── Helpers used by the typed-document list views ────────────────────────────

def _typed_list_qs(request, Model):
    """Return the doc list queryset filtered by org + optional search/status/date.

    Always pulls related Document + uploaded_by FKs in a single JOIN so the
    list templates (which render uploader name + filename next to each row)
    don't issue an N+1 query per record. With 1000 rows that's the difference
    between 1 query and 2001.
    """
    org = getattr(request.user, "organization", None)
    if not org:
        return Model.objects.none()
    qs = (
        Model.objects.filter(organization=org)
        .select_related("document", "uploaded_by", "organization")
        .order_by("-created_at")
    )

    p = request.GET
    if v := p.get("q"):
        # Most typed models have one of these "name-like" fields. Try them in order.
        from django.db.models import Q
        text_fields = [
            f.name for f in Model._meta.get_fields()
            if hasattr(f, "max_length") and f.max_length and f.get_internal_type() == "CharField"
        ]
        if text_fields:
            cond = Q()
            for fn in text_fields:
                cond |= Q(**{f"{fn}__icontains": v})
            qs = qs.filter(cond)
    if v := p.get("status"):
        if hasattr(Model, "audit_status") and v in {c[0] for c in Model._meta.get_field("audit_status").choices}:
            qs = qs.filter(audit_status=v)
    return qs


@login_required(login_url="/login/")
def _typed_list_render(request, Model, template_name, active_key, per_page=25):
    """Shared render path for the 7 typed-document list views.

    Builds the queryset, paginates, then hands the standard kwargs to the
    template. Replaces the duplicated 7-function block we used to have.
    """
    qs = _typed_list_qs(request, Model)
    page_kwargs = _paginate(qs, request, per_page=per_page)
    rows = list(page_kwargs["page_obj"].object_list)
    return render(request, template_name, _ctx(
        request, active_key,
        rows=rows,
        search=request.GET.get("q", ""),
        status_filter=request.GET.get("status", ""),
        **page_kwargs,
    ))


def purchase_orders(request):
    from apps.documents.typed_models import PurchaseOrder
    return _typed_list_render(request, PurchaseOrder,
                              "documents/purchase_orders.html", "purchase_orders")


@login_required(login_url="/login/")
def bank_statements(request):
    from apps.documents.typed_models import BankStatement
    return _typed_list_render(request, BankStatement,
                              "documents/bank_statements.html", "bank_statements")


@login_required(login_url="/login/")
def payroll(request):
    from apps.documents.typed_models import PayrollSheet
    return _typed_list_render(request, PayrollSheet,
                              "documents/payroll.html", "payroll")


@login_required(login_url="/login/")
def expense_reports(request):
    from apps.documents.typed_models import ExpenseReport
    return _typed_list_render(request, ExpenseReport,
                              "documents/expense_reports.html", "expense_reports")


@login_required(login_url="/login/")
def vat_returns(request):
    from apps.documents.typed_models import VATReturn
    return _typed_list_render(request, VATReturn,
                              "documents/vat_returns.html", "vat_returns")


@login_required(login_url="/login/")
def fixed_assets(request):
    from apps.documents.typed_models import FixedAsset
    return _typed_list_render(request, FixedAsset,
                              "documents/fixed_assets.html", "fixed_assets")


@login_required(login_url="/login/")
def sales_receipts(request):
    from apps.documents.typed_models import SalesReceipt
    return _typed_list_render(request, SalesReceipt,
                              "documents/sales_receipts.html", "sales_receipts")


@login_required(login_url="/login/")
def purchase_order_detail(request, pk):
    return render(request, "documents/detail/purchase_order_detail.html", _ctx(request, "purchase_orders", doc_id=str(pk)))


@login_required(login_url="/login/")
def bank_statement_detail(request, pk):
    return render(request, "documents/detail/bank_statement_detail.html", _ctx(request, "bank_statements", doc_id=str(pk)))


@login_required(login_url="/login/")
def payroll_detail(request, pk):
    return render(request, "documents/detail/payroll_detail.html", _ctx(request, "payroll", doc_id=str(pk)))


@login_required(login_url="/login/")
def expense_report_detail(request, pk):
    return render(request, "documents/detail/expense_report_detail.html", _ctx(request, "expense_reports", doc_id=str(pk)))


@login_required(login_url="/login/")
def vat_return_detail(request, pk):
    return render(request, "documents/detail/vat_return_detail.html", _ctx(request, "vat_returns", doc_id=str(pk)))


@login_required(login_url="/login/")
def fixed_asset_detail(request, pk):
    return render(request, "documents/detail/fixed_asset_detail.html", _ctx(request, "fixed_assets", doc_id=str(pk)))


@login_required(login_url="/login/")
def sales_receipt_detail(request, pk):
    return render(request, "documents/detail/sales_receipt_detail.html", _ctx(request, "sales_receipts", doc_id=str(pk)))


# ──────────────────────────────────────────────────────────────────────────────
# Phase-4 generic list/detail views for the 10 new doc types
# ──────────────────────────────────────────────────────────────────────────────

def _phase4_list(request, doc_type, Model, *,
                 columns, label, subtitle, icon, list_url_name, detail_url_name):
    """Render a typed-doc list page using the generic _typed_list.html template."""
    qs = (
        Model.objects.filter(organization=request.user.organization)
        .order_by("-created_at")
        if request.user.organization else Model.objects.none()
    )
    # Apply optional filters
    risk = (request.GET.get("risk") or "").strip().lower()
    if risk in ("low", "medium", "high", "critical"):
        qs = qs.filter(risk_level=risk)
    search = (request.GET.get("q") or "").strip()
    if search and columns:
        from django.db.models import Q
        # Search the first text column we know about
        first_text_col = next(
            (c["key"] for c in columns if c.get("kind") not in ("amount", "date")),
            None,
        )
        if first_text_col:
            qs = qs.filter(**{f"{first_text_col}__icontains": search})

    page_kwargs = _paginate(qs, request, per_page=25)
    rows = list(page_kwargs["page_obj"].object_list)
    return render(request, "documents/_typed_list.html", _ctx(
        request, "documents",
        doc_type_label=label,
        doc_type=doc_type,
        doc_type_subtitle=subtitle,
        doc_type_icon=icon,
        upload_url=f"/documents/upload/?type={doc_type}",
        report_url=f"/reports/invoice-audit/?type={doc_type}&kind=audit_summary",
        detail_url_name=detail_url_name,
        rows=rows,
        columns=columns,
        search=search,
        **page_kwargs,
    ))


def _phase4_detail(request, doc_type, Model, pk, *,
                   field_groups, label, icon, list_url_name):
    """Render the generic _typed_detail.html for a single typed-doc instance."""
    from django.shortcuts import get_object_or_404
    from apps.rule_engine.catalog.document_rules import ALL_RULES
    org = request.user.organization
    doc = get_object_or_404(Model, pk=pk, organization=org)

    # Enrich `failed_rule_codes` with catalog metadata for display
    rules_index = {r["rule_id"]: r for r in ALL_RULES}
    failed_rules = []
    for code in (doc.failed_rule_codes or []):
        meta = rules_index.get(code, {})
        failed_rules.append({
            "rule_code":      code,
            "name":           meta.get("name_ar") or meta.get("name_en") or code,
            "description":    meta.get("fail_message_ar") or meta.get("description_ar") or "",
            "recommendation": meta.get("recommendation_ar") or "",
            "severity":       meta.get("severity", "medium"),
        })

    # Reverse the list-url-name to a concrete path for back-link
    from django.urls import reverse
    try:
        list_url = reverse(f"frontend:{list_url_name}")
    except Exception:
        list_url = "/documents/"

    # Cross-doc links (PO ↔ GRN ↔ Invoice ↔ Payment + Phase-2 chains).
    # Returns {} for doc types the linker doesn't handle, so the panel hides
    # itself gracefully on those pages.
    try:
        from core.services.cross_doc_linker import find_links
        cross_links = find_links(doc_type, doc, org)
    except Exception:
        cross_links = {}

    return render(request, "documents/_typed_detail.html", _ctx(
        request, "documents",
        doc=doc,
        doc_type=doc_type,
        doc_type_label=label,
        doc_type_icon=icon,
        list_url=list_url,
        field_groups=field_groups,
        failed_rules=failed_rules,
        cross_links=cross_links,
    ))


# Per-type column / field-group schemas — drive the generic templates.
from django.utils.translation import gettext_lazy as _gl4

_SO_COLUMNS = [
    {"key": "so_number",     "label": _gl4("SO No.")},
    {"key": "so_date",       "label": _gl4("Date"),     "kind": "date"},
    {"key": "customer_name", "label": _gl4("Customer")},
    {"key": "total_amount",  "label": _gl4("Total"),    "kind": "amount"},
]
_SO_FIELDS = [
    {"title": _gl4("Header"), "fields": [
        {"key": "so_number",   "label": _gl4("SO Number")},
        {"key": "so_date",     "label": _gl4("Date"),     "kind": "date"},
        {"key": "customer_name",        "label": _gl4("Customer")},
        {"key": "customer_vat_number",  "label": _gl4("VAT Number")},
        {"key": "expected_delivery_date","label": _gl4("Expected Delivery"), "kind": "date"},
        {"key": "status",      "label": _gl4("Status")},
    ]},
    {"title": _gl4("Financials"), "fields": [
        {"key": "subtotal",        "label": _gl4("Subtotal"),     "kind": "amount"},
        {"key": "vat_amount",      "label": _gl4("VAT"),          "kind": "amount"},
        {"key": "total_amount",    "label": _gl4("Total"),        "kind": "amount"},
        {"key": "discount_amount", "label": _gl4("Discount"),     "kind": "amount"},
        {"key": "customer_credit_limit","label": _gl4("Credit Limit"), "kind": "amount"},
        {"key": "customer_outstanding", "label": _gl4("Outstanding"),  "kind": "amount"},
    ]},
]

_QT_COLUMNS = [
    {"key": "quotation_number", "label": _gl4("Quote No.")},
    {"key": "quotation_date",   "label": _gl4("Date"),    "kind": "date"},
    {"key": "expiry_date",      "label": _gl4("Expires"), "kind": "date"},
    {"key": "party_name",       "label": _gl4("Party")},
    {"key": "total_amount",     "label": _gl4("Total"),   "kind": "amount"},
]
_QT_FIELDS = [
    {"title": _gl4("Header"), "fields": [
        {"key": "quotation_number", "label": _gl4("Quote Number")},
        {"key": "quotation_date",   "label": _gl4("Date"),    "kind": "date"},
        {"key": "expiry_date",      "label": _gl4("Expiry"),  "kind": "date"},
        {"key": "party_type",       "label": _gl4("Party Type")},
        {"key": "party_name",       "label": _gl4("Party")},
        {"key": "status",           "label": _gl4("Status")},
    ]},
    {"title": _gl4("Financials"), "fields": [
        {"key": "subtotal",      "label": _gl4("Subtotal"),    "kind": "amount"},
        {"key": "discount_pct",  "label": _gl4("Discount %")},
        {"key": "vat_amount",    "label": _gl4("VAT"),         "kind": "amount"},
        {"key": "total_amount",  "label": _gl4("Total"),       "kind": "amount"},
    ]},
]

_PF_COLUMNS = [
    {"key": "proforma_number", "label": _gl4("Proforma No.")},
    {"key": "proforma_date",   "label": _gl4("Date"),     "kind": "date"},
    {"key": "customer_name",   "label": _gl4("Customer")},
    {"key": "total_amount",    "label": _gl4("Total"),    "kind": "amount"},
]
_PF_FIELDS = [
    {"title": _gl4("Header"), "fields": [
        {"key": "proforma_number",   "label": _gl4("Proforma Number")},
        {"key": "proforma_date",     "label": _gl4("Date"),       "kind": "date"},
        {"key": "validity_date",     "label": _gl4("Validity"),   "kind": "date"},
        {"key": "customer_name",     "label": _gl4("Customer")},
        {"key": "customer_vat_number","label": _gl4("VAT Number")},
        {"key": "is_marked_proforma","label": _gl4("Marked Proforma"), "kind": "bool"},
    ]},
    {"title": _gl4("Financials"), "fields": [
        {"key": "subtotal",     "label": _gl4("Subtotal"), "kind": "amount"},
        {"key": "vat_amount",   "label": _gl4("VAT"),      "kind": "amount"},
        {"key": "total_amount", "label": _gl4("Total"),    "kind": "amount"},
    ]},
]

_RV_COLUMNS = [
    {"key": "receipt_number", "label": _gl4("Receipt No.")},
    {"key": "receipt_date",   "label": _gl4("Date"),  "kind": "date"},
    {"key": "payer_name",     "label": _gl4("Payer")},
    {"key": "amount",         "label": _gl4("Amount"), "kind": "amount"},
]
_RV_FIELDS = [
    {"title": _gl4("Header"), "fields": [
        {"key": "receipt_number", "label": _gl4("Receipt Number")},
        {"key": "receipt_date",   "label": _gl4("Date"),       "kind": "date"},
        {"key": "payer_name",     "label": _gl4("Payer")},
        {"key": "receipt_method", "label": _gl4("Method")},
        {"key": "amount",         "label": _gl4("Amount"),     "kind": "amount"},
    ]},
    {"title": _gl4("References"), "fields": [
        {"key": "linked_invoice_number", "label": _gl4("Invoice No.")},
        {"key": "bank_reference",        "label": _gl4("Bank Reference")},
        {"key": "is_reconciled",         "label": _gl4("Reconciled"), "kind": "bool"},
        {"key": "is_duplicate",          "label": _gl4("Duplicate"),  "kind": "bool"},
    ]},
]

_CV_COLUMNS = [
    {"key": "voucher_number",    "label": _gl4("Voucher No.")},
    {"key": "voucher_date",      "label": _gl4("Date"), "kind": "date"},
    {"key": "movement_type",     "label": _gl4("Type")},
    {"key": "counterparty_name", "label": _gl4("Counterparty")},
    {"key": "amount",            "label": _gl4("Amount"), "kind": "amount"},
]
_CV_FIELDS = [
    {"title": _gl4("Header"), "fields": [
        {"key": "voucher_number",   "label": _gl4("Voucher Number")},
        {"key": "voucher_date",     "label": _gl4("Date"),  "kind": "date"},
        {"key": "movement_type",    "label": _gl4("Type")},
        {"key": "amount",           "label": _gl4("Amount"), "kind": "amount"},
        {"key": "reason",           "label": _gl4("Reason")},
        {"key": "has_attachment",   "label": _gl4("Has Attachment"), "kind": "bool"},
        {"key": "approval_status",  "label": _gl4("Approval")},
    ]},
]

_GL_COLUMNS = [
    {"key": "fiscal_year",   "label": _gl4("Fiscal Year")},
    {"key": "period_to",     "label": _gl4("Period End"),  "kind": "date"},
    {"key": "total_debit",   "label": _gl4("Total Debit"), "kind": "amount"},
    {"key": "total_credit",  "label": _gl4("Total Credit"),"kind": "amount"},
]
_GL_FIELDS = [
    {"title": _gl4("Period"), "fields": [
        {"key": "fiscal_year",     "label": _gl4("Fiscal Year")},
        {"key": "period_from",     "label": _gl4("Period From"), "kind": "date"},
        {"key": "period_to",       "label": _gl4("Period To"),   "kind": "date"},
    ]},
    {"title": _gl4("Totals"), "fields": [
        {"key": "total_debit",     "label": _gl4("Total Debit"),  "kind": "amount"},
        {"key": "total_credit",    "label": _gl4("Total Credit"), "kind": "amount"},
        {"key": "accounts_count",  "label": _gl4("Accounts")},
        {"key": "movements_count", "label": _gl4("Movements")},
        {"key": "is_balanced",     "label": _gl4("Balanced"), "kind": "bool"},
    ]},
]

_LDG_COLUMNS = [
    {"key": "account_number",    "label": _gl4("Account No.")},
    {"key": "account_name",      "label": _gl4("Account Name")},
    {"key": "period_to",         "label": _gl4("Period End"), "kind": "date"},
    {"key": "closing_balance",   "label": _gl4("Closing"),    "kind": "amount"},
]
_LDG_FIELDS = [
    {"title": _gl4("Account"), "fields": [
        {"key": "account_number", "label": _gl4("Account Number")},
        {"key": "account_name",   "label": _gl4("Account Name")},
        {"key": "account_type",   "label": _gl4("Type")},
        {"key": "period_from",    "label": _gl4("Period From"), "kind": "date"},
        {"key": "period_to",      "label": _gl4("Period To"),   "kind": "date"},
    ]},
    {"title": _gl4("Balances"), "fields": [
        {"key": "opening_balance", "label": _gl4("Opening Balance"), "kind": "amount"},
        {"key": "total_debit",     "label": _gl4("Total Debit"),     "kind": "amount"},
        {"key": "total_credit",    "label": _gl4("Total Credit"),    "kind": "amount"},
        {"key": "closing_balance", "label": _gl4("Closing Balance"), "kind": "amount"},
        {"key": "movements_count", "label": _gl4("Movements")},
    ]},
]

_CTR_COLUMNS = [
    {"key": "contract_number", "label": _gl4("Contract No.")},
    {"key": "title",           "label": _gl4("Title")},
    {"key": "party_b",         "label": _gl4("Counterparty")},
    {"key": "start_date",      "label": _gl4("Start"),  "kind": "date"},
    {"key": "end_date",        "label": _gl4("End"),    "kind": "date"},
    {"key": "contract_value",  "label": _gl4("Value"),  "kind": "amount"},
]
_CTR_FIELDS = [
    {"title": _gl4("Header"), "fields": [
        {"key": "contract_number", "label": _gl4("Contract Number")},
        {"key": "title",           "label": _gl4("Title")},
        {"key": "party_a",         "label": _gl4("Party A")},
        {"key": "party_b",         "label": _gl4("Party B")},
        {"key": "party_b_type",    "label": _gl4("Party Type")},
        {"key": "party_b_vat_number","label": _gl4("VAT Number")},
        {"key": "status",          "label": _gl4("Status")},
        {"key": "is_signed",       "label": _gl4("Signed"), "kind": "bool"},
    ]},
    {"title": _gl4("Dates & Value"), "fields": [
        {"key": "start_date",       "label": _gl4("Start Date"),  "kind": "date"},
        {"key": "end_date",         "label": _gl4("End Date"),    "kind": "date"},
        {"key": "signing_date",     "label": _gl4("Signed On"),   "kind": "date"},
        {"key": "contract_value",   "label": _gl4("Value"),       "kind": "amount"},
        {"key": "invoiced_to_date", "label": _gl4("Invoiced"),    "kind": "amount"},
        {"key": "payment_terms",    "label": _gl4("Payment Terms")},
    ]},
]

_SS_COLUMNS = [
    {"key": "supplier_name",   "label": _gl4("Supplier")},
    {"key": "period_to",       "label": _gl4("Period End"),    "kind": "date"},
    {"key": "opening_balance", "label": _gl4("Opening"),       "kind": "amount"},
    {"key": "closing_balance", "label": _gl4("Closing"),       "kind": "amount"},
    {"key": "balance_variance","label": _gl4("Variance"),      "kind": "amount"},
]
_SS_FIELDS = [
    {"title": _gl4("Supplier"), "fields": [
        {"key": "supplier_name",     "label": _gl4("Supplier")},
        {"key": "supplier_id",       "label": _gl4("Supplier ID")},
        {"key": "supplier_vat_number","label": _gl4("VAT Number")},
    ]},
    {"title": _gl4("Period"), "fields": [
        {"key": "period_from", "label": _gl4("From"), "kind": "date"},
        {"key": "period_to",   "label": _gl4("To"),   "kind": "date"},
    ]},
    {"title": _gl4("Balances"), "fields": [
        {"key": "opening_balance", "label": _gl4("Opening Balance"), "kind": "amount"},
        {"key": "total_invoiced",  "label": _gl4("Total Invoiced"),  "kind": "amount"},
        {"key": "total_paid",      "label": _gl4("Total Paid"),      "kind": "amount"},
        {"key": "closing_balance", "label": _gl4("Closing Balance"), "kind": "amount"},
        {"key": "balance_variance","label": _gl4("Variance"),        "kind": "amount"},
        {"key": "duplicate_count", "label": _gl4("Duplicate Count")},
    ]},
]

_CS_COLUMNS = [
    {"key": "customer_name",   "label": _gl4("Customer")},
    {"key": "period_to",       "label": _gl4("Period End"), "kind": "date"},
    {"key": "opening_balance", "label": _gl4("Opening"),    "kind": "amount"},
    {"key": "closing_balance", "label": _gl4("Closing"),    "kind": "amount"},
    {"key": "balance_variance","label": _gl4("Variance"),   "kind": "amount"},
]
_CS_FIELDS = [
    {"title": _gl4("Customer"), "fields": [
        {"key": "customer_name",     "label": _gl4("Customer")},
        {"key": "customer_id",       "label": _gl4("Customer ID")},
        {"key": "customer_vat_number","label": _gl4("VAT Number")},
    ]},
    {"title": _gl4("Period"), "fields": [
        {"key": "period_from", "label": _gl4("From"), "kind": "date"},
        {"key": "period_to",   "label": _gl4("To"),   "kind": "date"},
    ]},
    {"title": _gl4("Balances"), "fields": [
        {"key": "opening_balance",  "label": _gl4("Opening Balance"), "kind": "amount"},
        {"key": "total_invoiced",   "label": _gl4("Total Invoiced"),  "kind": "amount"},
        {"key": "total_received",   "label": _gl4("Total Received"),  "kind": "amount"},
        {"key": "closing_balance",  "label": _gl4("Closing Balance"), "kind": "amount"},
        {"key": "balance_variance", "label": _gl4("Variance"),        "kind": "amount"},
        {"key": "duplicate_count",  "label": _gl4("Duplicate Count")},
    ]},
]

# ── GRN (Goods Receipt Note) ──
_GRN_COLUMNS = [
    {"key": "grn_number",   "label": _gl4("GRN No.")},
    {"key": "grn_date",     "label": _gl4("Date"),    "kind": "date"},
    {"key": "po_number",    "label": _gl4("PO No.")},
    {"key": "vendor_name",  "label": _gl4("Vendor")},
    {"key": "total_amount", "label": _gl4("Total"),   "kind": "amount"},
]
_GRN_FIELDS = [
    {"title": _gl4("Header"), "fields": [
        {"key": "grn_number",   "label": _gl4("GRN Number")},
        {"key": "grn_date",     "label": _gl4("Date"),     "kind": "date"},
        {"key": "po_number",    "label": _gl4("PO Number")},
        {"key": "invoice_number","label": _gl4("Invoice Number")},
        {"key": "department",   "label": _gl4("Department")},
        {"key": "received_by",  "label": _gl4("Received By")},
        {"key": "warehouse_location", "label": _gl4("Warehouse")},
    ]},
    {"title": _gl4("Vendor"), "fields": [
        {"key": "vendor_name",       "label": _gl4("Vendor")},
        {"key": "vendor_vat_number", "label": _gl4("VAT Number")},
    ]},
    {"title": _gl4("Quantities"), "fields": [
        {"key": "total_ordered_qty",  "label": _gl4("Ordered Qty")},
        {"key": "total_received_qty", "label": _gl4("Received Qty")},
        {"key": "total_rejected_qty", "label": _gl4("Rejected Qty")},
        {"key": "rejection_rate_pct", "label": _gl4("Rejection %")},
    ]},
    {"title": _gl4("Amounts"), "fields": [
        {"key": "total_amount",   "label": _gl4("Total"),          "kind": "amount"},
        {"key": "invoice_amount", "label": _gl4("Invoice Amount"), "kind": "amount"},
        {"key": "currency",       "label": _gl4("Currency")},
    ]},
    {"title": _gl4("Status"), "fields": [
        {"key": "delivery_date",     "label": _gl4("Delivery Date"),     "kind": "date"},
        {"key": "delivery_overdue",  "label": _gl4("Overdue"),           "kind": "bool"},
        {"key": "quality_inspection_done", "label": _gl4("QC Done"),     "kind": "bool"},
        {"key": "approval_status",   "label": _gl4("Approval")},
    ]},
]

# ── Payment Voucher ──
_PV_COLUMNS = [
    {"key": "payment_number", "label": _gl4("Voucher No.")},
    {"key": "payment_date",   "label": _gl4("Date"),    "kind": "date"},
    {"key": "payee_name",     "label": _gl4("Payee")},
    {"key": "payment_method", "label": _gl4("Method")},
    {"key": "total_amount",   "label": _gl4("Total"),   "kind": "amount"},
]
_PV_FIELDS = [
    {"title": _gl4("Header"), "fields": [
        {"key": "payment_number", "label": _gl4("Voucher Number")},
        {"key": "payment_date",   "label": _gl4("Date"),     "kind": "date"},
        {"key": "payment_method", "label": _gl4("Method")},
    ]},
    {"title": _gl4("Payee"), "fields": [
        {"key": "payee_name",       "label": _gl4("Payee")},
        {"key": "payee_vat_number", "label": _gl4("VAT Number")},
        {"key": "payee_iban",       "label": _gl4("IBAN")},
    ]},
    {"title": _gl4("Amounts"), "fields": [
        {"key": "amount",       "label": _gl4("Amount"),     "kind": "amount"},
        {"key": "vat_amount",   "label": _gl4("VAT Amount"), "kind": "amount"},
        {"key": "total_amount", "label": _gl4("Total"),      "kind": "amount"},
        {"key": "currency",     "label": _gl4("Currency")},
    ]},
    {"title": _gl4("References"), "fields": [
        {"key": "linked_invoice_number", "label": _gl4("Invoice Number")},
        {"key": "linked_po_number",      "label": _gl4("PO Number")},
        {"key": "bank_reference",        "label": _gl4("Bank Reference")},
    ]},
    {"title": _gl4("Control"), "fields": [
        {"key": "approval_status",     "label": _gl4("Approval Status")},
        {"key": "is_duplicate",        "label": _gl4("Duplicate"),         "kind": "bool"},
        {"key": "exceeds_threshold",   "label": _gl4("Exceeds Threshold"), "kind": "bool"},
        {"key": "is_advance_payment",  "label": _gl4("Advance"),           "kind": "bool"},
        {"key": "days_since_invoice",  "label": _gl4("Days vs. Invoice")},
    ]},
]

# ── Journal Entry (lightweight model — Phase 6 addition) ──
_JE_COLUMNS = [
    {"key": "entry_number",  "label": _gl4("Entry No.")},
    {"key": "entry_date",    "label": _gl4("Date"),    "kind": "date"},
    {"key": "description",   "label": _gl4("Description")},
    {"key": "total_debit",   "label": _gl4("Debit"),   "kind": "amount"},
    {"key": "total_credit",  "label": _gl4("Credit"),  "kind": "amount"},
]
_JE_FIELDS = [
    {"title": _gl4("Header"), "fields": [
        {"key": "entry_number",     "label": _gl4("Entry Number")},
        {"key": "entry_date",       "label": _gl4("Entry Date"), "kind": "date"},
        {"key": "description",      "label": _gl4("Description")},
        {"key": "fiscal_period",    "label": _gl4("Fiscal Period")},
    ]},
    {"title": _gl4("Totals"), "fields": [
        {"key": "total_debit",  "label": _gl4("Total Debit"),  "kind": "amount"},
        {"key": "total_credit", "label": _gl4("Total Credit"), "kind": "amount"},
        {"key": "is_balanced",  "label": _gl4("Balanced"),     "kind": "bool"},
        {"key": "lines_count",  "label": _gl4("Lines")},
    ]},
    {"title": _gl4("Control"), "fields": [
        {"key": "is_manual",        "label": _gl4("Manual Entry"), "kind": "bool"},
        {"key": "has_attachment",   "label": _gl4("Attachment"),   "kind": "bool"},
        {"key": "is_period_close",  "label": _gl4("Period Close"), "kind": "bool"},
        {"key": "approval_status",  "label": _gl4("Approval")},
    ]},
]


# ── Phase-4 list views ─────────────────────────────────────────────────────
@login_required(login_url="/login/")
def sales_orders(request):
    from apps.documents.models import SalesOrder
    return _phase4_list(request, "sales_order", SalesOrder,
        columns=_SO_COLUMNS, label=_gl4("Sales Orders"),
        subtitle=_gl4("Customer purchase orders, fulfilment status, credit checks."),
        icon="shopping-cart",
        list_url_name="sales_orders", detail_url_name="frontend:sales_order_detail")

@login_required(login_url="/login/")
def quotations(request):
    from apps.documents.models import Quotation
    return _phase4_list(request, "quotation", Quotation,
        columns=_QT_COLUMNS, label=_gl4("Quotations"),
        subtitle=_gl4("Price quotations issued to customers or received from suppliers."),
        icon="file-text",
        list_url_name="quotations", detail_url_name="frontend:quotation_detail")

@login_required(login_url="/login/")
def proforma_invoices(request):
    from apps.documents.models import ProformaInvoice
    return _phase4_list(request, "proforma_invoice", ProformaInvoice,
        columns=_PF_COLUMNS, label=_gl4("Proforma Invoices"),
        subtitle=_gl4("Pre-tax / preliminary invoices, not yet posted as revenue."),
        icon="file-text",
        list_url_name="proforma_invoices", detail_url_name="frontend:proforma_invoice_detail")

@login_required(login_url="/login/")
def receipt_vouchers(request):
    from apps.documents.models import ReceiptVoucher
    return _phase4_list(request, "receipt_voucher", ReceiptVoucher,
        columns=_RV_COLUMNS, label=_gl4("Receipt Vouchers"),
        subtitle=_gl4("Cash + bank receipts collected from customers / payers."),
        icon="receipt",
        list_url_name="receipt_vouchers", detail_url_name="frontend:receipt_voucher_detail")

@login_required(login_url="/login/")
def cash_vouchers(request):
    from apps.documents.models import CashVoucher
    return _phase4_list(request, "cash_voucher", CashVoucher,
        columns=_CV_COLUMNS, label=_gl4("Cash Vouchers"),
        subtitle=_gl4("Petty-cash and on-the-spot cash movements."),
        icon="banknote",
        list_url_name="cash_vouchers", detail_url_name="frontend:cash_voucher_detail")

@login_required(login_url="/login/")
def general_ledgers(request):
    from apps.documents.models import GeneralLedger
    return _phase4_list(request, "general_ledger", GeneralLedger,
        columns=_GL_COLUMNS, label=_gl4("General Ledgers"),
        subtitle=_gl4("Period-level GL snapshots — debit/credit totals + balanced check."),
        icon="book-open",
        list_url_name="general_ledgers", detail_url_name="frontend:general_ledger_detail")

@login_required(login_url="/login/")
def ledgers(request):
    from apps.documents.models import Ledger
    return _phase4_list(request, "ledger", Ledger,
        columns=_LDG_COLUMNS, label=_gl4("Ledgers"),
        subtitle=_gl4("Account-level ledgers with movements + closing balance."),
        icon="book",
        list_url_name="ledgers", detail_url_name="frontend:ledger_detail")

@login_required(login_url="/login/")
def contracts(request):
    from apps.documents.models import Contract
    return _phase4_list(request, "contract", Contract,
        columns=_CTR_COLUMNS, label=_gl4("Contracts"),
        subtitle=_gl4("Vendor / customer contracts — value, signing status, modifications."),
        icon="file-signature",
        list_url_name="contracts", detail_url_name="frontend:contract_detail")

@login_required(login_url="/login/")
def supplier_statements(request):
    from apps.documents.models import SupplierStatement
    return _phase4_list(request, "supplier_statement", SupplierStatement,
        columns=_SS_COLUMNS, label=_gl4("Supplier Statements"),
        subtitle=_gl4("Reconcile our payables ledger against the supplier's statement."),
        icon="users",
        list_url_name="supplier_statements", detail_url_name="frontend:supplier_statement_detail")

@login_required(login_url="/login/")
def customer_statements(request):
    from apps.documents.models import CustomerStatement
    return _phase4_list(request, "customer_statement", CustomerStatement,
        columns=_CS_COLUMNS, label=_gl4("Customer Statements"),
        subtitle=_gl4("Reconcile our receivables ledger against the customer's view."),
        icon="user-check",
        list_url_name="customer_statements", detail_url_name="frontend:customer_statement_detail")


# ── Phase-4 detail views ───────────────────────────────────────────────────
@login_required(login_url="/login/")
def sales_order_detail(request, pk):
    from apps.documents.models import SalesOrder
    return _phase4_detail(request, "sales_order", SalesOrder, pk,
        field_groups=_SO_FIELDS, label=_gl4("Sales Order"),
        icon="shopping-cart", list_url_name="sales_orders")

@login_required(login_url="/login/")
def quotation_detail(request, pk):
    from apps.documents.models import Quotation
    return _phase4_detail(request, "quotation", Quotation, pk,
        field_groups=_QT_FIELDS, label=_gl4("Quotation"),
        icon="file-text", list_url_name="quotations")

@login_required(login_url="/login/")
def proforma_invoice_detail(request, pk):
    from apps.documents.models import ProformaInvoice
    return _phase4_detail(request, "proforma_invoice", ProformaInvoice, pk,
        field_groups=_PF_FIELDS, label=_gl4("Proforma Invoice"),
        icon="file-text", list_url_name="proforma_invoices")

@login_required(login_url="/login/")
def receipt_voucher_detail(request, pk):
    from apps.documents.models import ReceiptVoucher
    return _phase4_detail(request, "receipt_voucher", ReceiptVoucher, pk,
        field_groups=_RV_FIELDS, label=_gl4("Receipt Voucher"),
        icon="receipt", list_url_name="receipt_vouchers")

@login_required(login_url="/login/")
def cash_voucher_detail(request, pk):
    from apps.documents.models import CashVoucher
    return _phase4_detail(request, "cash_voucher", CashVoucher, pk,
        field_groups=_CV_FIELDS, label=_gl4("Cash Voucher"),
        icon="banknote", list_url_name="cash_vouchers")

@login_required(login_url="/login/")
def general_ledger_detail(request, pk):
    from apps.documents.models import GeneralLedger
    return _phase4_detail(request, "general_ledger", GeneralLedger, pk,
        field_groups=_GL_FIELDS, label=_gl4("General Ledger"),
        icon="book-open", list_url_name="general_ledgers")

@login_required(login_url="/login/")
def ledger_detail(request, pk):
    from apps.documents.models import Ledger
    return _phase4_detail(request, "ledger", Ledger, pk,
        field_groups=_LDG_FIELDS, label=_gl4("Ledger"),
        icon="book", list_url_name="ledgers")

@login_required(login_url="/login/")
def contract_detail(request, pk):
    from apps.documents.models import Contract
    return _phase4_detail(request, "contract", Contract, pk,
        field_groups=_CTR_FIELDS, label=_gl4("Contract"),
        icon="file-signature", list_url_name="contracts")

@login_required(login_url="/login/")
def supplier_statement_detail(request, pk):
    from apps.documents.models import SupplierStatement
    return _phase4_detail(request, "supplier_statement", SupplierStatement, pk,
        field_groups=_SS_FIELDS, label=_gl4("Supplier Statement"),
        icon="users", list_url_name="supplier_statements")

@login_required(login_url="/login/")
def customer_statement_detail(request, pk):
    from apps.documents.models import CustomerStatement
    return _phase4_detail(request, "customer_statement", CustomerStatement, pk,
        field_groups=_CS_FIELDS, label=_gl4("Customer Statement"),
        icon="user-check", list_url_name="customer_statements")


# ── GRN, Payment Voucher, Journal Entry — Phase-1 + new Phase-2 page wiring ──
@login_required(login_url="/login/")
def goods_receipt_notes(request):
    from apps.documents.typed_models import GoodsReceiptNote
    return _phase4_list(request, "goods_receipt_note", GoodsReceiptNote,
        columns=_GRN_COLUMNS, label=_gl4("Goods Receipts"),
        subtitle=_gl4("GRN matching with PO and invoice — quantity and quality checks."),
        icon="package-check",
        list_url_name="goods_receipt_notes",
        detail_url_name="frontend:goods_receipt_note_detail")

@login_required(login_url="/login/")
def goods_receipt_note_detail(request, pk):
    from apps.documents.typed_models import GoodsReceiptNote
    return _phase4_detail(request, "goods_receipt_note", GoodsReceiptNote, pk,
        field_groups=_GRN_FIELDS, label=_gl4("Goods Receipt"),
        icon="package-check", list_url_name="goods_receipt_notes")

@login_required(login_url="/login/")
def payment_vouchers(request):
    from apps.documents.typed_models import PaymentVoucher
    return _phase4_list(request, "payment_voucher", PaymentVoucher,
        columns=_PV_COLUMNS, label=_gl4("Payment Vouchers"),
        subtitle=_gl4("Vendor payments, approvals, three-way match."),
        icon="banknote",
        list_url_name="payment_vouchers",
        detail_url_name="frontend:payment_voucher_detail")

@login_required(login_url="/login/")
def payment_voucher_detail(request, pk):
    from apps.documents.typed_models import PaymentVoucher
    return _phase4_detail(request, "payment_voucher", PaymentVoucher, pk,
        field_groups=_PV_FIELDS, label=_gl4("Payment Voucher"),
        icon="banknote", list_url_name="payment_vouchers")

@login_required(login_url="/login/")
def journal_entries(request):
    from apps.documents.models import JournalEntry
    return _phase4_list(request, "journal_entry", JournalEntry,
        columns=_JE_COLUMNS, label=_gl4("Journal Entries"),
        subtitle=_gl4("Double-entry postings — debit/credit balance and approvals."),
        icon="book-open",
        list_url_name="journal_entries",
        detail_url_name="frontend:journal_entry_detail")

@login_required(login_url="/login/")
def journal_entry_detail(request, pk):
    from apps.documents.models import JournalEntry
    return _phase4_detail(request, "journal_entry", JournalEntry, pk,
        field_groups=_JE_FIELDS, label=_gl4("Journal Entry"),
        icon="book-open", list_url_name="journal_entries")


def robots_txt(request):
    from django.http import HttpResponse
    admin_url = getattr(django_settings, "ADMIN_URL", "admin/")
    content = (
        "User-agent: *\n"
        f"Disallow: /{admin_url}\n"
        "Disallow: /api/\n"
        "Disallow: /media/\n"
        "Allow: /\n\n"
        "Sitemap: https://www.tadgeeg.com/sitemap.xml\n"
    )
    return HttpResponse(content, content_type="text/plain")


def sitemap_xml(request):
    from django.http import HttpResponse
    content = """<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
  <url><loc>https://www.tadgeeg.com/</loc><changefreq>weekly</changefreq><priority>1.0</priority></url>
  <url><loc>https://www.tadgeeg.com/login/</loc><changefreq>monthly</changefreq><priority>0.8</priority></url>
  <url><loc>https://www.tadgeeg.com/register/</loc><changefreq>monthly</changefreq><priority>0.8</priority></url>
  <url><loc>https://www.tadgeeg.com/about/</loc><changefreq>monthly</changefreq><priority>0.6</priority></url>
  <url><loc>https://www.tadgeeg.com/contact/</loc><changefreq>monthly</changefreq><priority>0.6</priority></url>
</urlset>"""
    return HttpResponse(content, content_type="application/xml")


def page_not_found(request, exception=None):
    return render(request, "404.html", status=404)


def server_error(request):
    return render(request, "500.html", status=500)


def permission_denied(request, exception=None):
    return render(request, "403.html", status=403)
