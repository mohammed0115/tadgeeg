"""Tadgeeg AI report views."""

import logging
import json
from datetime import date
from urllib.parse import quote
from django.core.serializers.json import DjangoJSONEncoder
from django.http import HttpResponse, JsonResponse
from django.utils.translation import gettext as _
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from drf_spectacular.utils import extend_schema, OpenApiParameter
from django.db.models import Sum, Count, Avg, Q, Max, Min
from django.db.models.functions import TruncMonth
from core.services.ai_service import generate_audit_narrative
from apps.billing.ai_access import ai_access_response_or_none
from apps.authentication.permissions import IsSeniorAuditorOrAbove
from apps.transactions.models import Transaction
from apps.audit.models import AuditCase
from apps.documents.models import Document
from apps.compliance.models import ComplianceViolation
from apps.invoices.models import Invoice, InvoiceValidationResult, VendorProfile
from apps.reports.services.report_data_service import ReportDataService
from .models import Report

logger = logging.getLogger("finai")


def _attachment_disposition(filename: str) -> str:
    """Build an HTTP-safe Content-Disposition value for an arbitrary filename.

    Django wraps non-ASCII header values as RFC 2047 encoded-words
    (``=?utf-8?b?...?=``), which is the email-header format and is NOT
    parsed by browsers — they end up seeing the literal encoded-word as
    the entire header value, never recognise ``attachment;`` and so do
    not trigger a download. RFC 5987 / 6266 is the HTTP form: provide
    an ASCII ``filename="..."`` fallback plus ``filename*=UTF-8''<pct>``.
    """
    ascii_fallback = filename.encode("ascii", "ignore").decode("ascii").strip()
    # If stripping non-ASCII left only punctuation/extension (e.g. "__.pdf"),
    # fall back to a generic "report.<ext>" so legacy clients get a usable name.
    if not ascii_fallback or not any(ch.isalnum() for ch in ascii_fallback.rsplit(".", 1)[0]):
        ext = filename.rsplit(".", 1)[1] if "." in filename else "bin"
        ascii_fallback = f"report.{ext}"
    return f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{quote(filename)}"


def _render_report_pdf_bytes(html_str: str, base_url: str) -> bytes:
    from weasyprint import HTML as WP_HTML

    return WP_HTML(string=html_str, base_url=base_url).write_pdf()


def _render_report_pdf_cached(report_id, language: str, html_str: str, base_url: str) -> bytes:
    """Memoize PDF bytes for 1 hour keyed by (report_id, language).

    A typical financial report is 30-60 pages; WeasyPrint takes 3-15 seconds
    to render. Without a cache, every "Download PDF" click re-rendered the
    same bytes. Cache scope is per-language because the same `report_id`
    renders differently in ar vs en.

    Cache invalidates after 1h or when the saved Report row's `updated_at`
    changes (we include the report's last_modified epoch in the key so a
    re-run flushes the cache automatically).
    """
    import hashlib
    from django.core.cache import cache as _cache

    # The html_str hash bakes report data + version into the key — any change
    # to the report row's data flips this and invalidates the cache.
    digest = hashlib.sha256(html_str.encode("utf-8", errors="replace")).hexdigest()[:16]
    ck = f"reportpdf:v1:{report_id}:{language}:{digest}"
    blob = _cache.get(ck)
    if blob is not None:
        return blob

    blob = _render_report_pdf_bytes(html_str, base_url)
    # 1 hour — long enough that a user clicking Download a few times in a
    # session pays once; short enough that schema fixes propagate quickly.
    _cache.set(ck, blob, 60 * 60)
    return blob


def _json_safe(obj):
    """Recursively convert Decimal / UUID / date to JSON-safe Python types."""
    import decimal, uuid, datetime
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(i) for i in obj]
    if isinstance(obj, decimal.Decimal):
        return float(obj)
    if isinstance(obj, uuid.UUID):
        return str(obj)
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()
    return obj


# ─── Big Four & benchmark constants (mirrors audit/analytics views) ───────────

_BIG_FOUR_MAPPING = {
    "KPMG":     {"label": "KPMG — Invoice Completeness",       "groups": ["INV"],       "standard": "ISA 500"},
    "Deloitte": {"label": "Deloitte — Risk & Anomaly Detection","groups": ["ANO", "DUP"],"standard": "ISA 315"},
    "PwC":      {"label": "PwC — Internal Control Testing",    "groups": ["CTL"],       "standard": "ISA 330"},
    "EY":       {"label": "EY — Compliance Verification",      "groups": ["VAT", "DOC"],"standard": "ISA 250"},
}

_INDUSTRY_BENCHMARKS = {
    "finance":      {"label": "Financial Services & Banking",      "compliance_rate_pct": 94.2, "duplicate_rate_pct": 0.8, "avg_risk_score": 18.5, "vat_compliance_rate_pct": 97.8},
    "retail":       {"label": "Retail & E-Commerce",               "compliance_rate_pct": 88.7, "duplicate_rate_pct": 1.9, "avg_risk_score": 24.1, "vat_compliance_rate_pct": 93.4},
    "construction": {"label": "Construction & Real Estate",         "compliance_rate_pct": 85.3, "duplicate_rate_pct": 2.7, "avg_risk_score": 31.4, "vat_compliance_rate_pct": 89.2},
    "healthcare":   {"label": "Healthcare & Pharmaceuticals",       "compliance_rate_pct": 91.6, "duplicate_rate_pct": 1.1, "avg_risk_score": 20.8, "vat_compliance_rate_pct": 95.1},
    "manufacturing":{"label": "Manufacturing & Industry",           "compliance_rate_pct": 87.9, "duplicate_rate_pct": 2.3, "avg_risk_score": 27.6, "vat_compliance_rate_pct": 91.7},
    "services":     {"label": "Professional & Business Services",   "compliance_rate_pct": 90.1, "duplicate_rate_pct": 1.5, "avg_risk_score": 22.3, "vat_compliance_rate_pct": 94.0},
}


def _compute_big_four(vr_qs) -> dict:
    """Backward-compatible wrapper around the centralized report data service."""
    return ReportDataService().compute_big_four(vr_qs)


def _compute_benchmark(org, inv_qs) -> dict:
    """Backward-compatible wrapper around the centralized report data service."""
    return ReportDataService().compute_benchmark(org, inv_qs)


# ─── Helper: collect invoice section data ─────────────────────────────────────

def _collect_invoice_data(org, date_from=None, date_to=None, audit_session_id=None) -> dict:
    """Backward-compatible wrapper around the centralized report data service."""
    return ReportDataService().collect_invoice_data(
        org,
        date_from=date_from,
        date_to=date_to,
        audit_session_id=audit_session_id,
    )


def _collect_rule_engine_data(org, date_from=None, date_to=None) -> dict:
    """
    Collect aggregated statistics from the new rule engine (AuditRun / AuditResult).
    Returns an empty dict if no runs exist yet, so the caller can skip gracefully.
    """
    try:
        from apps.rule_engine.models.audit_execution import AuditRun, AuditResult
        from apps.rule_engine.models.rule_definition import RuleDefinitionTranslation
    except ImportError:
        return {}

    runs_qs = AuditRun.objects.filter(
        organization=org,
        status=AuditRun.Status.COMPLETED,
    )
    if date_from:
        runs_qs = runs_qs.filter(completed_at__date__gte=date_from)
    if date_to:
        runs_qs = runs_qs.filter(completed_at__date__lte=date_to)

    if not runs_qs.exists():
        return {}

    stats = runs_qs.aggregate(
        total_runs=Count("id"),
        avg_risk=Avg("risk_score"),
        blocking_count=Count("id", filter=Q(blocks_approval=True)),
        manual_review_count=Count("id", filter=Q(requires_manual_review=True)),
        total_passed=Sum("passed_rules"),
        total_failed=Sum("failed_rules"),
        total_warnings=Sum("warning_rules"),
        total_rules=Sum("total_rules"),
        critical_count=Count("id", filter=Q(risk_level="critical")),
        high_count=Count("id", filter=Q(risk_level="high")),
        medium_count=Count("id", filter=Q(risk_level="medium")),
        low_count=Count("id", filter=Q(risk_level="low")),
    )

    total_applied = int(stats["total_rules"] or 0)
    total_passed = int(stats["total_passed"] or 0)
    total_failed = int(stats["total_failed"] or 0)
    total_warnings = int(stats["total_warnings"] or 0)
    compliance_pct = round(total_passed / total_applied * 100, 1) if total_applied else 0.0

    # Top failed rules with Arabic names
    top_failures_qs = (
        AuditResult.objects.filter(audit_run__in=runs_qs, status="fail")
        .values("rule_code")
        .annotate(
            fail_count=Count("id"),
            avg_risk=Avg("risk_contribution"),
        )
        .order_by("-fail_count")[:10]
    )

    # Fetch Arabic translations for the top rule codes
    top_codes = [r["rule_code"] for r in top_failures_qs]
    translations: dict = {}
    for t in RuleDefinitionTranslation.objects.filter(
        rule__rule_code__in=top_codes
    ).select_related("rule"):
        code = t.rule.rule_code
        if code not in translations:
            translations[code] = {}
        translations[code][t.language] = t.name

    top_failed_rules = [
        {
            "rule_code": r["rule_code"],
            "fail_count": r["fail_count"],
            "avg_risk_contribution": round(float(r["avg_risk"] or 0), 2),
            "name_ar": translations.get(r["rule_code"], {}).get("ar", r["rule_code"]),
            "name_en": translations.get(r["rule_code"], {}).get("en", r["rule_code"]),
            "explanation_sample": (
                AuditResult.objects.filter(
                    audit_run__in=runs_qs,
                    status="fail",
                    rule_code=r["rule_code"],
                )
                .exclude(explanation="")
                .order_by("-executed_at")
                .values_list("explanation", flat=True)
                .first()
                or ""
            ),
            "evidence_count": AuditResult.objects.filter(
                audit_run__in=runs_qs,
                status="fail",
                rule_code=r["rule_code"],
                evidence_items__isnull=False,
            ).distinct().count(),
        }
        for r in top_failures_qs
    ]

    # Breakdown by document type
    by_doc_type = list(
        runs_qs.values("document_type")
        .annotate(
            run_count=Count("id"),
            avg_risk=Avg("risk_score"),
            blocking=Count("id", filter=Q(blocks_approval=True)),
            failed_rules_sum=Sum("failed_rules"),
        )
        .order_by("-run_count")
    )
    for row in by_doc_type:
        row["avg_risk"] = round(float(row["avg_risk"] or 0), 1)

    return {
        "total_runs": int(stats["total_runs"] or 0),
        "avg_risk_score": round(float(stats["avg_risk"] or 0), 1),
        "blocking_failures": int(stats["blocking_count"] or 0),
        "manual_review_required": int(stats["manual_review_count"] or 0),
        "total_rules_applied": total_applied,
        "total_passed": total_passed,
        "total_failed": total_failed,
        "total_warnings": total_warnings,
        "compliance_pct": compliance_pct,
        "risk_distribution": {
            "critical": int(stats["critical_count"] or 0),
            "high":     int(stats["high_count"]     or 0),
            "medium":   int(stats["medium_count"]   or 0),
            "low":      int(stats["low_count"]      or 0),
        },
        "top_failed_rules": top_failed_rules,
        "by_document_type": by_doc_type,
    }


def _extract_invoice_audit_section(report_data) -> dict:
    if isinstance(report_data, dict) and isinstance(report_data.get("invoice_audit"), dict):
        return report_data.get("invoice_audit") or {}
    return report_data or {}


def _weight_for_rule(rule_code, is_ar=True):
    prefix = str(rule_code or "").split("-", 1)[0]
    if prefix in {"DUP", "VAT", "ANO"}:
        return "مرتفع" if is_ar else "High"
    if prefix in {"INV", "CTL"}:
        return "متوسط" if is_ar else "Medium"
    return "منخفض" if is_ar else "Low"


def _failed_rules_with_invoice_refs(top_failed_rules, validations):
    top_codes = [r.get("rule_code") for r in (top_failed_rules or []) if r.get("rule_code")]
    top_codes_set = set(top_codes)
    buckets = {code: [] for code in top_codes}

    for vr in validations:
        inv_no = vr.invoice.invoice_number or str(vr.invoice_id)[:8]
        for code in (vr.failed_rule_codes or []):
            if code not in top_codes_set:
                continue
            if inv_no not in buckets[code]:
                buckets[code].append(inv_no)

    enriched = []
    for rule in (top_failed_rules or []):
        code = rule.get("rule_code")
        invoice_numbers = buckets.get(code, [])
        enriched.append({
            "rule_code": code,
            "description": rule.get("description") or code,
            "failure_count": int(rule.get("failures") or 0),
            "invoice_numbers": invoice_numbers,
        })
    return enriched


def _build_high_risk_violations(invoice_rows, validation_map, rule_catalog):
    rows = []
    for row in (invoice_rows or []):
        inv_id = row.get("id")
        vr = validation_map.get(str(inv_id))
        violations = []

        if vr:
            details = vr.validation_details or {}
            for code in (vr.failed_rule_codes or [])[:4]:
                detail = details.get(code, {}) if isinstance(details, dict) else {}
                reason = (
                    detail.get("reason")
                    or detail.get("message")
                    or detail.get("note")
                    or _("A rule violation was detected and requires detailed review.")
                )
                violations.append({
                    "rule_code": code,
                    "description": rule_catalog.get(code, code),
                    "reason": reason,
                })

        rows.append({
            "invoice_number": row.get("invoice_number") or "-",
            "amount": float(row.get("total_amount") or 0),
            "date": row.get("invoice_date") or "-",
            "risk_level": row.get("risk_level") or "low",
            "risk_score": float(row.get("risk_score") or 0),
            "violations": violations,
        })
    return rows


def _build_invoice_audit_report_ui(report, org):
    from apps.invoices.models import InvoiceValidationResult
    from core.services.invoice_validator import RULES

    source_data = _extract_invoice_audit_section(report.data or {})
    if not isinstance(source_data, dict):
        return None

    is_ar = str(getattr(report, "language", "") or "").lower().startswith("ar")

    def txt(ar_text, en_text):
        return ar_text if is_ar else en_text

    validation_qs = InvoiceValidationResult.objects.filter(invoice__organization=org).select_related("invoice")
    if report.period_from:
        validation_qs = validation_qs.filter(invoice__invoice_date__gte=report.period_from)
    if report.period_to:
        validation_qs = validation_qs.filter(invoice__invoice_date__lte=report.period_to)
    validations = list(validation_qs.only("invoice_id", "invoice__invoice_number", "failed_rule_codes", "validation_details"))
    validation_map = {str(vr.invoice_id): vr for vr in validations}

    overall = source_data.get("overall_stats", {})
    validation_summary = source_data.get("validation_summary", {})
    top_risk = source_data.get("top_risk_invoices", [])
    vendor_rows = source_data.get("vendor_analysis", [])
    top_failed = source_data.get("top_failed_rules", [])
    narrative = report.narrative or {}

    compliance_agg = validation_qs.aggregate(
        rules_applied=Sum("total_rules"),
        rules_passed=Sum("rules_passed"),
        rules_failed=Sum("rules_failed"),
    )
    rules_applied = int(compliance_agg.get("rules_applied") or 0)
    rules_passed = int(compliance_agg.get("rules_passed") or 0)
    rules_failed = int(compliance_agg.get("rules_failed") or 0)
    compliance_pct = round((rules_passed / rules_applied) * 100, 1) if rules_applied else 0

    risk_avg = float(overall.get("avg_risk_score") or 0)
    high_risk_count = int((overall.get("critical_count") or 0) + (overall.get("high_count") or 0))
    report_status = "high_risk" if risk_avg >= 70 or high_risk_count > 0 else "review" if risk_avg >= 40 else "compliant"

    failed_rules = _failed_rules_with_invoice_refs(top_failed, validations)
    high_risk_invoices = _build_high_risk_violations(top_risk, validation_map, RULES)
    dominant_vendor = vendor_rows[0] if vendor_rows else {}
    duplicate_count = int(overall.get("duplicate_count") or 0)
    missing_qr_count = int(overall.get("missing_qr_count") or 0)
    handwritten_count = int(overall.get("handwritten_count") or 0)
    new_vendor_count = int(overall.get("new_vendor_count") or 0)

    compliance_matrix = [
        {
            "rule_code": row.get("rule_code"),
            "status": "violated",
            "weight": _weight_for_rule(row.get("rule_code"), is_ar),
            "note": row.get("description") or txt(
                "إخفاق متكرر يتطلب إجراءً فوريًا.",
                "Repeated failure requiring immediate action.",
            ),
        }
        for row in top_failed[:6]
    ]
    if not compliance_matrix:
        compliance_matrix = [
            {
                "rule_code": "INV-001",
                "status": "compliant",
                "weight": txt("متوسط", "Medium"),
                "note": txt("توفر رقم الفاتورة في معظم العينات.", "Invoice number is present in most sampled records."),
            },
            {
                "rule_code": "VAT-002",
                "status": "compliant",
                "weight": txt("مرتفع", "High"),
                "note": txt("دقة جيدة في حساب ضريبة القيمة المضافة.", "Good VAT calculation accuracy."),
            },
            {
                "rule_code": "DUP-001",
                "status": "violated",
                "weight": txt("مرتفع", "High"),
                "note": txt("ظهرت حالات تكرار تحتاج تحقيق.", "Duplicate patterns were detected and require investigation."),
            },
        ]

    # ── Rule engine enrichment (if data exists in report) ─────────────────────
    from django.conf import settings as _settings
    re_raw = (report.data or {}).get("rule_engine") or {}
    if not re_raw and getattr(_settings, "USE_NEW_RULE_ENGINE", False):
        re_raw = _collect_rule_engine_data(org, report.period_from or None, report.period_to or None)

    re_runs        = int(re_raw.get("total_runs", 0))
    re_compliance  = float(re_raw.get("compliance_pct", compliance_pct))
    re_avg_risk    = float(re_raw.get("avg_risk_score", risk_avg))
    re_blocking    = int(re_raw.get("blocking_failures", 0))
    re_dist        = re_raw.get("risk_distribution", {})
    re_top_rules   = re_raw.get("top_failed_rules", [])
    re_by_doc      = re_raw.get("by_document_type", [])

    # Build human-readable key findings that incorporate rule engine data
    key_findings = [
        txt(f"عدد الفواتير عالية المخاطر: {high_risk_count}", f"High-risk invoices: {high_risk_count}"),
        txt(f"عدد الفواتير المكررة: {duplicate_count}", f"Duplicate invoices: {duplicate_count}"),
        txt(f"متوسط درجة المخاطر: {risk_avg:.1f}", f"Average risk score: {risk_avg:.1f}"),
    ]
    if re_runs:
        key_findings += [
            txt(
                f"إجمالي المستندات المدقوقة بمحرك القواعد الذكي: {re_runs}",
                f"Documents audited by the smart rule engine: {re_runs}",
            ),
            txt(
                f"إجمالي القواعد المطبقة: {re_raw.get('total_rules_applied', 0)} — نسبة امتثال {re_compliance:.1f}%",
                f"Total applied rules: {re_raw.get('total_rules_applied', 0)} — compliance {re_compliance:.1f}%",
            ),
            txt(f"حالات تعطيل الاعتماد: {re_blocking}", f"Blocking failures: {re_blocking}"),
        ]
        if re_top_rules:
            top_name = re_top_rules[0].get("name_ar") or re_top_rules[0].get("rule_code")
            key_findings.append(
                txt(
                    f"القاعدة الأكثر إخفاقًا: {top_name} ({re_top_rules[0]['fail_count']} إخفاق)",
                    f"Most failed rule: {top_name} ({re_top_rules[0]['fail_count']} failures)",
                )
            )

    # Determine effective compliance rate — prefer rule engine if available
    effective_compliance = re_compliance if re_runs else float(validation_summary.get("vat_compliance_pct") or compliance_pct)

    rule_engine_section = None
    if re_runs:
        rule_engine_section = {
            "total_runs":        re_runs,
            "compliance_pct":    re_compliance,
            "avg_risk_score":    re_avg_risk,
            "blocking_failures": re_blocking,
            "manual_review":     int(re_raw.get("manual_review_required", 0)),
            "total_rules_applied": int(re_raw.get("total_rules_applied", 0)),
            "total_passed":      int(re_raw.get("total_passed", 0)),
            "total_failed":      int(re_raw.get("total_failed", 0)),
            "total_warnings":    int(re_raw.get("total_warnings", 0)),
            "risk_distribution": {
                "critical": int(re_dist.get("critical", 0)),
                "high":     int(re_dist.get("high", 0)),
                "medium":   int(re_dist.get("medium", 0)),
                "low":      int(re_dist.get("low", 0)),
            },
            "top_failed_rules": re_top_rules,
            "by_document_type": re_by_doc,
        }

    return {
        "title": report.title or txt("تقرير تدقيق الفواتير", "Invoice Audit Report"),
        "organization_name": getattr(org, "name", "-") or "-",
        "generated_at": report.created_at,
        "report_type": "Invoice Audit Report",
        "status": report_status,
        "summary": {
            "total_invoices": int(overall.get("total_invoices") or 0),
            "total_amount": float(overall.get("total_amount") or 0),
            "compliance_rate": effective_compliance,
            "avg_risk_score": risk_avg,
            "high_risk_invoices": high_risk_count,
            "duplicate_count": duplicate_count,
        },
        "executive_summary": {
            "conclusion": narrative.get("executive_summary") or txt(
                "الصورة العامة تشير إلى ارتفاع نسبي في مخاطر الامتثال وتحتاج تدخلاً رقابياً سريعاً.",
                "The overall posture indicates elevated compliance risk requiring near-term control action.",
            ),
            "key_findings": key_findings,
            "recommendations": [
                txt("إغلاق المخالفات ذات الأولوية العالية خلال دورة المراجعة الحالية.", "Close high-priority violations within the current review cycle."),
                txt("تفعيل مراجعة استباقية على قواعد التكرار وضريبة القيمة المضافة.", "Apply proactive checks on duplicate and VAT-related rules."),
                txt("رفع جودة بيانات الموردين لتقليل الإخفاقات المتكررة.", "Improve vendor data quality to reduce repeated failures."),
            ],
        },
        "compliance_engine": {
            "rules_applied": rules_applied,
            "rules_passed": rules_passed,
            "rules_failed": rules_failed,
            "compliance_rate": compliance_pct,
            "rules_matrix": compliance_matrix,
        },
        "rule_engine": rule_engine_section,
        "high_risk_invoices": high_risk_invoices,
        "failed_rules": failed_rules,
        "risk_analysis": {
            "avg_risk_score": risk_avg,
            "high": int((overall.get("critical_count") or 0) + (overall.get("high_count") or 0)),
            "review": int(overall.get("medium_count") or 0),
            "safe": int(overall.get("low_count") or 0),
        },
        "duplicates_anomalies": {
            "duplicates": duplicate_count,
            "dominant_vendor": dominant_vendor.get("vendor_name") or "-",
            "vendor_dependency_pct": float(dominant_vendor.get("spend_share_pct") or 0),
            "patterns": [
                {"label": txt("غياب رمز QR", "Missing QR code"), "count": missing_qr_count, "severity": "high" if missing_qr_count else "low"},
                {"label": txt("فواتير بخط يدوي", "Handwritten invoices"), "count": handwritten_count, "severity": "review" if handwritten_count else "low"},
                {"label": txt("موردون جدد غير معروفين", "New unknown vendors"), "count": new_vendor_count, "severity": "review" if new_vendor_count else "low"},
            ],
        },
        "vendor_analysis": vendor_rows,
        "recommendations": {
            "immediate": [
                txt("تصحيح فواتير QR غير المتوافقة قبل الإقفال المالي.", "Fix non-compliant QR invoices before financial close."),
                txt("مراجعة المورد المسيطر وتدقيق معاملات العينة عالية القيمة.", "Review dominant-vendor exposure and high-value sampled transactions."),
                txt("التحقق من الرقم الضريبي للفواتير المرفوضة رقابياً.", "Verify VAT registration numbers for control-rejected invoices."),
            ],
            "future": [
                txt("تطبيق مراقبة لحظية لقواعد التكرار قبل اعتماد الفاتورة.", "Enable real-time duplicate-rule monitoring before invoice approval."),
                txt("تحسين جودة الإدخال وتدريب الفريق على متطلبات الامتثال.", "Improve data-entry quality and train the team on compliance controls."),
                txt("إضافة حدود تنبيه تلقائية عند ارتفاع تركز الإنفاق على مورد واحد.", "Add threshold alerts for high spend concentration on a single vendor."),
            ],
        },
    }


# ─── Views ────────────────────────────────────────────────────────────────────

class GenerateAuditReportView(APIView):
    """Generate a comprehensive AI audit report including all invoice audit data."""
    permission_classes = [IsAuthenticated, IsSeniorAuditorOrAbove]

    @extend_schema(
        tags=["Reports"],
        summary="Generate a full audit report — includes invoices, transactions, and AI narrative",
        request={
            "type": "object",
            "properties": {
                "report_type": {
                    "type": "string",
                    "enum": [
                        "executive_summary",
                        "invoice_audit",
                        "detailed_findings",
                        "compliance",
                        "risk_assessment",
                        "vendor_analysis",
                        "spend_analysis",
                    ],
                    "default": "executive_summary",
                },
                "date_from": {"type": "string", "format": "date"},
                "date_to":   {"type": "string", "format": "date"},
                "language":  {"type": "string", "enum": ["en", "ar"], "default": "en"},
                "audit_session_id": {"type": "string", "format": "uuid"},
                "include_invoices":     {"type": "boolean", "default": True},
                "include_transactions": {"type": "boolean", "default": True},
            },
        },
    )
    def post(self, request):
        org = request.user.organization
        if not org:
            return Response({"error": "No organization."}, status=400)

        report_type       = request.data.get("report_type", "executive_summary")
        date_from         = request.data.get("date_from")
        date_to           = request.data.get("date_to") or str(date.today())
        language          = request.data.get("language", "en")
        audit_session_id  = request.data.get("audit_session_id")
        include_invoices  = request.data.get("include_invoices", True)
        include_tx        = request.data.get("include_transactions", True)

        audit_data = {
            "organization": {
                "name":     org.name,
                "country":  org.get_country_display(),
                "industry": org.industry,
                "vat_number": org.vat_number,
                "currency": org.currency,
            },
            "audit_period": {"from": date_from, "to": date_to},
            "report_type":  report_type,
            "audit_session_id": audit_session_id,
        }

        # ── Invoice Audit Section ──────────────────────────────────────────────
        if include_invoices:
            audit_data["invoice_audit"] = _collect_invoice_data(org, date_from, date_to, audit_session_id=audit_session_id)

        # ── Rule Engine Section (40-rule smart audit) ──────────────────────────
        from django.conf import settings as _settings
        if getattr(_settings, "USE_NEW_RULE_ENGINE", False):
            re_data = _collect_rule_engine_data(org, date_from, date_to)
            if re_data:
                audit_data["rule_engine"] = re_data

        # ── Transaction Section ────────────────────────────────────────────────
        if include_tx:
            tx_qs = Transaction.objects.filter(organization=org)
            if date_from:
                tx_qs = tx_qs.filter(transaction_date__gte=date_from)
            tx_qs = tx_qs.filter(transaction_date__lte=date_to)

            tx_stats = tx_qs.aggregate(
                total_income=Sum("amount", filter=Q(transaction_type="income")),
                total_expense=Sum("amount", filter=Q(transaction_type="expense")),
                total_vat=Sum("vat_amount"),
                tx_count=Count("id"),
                flagged_count=Count("id", filter=Q(is_flagged=True)),
                avg_risk=Avg("risk_score"),
                critical_count=Count("id", filter=Q(risk_level="critical")),
                high_count=Count("id", filter=Q(risk_level="high")),
            )
            top_risk_tx = list(
                tx_qs.filter(risk_score__gt=50).order_by("-risk_score").values(
                    "id", "transaction_type", "amount", "currency", "vendor_name",
                    "description", "risk_score", "risk_level", "flag_reason", "transaction_date"
                )[:15]
            )
            for t in top_risk_tx:
                t["id"] = str(t["id"])
                t["amount"] = float(t["amount"])
                t["transaction_date"] = str(t["transaction_date"])

            audit_data["financial_summary"] = {
                "total_income":    float(tx_stats["total_income"]  or 0),
                "total_expense":   float(tx_stats["total_expense"] or 0),
                "net_profit":      float((tx_stats["total_income"] or 0) - (tx_stats["total_expense"] or 0)),
                "total_vat":       float(tx_stats["total_vat"]     or 0),
                "transaction_count": tx_stats["tx_count"],
                "flagged_transactions": tx_stats["flagged_count"],
                "flag_rate_pct":   round(tx_stats["flagged_count"] / (tx_stats["tx_count"] or 1) * 100, 2),
                "avg_risk_score":  round(float(tx_stats["avg_risk"] or 0), 2),
                "critical_transactions": tx_stats["critical_count"],
                "top_risk_transactions": top_risk_tx,
            }

        # ── Audit Cases Section ────────────────────────────────────────────────
        cases_qs = AuditCase.objects.filter(organization=org)
        if audit_session_id:
            cases_qs = cases_qs.filter(invoice__audit_session_id=audit_session_id)
        if date_from:
            cases_qs = cases_qs.filter(created_at__date__gte=date_from)
        cases_qs = cases_qs.filter(created_at__date__lte=date_to)
        audit_data["audit_cases"] = cases_qs.aggregate(
            total=Count("id"),
            open=Count("id", filter=Q(status="open")),
            resolved=Count("id", filter=Q(status="resolved")),
            critical=Count("id", filter=Q(priority="critical")),
            high=Count("id", filter=Q(priority="high")),
        )

        # ── AI Narrative ───────────────────────────────────────────────────────
        blocked = ai_access_response_or_none(org)
        if blocked is not None:
            return blocked
        audit_data = _json_safe(audit_data)
        narrative = generate_audit_narrative(audit_data, language=language)

        # ── Save Report ────────────────────────────────────────────────────────
        report = Report.objects.create(
            organization=org,
            generated_by=request.user,
            report_type=report_type,
            language=language,
            period_from=date_from or "",
            period_to=date_to or str(date.today()),
            data=audit_data,
            narrative=narrative,
            title=f"{report_type.replace('_', ' ').title()} Report — {date_to}",
        )

        return Response({
            "id":           str(report.id),
            "report_id":    str(report.id),
            "title":        report.title,
            "report_type":  report_type,
            "language":     language,
            "generated_at": report.created_at.isoformat(),
            "data":         audit_data,
            "narrative":    narrative,
        })


class InvoiceAuditReportView(APIView):
    """Dedicated invoice audit report with all 30 rule statistics."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Reports"],
        summary="Invoice audit report — full 30-rule breakdown, vendor risk, duplicates, spend analysis",
        parameters=[
            OpenApiParameter("date_from"),
            OpenApiParameter("date_to"),
            OpenApiParameter("language", description="en or ar"),
        ],
    )
    def get(self, request):
        org = request.user.organization
        if not org:
            return Response({"error": "No organization."}, status=400)

        date_from = request.query_params.get("date_from")
        date_to   = request.query_params.get("date_to") or str(date.today())
        language  = request.query_params.get("language", "en")

        data = _json_safe(_collect_invoice_data(org, date_from, date_to))

        # AI narrative for invoice audit specifically
        blocked = ai_access_response_or_none(org)
        if blocked is not None:
            return blocked
        narrative = generate_audit_narrative(
            {
                "organization": {"name": org.name, "country": org.country, "currency": org.currency},
                "audit_period": {"from": date_from, "to": date_to},
                "report_type":  "invoice_audit",
                "invoice_audit": data,
            },
            language=language,
        )

        report = Report.objects.create(
            organization=org,
            generated_by=request.user,
            report_type="invoice_audit",
            language=language,
            period_from=date_from or "",
            period_to=date_to,
            data=data,
            narrative=narrative,
            title=f"Invoice Audit Report — {date_to}",
        )

        return Response({
            "id":           str(report.id),
            "report_id":    str(report.id),
            "title":        report.title,
            "generated_at": report.created_at.isoformat(),
            "language":     language,
            "data":         data,
            "narrative":    narrative,
        })


class ValidationRulesSummaryReportView(APIView):
    """Report: how many invoices passed/failed each of the 30 rules."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Reports"],
        summary="30-rule pass/fail breakdown across all invoices",
        parameters=[
            OpenApiParameter("date_from"),
            OpenApiParameter("date_to"),
        ],
    )
    def get(self, request):
        from core.services.invoice_validator import RULES
        org = request.user.organization
        inv_qs = Invoice.objects.filter(organization=org)
        if v := request.query_params.get("date_from"):
            inv_qs = inv_qs.filter(invoice_date__gte=v)
        if v := request.query_params.get("date_to"):
            inv_qs = inv_qs.filter(invoice_date__lte=v)

        vr_qs = InvoiceValidationResult.objects.filter(invoice__in=inv_qs)
        total = vr_qs.count()

        rule_breakdown = []
        for code, description in RULES.items():
            failures = sum(
                1 for vr in vr_qs.only("failed_rule_codes")
                if code in (vr.failed_rule_codes or [])
            )
            rule_breakdown.append({
                "rule_code":    code,
                "description":  description,
                "group":        code.split("-")[0],
                "failures":     failures,
                "passes":       total - failures,
                "pass_rate_pct": round((total - failures) / total * 100, 1) if total else 0,
                "fail_rate_pct": round(failures / total * 100, 1) if total else 0,
            })

        # Sort by fail rate descending
        rule_breakdown.sort(key=lambda x: -x["fail_rate_pct"])

        return Response({
            "report_type":  "30_rules_summary",
            "total_invoices_analyzed": total,
            "generated_at": str(date.today()),
            "rules":        rule_breakdown,
            "most_problematic": rule_breakdown[:5],
        })


class ReportListView(APIView):
    """List all generated reports."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Reports"],
        summary="List all generated reports",
        parameters=[OpenApiParameter("report_type", description="Filter by type")],
    )
    def get(self, request):
        from .serializers import ReportSerializer
        qs = Report.objects.filter(organization=request.user.organization).order_by("-created_at")
        if v := request.query_params.get("report_type"):
            qs = qs.filter(report_type=v)
        return Response(ReportSerializer(qs[:50], many=True).data)


class ReportDetailView(APIView):
    """Retrieve a specific report."""
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Reports"], summary="Get a specific report with full data and narrative")
    def get(self, request, pk):
        try:
            report = Report.objects.get(pk=pk, organization=request.user.organization)
        except Report.DoesNotExist:
            return Response({"error": "Report not found."}, status=404)

        return Response({
            "id":           str(report.id),
            "title":        report.title,
            "report_type":  report.report_type,
            "language":     report.language,
            "period_from":  report.period_from,
            "period_to":    report.period_to,
            "generated_by": report.generated_by.full_name if report.generated_by else None,
            "created_at":   report.created_at.isoformat(),
            "data":         report.data,
            "narrative":    report.narrative,
        })


class ReportExportView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Reports"], summary="Export organization data as JSON")
    def get(self, request):
        org = request.user.organization
        if not org:
            return Response({"error": "No organization."}, status=400)

        try:
            org_settings = getattr(org, "settings", None) or getattr(org, "finai_settings", None)
        except Exception:
            org_settings = None

        payload = {
            "exported_at": date.today().isoformat(),
            "organization": {
                "id": str(org.id),
                "name": org.name,
                "name_ar": org.name_ar,
                "country": org.country,
                "currency": org.currency,
                "vat_number": org.vat_number,
                "cr_number": org.cr_number,
                "vat_rate": org.vat_rate,
                "address": org.address,
                "website": getattr(org, "website", ""),
            },
            "settings": {
                "financial": getattr(org_settings, "financial", {}),
                "notifications": getattr(org_settings, "notifications", {}),
            },
            "reports": list(Report.objects.filter(organization=org).values()),
            "invoices": list(Invoice.objects.filter(organization=org).values()),
            "transactions": list(Transaction.objects.filter(organization=org).values()),
            "documents": list(Document.objects.filter(organization=org).values()),
            "audit_cases": list(AuditCase.objects.filter(organization=org).values()),
            "compliance_violations": list(ComplianceViolation.objects.filter(organization=org).values()),
        }

        response = HttpResponse(
            json.dumps(payload, cls=DjangoJSONEncoder, ensure_ascii=False, indent=2),
            content_type="application/json; charset=utf-8",
        )
        response["Content-Disposition"] = _attachment_disposition(f"tadgeeg-ai-export-{org.id}.json")
        return response


class ReportPDFView(APIView):
    """Render a saved report as a downloadable PDF."""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            report = Report.objects.get(pk=pk, organization=request.user.organization)
        except Report.DoesNotExist:
            return HttpResponse(_("Report not found."), status=404)

        from django.template.loader import render_to_string
        template_name = "reports/invoice_audit_report_pdf.html" if report.report_type == "invoice_audit" else "reports/report_pdf.html"
        html_str = render_to_string(template_name, {
            "report": report,
            "narrative": report.narrative or {},
            "data": report.data or {},
            "org": request.user.organization,
            "invoice_audit_ui": _build_invoice_audit_report_ui(report, request.user.organization) if report.report_type == "invoice_audit" else None,
        }, request=request)

        safe_title = (report.title or "report").replace(" ", "_").replace("/", "-")[:60]

        from django.utils.translation import get_language
        try:
            pdf_bytes = _render_report_pdf_cached(
                report.id,
                (get_language() or "ar")[:2],
                html_str,
                request.build_absolute_uri("/"),
            )
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Length"] = str(len(pdf_bytes))
            response["Content-Disposition"] = _attachment_disposition(f"{safe_title}.pdf")
            return response
        except (ModuleNotFoundError, OSError) as exc:
            # ModuleNotFoundError = WeasyPrint missing (server misconfigured).
            # OSError            = GTK / Pango / cairo libs missing (common
            #                      on Windows dev boxes).
            # In either case the server CAN still serve the HTML — let the
            # browser do Ctrl+P → Save as PDF — instead of blowing up.
            logger.warning(
                "WeasyPrint unavailable (%s) — serving HTML fallback: %s",
                type(exc).__name__, exc,
                extra={"report_id": str(report.id)},
            )
        except Exception:
            logger.exception("Report PDF generation failed", extra={"report_id": str(report.id)})
            return JsonResponse(
                {"error": _("Failed to generate PDF for this report. Check the server log and try again.")},
                status=500,
            )

        # HTML fallback: browser can use Ctrl+P → Save as PDF
        response = HttpResponse(html_str, content_type="text/html; charset=utf-8")
        response["Content-Disposition"] = _attachment_disposition(f"{safe_title}.html")
        return response


class ReportExcelExportView(APIView):
    """Export a saved report as an .xlsx workbook (FR-7 gap fix)."""
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Reports"], summary="Export report as Excel (.xlsx)")
    def get(self, request, pk):
        try:
            report = Report.objects.get(pk=pk, organization=request.user.organization)
        except Report.DoesNotExist:
            return HttpResponse(_("Report not found."), status=404)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return JsonResponse({"error": _("openpyxl not available on server.")}, status=500)

        import io
        wb = openpyxl.Workbook()
        is_ar = str(getattr(report, "language", "") or "").lower().startswith("ar")

        def label(ar_text, en_text):
            return ar_text if is_ar else en_text

        # ── Sheet 1: Summary ────────────────────────────────────────────────
        ws = wb.active
        ws.title = label("الملخص", "Summary")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")

        def _hdr(cell_ref, value):
            cell = ws[cell_ref]
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        org = request.user.organization
        rows = [
            (label("عنوان التقرير", "Report Title"), report.title),
            (label("نوع التقرير", "Report Type"), report.report_type),
            (label("المنظمة", "Organization"), org.name),
            (label("الرقم الضريبي", "VAT Number"), org.vat_number or ""),
            (label("الفترة من", "Period From"), str(report.period_from or "")),
            (label("الفترة إلى", "Period To"), str(report.period_to or "")),
            (label("أُنشئ بواسطة", "Generated By"), report.generated_by.full_name if report.generated_by else ""),
            (label("تاريخ الإنشاء", "Generated At"), report.created_at.strftime("%Y-%m-%d %H:%M")),
            (label("اللغة", "Language"), report.language),
        ]
        _hdr("A1", label("الحقل", "Field"))
        _hdr("B1", label("القيمة", "Value"))
        for i, (k, v) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 40

        # Narrative KPIs if present
        narrative = report.narrative or {}
        if narrative:
            ws.append([])
            ws.append([label("── مؤشرات السرد ──", "── Narrative KPIs ──")])
            for key, val in narrative.items():
                if not isinstance(val, (dict, list)):
                    ws.append([str(key), str(val)])

        # ── Sheet 2: Invoices ────────────────────────────────────────────────
        data = report.data or {}
        invoices_data = data.get("invoices", [])
        ws2 = wb.create_sheet(label("الفواتير", "Invoices"))
        inv_headers = [
            label("رقم الفاتورة", "Invoice Number"),
            label("المورّد", "Supplier"),
            label("التاريخ", "Date"),
            label("المبلغ الأساسي", "Base Amount"),
            label("الضريبة", "Tax Amount"),
            label("الإجمالي", "Total Amount"),
            label("العملة", "Currency"),
            label("الحالة", "Status"),
            label("حالة الامتثال", "Compliance Status"),
        ]
        for col, h in enumerate(inv_headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        if isinstance(invoices_data, list):
            for row_i, inv in enumerate(invoices_data, start=2):
                if isinstance(inv, dict):
                    ws2.cell(row=row_i, column=1, value=inv.get("invoice_number", ""))
                    ws2.cell(row=row_i, column=2, value=inv.get("supplier_name", ""))
                    ws2.cell(row=row_i, column=3, value=str(inv.get("invoice_date", "")))
                    ws2.cell(row=row_i, column=4, value=float(inv.get("total_amount", 0) or 0))
                    ws2.cell(row=row_i, column=5, value=float(inv.get("tax_amount", 0) or 0))
                    ws2.cell(row=row_i, column=6, value=float(inv.get("total_amount", 0) or 0))
                    ws2.cell(row=row_i, column=7, value=inv.get("currency", "SAR"))
                    ws2.cell(row=row_i, column=8, value=inv.get("status", ""))
                    ws2.cell(row=row_i, column=9, value=inv.get("compliance_status", ""))
        for col in range(1, len(inv_headers) + 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

        # ── Sheet 3: Audit Cases / Findings ─────────────────────────────────
        ws3 = wb.create_sheet(label("الملاحظات", "Findings"))
        cases_qs = AuditCase.objects.filter(
            organization=request.user.organization
        ).values("case_number", "case_type", "priority", "status", "title",
                 "description", "created_at", "resolved_at")
        case_headers = [
            label("رقم الحالة", "Case #"),
            label("النوع", "Type"),
            label("الأولوية", "Priority"),
            label("الحالة", "Status"),
            label("العنوان", "Title"),
            label("الوصف", "Description"),
            label("تاريخ الإنشاء", "Created At"),
            label("تاريخ الإغلاق", "Resolved At"),
        ]
        for col, h in enumerate(case_headers, 1):
            c = ws3.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        for row_i, case in enumerate(cases_qs, start=2):
            ws3.cell(row=row_i, column=1, value=case["case_number"])
            ws3.cell(row=row_i, column=2, value=case["case_type"])
            ws3.cell(row=row_i, column=3, value=case["priority"])
            ws3.cell(row=row_i, column=4, value=case["status"])
            ws3.cell(row=row_i, column=5, value=case["title"])
            ws3.cell(row=row_i, column=6, value=case["description"] or "")
            ws3.cell(row=row_i, column=7, value=str(case["created_at"])[:19] if case["created_at"] else "")
            ws3.cell(row=row_i, column=8, value=str(case["resolved_at"])[:19] if case["resolved_at"] else "")
        for col in range(1, len(case_headers) + 1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # ── Sheet 4: Rule Summary ────────────────────────────────────────────
        ws4 = wb.create_sheet(label("ملخص القواعد", "Rule Summary"))
        rule_summary = data.get("rule_group_summary", [])
        rs_headers = [
            label("مجموعة القواعد", "Rule Group"),
            label("ناجح", "Passed"),
            label("فاشل", "Failed"),
            label("الإجمالي", "Total"),
            label("نسبة النجاح %", "Pass Rate %"),
        ]
        for col, h in enumerate(rs_headers, 1):
            c = ws4.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        if isinstance(rule_summary, list):
            for row_i, rs in enumerate(rule_summary, start=2):
                if isinstance(rs, dict):
                    ws4.cell(row=row_i, column=1, value=rs.get("label", rs.get("code", "")))
                    ws4.cell(row=row_i, column=2, value=rs.get("passed", 0))
                    ws4.cell(row=row_i, column=3, value=rs.get("failed", 0))
                    ws4.cell(row=row_i, column=4, value=rs.get("total", 0))
                    ws4.cell(row=row_i, column=5, value=rs.get("pct", 0))

        # ── Serialize and return ─────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_title = (report.title or "report").replace(" ", "_").replace("/", "-")[:50]
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = _attachment_disposition(f"{safe_title}.xlsx")
        return response


class ReportEmailView(APIView):
    """Email a report PDF to the requesting user (FR-7 gap fix)."""
    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Reports"], summary="Email report PDF to current user")
    def post(self, request, pk):
        try:
            report = Report.objects.get(pk=pk, organization=request.user.organization)
        except Report.DoesNotExist:
            return Response({"error": _("Report not found.")}, status=404)

        from django.core.mail import EmailMessage
        from django.template.loader import render_to_string
        from django.conf import settings as django_settings

        template_name = "reports/invoice_audit_report_pdf.html" if report.report_type == "invoice_audit" else "reports/report_pdf.html"
        html_str = render_to_string(template_name, {
            "report": report,
            "narrative": report.narrative or {},
            "data": report.data or {},
            "org": request.user.organization,
            "invoice_audit_ui": _build_invoice_audit_report_ui(report, request.user.organization) if report.report_type == "invoice_audit" else None,
        }, request=request)

        safe_title = (report.title or "report").replace(" ", "_").replace("/", "-")[:60]
        recipient = request.user.email

        try:
            pdf_bytes = _render_report_pdf_bytes(html_str, request.build_absolute_uri("/"))
            msg = EmailMessage(
                subject=f"[Tadgeeg AI] {report.title}",
                body=_("Attached report: %(title)s\n\nGenerated at %(dt)s") % {"title": report.title, "dt": report.created_at.strftime("%Y-%m-%d %H:%M")},
                from_email=getattr(django_settings, "DEFAULT_FROM_EMAIL", "noreply@tadgeeg.com"),
                to=[recipient],
            )
            msg.attach(f"{safe_title}.pdf", pdf_bytes, "application/pdf")
            msg.send()
            return Response({"success": True, "message": _("Report sent to %(recipient)s") % {"recipient": recipient}})
        except Exception as e:
            logger.exception("Report email failed for report %s", pk)
            return Response({"error": _("Failed to send email. Check SMTP settings.")}, status=500)


# ─── Async PDF rendering ──────────────────────────────────────────────────────
# The sync ReportPDFView caches its output, so repeat downloads are free —
# but the first render of a 30-60 page audit report can take 5-15s and
# exceed the gunicorn worker timeout. These two views give clients a way to
# pre-warm the PDF cache via Celery and then download from the existing
# sync endpoint with an instant cache hit.

class ReportPDFAsyncView(APIView):
    """POST: enqueue async PDF rendering. Returns task_id for polling."""

    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Reports"], summary="Queue async PDF render for a report")
    def post(self, request, pk):
        try:
            report = Report.objects.get(pk=pk, organization=request.user.organization)
        except Report.DoesNotExist:
            return Response({"error": _("Report not found.")}, status=404)

        from django.utils.translation import get_language
        from apps.reports.tasks import render_report_pdf

        language = (get_language() or "ar")[:2]
        try:
            task = render_report_pdf.delay(report_id=str(report.id), language=language)
        except Exception as exc:
            logger.warning("Could not enqueue PDF render task: %s", exc)
            return Response({
                "error": _("Async rendering backend unavailable. "
                           "Download directly via /pdf/ instead."),
            }, status=503)

        return Response({
            "task_id":      task.id,
            "status":       "queued",
            "report_id":    str(report.id),
            "language":     language,
            "poll_url":     f"/api/v1/reports/pdf-jobs/{task.id}/",
            "download_url": f"/api/v1/reports/{report.id}/pdf/",
        }, status=202)


class ReportPDFStatusView(APIView):
    """GET: poll status of an async PDF render task."""

    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Reports"], summary="Poll async PDF render status")
    def get(self, request, task_id: str):
        from celery.result import AsyncResult
        from finai_backend.celery import app as celery_app

        result = AsyncResult(task_id, app=celery_app)
        state = result.state  # PENDING, STARTED, RETRY, FAILURE, SUCCESS

        payload = {
            "task_id": task_id,
            "state":   state,
        }
        if state == "SUCCESS":
            info = result.result or {}
            payload["status"] = info.get("status", "done")
            payload["report_id"] = info.get("report_id")
            payload["download_url"] = (
                f"/api/v1/reports/{info.get('report_id')}/pdf/"
                if info.get("status") == "done" else None
            )
            return Response(payload)
        if state in ("FAILURE", "REVOKED"):
            payload["status"] = "failed"
            payload["error"] = str(result.result) if result.result else "unknown"
            return Response(payload, status=500)
        # PENDING / STARTED / RETRY — caller should keep polling
        payload["status"] = state.lower()
        return Response(payload, status=202)


# ─── Profit & Loss + Balance Sheet ────────────────────────────────────────────
# Both reports derive directly from the existing Ledger rows in the org. Each
# Ledger carries `account_type` (free-text, e.g. "Asset" / "Liability" /
# "Equity" / "Revenue" / "Expense"), opening_balance, closing_balance, and
# the period bounds.
#
# The classification map below is intentionally permissive — many orgs ship
# Arabic synonyms, and the .account_type column is free-text. Anything that
# doesn't match any bucket falls into "unclassified", which the report
# surfaces so operators can fix the source data without losing visibility.

_PNL_REVENUE  = {"revenue", "income", "sales", "إيراد", "إيرادات", "مبيعات"}
_PNL_EXPENSE  = {"expense", "expenses", "cost", "cogs", "مصروف", "مصاريف", "تكلفة"}
_BS_ASSET     = {"asset", "assets", "أصل", "أصول"}
_BS_LIABILITY = {"liability", "liabilities", "payable", "خصم", "خصوم", "التزام"}
_BS_EQUITY    = {"equity", "capital", "حقوق ملكية", "رأس مال", "حقوق المساهمين"}


def _classify_account_type(value: str) -> str:
    key = (value or "").strip().lower()
    if key in _PNL_REVENUE:  return "revenue"
    if key in _PNL_EXPENSE:  return "expense"
    if key in _BS_ASSET:     return "asset"
    if key in _BS_LIABILITY: return "liability"
    if key in _BS_EQUITY:    return "equity"
    return "unclassified"


def _ledger_rows(org, period_from=None, period_to=None):
    """Tenant-scoped Ledger rows in the requested period, only the fields
    these reports need so we don't drag in the heavy JSON columns."""
    from apps.documents.typed_models_v2 import Ledger

    qs = (Ledger.objects.filter(organization=org)
                         .only("id", "account_number", "account_name",
                               "account_type", "currency",
                               "opening_balance", "closing_balance",
                               "period_from", "period_to"))
    if period_from:
        qs = qs.filter(period_to__gte=period_from)
    if period_to:
        qs = qs.filter(period_from__lte=period_to)
    return qs


class ProfitAndLossReportView(APIView):
    """GET /api/v1/reports/profit-and-loss/ — P&L derived from Ledger rows."""

    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Reports"],
        summary="Profit & Loss statement (derived from per-account ledgers)",
        parameters=[
            OpenApiParameter("period_from", description="ISO date (inclusive)"),
            OpenApiParameter("period_to",   description="ISO date (inclusive)"),
        ],
    )
    def get(self, request):
        from datetime import date as _date
        try:
            period_from = _date.fromisoformat(request.query_params["period_from"]) \
                if request.query_params.get("period_from") else None
            period_to = _date.fromisoformat(request.query_params["period_to"]) \
                if request.query_params.get("period_to") else None
        except ValueError:
            return Response({"error": _("Invalid period dates")}, status=400)

        rows = _ledger_rows(request.user.organization, period_from, period_to)
        revenue_lines = []
        expense_lines = []
        unclassified  = []
        total_rev = 0.0
        total_exp = 0.0

        for r in rows:
            bucket = _classify_account_type(r.account_type or "")
            # P&L convention: revenue = credit balance (we approximate via the
            # change between opening and closing for both revenue and expense).
            opening = float(r.opening_balance or 0)
            closing = float(r.closing_balance or 0)
            movement = closing - opening
            line = {
                "account_number": r.account_number,
                "account_name":   r.account_name,
                "account_type":   r.account_type,
                "movement":       round(movement, 2),
                "opening":        round(opening, 2),
                "closing":        round(closing, 2),
                "currency":       r.currency or "SAR",
            }
            if bucket == "revenue":
                revenue_lines.append(line)
                total_rev += movement
            elif bucket == "expense":
                expense_lines.append(line)
                total_exp += movement
            elif bucket in ("asset", "liability", "equity"):
                continue  # belongs on the Balance Sheet, not P&L
            else:
                unclassified.append(line)

        net_income = total_rev - total_exp
        return Response({
            "period_from":     str(period_from) if period_from else None,
            "period_to":       str(period_to)   if period_to   else None,
            "revenue": {
                "total": round(total_rev, 2),
                "lines": revenue_lines,
            },
            "expenses": {
                "total": round(total_exp, 2),
                "lines": expense_lines,
            },
            "net_income": round(net_income, 2),
            "unclassified": unclassified,
            "notes": (
                "Derived from per-account Ledger rows in the requested period. "
                "Movement = closing_balance - opening_balance. Rows whose "
                "account_type does not match a revenue/expense/asset/liability/"
                "equity synonym appear under 'unclassified' — fix the source "
                "Ledger.account_type to clear them."
            ),
        })


class BalanceSheetReportView(APIView):
    """GET /api/v1/reports/balance-sheet/ — Balance Sheet from Ledger rows."""

    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Reports"],
        summary="Balance Sheet (derived from per-account ledgers)",
        parameters=[
            OpenApiParameter("as_of", description="ISO date (snapshot date)"),
        ],
    )
    def get(self, request):
        from datetime import date as _date
        as_of = None
        if v := request.query_params.get("as_of"):
            try:
                as_of = _date.fromisoformat(v)
            except ValueError:
                return Response({"error": _("Invalid as_of date")}, status=400)

        # We pull every ledger row whose period_to ≤ as_of (latest snapshot per
        # account). For the simple derivation we trust the most recent
        # closing_balance per account_number.
        rows = _ledger_rows(request.user.organization, period_to=as_of)

        latest_per_account: dict = {}
        for r in rows:
            key = r.account_number or str(r.id)
            existing = latest_per_account.get(key)
            if existing is None or (r.period_to and existing.period_to
                                    and r.period_to > existing.period_to):
                latest_per_account[key] = r

        assets, liabilities, equity, unclassified = [], [], [], []
        total_a = total_l = total_e = 0.0
        for r in latest_per_account.values():
            bucket = _classify_account_type(r.account_type or "")
            closing = float(r.closing_balance or 0)
            line = {
                "account_number": r.account_number,
                "account_name":   r.account_name,
                "account_type":   r.account_type,
                "balance":        round(closing, 2),
                "currency":       r.currency or "SAR",
            }
            if bucket == "asset":
                assets.append(line); total_a += closing
            elif bucket == "liability":
                liabilities.append(line); total_l += closing
            elif bucket == "equity":
                equity.append(line); total_e += closing
            elif bucket in ("revenue", "expense"):
                continue  # belongs on the P&L
            else:
                unclassified.append(line)

        balanced = abs(total_a - (total_l + total_e)) < 1.00
        return Response({
            "as_of": str(as_of) if as_of else None,
            "assets":      {"total": round(total_a, 2), "lines": assets},
            "liabilities": {"total": round(total_l, 2), "lines": liabilities},
            "equity":      {"total": round(total_e, 2), "lines": equity},
            "check": {
                "assets":              round(total_a, 2),
                "liabilities_plus_equity": round(total_l + total_e, 2),
                "difference":          round(total_a - (total_l + total_e), 2),
                "balanced":            balanced,
            },
            "unclassified": unclassified,
            "notes": (
                "Latest closing_balance per account_number on or before as_of. "
                "Assets should equal Liabilities + Equity; the 'check' object "
                "surfaces the difference. Rows whose account_type does not "
                "match a known bucket appear under 'unclassified'."
            ),
        })


# ─── Three-Way Match Report ────────────────────────────────────────────────────
# The ThreeWayMatchService at apps/rule_engine/services/three_way_match.py has
# been doing all the work; what was missing was an HTTP endpoint that runs it
# across an organization and returns a digest. The spec lists this as a P0
# report. JSON-only for now — PDF export can follow the same template pattern
# as the other report types when needed.

class ThreeWayMatchReportView(APIView):
    """Generate a 3-way match (PO ↔ Invoice ↔ GRN) report for the current org.

    For every PO that has both a matching Invoice and matching GRN (linked by
    po_number), runs the match service and aggregates a summary. Returns:

      {
        "summary": {
          "scanned_pos":          int,
          "triplets_evaluated":   int,
          "matched":              int,
          "partial":              int,
          "mismatched":           int,
          "missing_invoice":      int,
          "missing_grn":          int,
        },
        "discrepancies": [ ... up to 50 worst rows ... ],
      }
    """

    permission_classes = [IsAuthenticated]

    @extend_schema(tags=["Reports"], summary="Three-way match (PO ↔ Invoice ↔ GRN) report")
    def get(self, request):
        from apps.documents.typed_models import PurchaseOrder, GoodsReceiptNote
        from apps.invoices.models import Invoice
        from apps.rule_engine.services.three_way_match import (
            ThreeWayMatchService, MatchStatus,
        )

        org = request.user.organization
        # Tenant-scoped POs that carry a po_number. Linking direction:
        #   • PO → Invoice via `PurchaseOrder.linked_invoice_id` (UUID FK column).
        #   • PO → GRN     via `GoodsReceiptNote.linked_po_number` (string column).
        # Invoice itself has no `po_number` column — link is one-way from PO.
        po_qs = (PurchaseOrder.objects.filter(organization=org)
                                       .exclude(po_number="")
                                       .only("id", "po_number", "linked_invoice_id"))

        summary = {
            "scanned_pos":        po_qs.count(),
            "triplets_evaluated": 0,
            "matched":            0,
            "partial":            0,
            "mismatched":         0,
            "missing_invoice":    0,
            "missing_grn":        0,
        }
        worst: list[dict] = []
        svc = ThreeWayMatchService()
        org_id = str(org.id)

        for po in po_qs.iterator():
            inv = None
            if po.linked_invoice_id:
                inv = (Invoice.objects.filter(organization=org, id=po.linked_invoice_id)
                                      .only("id", "invoice_number", "total_amount",
                                            "vendor_name", "currency")
                                      .first())
            if inv is None:
                summary["missing_invoice"] += 1
                continue
            grn = (GoodsReceiptNote.objects.filter(organization=org,
                                                   linked_po_number=po.po_number)
                                            .only("id", "grn_number", "vendor_name")
                                            .first())
            if grn is None:
                summary["missing_grn"] += 1
                continue

            try:
                result = svc.match(po=po, invoice=inv, grn=grn, organization_id=org_id)
            except Exception as exc:
                logger.warning("3WM failed po=%s: %s", po.po_number, exc)
                continue

            summary["triplets_evaluated"] += 1
            status_str = result.status.value if hasattr(result.status, "value") else str(result.status)
            if status_str in ("match", "matched", MatchStatus.MATCH.value if hasattr(MatchStatus, "MATCH") else "match"):
                summary["matched"] += 1
            elif status_str in ("partial", "partial_match", MatchStatus.PARTIAL.value if hasattr(MatchStatus, "PARTIAL") else "partial"):
                summary["partial"] += 1
                worst.append(self._discrepancy_row(result, "partial"))
            else:
                summary["mismatched"] += 1
                worst.append(self._discrepancy_row(result, status_str))

        # Sort worst by match_score ascending (lowest first), trim to 50.
        worst.sort(key=lambda r: r.get("match_score", 0))
        return Response({
            "summary":       summary,
            "discrepancies": worst[:50],
        })

    @staticmethod
    def _discrepancy_row(result, status_str: str) -> dict:
        return {
            "status":         status_str,
            "match_score":    float(getattr(result, "match_score", 0) or 0),
            "po_id":          str(getattr(result, "po_id", "") or ""),
            "po_number":      getattr(result, "po_number", "") or "",
            "invoice_id":     str(getattr(result, "invoice_id", "") or ""),
            "invoice_number": getattr(result, "invoice_number", "") or "",
            "grn_id":         str(getattr(result, "grn_id", "") or ""),
            "grn_number":     getattr(result, "grn_number", "") or "",
            "vendor_name":    getattr(result, "vendor_name", "") or "",
            "currency":       getattr(result, "currency", "") or "",
            "discrepancies":  list(getattr(result, "discrepancies", []) or [])[:10],
            "explanation":    getattr(result, "explanation", "") or "",
            "explanation_ar": getattr(result, "explanation_ar", "") or "",
        }
