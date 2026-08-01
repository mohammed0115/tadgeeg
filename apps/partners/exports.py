"""Application exports (§F). Staff-only, and filter-respecting by construction.

Both renderers receive the already-filtered queryset from the caller. An export
that re-derives its own queryset is how "export ignores the active filters and
dumps every row" happens — a data-exposure bug, as established in Phase 1.

Neither export includes ``submitted_ip`` or internal notes.
"""

from __future__ import annotations

import io

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from .selectors import application_row

COLUMNS = (
    ("Company", "company_name"),
    ("Contact", "contact_name"),
    ("Email", "email"),
    ("Mobile", "mobile"),
    ("Country", "country"),
    ("City", "city"),
    ("Website", "website"),
    ("Requested type", "requested_partner_type"),
    ("Status", "status"),
    ("Attachments", "attachment_count"),
    ("Submitted", "created_at"),
    ("Reviewed", "reviewed_at"),
)

MAX_EXPORT_ROWS = 5000


def _rows(queryset):
    rows, truncated = [], False
    for index, application in enumerate(queryset.iterator(chunk_size=200)):
        if index >= MAX_EXPORT_ROWS:
            truncated = True
            break
        row = application_row(application)
        row["business_areas"] = ", ".join(row.get("business_areas") or [])
        rows.append(row)
    return rows, truncated


def _filename(extension: str) -> str:
    return f"partner-applications-{timezone.now():%Y%m%d-%H%M}.{extension}"


def export_applications_xlsx(queryset) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    rows, truncated = _rows(queryset)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Applications"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="003366", end_color="003366")

    for column_index, (label, _key) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(column_index)
        ].width = 22

    for row_index, row in enumerate(rows, start=2):
        for column_index, (_label, key) in enumerate(COLUMNS, start=1):
            sheet.cell(row=row_index, column=column_index, value=row.get(key, ""))

    sheet.freeze_panes = "A2"
    if truncated:
        sheet.cell(row=len(rows) + 3, column=1,
                   value=f"Truncated at {MAX_EXPORT_ROWS} rows. Narrow the filters.")

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_filename("xlsx")}"'
    response["X-Export-Truncated"] = "1" if truncated else "0"
    return response


def export_applications_pdf(queryset):
    """Returns None when the renderer is unavailable so the caller can 503."""
    try:
        from weasyprint import HTML
    except Exception:                                   # noqa: BLE001
        return None

    rows, truncated = _rows(queryset)
    html = render_to_string(
        "platform_admin/partner_applications_export.html",
        {
            "headers": [label for label, _key in COLUMNS],
            "table_rows": [[row.get(key, "") for _label, key in COLUMNS] for row in rows],
            "row_count": len(rows),
            "truncated": truncated,
            "max_rows": MAX_EXPORT_ROWS,
            "generated_at": timezone.now(),
        },
    )
    try:
        pdf_bytes = HTML(string=html).write_pdf()
    except Exception:                                   # noqa: BLE001
        return None

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_filename("pdf")}"'
    response["X-Export-Truncated"] = "1" if truncated else "0"
    return response
