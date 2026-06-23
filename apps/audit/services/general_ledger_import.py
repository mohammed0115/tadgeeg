"""General Ledger import parser + validation service (TADGEEG-FIN-AUDIT-2A).

Reads an uploaded CSV/XLSX general ledger, maps client headers (Arabic/English)
to canonical columns, normalises amounts and dates, validates row- and
batch-level, groups by journal to detect unbalanced journals, links each row to
an existing ``AccountMapping`` (classification only), optionally checks
alignment against a linked ``TrialBalanceImport``, and persists
``GeneralLedgerRow`` staging rows + a summary.

It NEVER writes to ``apps.ledger`` and NEVER posts journal entries.
"""
from __future__ import annotations

import csv
import datetime
import hashlib
import io
from decimal import Decimal, InvalidOperation
from typing import IO

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from apps.audit.general_ledger_models import GeneralLedgerImport, GeneralLedgerRow
from apps.audit.trial_balance_models import AccountMapping

# A GL is a flat list of lines; "balanced" means total debits == total credits
# across the batch (and per journal when a journal key is present).
BALANCE_TOLERANCE = Decimal("0.01")
_ZERO = Decimal("0")

SUPPORTED_FORMATS = {"csv", "xlsx", "xls"}

HEADER_ALIASES: dict[str, list[str]] = {
    "account_code": ["account code", "account_code", "account no", "account number",
                     "gl code", "code", "كود الحساب", "رقم الحساب", "الحساب"],
    "account_name": ["account name", "account_name", "account title", "name",
                     "اسم الحساب"],
    "debit":  ["debit", "dr", "مدين"],
    "credit": ["credit", "cr", "دائن"],
    "amount": ["amount", "signed amount", "net amount", "value", "المبلغ", "القيمة"],
    "transaction_date": ["transaction date", "trans date", "date", "doc date",
                         "التاريخ", "تاريخ المعاملة", "تاريخ"],
    "posting_date": ["posting date", "post date", "gl date", "تاريخ الترحيل", "تاريخ القيد"],
    "journal_number": ["journal number", "journal no", "journal", "entry number",
                       "entry reference", "voucher", "voucher no", "je number",
                       "رقم القيد", "رقم اليومية", "رقم السند", "القيد"],
    "line_number": ["line number", "line no", "line", "seq", "رقم السطر"],
    "description": ["description", "narration", "memo", "details", "البيان", "الوصف", "بيان"],
    "document_number": ["document number", "document no", "doc number", "invoice number",
                        "invoice no", "رقم المستند", "رقم الفاتورة"],
    "reference": ["reference", "ref", "ref no", "المرجع", "مرجع"],
    "counterparty": ["counterparty", "customer", "supplier", "vendor", "party",
                     "الطرف المقابل", "العميل", "المورد"],
    "cost_center": ["cost center", "cost centre", "costcenter", "مركز التكلفة"],
    "department": ["department", "dept", "القسم", "الإدارة"],
    "entered_by": ["entered by", "user", "user name", "username", "created by",
                   "posted by", "المستخدم", "أدخله", "المستخدم المدخل"],
    "source_system": ["source system", "source", "system", "module", "النظام", "المصدر"],
    "currency": ["currency", "ccy", "العملة"],
}

_DATE_FORMATS = (
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d",
    "%d.%m.%Y", "%d %b %Y", "%d %B %Y", "%Y-%m-%d %H:%M:%S",
)


class GeneralLedgerImportError(Exception):
    """Raised for unrecoverable problems (bad format, unreadable file)."""


def _norm_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _build_column_mapping(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for original in headers:
        key = _norm_header(original)
        for canonical, aliases in HEADER_ALIASES.items():
            if canonical in mapping:
                continue
            if key == canonical or key in aliases:
                mapping[canonical] = original
                break
    return mapping


def _normalize_amount(value) -> Decimal | None:
    if value is None:
        return _ZERO
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip()
    if text == "":
        return _ZERO
    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]
    for ch in (",", " ", " ", "ر.س", "SAR", "$", "ريال"):
        text = text.replace(ch, "")
    if text in ("", "-"):
        return _ZERO
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def _normalize_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # Trim a trailing time component for "YYYY-MM-DD HH:MM:SS"-style values.
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return "INVALID"  # sentinel: present but unparseable


def _read_csv(file_obj: IO[bytes]) -> list[dict]:
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(raw))
    return [dict(r) for r in reader]


def _read_xlsx(file_obj: IO[bytes]) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise GeneralLedgerImportError("openpyxl is required to read .xlsx files") from exc
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for values in rows[1:]:
        out.append({headers[i]: values[i] if i < len(values) else None
                    for i in range(len(headers))})
    return out


def parse_and_validate(import_batch: GeneralLedgerImport, file_obj: IO[bytes] | None = None):
    """Parse + validate the GL file, persist rows, update the import summary.

    Returns the refreshed ``import_batch``. Raises ``GeneralLedgerImportError``
    for unrecoverable issues and ``ValidationError`` for cross-tenant imports.
    """
    if import_batch.engagement.organization_id != import_batch.organization_id:
        raise ValidationError(
            "cross-tenant import denied: organization != engagement.organization"
        )
    if import_batch.related_trial_balance_import_id:
        tb = import_batch.related_trial_balance_import
        if (tb.engagement_id != import_batch.engagement_id
                or tb.organization_id != import_batch.organization_id):
            raise ValidationError(
                "related_trial_balance_import must belong to the same engagement/org."
            )

    fmt = (import_batch.source_format or "").lower()
    if fmt not in SUPPORTED_FORMATS:
        import_batch.status = GeneralLedgerImport.Status.VALIDATION_FAILED
        import_batch.errors = [f"unsupported format: {import_batch.source_format!r}"]
        import_batch.save(update_fields=["status", "errors", "updated_at"])
        raise GeneralLedgerImportError(f"unsupported format: {import_batch.source_format!r}")

    if file_obj is None:
        if not import_batch.uploaded_file:
            raise GeneralLedgerImportError("no file attached to import")
        file_obj = import_batch.uploaded_file.open("rb")

    data = file_obj.read()
    if isinstance(data, str):
        data = data.encode("utf-8")
    import_batch.file_sha256 = hashlib.sha256(data).hexdigest()
    import_batch.status = GeneralLedgerImport.Status.VALIDATING
    import_batch.save(update_fields=["status", "file_sha256", "updated_at"])

    buf = io.BytesIO(data)
    try:
        records = _read_csv(buf) if fmt == "csv" else _read_xlsx(buf)
    except GeneralLedgerImportError:
        raise
    except Exception as exc:
        import_batch.status = GeneralLedgerImport.Status.VALIDATION_FAILED
        import_batch.errors = [f"could not parse file: {exc}"]
        import_batch.save(update_fields=["status", "errors", "updated_at"])
        raise GeneralLedgerImportError(str(exc)) from exc

    headers = list(records[0].keys()) if records else []
    mapping = _build_column_mapping(headers)

    import_errors: list[str] = []
    import_warnings: list[str] = []
    if "account_code" not in mapping:
        import_errors.append("required column 'account_code' not found")
    has_amount = ("debit" in mapping or "credit" in mapping or "amount" in mapping)
    if not has_amount:
        import_errors.append("no debit/credit or amount column found")
    if "transaction_date" not in mapping and "posting_date" not in mapping:
        import_warnings.append("no transaction_date or posting_date column detected")

    if (import_batch.period_start and import_batch.period_end
            and import_batch.period_start > import_batch.period_end):
        import_errors.append("period_start is after period_end")

    # Pre-load engagement account mappings for classification linking.
    mappings_by_code = {
        m.account_code: m for m in
        AccountMapping.objects.filter(engagement=import_batch.engagement)
    }

    def col(rec, key):
        c = mapping.get(key)
        return rec.get(c) if c else None

    total_debit = _ZERO
    total_credit = _ZERO
    valid_count = 0
    invalid_count = 0
    row_number = 0
    new_rows: list[GeneralLedgerRow] = []
    # journal_key -> [debit_total, credit_total]
    journals: dict[str, list[Decimal]] = {}
    unmapped_codes: set[str] = set()
    gl_codes: set[str] = set()

    for rec in records:
        if not any(str(v).strip() for v in rec.values() if v is not None):
            continue
        row_number += 1
        row_errors: list[str] = []

        account_code = str(col(rec, "account_code") or "").strip()
        if not account_code:
            row_errors.append("missing account_code")

        # Amounts: explicit debit/credit, else signed amount.
        debit = credit = _ZERO
        if "debit" in mapping or "credit" in mapping:
            d = _normalize_amount(col(rec, "debit")) if "debit" in mapping else _ZERO
            c = _normalize_amount(col(rec, "credit")) if "credit" in mapping else _ZERO
            if d is None:
                row_errors.append("non-numeric debit"); d = _ZERO
            if c is None:
                row_errors.append("non-numeric credit"); c = _ZERO
            debit, credit = d, c
        elif "amount" in mapping:
            amt = _normalize_amount(col(rec, "amount"))
            if amt is None:
                row_errors.append("non-numeric amount"); amt = _ZERO
            if amt >= 0:
                debit = amt
            else:
                credit = -amt
        signed_amount = debit - credit

        tx_date = _normalize_date(col(rec, "transaction_date")) if "transaction_date" in mapping else None
        post_date = _normalize_date(col(rec, "posting_date")) if "posting_date" in mapping else None
        if tx_date == "INVALID":
            row_errors.append("unparseable transaction_date"); tx_date = None
        if post_date == "INVALID":
            row_errors.append("unparseable posting_date"); post_date = None

        journal_number = str(col(rec, "journal_number") or "").strip()
        is_valid = not row_errors

        if is_valid:
            valid_count += 1
            total_debit += debit
            total_credit += credit
            if journal_number:
                bucket = journals.setdefault(journal_number, [_ZERO, _ZERO])
                bucket[0] += debit
                bucket[1] += credit
            if account_code:
                gl_codes.add(account_code)
        else:
            invalid_count += 1

        mapped = mappings_by_code.get(account_code) if account_code else None
        if account_code and mapped is None:
            unmapped_codes.add(account_code)

        new_rows.append(GeneralLedgerRow(
            import_batch=import_batch,
            engagement=import_batch.engagement,
            organization=import_batch.organization,
            row_number=row_number,
            journal_number=journal_number[:64],
            line_number=str(col(rec, "line_number") or "").strip()[:32],
            transaction_date=tx_date,
            posting_date=post_date,
            account_code=account_code[:64],
            account_name=str(col(rec, "account_name") or "").strip()[:255],
            mapped_account=mapped,
            debit=debit, credit=credit, signed_amount=signed_amount,
            currency=(str(col(rec, "currency") or "").strip()[:8] or import_batch.currency),
            description=str(col(rec, "description") or "").strip(),
            document_number=str(col(rec, "document_number") or "").strip()[:128],
            reference=str(col(rec, "reference") or "").strip()[:128],
            counterparty=str(col(rec, "counterparty") or "").strip()[:255],
            cost_center=str(col(rec, "cost_center") or "").strip()[:64],
            department=str(col(rec, "department") or "").strip()[:64],
            entered_by=str(col(rec, "entered_by") or "").strip()[:128],
            source_system=str(col(rec, "source_system") or "").strip()[:64],
            is_valid=is_valid,
            validation_errors=row_errors,
            raw_data={str(k): (str(v) if v is not None else None) for k, v in rec.items()},
        ))

    difference = total_debit - total_credit
    is_balanced = abs(difference) <= BALANCE_TOLERANCE

    balanced_journals = sum(
        1 for d, c in journals.values() if abs(d - c) <= BALANCE_TOLERANCE
    )
    unbalanced_journals = len(journals) - balanced_journals
    if unbalanced_journals:
        import_warnings.append(f"{unbalanced_journals} unbalanced journal(s) detected")
    if not is_balanced:
        import_warnings.append("GL total debits != total credits")

    # ── Trial Balance alignment (warnings, not hard failures) ───────────
    tb_alignment: dict = {}
    if import_batch.related_trial_balance_import_id:
        tb_codes = set(
            import_batch.related_trial_balance_import.rows
            .exclude(account_code="").values_list("account_code", flat=True)
        )
        gl_not_in_tb = sorted(gl_codes - tb_codes)
        tb_not_in_gl = sorted(tb_codes - gl_codes)
        if gl_not_in_tb:
            import_warnings.append(f"{len(gl_not_in_tb)} GL account(s) missing from the trial balance")
        if tb_not_in_gl:
            import_warnings.append(f"{len(tb_not_in_gl)} trial-balance account(s) with no GL activity")
        tb_alignment = {
            "gl_accounts_missing_from_tb": gl_not_in_tb[:200],
            "tb_accounts_without_gl_activity": tb_not_in_gl[:200],
            "gl_accounts_missing_from_tb_count": len(gl_not_in_tb),
            "tb_accounts_without_gl_activity_count": len(tb_not_in_gl),
        }
        # Period/fiscal-year compatibility (warning).
        tb = import_batch.related_trial_balance_import
        if (tb.fiscal_year and import_batch.fiscal_year
                and tb.fiscal_year != import_batch.fiscal_year):
            import_warnings.append("GL fiscal_year differs from the linked trial balance")

    import_batch.row_count = row_number
    import_batch.valid_row_count = valid_count
    import_batch.invalid_row_count = invalid_count
    import_batch.journal_count = len(journals)
    import_batch.balanced_journal_count = balanced_journals
    import_batch.unbalanced_journal_count = unbalanced_journals
    import_batch.total_debit = total_debit
    import_batch.total_credit = total_credit
    import_batch.difference = difference
    import_batch.is_balanced = is_balanced
    import_batch.column_mapping = mapping
    import_batch.errors = import_errors
    import_batch.warnings = import_warnings
    import_batch.validation_summary = {
        "row_count": row_number,
        "valid_row_count": valid_count,
        "invalid_row_count": invalid_count,
        "journal_count": len(journals),
        "balanced_journal_count": balanced_journals,
        "unbalanced_journal_count": unbalanced_journals,
        "total_debit": str(total_debit),
        "total_credit": str(total_credit),
        "difference": str(difference),
        "is_balanced": is_balanced,
        "balance_tolerance": str(BALANCE_TOLERANCE),
        "columns_detected": list(mapping.keys()),
        "unmapped_account_count": len(unmapped_codes),
        "distinct_gl_accounts": len(gl_codes),
        "trial_balance_alignment": tb_alignment,
    }

    # Hard errors or invalid rows → failed; else validated (journal imbalance
    # and TB-alignment gaps are warnings, not hard failures).
    if import_errors or invalid_count:
        import_batch.status = GeneralLedgerImport.Status.VALIDATION_FAILED
    else:
        import_batch.status = GeneralLedgerImport.Status.VALIDATED
        import_batch.validated_at = timezone.now()

    with transaction.atomic():
        import_batch.rows.all().delete()  # idempotent re-validation
        GeneralLedgerRow.objects.bulk_create(new_rows)
        import_batch.save()
    return import_batch
